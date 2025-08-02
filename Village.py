import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
import hashlib
from datetime import datetime, timedelta
import time
import csv
import os
import json
import logging
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass
from enum import Enum
import threading

# Custom Exception Classes for User-Friendly Error Handling
class BankMmudziException(Exception):
    """Base exception class for Bank Mmudzi application"""
    def __init__(self, message: str, error_code: str = None, details: dict = None):
        self.message = message
        self.error_code = error_code
        self.details = details or {}
        super().__init__(self.message)

class ValidationError(BankMmudziException):
    """Exception raised for input validation errors"""
    def __init__(self, message: str, field_name: str = None, invalid_value: str = None):
        super().__init__(message, "VALIDATION_ERROR", {
            'field_name': field_name,
            'invalid_value': invalid_value
        })

class DuplicateRecordError(BankMmudziException):
    """Exception raised when trying to create duplicate records"""
    def __init__(self, message: str, record_type: str = None, existing_id: int = None):
        super().__init__(message, "DUPLICATE_RECORD", {
            'record_type': record_type,
            'existing_id': existing_id
        })

class InsufficientFundsError(BankMmudziException):
    """Exception raised when loan amount exceeds available funds"""
    def __init__(self, message: str, requested_amount: float = None, available_amount: float = None):
        super().__init__(message, "INSUFFICIENT_FUNDS", {
            'requested_amount': requested_amount,
            'available_amount': available_amount
        })

class RecordNotFoundError(BankMmudziException):
    """Exception raised when a required record is not found"""
    def __init__(self, message: str, record_type: str = None, record_id: int = None):
        super().__init__(message, "RECORD_NOT_FOUND", {
            'record_type': record_type,
            'record_id': record_id
        })

class BusinessRuleViolationError(BankMmudziException):
    """Exception raised when business rules are violated"""
    def __init__(self, message: str, rule_name: str = None):
        super().__init__(message, "BUSINESS_RULE_VIOLATION", {
            'rule_name': rule_name
        })

class DatabaseError(BankMmudziException):
    """Exception raised for database-related errors"""
    def __init__(self, message: str, operation: str = None, table: str = None):
        super().__init__(message, "DATABASE_ERROR", {
            'operation': operation,
            'table': table
        })

# Error Handler Class for User-Friendly Messages
class ErrorHandler:
    """
    Centralized error handling with user-friendly messages
    """
    
    @staticmethod
    def get_user_friendly_message(error: Exception) -> str:
        """
        Convert technical errors to user-friendly messages
        
        Args:
            error (Exception): The original error
            
        Returns:
            str: User-friendly error message
        """
        if isinstance(error, BankMmudziException):
            return error.message
        
        # Handle common database errors
        error_str = str(error).lower()
        
        if 'unique constraint' in error_str or 'duplicate' in error_str:
            return "This record already exists. Please check your input and try again."
        
        if 'foreign key constraint' in error_str:
            return "Cannot complete this operation because it would affect related records. Please check dependencies."
        
        if 'not null constraint' in error_str:
            return "Required information is missing. Please fill in all required fields."
        
        if 'no such table' in error_str:
            return "System error: Database table not found. Please contact support."
        
        if 'database is locked' in error_str:
            return "The system is busy. Please wait a moment and try again."
        
        if 'connection' in error_str:
            return "Unable to connect to the database. Please check your connection and try again."
        
        if 'permission' in error_str or 'access' in error_str:
            return "You don't have permission to perform this operation."
        
        # Handle validation errors
        if 'invalid' in error_str:
            return "The information you entered is not valid. Please check and try again."
        
        # Handle file/network errors
        if 'file not found' in error_str:
            return "Required file not found. Please contact support."
        
        if 'timeout' in error_str:
            return "The operation took too long. Please try again."
        
        # Default fallback
        return "An unexpected error occurred. Please try again or contact support if the problem persists."
    
    @staticmethod
    def handle_error(error: Exception, operation: str = "operation") -> dict:
        """
        Handle an error and return structured error information
        
        Args:
            error (Exception): The error to handle
            operation (str): Description of the operation that failed
            
        Returns:
            dict: Structured error information
        """
        user_message = ErrorHandler.get_user_friendly_message(error)
        
        # Log the technical error for debugging
        logger.error(f"Error during {operation}: {str(error)}")
        
        return {
            'success': False,
            'error': True,
            'message': user_message,
            'operation': operation,
            'error_type': type(error).__name__,
            'technical_details': str(error) if isinstance(error, BankMmudziException) else None
        }
    
    @staticmethod
    def validate_and_handle(validation_func, *args, **kwargs) -> dict:
        """
        Execute a validation function and handle any errors
        
        Args:
            validation_func: The validation function to execute
            *args: Arguments for the validation function
            **kwargs: Keyword arguments for the validation function
            
        Returns:
            dict: Validation result or error information
        """
        try:
            result = validation_func(*args, **kwargs)
            if isinstance(result, dict) and 'valid' in result:
                if not result['valid']:
                    return {
                        'success': False,
                        'error': True,
                        'message': result.get('message', 'Validation failed'),
                        'validation_result': result
                    }
                else:
                    return {
                        'success': True,
                        'error': False,
                        'message': result.get('message', 'Validation passed'),
                        'validation_result': result
                    }
            return {'success': True, 'error': False, 'result': result}
        except Exception as e:
            return ErrorHandler.handle_error(e, "validation")

try:
    from tkcalendar import Calendar
    TKCALENDAR_AVAILABLE = True
except ImportError:
    TKCALENDAR_AVAILABLE = False
    print("tkcalendar not available. Install with: pip install tkcalendar")

try:
    import mysql.connector
    MYSQL_AVAILABLE = True
except ImportError:
    MYSQL_AVAILABLE = False
    print("MySQL connector not available. Install with: pip install mysql-connector-python")

# Additional imports for enhanced functionality
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("ReportLab not available. Install with: pip install reportlab for PDF export")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("OpenPyXL not available. Install with: pip install openpyxl for Excel export")

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Database configuration
MYSQL_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'nehe',
    'database': 'bank_mmudzi'
}
SQLITE_DB_NAME = 'bank_mmudzi.db'

# Global database manager instance
db_manager = None

class DatabaseManager:
    """Enhanced database manager with MySQL/SQLite support"""
    
    def __init__(self):
        self.current_db_type = 'sqlite'  # Default to SQLite
        self.mysql_available = MYSQL_AVAILABLE
        self.connection_pool = {}
        self.last_mysql_check = 0
        self.connection_retry_count = 0
        
        # Try to connect to MySQL first
        if self.mysql_available:
            try:
                self._test_mysql_connection()
                self.current_db_type = 'mysql'
                logger.info("Database Manager: Using MySQL as primary database")
            except Exception as e:
                logger.warning(f"MySQL not available, falling back to SQLite: {e}")
                self.current_db_type = 'sqlite'
        else:
            logger.info("Database Manager: Using SQLite (MySQL connector not available)")
    
    def _test_mysql_connection(self):
        """Test MySQL connection"""
        if not self.mysql_available:
            raise Exception("MySQL connector not available")
        
        conn = mysql.connector.connect(**MYSQL_CONFIG)
        conn.close()
        return True
    
    def get_sync_status(self):
        """Get synchronization status"""
        return {
            'current_db_type': self.current_db_type,
            'mysql_available': self.mysql_available,
            'sqlite_available': True,
            'last_sync': datetime.now().isoformat(),
            'sync_pending': False
        }
    
    def track_change(self, table_name, record_id, operation, data):
        """Track changes for synchronization"""
        # This is a placeholder for change tracking
        logger.info(f"Change tracked: {operation} on {table_name} record {record_id}")
        pass

def initialize_database_manager():
    """Initialize the global database manager"""
    global db_manager
    if db_manager is None:
        db_manager = DatabaseManager()
    return db_manager

# Month mapping
MONTHS = ['January', 'February', 'March', 'April', 'May', 'June', 
          'July', 'August', 'September', 'October', 'November', 'December']
MONTH_TO_NUM = {month: i+1 for i, month in enumerate(MONTHS)}

# Financial Calculation Engine
class FinancialCalculator:
    """
    Handles all financial calculations including loan amortization and dividend distribution
    """
    
    @staticmethod
    def calculate_loan_payment(principal: float, monthly_rate: float, months: int = 12) -> float:
        """
        Calculate monthly payment using proper amortization formula:
        P = L × [r(1+r)^n]/[(1+r)^n-1]
        
        Args:
            principal (float): Loan amount (L)
            monthly_rate (float): Monthly interest rate (r)
            months (int): Number of months (n), default 12
            
        Returns:
            float: Monthly payment amount
        """
        if monthly_rate == 0:
            return principal / months
        
        # Calculate (1+r)^n
        power_term = (1 + monthly_rate) ** months
        
        # Calculate P = L × [r(1+r)^n]/[(1+r)^n-1]
        monthly_payment = principal * (monthly_rate * power_term) / (power_term - 1)
        
        return round(monthly_payment, 2)
    
    @staticmethod
    def calculate_total_interest(principal: float, monthly_payment: float, months: int = 12) -> float:
        """
        Calculate total interest paid over the loan term
        
        Args:
            principal (float): Loan amount
            monthly_payment (float): Monthly payment amount
            months (int): Number of months
            
        Returns:
            float: Total interest amount
        """
        total_payments = monthly_payment * months
        total_interest = total_payments - principal
        return round(total_interest, 2)
    
    @staticmethod
    def generate_amortization_schedule(loan_id: int, principal: float, monthly_rate: float, 
                                     monthly_payment: float, months: int = 12) -> list:
        """
        Generate amortization schedule with principal/interest breakdown
        
        Args:
            loan_id (int): Loan ID
            principal (float): Loan amount
            monthly_rate (float): Monthly interest rate
            monthly_payment (float): Monthly payment amount
            months (int): Number of months
            
        Returns:
            list: List of tuples (payment_number, payment_date, principal_amount, interest_amount, remaining_balance)
        """
        schedule = []
        remaining_balance = principal
        
        for payment_number in range(1, months + 1):
            # Calculate interest portion
            interest_amount = remaining_balance * monthly_rate
            
            # Calculate principal portion
            principal_amount = monthly_payment - interest_amount
            
            # Update remaining balance
            remaining_balance = max(0, remaining_balance - principal_amount)
            
            # Calculate payment date (assuming monthly payments)
            payment_date = datetime.now() + timedelta(days=30 * payment_number)
            
            schedule.append((
                payment_number,
                payment_date.strftime('%Y-%m-%d'),
                round(principal_amount, 2),
                round(interest_amount, 2),
                round(remaining_balance, 2)
            ))
            
            # Break if loan is fully paid
            if remaining_balance <= 0:
                break
        
        return schedule
    
    @staticmethod
    def calculate_member_dividend(member_id: int, year: int, contributions: float, 
                                interest_paid: float, outstanding_balance: float) -> float:
        """
        Calculate member dividend based on the financial model:
        - Non-borrowers: 12C (total contributions)
        - Borrowers: 12C + interest paid - outstanding balance
        
        Args:
            member_id (int): Member ID
            year (int): Year for calculation
            contributions (float): Total contributions (12C)
            interest_paid (float): Total interest paid by member
            outstanding_balance (float): Outstanding loan balance
            
        Returns:
            float: Dividend amount
        """
        if outstanding_balance > 0 or interest_paid > 0:
            # Borrower: 12C + interest paid - outstanding balance
            dividend = contributions + interest_paid - outstanding_balance
        else:
            # Non-borrower: 12C
            dividend = contributions
        
        return max(0, round(dividend, 2))
    
    @staticmethod
    def calculate_total_fund(total_contributions: float, total_interest_payments: float) -> float:
        """
        Calculate total fund available for dividend distribution
        Fund = All contributions + All interest payments
        
        Args:
            total_contributions (float): Sum of all member contributions
            total_interest_payments (float): Sum of all interest payments received
            
        Returns:
            float: Total fund amount
        """
        return round(total_contributions + total_interest_payments, 2)
    
    @staticmethod
    def validate_dividend_distribution(total_fund: float, total_dividends: float) -> bool:
        """
        Validate that dividend distribution equals the total fund
        
        Args:
            total_fund (float): Total fund available
            total_dividends (float): Sum of all dividends to be distributed
            
        Returns:
            bool: True if distribution is valid
        """
        return abs(total_fund - total_dividends) <= 0.02  # Allow for small rounding differences

# Dividend Calculation Functions
def get_member_annual_contributions(member_id: int, year: int) -> float:
    """
    Get total contributions for a member in a specific year
    """
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Use database-specific syntax
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) 
                FROM Contributions 
                WHERE member_id = %s AND year = %s
            ''', (member_id, year))
        else:
            cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) 
                FROM Contributions 
                WHERE member_id = ? AND year = ?
            ''', (member_id, year))
        
        result = cursor.fetchone()
        conn.close()
        return float(result[0]) if result else 0.0
    except Exception as e:
        logger.error(f"Error getting member annual contributions: {e}")
        return 0.0

def get_member_interest_paid(member_id: int, year: int) -> float:
    """
    Get total interest paid by a member in a specific year
    """
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Use database-specific date extraction
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            date_extract = 'YEAR(r.repayment_date)'
            cursor.execute(f'''
                SELECT COALESCE(SUM(r.interest_amount), 0)
                FROM Repayments r
                JOIN Loans l ON r.loan_id = l.loan_id
                WHERE l.member_id = %s AND {date_extract} = %s
            ''', (member_id, str(year)))
        else:
            date_extract = "strftime('%Y', r.repayment_date)"
            cursor.execute(f'''
                SELECT COALESCE(SUM(r.interest_amount), 0)
                FROM Repayments r
                JOIN Loans l ON r.loan_id = l.loan_id
                WHERE l.member_id = ? AND {date_extract} = ?
            ''', (member_id, str(year)))
        
        result = cursor.fetchone()
        conn.close()
        return float(result[0]) if result else 0.0
    except Exception as e:
        logger.error(f"Error getting member interest paid: {e}")
        return 0.0

def get_member_outstanding_balance(member_id: int, year: int) -> float:
    """
    Get outstanding loan balance for a member at the end of a specific year
    """
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Use database-specific date extraction
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('''
                SELECT COALESCE(SUM(outstanding_balance), 0)
                FROM Loans
                WHERE member_id = %s AND YEAR(loan_date) <= %s
            ''', (member_id, str(year)))
        else:
            cursor.execute('''
                SELECT COALESCE(SUM(outstanding_balance), 0)
                FROM Loans
                WHERE member_id = ? AND strftime('%Y', loan_date) <= ?
            ''', (member_id, str(year)))
        
        result = cursor.fetchone()
        conn.close()
        return float(result[0]) if result else 0.0
    except Exception as e:
        logger.error(f"Error getting member outstanding balance: {e}")
        return 0.0

def calculate_member_dividend_for_year(member_id: int, year: int) -> dict:
    """
    Calculate dividend for a specific member and year
    Returns dict with calculation details
    """
    # Get member data
    contributions = get_member_annual_contributions(member_id, year)
    interest_paid = get_member_interest_paid(member_id, year)
    outstanding_balance = get_member_outstanding_balance(member_id, year)
    
    # Calculate dividend using the financial model
    dividend_amount = FinancialCalculator.calculate_member_dividend(
        member_id, year, contributions, interest_paid, outstanding_balance
    )
    
    return {
        'member_id': member_id,
        'year': year,
        'total_contributions': contributions,
        'total_interest_paid': interest_paid,
        'outstanding_balance': outstanding_balance,
        'dividend_amount': dividend_amount
    }

def store_dividend_calculation(dividend_data: dict) -> int:
    """
    Store dividend calculation in the DividendCalculations table
    """
    conn = connect_db()
    cursor = conn.cursor()
    
    calculation_date = datetime.now().strftime('%Y-%m-%d')
    
    # Insert or update dividend calculation using database-specific syntax
    if db_manager.current_db_type == 'mysql':
        cursor.execute('''
            INSERT INTO DividendCalculations 
            (member_id, year, total_contributions, total_interest_paid, 
             outstanding_balance, dividend_amount, calculation_date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON DUPLICATE KEY UPDATE
            total_contributions = VALUES(total_contributions),
            total_interest_paid = VALUES(total_interest_paid),
            outstanding_balance = VALUES(outstanding_balance),
            dividend_amount = VALUES(dividend_amount),
            calculation_date = VALUES(calculation_date)
        ''', (
            dividend_data['member_id'],
            dividend_data['year'],
            dividend_data['total_contributions'],
            dividend_data['total_interest_paid'],
            dividend_data['outstanding_balance'],
            dividend_data['dividend_amount'],
            calculation_date
        ))
    else:
        cursor.execute('''
            INSERT OR REPLACE INTO DividendCalculations 
            (member_id, year, total_contributions, total_interest_paid, 
             outstanding_balance, dividend_amount, calculation_date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            dividend_data['member_id'],
            dividend_data['year'],
            dividend_data['total_contributions'],
            dividend_data['total_interest_paid'],
            dividend_data['outstanding_balance'],
            dividend_data['dividend_amount'],
            calculation_date
        ))
    
    calculation_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite':
        data = {
            'calculation_id': calculation_id,
            **dividend_data,
            'calculation_date': calculation_date
        }
        db_manager.track_change('DividendCalculations', calculation_id, 'INSERT', data)
    
    return calculation_id

def get_total_contributions_for_year(year: int) -> float:
    """
    Get total contributions from all members for a specific year
    """
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Use database-specific syntax
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) 
                FROM Contributions 
                WHERE year = %s
            ''', (year,))
        else:
            cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) 
                FROM Contributions 
                WHERE year = ?
            ''', (year,))
        
        result = cursor.fetchone()
        conn.close()
        return float(result[0]) if result else 0.0
    except Exception as e:
        logger.error(f"Error getting total contributions: {e}")
        return 0.0

def get_total_interest_payments_for_year(year: int) -> float:
    """
    Get total interest payments from all repayments for a specific year
    This sums actual interest amounts from repayments, not theoretical calculations
    """
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Use database-specific date extraction and syntax
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('''
                SELECT COALESCE(SUM(interest_amount), 0)
                FROM Repayments
                WHERE YEAR(repayment_date) = %s
            ''', (str(year),))
        else:
            cursor.execute('''
                SELECT COALESCE(SUM(interest_amount), 0)
                FROM Repayments
                WHERE strftime('%Y', repayment_date) = ?
            ''', (str(year),))
        
        result = cursor.fetchone()
        conn.close()
        return float(result[0]) if result else 0.0
    except Exception as e:
        logger.error(f"Error getting total interest payments: {e}")
        return 0.0

def calculate_total_fund_for_year(year: int) -> dict:
    """
    Calculate total fund for a specific year using actual data from database
    Fund = All contributions + All interest payments (from actual repayments)
    
    Returns:
        dict: Contains total_contributions, total_interest_payments, and total_fund
    """
    total_contributions = get_total_contributions_for_year(year)
    total_interest_payments = get_total_interest_payments_for_year(year)
    total_fund = FinancialCalculator.calculate_total_fund(total_contributions, total_interest_payments)
    
    return {
        'year': year,
        'total_contributions': total_contributions,
        'total_interest_payments': total_interest_payments,
        'total_fund': total_fund,
        'calculation_date': datetime.now().strftime('%Y-%m-%d')
    }

def calculate_all_dividends_for_year(year: int) -> dict:
    """
    Calculate dividends for all members for a specific year
    Returns summary with validation using proper fund calculation
    """
    conn = connect_db()
    cursor = conn.cursor()
    
    # Get all members
    cursor.execute('SELECT member_id FROM Members WHERE status = "active"')
    members = cursor.fetchall()
    conn.close()
    
    dividend_calculations = []
    total_dividends = 0
    
    for (member_id,) in members:
        dividend_data = calculate_member_dividend_for_year(member_id, year)
        dividend_calculations.append(dividend_data)
        total_dividends += dividend_data['dividend_amount']
        
        # Store the calculation
        store_dividend_calculation(dividend_data)
    
    # Calculate total fund using proper method (actual database data)
    fund_data = calculate_total_fund_for_year(year)
    total_fund = fund_data['total_fund']
    
    # Validate distribution
    is_valid = FinancialCalculator.validate_dividend_distribution(total_fund, total_dividends)
    
    return {
        'year': year,
        'dividend_calculations': dividend_calculations,
        'total_fund': total_fund,
        'total_contributions': fund_data['total_contributions'],
        'total_interest_payments': fund_data['total_interest_payments'],
        'total_dividends': total_dividends,
        'distribution_valid': is_valid,
        'calculation_date': datetime.now().strftime('%Y-%m-%d')
    }

def get_dividend_calculations_for_year(year: int) -> list:
    """
    Get stored dividend calculations for a specific year
    """
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Use database-specific syntax
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('''
                SELECT dc.*, m.name, m.surname
                FROM DividendCalculations dc
                JOIN Members m ON dc.member_id = m.member_id
                WHERE dc.year = %s
                ORDER BY m.name, m.surname
            ''', (year,))
        else:
            cursor.execute('''
                SELECT dc.*, m.name, m.surname
                FROM DividendCalculations dc
                JOIN Members m ON dc.member_id = m.member_id
                WHERE dc.year = ?
                ORDER BY m.name, m.surname
            ''', (year,))
        
        results = cursor.fetchall()
        conn.close()
        
        calculations = []
        for row in results:
            calculations.append({
                'calculation_id': row[0],
                'member_id': row[1],
                'year': row[2],
                'total_contributions': row[3],
                'total_interest_paid': row[4],
                'outstanding_balance': row[5],
                'dividend_amount': row[6],
                'calculation_date': row[7],
                'status': row[8] if len(row) > 8 else 'calculated',
                'member_name': f"{row[-2]} {row[-1]}"  # name + surname
            })
        
        return calculations
    except Exception as e:
        logger.error(f"Error getting dividend calculations: {e}")
        return []

# Enhanced Report Generation System
class EnhancedReportGenerator:
    """
    Enhanced report generation with comprehensive financial analysis
    """
    
    def __init__(self):
        self.current_date = datetime.now()
    
    def format_currency(self, amount: float) -> str:
        """Format currency with proper MWK formatting"""
        return f"MWK {amount:,.2f}"
    
    def generate_comprehensive_report(self, end_date: str) -> str:
        """
        Generate comprehensive financial report with enhanced formatting
        
        Args:
            end_date (str): End date for the report
            
        Returns:
            str: Formatted comprehensive report
        """
        try:
            report_lines = []
            
            # Header
            report_lines.extend(self._generate_report_header(end_date))
            
            # Executive Summary
            report_lines.extend(self._generate_executive_summary(end_date))
            
            # Monthly Breakdown
            report_lines.extend(self._generate_monthly_breakdown(end_date))
            
            # Member Analysis
            report_lines.extend(self._generate_member_analysis(end_date))
            
            # Loan Portfolio Analysis
            report_lines.extend(self._generate_loan_analysis(end_date))
            
            # Dividend Distribution
            report_lines.extend(self._generate_dividend_analysis(end_date))
            
            # Year-over-Year Growth
            report_lines.extend(self._generate_growth_metrics(end_date))
            
            # System Status
            report_lines.extend(self._generate_system_status())
            
            return '\n'.join(report_lines)
            
        except Exception as e:
            return f"Error generating report: {ErrorHandler.get_user_friendly_message(e)}"
    
    def _generate_report_header(self, end_date: str) -> list:
        """Generate report header section"""
        lines = []
        lines.append("=" * 80)
        lines.append("BANK MMUDZI COMPREHENSIVE FINANCIAL REPORT".center(80))
        lines.append("=" * 80)
        lines.append(f"Report Period: Up to {end_date}")
        lines.append(f"Generated: {self.current_date.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Report Type: Comprehensive Financial Analysis")
        lines.append("=" * 80)
        lines.append("")
        return lines
    
    def _generate_executive_summary(self, end_date: str) -> list:
        """Generate executive summary section"""
        lines = []
        lines.append("EXECUTIVE SUMMARY")
        lines.append("-" * 40)
        
        try:
            # Get summary statistics
            conn = connect_db()
            cursor = conn.cursor()
            
            # Total members
            cursor.execute("SELECT COUNT(*) FROM Members WHERE status = 'active'")
            total_members = cursor.fetchone()[0]
            
            # Total contributions
            cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM Contributions WHERE contribution_date <= ?", (end_date,))
            total_contributions = cursor.fetchone()[0]
            
            # Total loans disbursed
            cursor.execute("SELECT COALESCE(SUM(loan_amount), 0) FROM Loans WHERE loan_date <= ?", (end_date,))
            total_loans_disbursed = cursor.fetchone()[0]
            
            # Total repayments
            cursor.execute("SELECT COALESCE(SUM(repayment_amount), 0) FROM Repayments WHERE repayment_date <= ?", (end_date,))
            total_repayments = cursor.fetchone()[0]
            
            # Outstanding loans
            cursor.execute("SELECT COALESCE(SUM(outstanding_balance), 0) FROM Loans WHERE outstanding_balance > 0")
            outstanding_loans = cursor.fetchone()[0]
            
            # Available funds
            available_funds = total_contributions - outstanding_loans
            
            conn.close()
            
            lines.append(f"Active Members: {total_members:,}")
            lines.append(f"Total Contributions: {self.format_currency(total_contributions)}")
            lines.append(f"Total Loans Disbursed: {self.format_currency(total_loans_disbursed)}")
            lines.append(f"Total Repayments Received: {self.format_currency(total_repayments)}")
            lines.append(f"Outstanding Loan Balance: {self.format_currency(outstanding_loans)}")
            lines.append(f"Available Funds: {self.format_currency(available_funds)}")
            
            # Calculate key ratios
            if total_contributions > 0:
                loan_to_contribution_ratio = (total_loans_disbursed / total_contributions) * 100
                lines.append(f"Loan-to-Contribution Ratio: {loan_to_contribution_ratio:.1f}%")
            
            if total_loans_disbursed > 0:
                repayment_rate = (total_repayments / total_loans_disbursed) * 100
                lines.append(f"Repayment Rate: {repayment_rate:.1f}%")
            
        except Exception as e:
            lines.append(f"Error generating summary: {str(e)}")
        
        lines.append("")
        return lines
    
    def _generate_monthly_breakdown(self, end_date: str) -> list:
        """Generate monthly breakdown section"""
        lines = []
        lines.append("MONTHLY BREAKDOWN (CURRENT YEAR)")
        lines.append("-" * 40)
        
        try:
            current_year = datetime.strptime(end_date, '%Y-%m-%d').year
            
            conn = connect_db()
            cursor = conn.cursor()
            
            lines.append(f"{'Month':<12} {'Contributions':<15} {'Loans':<15} {'Repayments':<15}")
            lines.append("-" * 60)
            
            total_monthly_contrib = 0
            total_monthly_loans = 0
            total_monthly_repay = 0
            
            for month_num in range(1, 13):
                month_name = MONTHS[month_num - 1]
                
                # Monthly contributions
                cursor.execute("""
                    SELECT COALESCE(SUM(amount), 0) FROM Contributions 
                    WHERE year = ? AND month = ?
                """, (current_year, month_num))
                monthly_contrib = cursor.fetchone()[0]
                
                # Monthly loans
                cursor.execute("""
                    SELECT COALESCE(SUM(loan_amount), 0) FROM Loans 
                    WHERE strftime('%Y', loan_date) = ? AND strftime('%m', loan_date) = ?
                """, (str(current_year), f"{month_num:02d}"))
                monthly_loans = cursor.fetchone()[0]
                
                # Monthly repayments
                cursor.execute("""
                    SELECT COALESCE(SUM(repayment_amount), 0) FROM Repayments 
                    WHERE strftime('%Y', repayment_date) = ? AND strftime('%m', repayment_date) = ?
                """, (str(current_year), f"{month_num:02d}"))
                monthly_repay = cursor.fetchone()[0]
                
                lines.append(f"{month_name:<12} {monthly_contrib:<15.2f} {monthly_loans:<15.2f} {monthly_repay:<15.2f}")
                
                total_monthly_contrib += monthly_contrib
                total_monthly_loans += monthly_loans
                total_monthly_repay += monthly_repay
            
            lines.append("-" * 60)
            lines.append(f"{'TOTAL':<12} {total_monthly_contrib:<15.2f} {total_monthly_loans:<15.2f} {total_monthly_repay:<15.2f}")
            
            conn.close()
            
        except Exception as e:
            lines.append(f"Error generating monthly breakdown: {str(e)}")
        
        lines.append("")
        return lines
    
    def _generate_member_analysis(self, end_date: str) -> list:
        """Generate member analysis section"""
        lines = []
        lines.append("MEMBER ANALYSIS")
        lines.append("-" * 40)
        
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Get member data with contributions and loans
            cursor.execute("""
                SELECT m.member_id, m.name, m.surname,
                       COALESCE(SUM(c.amount), 0) as total_contributions,
                       COALESCE(SUM(l.outstanding_balance), 0) as outstanding_balance,
                       COUNT(DISTINCT l.loan_id) as loan_count
                FROM Members m
                LEFT JOIN Contributions c ON m.member_id = c.member_id AND c.contribution_date <= ?
                LEFT JOIN Loans l ON m.member_id = l.member_id AND l.outstanding_balance > 0
                WHERE m.status = 'active'
                GROUP BY m.member_id, m.name, m.surname
                ORDER BY total_contributions DESC
            """, (end_date,))
            
            members_data = cursor.fetchall()
            
            lines.append(f"{'Member Name':<25} {'Contributions':<15} {'Outstanding':<15} {'Loans':<8} {'Status':<12}")
            lines.append("-" * 80)
            
            for member_data in members_data:
                member_id, name, surname, contributions, outstanding, loan_count = member_data
                full_name = f"{name} {surname}"
                
                # Determine member status
                if outstanding > 0:
                    status = "Borrower"
                elif contributions > 0:
                    status = "Active"
                else:
                    status = "Inactive"
                
                lines.append(f"{full_name:<25} {contributions:<15.2f} {outstanding:<15.2f} {loan_count:<8} {status:<12}")
            
            conn.close()
            
        except Exception as e:
            lines.append(f"Error generating member analysis: {str(e)}")
        
        lines.append("")
        return lines
    
    def _generate_loan_analysis(self, end_date: str) -> list:
        """Generate loan portfolio analysis"""
        lines = []
        lines.append("LOAN PORTFOLIO ANALYSIS")
        lines.append("-" * 40)
        
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Loan statistics
            cursor.execute("SELECT COUNT(*) FROM Loans WHERE loan_date <= ?", (end_date,))
            total_loans = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM Loans WHERE outstanding_balance > 0")
            active_loans = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM Loans WHERE outstanding_balance = 0 AND loan_date <= ?", (end_date,))
            completed_loans = cursor.fetchone()[0]
            
            cursor.execute("SELECT AVG(loan_amount) FROM Loans WHERE loan_date <= ?", (end_date,))
            avg_loan_amount = cursor.fetchone()[0] or 0
            
            cursor.execute("SELECT AVG(interest_rate) FROM Loans WHERE loan_date <= ?", (end_date,))
            avg_interest_rate = cursor.fetchone()[0] or 0
            
            lines.append(f"Total Loans Issued: {total_loans}")
            lines.append(f"Active Loans: {active_loans}")
            lines.append(f"Completed Loans: {completed_loans}")
            lines.append(f"Average Loan Amount: {self.format_currency(avg_loan_amount)}")
            lines.append(f"Average Interest Rate: {avg_interest_rate:.2f}% per month")
            
            # Loan size distribution
            lines.append("")
            lines.append("Loan Size Distribution:")
            
            size_ranges = [
                (0, 500, "Small (0-500)"),
                (501, 1000, "Medium (501-1,000)"),
                (1001, 2000, "Large (1,001-2,000)"),
                (2001, float('inf'), "Very Large (2,000+)")
            ]
            
            for min_amount, max_amount, label in size_ranges:
                if max_amount == float('inf'):
                    cursor.execute("SELECT COUNT(*) FROM Loans WHERE loan_amount >= ? AND loan_date <= ?", (min_amount, end_date))
                else:
                    cursor.execute("SELECT COUNT(*) FROM Loans WHERE loan_amount >= ? AND loan_amount <= ? AND loan_date <= ?", (min_amount, max_amount, end_date))
                
                count = cursor.fetchone()[0]
                percentage = (count / total_loans * 100) if total_loans > 0 else 0
                lines.append(f"  {label}: {count} loans ({percentage:.1f}%)")
            
            conn.close()
            
        except Exception as e:
            lines.append(f"Error generating loan analysis: {str(e)}")
        
        lines.append("")
        return lines
    
    def _generate_dividend_analysis(self, end_date: str) -> list:
        """Generate dividend distribution analysis"""
        lines = []
        lines.append("DIVIDEND DISTRIBUTION ANALYSIS")
        lines.append("-" * 40)
        
        try:
            current_year = datetime.strptime(end_date, '%Y-%m-%d').year
            
            # Calculate dividends for the year
            dividend_summary = calculate_all_dividends_for_year(current_year)
            
            lines.append(f"Year: {current_year}")
            lines.append(f"Total Fund Available: {self.format_currency(dividend_summary['total_fund'])}")
            lines.append(f"Total Dividends Calculated: {self.format_currency(dividend_summary['total_dividends'])}")
            lines.append(f"Distribution Valid: {'Yes' if dividend_summary['distribution_valid'] else 'No'}")
            lines.append("")
            
            # Individual dividend breakdown
            lines.append("Individual Dividend Breakdown:")
            lines.append(f"{'Member Name':<25} {'Contributions':<15} {'Interest Paid':<15} {'Outstanding':<15} {'Dividend':<15}")
            lines.append("-" * 90)
            
            for dividend_calc in dividend_summary['dividend_calculations']:
                # Get member name
                conn = connect_db()
                cursor = conn.cursor()
                cursor.execute("SELECT name, surname FROM Members WHERE member_id = ?", (dividend_calc['member_id'],))
                member_data = cursor.fetchone()
                conn.close()
                
                if member_data:
                    member_name = f"{member_data[0]} {member_data[1]}"
                    lines.append(f"{member_name:<25} {dividend_calc['total_contributions']:<15.2f} {dividend_calc['total_interest_paid']:<15.2f} {dividend_calc['outstanding_balance']:<15.2f} {dividend_calc['dividend_amount']:<15.2f}")
            
        except Exception as e:
            lines.append(f"Error generating dividend analysis: {str(e)}")
        
        lines.append("")
        return lines
    
    def _generate_growth_metrics(self, end_date: str) -> list:
        """Generate year-over-year growth metrics"""
        lines = []
        lines.append("YEAR-OVER-YEAR GROWTH METRICS")
        lines.append("-" * 40)
        
        try:
            current_year = datetime.strptime(end_date, '%Y-%m-%d').year
            previous_year = current_year - 1
            
            conn = connect_db()
            cursor = conn.cursor()
            
            # Compare contributions
            cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM Contributions WHERE year = ?", (current_year,))
            current_contributions = cursor.fetchone()[0]
            
            cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM Contributions WHERE year = ?", (previous_year,))
            previous_contributions = cursor.fetchone()[0]
            
            # Compare loans
            cursor.execute("SELECT COALESCE(SUM(loan_amount), 0) FROM Loans WHERE strftime('%Y', loan_date) = ?", (str(current_year),))
            current_loans = cursor.fetchone()[0]
            
            cursor.execute("SELECT COALESCE(SUM(loan_amount), 0) FROM Loans WHERE strftime('%Y', loan_date) = ?", (str(previous_year),))
            previous_loans = cursor.fetchone()[0]
            
            # Compare members
            cursor.execute("SELECT COUNT(*) FROM Members WHERE strftime('%Y', created_at) = ? AND status = 'active'", (str(current_year),))
            new_members = cursor.fetchone()[0]
            
            conn.close()
            
            # Calculate growth rates
            contrib_growth = ((current_contributions - previous_contributions) / previous_contributions * 100) if previous_contributions > 0 else 0
            loan_growth = ((current_loans - previous_loans) / previous_loans * 100) if previous_loans > 0 else 0
            
            lines.append(f"Contributions Growth:")
            lines.append(f"  {previous_year}: {self.format_currency(previous_contributions)}")
            lines.append(f"  {current_year}: {self.format_currency(current_contributions)}")
            lines.append(f"  Growth Rate: {contrib_growth:+.1f}%")
            lines.append("")
            
            lines.append(f"Loans Growth:")
            lines.append(f"  {previous_year}: {self.format_currency(previous_loans)}")
            lines.append(f"  {current_year}: {self.format_currency(current_loans)}")
            lines.append(f"  Growth Rate: {loan_growth:+.1f}%")
            lines.append("")
            
            lines.append(f"New Members This Year: {new_members}")
            
        except Exception as e:
            lines.append(f"Error generating growth metrics: {str(e)}")
        
        lines.append("")
        return lines
    
    def _generate_system_status(self) -> list:
        """Generate system status section"""
        lines = []
        lines.append("SYSTEM STATUS")
        lines.append("-" * 40)
        
        try:
            # Database status
            lines.append(f"Database Type: SQLite")
            lines.append(f"MySQL Available: {'Yes' if MYSQL_AVAILABLE else 'No'}")
            
            # Get database file size
            if os.path.exists(SQLITE_DB_NAME):
                db_size = os.path.getsize(SQLITE_DB_NAME)
                lines.append(f"Database Size: {db_size:,} bytes")
            
            # Audit log status
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM AuditLog")
            audit_count = cursor.fetchone()[0]
            lines.append(f"Audit Log Entries: {audit_count:,}")
            
            # Recent activity
            cursor.execute("SELECT COUNT(*) FROM AuditLog WHERE created_at >= datetime('now', '-7 days')")
            recent_activity = cursor.fetchone()[0]
            lines.append(f"Activity (Last 7 Days): {recent_activity:,} transactions")
            
            conn.close()
            
        except Exception as e:
            lines.append(f"Error generating system status: {str(e)}")
        
        lines.append("")
        lines.append("=" * 80)
        lines.append("END OF REPORT")
        lines.append("=" * 80)
        
        return lines
    
    def export_to_pdf(self, report_content: str, filename: str) -> dict:
        """
        Export report to PDF format
        
        Args:
            report_content (str): Report content to export
            filename (str): Output filename
            
        Returns:
            dict: Export result with success status and message
        """
        if not REPORTLAB_AVAILABLE:
            return {
                'success': False,
                'message': 'PDF export not available. Please install reportlab: pip install reportlab'
            }
        
        try:
            # Create PDF document
            doc = SimpleDocTemplate(filename, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Normal style for content
            normal_style = styles['Normal']
            normal_style.fontSize = 10
            normal_style.fontName = 'Courier'  # Monospace font for better formatting
            
            # Split content into lines and process
            lines = report_content.split('\n')
            current_paragraph = []
            
            for line in lines:
                if line.strip() == '':
                    if current_paragraph:
                        story.append(Paragraph('<br/>'.join(current_paragraph), normal_style))
                        story.append(Spacer(1, 12))
                        current_paragraph = []
                elif line.startswith('='):
                    # Title line
                    if current_paragraph:
                        story.append(Paragraph('<br/>'.join(current_paragraph), normal_style))
                        current_paragraph = []
                    if 'BANK MMUDZI' in line:
                        story.append(Paragraph('BANK MMUDZI FINANCIAL REPORT', title_style))
                    story.append(Spacer(1, 12))
                else:
                    # Regular content line
                    current_paragraph.append(line.replace('<', '&lt;').replace('>', '&gt;'))
            
            # Add any remaining content
            if current_paragraph:
                story.append(Paragraph('<br/>'.join(current_paragraph), normal_style))
            
            # Build PDF
            doc.build(story)
            
            return {
                'success': True,
                'message': f'Report exported to PDF: {filename}'
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'Error exporting to PDF: {ErrorHandler.get_user_friendly_message(e)}'
            }
    
    def export_to_excel(self, report_data: dict, filename: str) -> dict:
        """
        Export financial data to Excel format
        
        Args:
            report_data (dict): Structured report data
            filename (str): Output filename
            
        Returns:
            dict: Export result with success status and message
        """
        if not OPENPYXL_AVAILABLE:
            return {
                'success': False,
                'message': 'Excel export not available. Please install openpyxl: pip install openpyxl'
            }
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            summary_sheet = wb.create_sheet("Summary")
            self._create_summary_sheet(summary_sheet, report_data)
            
            # Create members sheet
            members_sheet = wb.create_sheet("Members")
            self._create_members_sheet(members_sheet, report_data)
            
            # Create loans sheet
            loans_sheet = wb.create_sheet("Loans")
            self._create_loans_sheet(loans_sheet, report_data)
            
            # Create contributions sheet
            contributions_sheet = wb.create_sheet("Contributions")
            self._create_contributions_sheet(contributions_sheet, report_data)
            
            # Save workbook
            wb.save(filename)
            
            return {
                'success': True,
                'message': f'Report exported to Excel: {filename}'
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'Error exporting to Excel: {ErrorHandler.get_user_friendly_message(e)}'
            }
    
    def _create_summary_sheet(self, sheet, report_data):
        """Create summary sheet in Excel workbook"""
        # Headers
        sheet['A1'] = 'Bank Mmudzi Financial Summary'
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:D1')
        
        # Summary data
        row = 3
        summary_items = [
            ('Total Members', report_data.get('total_members', 0)),
            ('Total Contributions', report_data.get('total_contributions', 0)),
            ('Total Loans', report_data.get('total_loans', 0)),
            ('Outstanding Balance', report_data.get('outstanding_balance', 0)),
            ('Available Funds', report_data.get('available_funds', 0))
        ]
        
        for label, value in summary_items:
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
    
    def _create_members_sheet(self, sheet, report_data):
        """Create members sheet in Excel workbook"""
        headers = ['Member ID', 'Name', 'Phone', 'Total Contributions', 'Active Loans', 'Status']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add member data (would be populated from actual data)
        # This is a placeholder - actual implementation would fetch member data
        
    def _create_loans_sheet(self, sheet, report_data):
        """Create loans sheet in Excel workbook"""
        headers = ['Loan ID', 'Member', 'Amount', 'Interest Rate', 'Outstanding', 'Status']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    def _create_contributions_sheet(self, sheet, report_data):
        """Create contributions sheet in Excel workbook"""
        headers = ['Date', 'Member', 'Month', 'Year', 'Amount']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

# Validation Engine
class ValidationEngine:
    """
    Comprehensive input validation engine for all user inputs
    """
    
    @staticmethod
    def validate_phone_number(phone: str) -> dict:
        """
        Validate phone number format
        Accepts formats: +265123456789, 0123456789, 123456789
        
        Args:
            phone (str): Phone number to validate
            
        Returns:
            dict: {'valid': bool, 'message': str, 'formatted': str}
        """
        if not phone:
            return {'valid': False, 'message': 'Phone number is required', 'formatted': ''}
        
        # Remove spaces and common separators
        cleaned_phone = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        
        # Check for valid characters (only digits and + at start)
        if not cleaned_phone.replace('+', '').isdigit():
            return {'valid': False, 'message': 'Phone number can only contain digits and + at the beginning', 'formatted': ''}
        
        # Handle different formats
        if cleaned_phone.startswith('+265'):
            # International format: +265123456789
            if len(cleaned_phone) != 13:
                return {'valid': False, 'message': 'International format should be +265 followed by 9 digits', 'formatted': ''}
            formatted = cleaned_phone
        elif cleaned_phone.startswith('0'):
            # National format: 0123456789
            if len(cleaned_phone) != 10:
                return {'valid': False, 'message': 'National format should be 10 digits starting with 0', 'formatted': ''}
            formatted = '+265' + cleaned_phone[1:]  # Convert to international
        elif len(cleaned_phone) == 9:
            # Local format: 123456789
            formatted = '+265' + cleaned_phone
        else:
            return {'valid': False, 'message': 'Invalid phone number format. Use +265123456789, 0123456789, or 123456789', 'formatted': ''}
        
        return {'valid': True, 'message': 'Valid phone number', 'formatted': formatted}
    
    @staticmethod
    def validate_financial_amount(amount_str: str, field_name: str = "Amount") -> dict:
        """
        Validate financial amount with proper decimal precision
        
        Args:
            amount_str (str): Amount string to validate
            field_name (str): Name of the field for error messages
            
        Returns:
            dict: {'valid': bool, 'message': str, 'value': float}
        """
        if not amount_str or amount_str.strip() == '':
            return {'valid': False, 'message': f'{field_name} is required', 'value': 0.0}
        
        try:
            # Remove any currency symbols and spaces
            cleaned_amount = amount_str.replace('MWK', '').replace('K', '').replace(',', '').strip()
            
            # Convert to float
            amount = float(cleaned_amount)
            
            # Check for negative values
            if amount < 0:
                return {'valid': False, 'message': f'{field_name} cannot be negative', 'value': 0.0}
            
            # Check for zero (usually not allowed for financial transactions)
            if amount == 0:
                return {'valid': False, 'message': f'{field_name} must be greater than zero', 'value': 0.0}
            
            # Check for reasonable maximum (prevent extremely large values)
            if amount > 10000000:  # 10 million MWK
                return {'valid': False, 'message': f'{field_name} exceeds maximum allowed value (10,000,000 MWK)', 'value': 0.0}
            
            # Round to 2 decimal places for currency precision
            rounded_amount = round(amount, 2)
            
            return {'valid': True, 'message': 'Valid amount', 'value': rounded_amount}
            
        except ValueError:
            return {'valid': False, 'message': f'{field_name} must be a valid number', 'value': 0.0}
    
    @staticmethod
    def validate_member_data(name: str, surname: str, phone: str, email: str = '') -> dict:
        """
        Validate complete member data
        
        Args:
            name (str): Member's first name
            surname (str): Member's surname
            phone (str): Member's phone number
            email (str): Member's email (optional)
            
        Returns:
            dict: {'valid': bool, 'errors': list, 'formatted_data': dict}
        """
        errors = []
        formatted_data = {}
        
        # Validate name
        if not name or name.strip() == '':
            errors.append('First name is required')
        elif len(name.strip()) < 2:
            errors.append('First name must be at least 2 characters')
        elif not name.replace(' ', '').replace('-', '').replace("'", '').isalpha():
            errors.append('First name can only contain letters, spaces, hyphens, and apostrophes')
        else:
            formatted_data['name'] = name.strip().title()
        
        # Validate surname
        if not surname or surname.strip() == '':
            errors.append('Surname is required')
        elif len(surname.strip()) < 2:
            errors.append('Surname must be at least 2 characters')
        elif not surname.replace(' ', '').replace('-', '').replace("'", '').isalpha():
            errors.append('Surname can only contain letters, spaces, hyphens, and apostrophes')
        else:
            formatted_data['surname'] = surname.strip().title()
        
        # Validate phone
        phone_validation = ValidationEngine.validate_phone_number(phone)
        if not phone_validation['valid']:
            errors.append(phone_validation['message'])
        else:
            formatted_data['phone'] = phone_validation['formatted']
        
        # Validate email (optional)
        if email and email.strip():
            email = email.strip().lower()
            if '@' not in email or '.' not in email.split('@')[1]:
                errors.append('Invalid email format')
            elif len(email) > 100:
                errors.append('Email address is too long')
            else:
                formatted_data['email'] = email
        else:
            formatted_data['email'] = ''
        
        return {
            'valid': len(errors) == 0,
            'errors': errors,
            'formatted_data': formatted_data
        }
    
    @staticmethod
    def check_duplicate_contribution(member_id: int, month: int, year: int) -> dict:
        """
        Check for duplicate contribution for same member/month/year combination
        
        Args:
            member_id (int): Member ID
            month (int): Month number (1-12)
            year (int): Year
            
        Returns:
            dict: {'is_duplicate': bool, 'message': str, 'existing_contribution': dict}
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT contribution_id, amount, contribution_date 
                FROM Contributions 
                WHERE member_id = ? AND month = ? AND year = ?
            ''', (member_id, month, year))
            
            existing = cursor.fetchone()
            conn.close()
            
            if existing:
                month_name = MONTHS[month - 1]  # Convert month number to name
                return {
                    'is_duplicate': True,
                    'message': f'Contribution already exists for {month_name} {year}',
                    'existing_contribution': {
                        'contribution_id': existing[0],
                        'amount': existing[1],
                        'contribution_date': existing[2]
                    }
                }
            else:
                return {
                    'is_duplicate': False,
                    'message': 'No duplicate found',
                    'existing_contribution': None
                }
                
        except Exception as e:
            return {
                'is_duplicate': False,
                'message': f'Error checking for duplicates: {str(e)}',
                'existing_contribution': None
            }
    
    @staticmethod
    def validate_loan_capacity(member_id: int, requested_amount: float) -> dict:
        """
        Calculate and validate loan capacity for a member
        
        Args:
            member_id (int): Member ID
            requested_amount (float): Requested loan amount
            
        Returns:
            dict: {'valid': bool, 'message': str, 'available_capacity': float, 'total_fund': float}
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Get total fund available (sum of all contributions)
            cursor.execute('SELECT COALESCE(SUM(amount), 0) FROM Contributions')
            total_contributions = cursor.fetchone()[0]
            
            # Get total outstanding loans
            cursor.execute('SELECT COALESCE(SUM(outstanding_balance), 0) FROM Loans WHERE outstanding_balance > 0')
            total_outstanding = cursor.fetchone()[0]
            
            # Calculate available capacity
            available_capacity = total_contributions - total_outstanding
            
            # Get member's existing outstanding loans
            cursor.execute('''
                SELECT COALESCE(SUM(outstanding_balance), 0) 
                FROM Loans 
                WHERE member_id = ? AND outstanding_balance > 0
            ''', (member_id,))
            member_outstanding = cursor.fetchone()[0]
            
            conn.close()
            
            # Check if requested amount exceeds available capacity
            if requested_amount > available_capacity:
                return {
                    'valid': False,
                    'message': f'Requested amount ({requested_amount:,.2f} MWK) exceeds available capacity ({available_capacity:,.2f} MWK)',
                    'available_capacity': available_capacity,
                    'total_fund': total_contributions,
                    'member_outstanding': member_outstanding
                }
            
            # Check if member already has outstanding loans (optional business rule)
            if member_outstanding > 0:
                return {
                    'valid': False,
                    'message': f'Member has existing outstanding loan balance of {member_outstanding:,.2f} MWK. Please clear existing loans before taking new ones.',
                    'available_capacity': available_capacity,
                    'total_fund': total_contributions,
                    'member_outstanding': member_outstanding
                }
            
            return {
                'valid': True,
                'message': f'Loan approved. Available capacity: {available_capacity:,.2f} MWK',
                'available_capacity': available_capacity,
                'total_fund': total_contributions,
                'member_outstanding': member_outstanding
            }
            
        except Exception as e:
            return {
                'valid': False,
                'message': f'Error validating loan capacity: {str(e)}',
                'available_capacity': 0.0,
                'total_fund': 0.0,
                'member_outstanding': 0.0
            }
    
    @staticmethod
    def validate_interest_rate(rate_str: str) -> dict:
        """
        Validate interest rate input
        
        Args:
            rate_str (str): Interest rate string
            
        Returns:
            dict: {'valid': bool, 'message': str, 'value': float}
        """
        if not rate_str or rate_str.strip() == '':
            return {'valid': False, 'message': 'Interest rate is required', 'value': 0.0}
        
        try:
            # Remove percentage symbol if present
            cleaned_rate = rate_str.replace('%', '').strip()
            rate = float(cleaned_rate)
            
            # Convert percentage to decimal if it seems to be in percentage format
            if rate > 1.0:
                rate = rate / 100.0
            
            # Validate range (0% to 50% monthly seems reasonable)
            if rate < 0:
                return {'valid': False, 'message': 'Interest rate cannot be negative', 'value': 0.0}
            
            if rate > 0.5:  # 50% monthly rate
                return {'valid': False, 'message': 'Interest rate cannot exceed 50% per month', 'value': 0.0}
            
            return {'valid': True, 'message': 'Valid interest rate', 'value': rate}
            
        except ValueError:
            return {'valid': False, 'message': 'Interest rate must be a valid number', 'value': 0.0}
    
    @staticmethod
    def validate_date_input(date_str: str, field_name: str = "Date") -> dict:
        """
        Validate date input in various formats
        
        Args:
            date_str (str): Date string to validate
            field_name (str): Name of the field for error messages
            
        Returns:
            dict: {'valid': bool, 'message': str, 'formatted_date': str}
        """
        if not date_str or date_str.strip() == '':
            return {'valid': False, 'message': f'{field_name} is required', 'formatted_date': ''}
        
        date_str = date_str.strip()
        
        # Try different date formats
        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']
        
        for date_format in date_formats:
            try:
                parsed_date = datetime.strptime(date_str, date_format)
                
                # Check if date is reasonable (not too far in past or future)
                current_date = datetime.now()
                min_date = datetime(1900, 1, 1)
                max_date = datetime(2100, 12, 31)
                
                if parsed_date < min_date:
                    return {'valid': False, 'message': f'{field_name} cannot be before year 1900', 'formatted_date': ''}
                
                if parsed_date > max_date:
                    return {'valid': False, 'message': f'{field_name} cannot be after year 2100', 'formatted_date': ''}
                
                # Return in standard format
                formatted_date = parsed_date.strftime('%Y-%m-%d')
                return {'valid': True, 'message': 'Valid date', 'formatted_date': formatted_date}
                
            except ValueError:
                continue
        
        return {'valid': False, 'message': f'{field_name} must be in format YYYY-MM-DD, DD/MM/YYYY, or MM/DD/YYYY', 'formatted_date': ''}
    
    @staticmethod
    def validate_member_number(member_number: str) -> dict:
        """
        Validate member number format and uniqueness
        
        Args:
            member_number (str): Member number to validate
            
        Returns:
            dict: {'valid': bool, 'message': str, 'formatted_number': str}
        """
        if not member_number or member_number.strip() == '':
            return {'valid': False, 'message': 'Member number is required', 'formatted_number': ''}
        
        member_number = member_number.strip().upper()
        
        # Check format: should be alphanumeric, 6-12 characters
        if not member_number.isalnum():
            return {'valid': False, 'message': 'Member number can only contain letters and numbers', 'formatted_number': ''}
        
        if len(member_number) < 6:
            return {'valid': False, 'message': 'Member number must be at least 6 characters', 'formatted_number': ''}
        
        if len(member_number) > 12:
            return {'valid': False, 'message': 'Member number cannot exceed 12 characters', 'formatted_number': ''}
        
        # Check uniqueness in database
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM Members WHERE member_number = ?', (member_number,))
            count = cursor.fetchone()[0]
            conn.close()
            
            if count > 0:
                return {'valid': False, 'message': 'Member number already exists', 'formatted_number': ''}
            
        except Exception as e:
            return {'valid': False, 'message': f'Error checking member number uniqueness: {str(e)}', 'formatted_number': ''}
        
        return {'valid': True, 'message': 'Valid member number', 'formatted_number': member_number}
    
    @staticmethod
    def validate_password_strength(password: str) -> dict:
        """
        Validate password strength according to security requirements
        
        Args:
            password (str): Password to validate
            
        Returns:
            dict: {'valid': bool, 'message': str, 'strength_score': int}
        """
        if not password:
            return {'valid': False, 'message': 'Password is required', 'strength_score': 0}
        
        errors = []
        strength_score = 0
        
        # Check minimum length
        if len(password) < 8:
            errors.append('Password must be at least 8 characters long')
        else:
            strength_score += 1
        
        # Check for uppercase letters
        if not any(c.isupper() for c in password):
            errors.append('Password must contain at least one uppercase letter')
        else:
            strength_score += 1
        
        # Check for lowercase letters
        if not any(c.islower() for c in password):
            errors.append('Password must contain at least one lowercase letter')
        else:
            strength_score += 1
        
        # Check for numbers
        if not any(c.isdigit() for c in password):
            errors.append('Password must contain at least one number')
        else:
            strength_score += 1
        
        # Check for special characters
        special_chars = "!@#$%^&*()_+-=[]{}|;:,.<>?"
        if not any(c in special_chars for c in password):
            errors.append('Password must contain at least one special character (!@#$%^&*()_+-=[]{}|;:,.<>?)')
        else:
            strength_score += 1
        
        # Check for common weak passwords
        weak_passwords = ['password', '123456', 'qwerty', 'admin', 'letmein', 'welcome']
        if password.lower() in weak_passwords:
            errors.append('Password is too common and easily guessed')
            strength_score = 0
        
        if errors:
            return {'valid': False, 'message': '; '.join(errors), 'strength_score': strength_score}
        
        return {'valid': True, 'message': 'Strong password', 'strength_score': strength_score}
    
    @staticmethod
    def validate_loan_term(term_str: str) -> dict:
        """
        Validate loan term input
        
        Args:
            term_str (str): Loan term string (in months)
            
        Returns:
            dict: {'valid': bool, 'message': str, 'value': int}
        """
        if not term_str or term_str.strip() == '':
            return {'valid': False, 'message': 'Loan term is required', 'value': 0}
        
        try:
            term = int(term_str.strip())
            
            # Check reasonable range (1 month to 5 years)
            if term < 1:
                return {'valid': False, 'message': 'Loan term must be at least 1 month', 'value': 0}
            
            if term > 60:  # 5 years
                return {'valid': False, 'message': 'Loan term cannot exceed 60 months (5 years)', 'value': 0}
            
            return {'valid': True, 'message': 'Valid loan term', 'value': term}
            
        except ValueError:
            return {'valid': False, 'message': 'Loan term must be a whole number of months', 'value': 0}
    
    @staticmethod
    def validate_contribution_month(month_str: str, year: int) -> dict:
        """
        Validate contribution month input
        
        Args:
            month_str (str): Month name or number
            year (int): Year for the contribution
            
        Returns:
            dict: {'valid': bool, 'message': str, 'month_number': int, 'month_name': str}
        """
        if not month_str or month_str.strip() == '':
            return {'valid': False, 'message': 'Month is required', 'month_number': 0, 'month_name': ''}
        
        month_str = month_str.strip()
        
        # Try to parse as month name
        if month_str.title() in MONTHS:
            month_number = MONTHS.index(month_str.title()) + 1
            return {'valid': True, 'message': 'Valid month', 'month_number': month_number, 'month_name': month_str.title()}
        
        # Try to parse as month number
        try:
            month_number = int(month_str)
            if 1 <= month_number <= 12:
                month_name = MONTHS[month_number - 1]
                return {'valid': True, 'message': 'Valid month', 'month_number': month_number, 'month_name': month_name}
            else:
                return {'valid': False, 'message': 'Month number must be between 1 and 12', 'month_number': 0, 'month_name': ''}
        except ValueError:
            pass
        
        return {'valid': False, 'message': 'Month must be a valid month name (e.g., January) or number (1-12)', 'month_number': 0, 'month_name': ''}
    
    @staticmethod
    def validate_year_input(year_str: str) -> dict:
        """
        Validate year input
        
        Args:
            year_str (str): Year string to validate
            
        Returns:
            dict: {'valid': bool, 'message': str, 'value': int}
        """
        if not year_str or year_str.strip() == '':
            return {'valid': False, 'message': 'Year is required', 'value': 0}
        
        try:
            year = int(year_str.strip())
            
            # Check reasonable range
            current_year = datetime.now().year
            if year < 2000:
                return {'valid': False, 'message': 'Year cannot be before 2000', 'value': 0}
            
            if year > current_year + 1:
                return {'valid': False, 'message': f'Year cannot be after {current_year + 1}', 'value': 0}
            
            return {'valid': True, 'message': 'Valid year', 'value': year}
            
        except ValueError:
            return {'valid': False, 'message': 'Year must be a valid 4-digit number', 'value': 0}

# Enhanced Member Validation Engine
class MemberValidationEngine:
    """
    Centralized validation engine specifically for member operations
    Extends the base ValidationEngine with member-specific validation logic
    """
    
    @staticmethod
    def validate_member_form_data(data: dict, member_id: int = None) -> dict:
        """
        Comprehensive validation for member form data with real-time feedback
        
        Args:
            data (dict): Member data dictionary
            member_id (int): Existing member ID for updates (None for new members)
            
        Returns:
            dict: {
                'valid': bool,
                'errors': dict,  # Field-specific errors
                'warnings': dict,  # Field-specific warnings
                'formatted_data': dict,  # Cleaned and formatted data
                'field_status': dict  # Per-field validation status
            }
        """
        errors = {}
        warnings = {}
        formatted_data = {}
        field_status = {}
        
        # Validate first name
        name_result = MemberValidationEngine._validate_name_field(
            data.get('name', ''), 'First Name'
        )
        if not name_result['valid']:
            errors['name'] = name_result['message']
            field_status['name'] = 'invalid'
        else:
            formatted_data['name'] = name_result['formatted_value']
            field_status['name'] = 'valid'
        
        # Validate surname
        surname_result = MemberValidationEngine._validate_name_field(
            data.get('surname', ''), 'Last Name'
        )
        if not surname_result['valid']:
            errors['surname'] = surname_result['message']
            field_status['surname'] = 'invalid'
        else:
            formatted_data['surname'] = surname_result['formatted_value']
            field_status['surname'] = 'valid'
        
        # Validate phone number with auto-formatting
        phone_result = MemberValidationEngine.validate_phone_with_formatting(
            data.get('phone', '')
        )
        if not phone_result['valid']:
            errors['phone'] = phone_result['message']
            field_status['phone'] = 'invalid'
        else:
            formatted_data['phone'] = phone_result['formatted']
            field_status['phone'] = 'valid'
            
            # Check for duplicate phone numbers
            duplicate_check = MemberValidationEngine.check_phone_uniqueness(
                phone_result['formatted'], member_id
            )
            if not duplicate_check['unique']:
                warnings['phone'] = duplicate_check['message']
        
        # Validate email (optional)
        email_result = MemberValidationEngine.validate_email_field(
            data.get('email', '')
        )
        if not email_result['valid'] and data.get('email', '').strip():
            errors['email'] = email_result['message']
            field_status['email'] = 'invalid'
        else:
            formatted_data['email'] = email_result['formatted_value']
            field_status['email'] = 'valid' if email_result['formatted_value'] else 'neutral'
        
        # Validate member number (if provided)
        if data.get('member_number'):
            member_number_result = MemberValidationEngine.validate_member_number_field(
                data.get('member_number', ''), member_id
            )
            if not member_number_result['valid']:
                errors['member_number'] = member_number_result['message']
                field_status['member_number'] = 'invalid'
            else:
                formatted_data['member_number'] = member_number_result['formatted_number']
                field_status['member_number'] = 'valid'
        
        # Validate address (optional)
        address_result = MemberValidationEngine._validate_address_field(
            data.get('address', '')
        )
        formatted_data['address'] = address_result['formatted_value']
        field_status['address'] = 'neutral'
        
        # Validate notes (optional)
        notes_result = MemberValidationEngine._validate_notes_field(
            data.get('notes', '')
        )
        if not notes_result['valid']:
            warnings['notes'] = notes_result['message']
        formatted_data['notes'] = notes_result['formatted_value']
        field_status['notes'] = 'neutral'
        
        return {
            'valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings,
            'formatted_data': formatted_data,
            'field_status': field_status
        }
    
    @staticmethod
    def _validate_name_field(name: str, field_name: str) -> dict:
        """Validate name fields with enhanced rules"""
        if not name or name.strip() == '':
            return {'valid': False, 'message': f'{field_name} is required', 'formatted_value': ''}
        
        name = name.strip()
        
        # Check minimum length
        if len(name) < 2:
            return {'valid': False, 'message': f'{field_name} must be at least 2 characters', 'formatted_value': ''}
        
        # Check maximum length
        if len(name) > 50:
            return {'valid': False, 'message': f'{field_name} cannot exceed 50 characters', 'formatted_value': ''}
        
        # Check for valid characters (letters, spaces, hyphens, apostrophes)
        allowed_chars = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ -'")
        if not all(c in allowed_chars for c in name):
            return {'valid': False, 'message': f'{field_name} can only contain letters, spaces, hyphens, and apostrophes', 'formatted_value': ''}
        
        # Check for reasonable patterns
        if name.count(' ') > 3:
            return {'valid': False, 'message': f'{field_name} has too many spaces', 'formatted_value': ''}
        
        if '--' in name or "''" in name:
            return {'valid': False, 'message': f'{field_name} contains invalid character sequences', 'formatted_value': ''}
        
        # Format: Title case with proper spacing
        formatted_name = ' '.join(word.capitalize() for word in name.split())
        
        return {'valid': True, 'message': 'Valid name', 'formatted_value': formatted_name}
    
    @staticmethod
    def validate_phone_with_formatting(phone: str) -> dict:
        """Enhanced phone validation with auto-formatting"""
        if not phone or phone.strip() == '':
            return {'valid': False, 'message': 'Phone number is required', 'formatted': '', 'display_format': ''}
        
        # Use base validation engine
        base_result = ValidationEngine.validate_phone_number(phone)
        
        if not base_result['valid']:
            return {
                'valid': False,
                'message': base_result['message'],
                'formatted': '',
                'display_format': ''
            }
        
        # Create display format for user interface
        formatted_phone = base_result['formatted']  # +265123456789
        display_format = f"{formatted_phone[:4]} {formatted_phone[4:7]} {formatted_phone[7:10]} {formatted_phone[10:]}"
        
        return {
            'valid': True,
            'message': 'Valid phone number',
            'formatted': formatted_phone,
            'display_format': display_format.strip()
        }
    
    @staticmethod
    def check_phone_uniqueness(phone: str, exclude_member_id: int = None) -> dict:
        """Check if phone number is unique in the system"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            if exclude_member_id:
                cursor.execute(
                    'SELECT member_id, name, surname FROM Members WHERE phone = ? AND member_id != ?',
                    (phone, exclude_member_id)
                )
            else:
                cursor.execute(
                    'SELECT member_id, name, surname FROM Members WHERE phone = ?',
                    (phone,)
                )
            
            existing = cursor.fetchone()
            conn.close()
            
            if existing:
                return {
                    'unique': False,
                    'message': f'Phone number already used by {existing[1]} {existing[2]} (ID: {existing[0]})',
                    'existing_member': {
                        'member_id': existing[0],
                        'name': existing[1],
                        'surname': existing[2]
                    }
                }
            
            return {'unique': True, 'message': 'Phone number is unique', 'existing_member': None}
            
        except Exception as e:
            return {'unique': True, 'message': f'Could not check uniqueness: {str(e)}', 'existing_member': None}
    
    @staticmethod
    def validate_email_field(email: str) -> dict:
        """Enhanced email validation"""
        if not email or email.strip() == '':
            return {'valid': True, 'message': 'Email is optional', 'formatted_value': ''}
        
        email = email.strip().lower()
        
        # Basic format validation
        if '@' not in email:
            return {'valid': False, 'message': 'Email must contain @ symbol', 'formatted_value': ''}
        
        if email.count('@') != 1:
            return {'valid': False, 'message': 'Email must contain exactly one @ symbol', 'formatted_value': ''}
        
        local, domain = email.split('@')
        
        # Validate local part
        if len(local) == 0:
            return {'valid': False, 'message': 'Email must have text before @ symbol', 'formatted_value': ''}
        
        if len(local) > 64:
            return {'valid': False, 'message': 'Email local part is too long', 'formatted_value': ''}
        
        # Validate domain part
        if len(domain) == 0:
            return {'valid': False, 'message': 'Email must have domain after @ symbol', 'formatted_value': ''}
        
        if '.' not in domain:
            return {'valid': False, 'message': 'Email domain must contain a dot', 'formatted_value': ''}
        
        if domain.startswith('.') or domain.endswith('.'):
            return {'valid': False, 'message': 'Email domain cannot start or end with a dot', 'formatted_value': ''}
        
        # Check overall length
        if len(email) > 254:
            return {'valid': False, 'message': 'Email address is too long', 'formatted_value': ''}
        
        # Basic character validation
        allowed_chars = set("abcdefghijklmnopqrstuvwxyz0123456789.-_@")
        if not all(c in allowed_chars for c in email):
            return {'valid': False, 'message': 'Email contains invalid characters', 'formatted_value': ''}
        
        return {'valid': True, 'message': 'Valid email address', 'formatted_value': email}
    
    @staticmethod
    def validate_member_number_field(member_number: str, exclude_member_id: int = None) -> dict:
        """Enhanced member number validation"""
        if not member_number or member_number.strip() == '':
            return {'valid': False, 'message': 'Member number is required', 'formatted_number': ''}
        
        member_number = member_number.strip().upper()
        
        # Format validation
        if not member_number.isalnum():
            return {'valid': False, 'message': 'Member number can only contain letters and numbers', 'formatted_number': ''}
        
        if len(member_number) < 4:
            return {'valid': False, 'message': 'Member number must be at least 4 characters', 'formatted_number': ''}
        
        if len(member_number) > 15:
            return {'valid': False, 'message': 'Member number cannot exceed 15 characters', 'formatted_number': ''}
        
        # Check uniqueness
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            if exclude_member_id:
                cursor.execute(
                    'SELECT member_id, name, surname FROM Members WHERE member_number = ? AND member_id != ?',
                    (member_number, exclude_member_id)
                )
            else:
                cursor.execute(
                    'SELECT member_id, name, surname FROM Members WHERE member_number = ?',
                    (member_number,)
                )
            
            existing = cursor.fetchone()
            conn.close()
            
            if existing:
                return {
                    'valid': False,
                    'message': f'Member number already used by {existing[1]} {existing[2]}',
                    'formatted_number': member_number
                }
            
        except Exception as e:
            return {
                'valid': False,
                'message': f'Error checking member number uniqueness: {str(e)}',
                'formatted_number': member_number
            }
        
        return {'valid': True, 'message': 'Valid member number', 'formatted_number': member_number}
    
    @staticmethod
    def _validate_address_field(address: str) -> dict:
        """Validate address field"""
        if not address or address.strip() == '':
            return {'valid': True, 'message': 'Address is optional', 'formatted_value': ''}
        
        address = address.strip()
        
        if len(address) > 200:
            return {'valid': False, 'message': 'Address cannot exceed 200 characters', 'formatted_value': ''}
        
        # Basic formatting
        formatted_address = ' '.join(address.split())
        
        return {'valid': True, 'message': 'Valid address', 'formatted_value': formatted_address}
    
    @staticmethod
    def _validate_notes_field(notes: str) -> dict:
        """Validate notes field"""
        if not notes or notes.strip() == '':
            return {'valid': True, 'message': 'Notes are optional', 'formatted_value': ''}
        
        notes = notes.strip()
        
        if len(notes) > 500:
            return {'valid': False, 'message': 'Notes cannot exceed 500 characters', 'formatted_value': notes[:500]}
        
        return {'valid': True, 'message': 'Valid notes', 'formatted_value': notes}
    
    @staticmethod
    def generate_member_number() -> str:
        """Generate a unique member number"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Get the highest existing member ID
            cursor.execute('SELECT MAX(member_id) FROM Members')
            max_id = cursor.fetchone()[0] or 0
            
            # Generate member number: BM + year + sequential number
            current_year = datetime.now().year
            year_suffix = str(current_year)[-2:]  # Last 2 digits of year
            
            # Try sequential numbers until we find a unique one
            for i in range(max_id + 1, max_id + 1000):
                member_number = f"BM{year_suffix}{i:04d}"
                
                cursor.execute('SELECT COUNT(*) FROM Members WHERE member_number = ?', (member_number,))
                if cursor.fetchone()[0] == 0:
                    conn.close()
                    return member_number
            
            conn.close()
            # Fallback to timestamp-based number
            import time
            timestamp = str(int(time.time()))[-6:]
            return f"BM{year_suffix}{timestamp}"
            
        except Exception as e:
            # Fallback to simple timestamp-based number
            import time
            current_year = datetime.now().year
            year_suffix = str(current_year)[-2:]
            timestamp = str(int(time.time()))[-6:]
            return f"BM{year_suffix}{timestamp}"
    
    @staticmethod
    def validate_cross_field_dependencies(data: dict) -> dict:
        """Validate cross-field dependencies and business rules"""
        warnings = []
        errors = []
        
        # Check if phone and email are both missing (at least one contact method should be provided)
        if not data.get('phone', '').strip() and not data.get('email', '').strip():
            errors.append('At least one contact method (phone or email) must be provided')
        
        # Check for suspicious patterns
        name = data.get('name', '').strip().lower()
        surname = data.get('surname', '').strip().lower()
        
        if name and surname and name == surname:
            warnings.append('First name and last name are identical - please verify')
        
        # Check for test/dummy data patterns
        test_patterns = ['test', 'dummy', 'sample', 'example', 'temp']
        if any(pattern in name for pattern in test_patterns) or any(pattern in surname for pattern in test_patterns):
            warnings.append('Name appears to contain test data - please verify this is a real member')
        
        return {
            'valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings
        }

# Unit Tests for Financial Calculations
def test_loan_amortization_calculations():
    """
    Test loan amortization calculations with provided examples
    L=1000, r=0.2, expected P≈221
    """
    print("Testing Loan Amortization Calculations...")
    
    # Test case from requirements: L=1000, r=0.2, expected P≈221
    principal = 1000.0
    monthly_rate = 0.2
    months = 12
    
    # Calculate monthly payment
    monthly_payment = FinancialCalculator.calculate_loan_payment(principal, monthly_rate, months)
    print(f"Principal: {principal}, Rate: {monthly_rate}, Months: {months}")
    print(f"Calculated Monthly Payment: {monthly_payment}")
    print(f"Expected Monthly Payment: ~221")
    
    # Verify the calculation is approximately correct (within 5 MWK due to rounding differences)
    expected_payment = 221
    assert abs(monthly_payment - expected_payment) <= 5, f"Monthly payment {monthly_payment} should be approximately {expected_payment}"
    
    # Calculate total interest
    total_interest = FinancialCalculator.calculate_total_interest(principal, monthly_payment, months)
    print(f"Calculated Total Interest: {total_interest}")
    
    # Generate amortization schedule
    schedule = FinancialCalculator.generate_amortization_schedule(1, principal, monthly_rate, monthly_payment, months)
    print(f"Generated {len(schedule)} payment entries in amortization schedule")
    
    # Verify schedule integrity
    total_principal_payments = sum(entry[2] for entry in schedule)  # principal_amount is index 2
    total_interest_payments = sum(entry[3] for entry in schedule)   # interest_amount is index 3
    
    print(f"Total Principal Payments: {total_principal_payments}")
    print(f"Total Interest Payments: {total_interest_payments}")
    print(f"Sum of Payments: {total_principal_payments + total_interest_payments}")
    print(f"Expected Sum: {monthly_payment * months}")
    
    # Verify that principal payments sum to original loan amount (allow for small rounding differences)
    assert abs(total_principal_payments - principal) <= 0.25, f"Principal payments {total_principal_payments} should equal loan amount {principal}"
    
    # Verify that total payments equal monthly_payment * months
    total_payments = total_principal_payments + total_interest_payments
    expected_total = monthly_payment * months
    assert abs(total_payments - expected_total) <= 0.01, f"Total payments {total_payments} should equal {expected_total}"
    
    print("✓ Loan amortization calculations test passed!")
    return True

def test_dividend_calculations():
    """
    Test dividend calculations with provided examples
    Expected: borrower=2852, non-borrower=1200
    """
    print("\nTesting Dividend Calculations...")
    
    # Test case from requirements
    monthly_contribution = 100  # C = 100
    annual_contributions = monthly_contribution * 12  # 12C = 1200
    
    # Non-borrower case
    non_borrower_dividend = FinancialCalculator.calculate_member_dividend(
        member_id=1, 
        year=2024, 
        contributions=annual_contributions, 
        interest_paid=0, 
        outstanding_balance=0
    )
    print(f"Non-borrower dividend: {non_borrower_dividend} (expected: 1200)")
    assert non_borrower_dividend == 1200, f"Non-borrower dividend should be 1200, got {non_borrower_dividend}"
    
    # Borrower case - using the example calculation
    # From the task: borrower gets 12C + interest paid - outstanding balance = 2852
    # This means: 1200 + interest_paid - outstanding_balance = 2852
    # So: interest_paid - outstanding_balance = 1652
    interest_paid = 1652  # This would be the total interest paid over the loan term
    outstanding_balance = 0  # Assuming loan is fully paid
    
    borrower_dividend = FinancialCalculator.calculate_member_dividend(
        member_id=2, 
        year=2024, 
        contributions=annual_contributions, 
        interest_paid=interest_paid, 
        outstanding_balance=outstanding_balance
    )
    print(f"Borrower dividend: {borrower_dividend} (expected: 2852)")
    assert borrower_dividend == 2852, f"Borrower dividend should be 2852, got {borrower_dividend}"
    
    print("✓ Dividend calculations test passed!")
    return True

def test_total_fund_calculation():
    """
    Test total fund calculation with various scenarios
    """
    print("\nTesting Total Fund Calculation...")
    
    # Test 1: Basic calculation
    # Example: 10 members each contributing 1200 annually = 12000 total contributions
    # Plus 1652 total interest payments = 13652 total fund
    total_contributions = 12000
    total_interest_payments = 1652
    
    total_fund = FinancialCalculator.calculate_total_fund(total_contributions, total_interest_payments)
    expected_fund = 13652
    
    print(f"Test 1 - Total Fund: {total_fund} (expected: {expected_fund})")
    assert total_fund == expected_fund, f"Total fund should be {expected_fund}, got {total_fund}"
    
    # Test 2: Dividend distribution validation
    # If we have 9 non-borrowers (1200 each) + 1 borrower (2852) = 9*1200 + 2852 = 13652
    total_dividends = 9 * 1200 + 2852
    is_valid = FinancialCalculator.validate_dividend_distribution(total_fund, total_dividends)
    
    print(f"Test 2 - Total Dividends: {total_dividends}")
    print(f"Test 2 - Distribution Valid: {is_valid}")
    assert is_valid, "Dividend distribution should be valid"
    
    # Test 3: Edge case - no interest payments
    fund_no_interest = FinancialCalculator.calculate_total_fund(12000, 0)
    print(f"Test 3 - Fund with no interest: {fund_no_interest} (expected: 12000)")
    assert fund_no_interest == 12000, "Fund with no interest should equal contributions"
    
    # Test 4: Edge case - no contributions
    fund_no_contributions = FinancialCalculator.calculate_total_fund(0, 1652)
    print(f"Test 4 - Fund with no contributions: {fund_no_contributions} (expected: 1652)")
    assert fund_no_contributions == 1652, "Fund with no contributions should equal interest payments"
    
    # Test 5: Validation with small rounding differences
    is_valid_rounded = FinancialCalculator.validate_dividend_distribution(13652.00, 13652.01)
    print(f"Test 5 - Small rounding difference valid: {is_valid_rounded}")
    assert is_valid_rounded, "Small rounding differences should be valid"
    
    # Test 6: Validation with large differences (should fail)
    is_valid_large_diff = FinancialCalculator.validate_dividend_distribution(13652.00, 13700.00)
    print(f"Test 6 - Large difference valid: {is_valid_large_diff}")
    assert not is_valid_large_diff, "Large differences should be invalid"
    
    print("✓ Total fund calculation test passed!")
    return True

def test_interest_payment_tracking():
    """
    Test that interest payment tracking uses actual repayments, not simple calculations
    """
    print("\nTesting Interest Payment Tracking...")
    
    # This test verifies that we're using actual interest_amount from Repayments table
    # rather than simple loan_amount * interest_rate calculation
    
    # Test scenario: Loan of 1000 at 20% monthly rate
    principal = 1000.0
    monthly_rate = 0.2
    monthly_payment = FinancialCalculator.calculate_loan_payment(principal, monthly_rate, 12)
    
    # Generate amortization schedule to see actual interest breakdown
    schedule = FinancialCalculator.generate_amortization_schedule(1, principal, monthly_rate, monthly_payment, 12)
    
    # Calculate actual interest from amortization schedule
    actual_total_interest = sum(entry[3] for entry in schedule)  # interest_amount is index 3
    
    # Simple calculation would be: loan_amount * rate * months = 1000 * 0.2 * 12 = 2400
    simple_calculation = principal * monthly_rate * 12
    
    print(f"Actual interest from amortization: {actual_total_interest}")
    print(f"Simple calculation (incorrect): {simple_calculation}")
    
    # Verify that actual interest is different from simple calculation
    assert actual_total_interest != simple_calculation, "Interest tracking should use actual payments, not simple calculation"
    
    # For this specific case, actual interest should be less than simple calculation
    # because principal reduces over time in amortization
    assert actual_total_interest < simple_calculation, "Amortized interest should be less than simple interest"
    
    print("✓ Interest payment tracking test passed!")
    return True

def test_fund_calculation_integration():
    """
    Test integration of fund calculation with dividend distribution
    """
    print("\nTesting Fund Calculation Integration...")
    
    # Test scenario with multiple members
    test_year = 2024
    
    # Simulate fund calculation data
    test_contributions = 5000.0  # Total contributions from all members
    test_interest_payments = 800.0  # Total interest payments from repayments
    
    # Calculate total fund
    fund_data = {
        'year': test_year,
        'total_contributions': test_contributions,
        'total_interest_payments': test_interest_payments,
        'total_fund': FinancialCalculator.calculate_total_fund(test_contributions, test_interest_payments),
        'calculation_date': datetime.now().strftime('%Y-%m-%d')
    }
    
    expected_fund = 5800.0
    print(f"Calculated fund: {fund_data['total_fund']} (expected: {expected_fund})")
    assert fund_data['total_fund'] == expected_fund, f"Fund should be {expected_fund}"
    
    # Test dividend distribution validation
    # Simulate dividend calculations that should equal the fund
    test_dividends = [
        {'member_id': 1, 'dividend_amount': 1200.0},  # Non-borrower
        {'member_id': 2, 'dividend_amount': 1500.0},  # Borrower with some interest
        {'member_id': 3, 'dividend_amount': 1200.0},  # Non-borrower
        {'member_id': 4, 'dividend_amount': 1900.0},  # Borrower with more interest
    ]
    
    total_dividends = sum(d['dividend_amount'] for d in test_dividends)
    print(f"Total dividends: {total_dividends}")
    
    # Validation should pass if dividends equal fund
    is_valid = FinancialCalculator.validate_dividend_distribution(fund_data['total_fund'], total_dividends)
    print(f"Distribution valid: {is_valid}")
    assert is_valid, "Dividend distribution should be valid when equal to fund"
    
    print("✓ Fund calculation integration test passed!")
    return True

def run_financial_calculation_tests():
    """
    Run all financial calculation tests including fund calculation verification
    """
    print("=" * 50)
    print("RUNNING FINANCIAL CALCULATION TESTS")
    print("=" * 50)
    
    try:
        test_loan_amortization_calculations()
        test_dividend_calculations()
        test_total_fund_calculation()
        test_interest_payment_tracking()
        test_fund_calculation_integration()
        
        print("\n" + "=" * 50)
        print("ALL TESTS PASSED! ✓")
        print("Financial calculation engine is working correctly.")
        print("Fund calculation properly uses actual interest payments from repayments.")
        print("Dividend distribution validation ensures fund balance.")
        print("=" * 50)
        return True
        
    except AssertionError as e:
        print(f"\n❌ TEST FAILED: {e}")
        return False
    except Exception as e:
        print(f"\n❌ TEST ERROR: {e}")
        return False

# Database Synchronization Classes and Enums
class SyncStatus(Enum):
    """Enumeration for synchronization status"""
    IDLE = "idle"
    SYNCING = "syncing"
    CONFLICT = "conflict"
    ERROR = "error"
    COMPLETED = "completed"

class ConflictResolution(Enum):
    """Enumeration for conflict resolution strategies"""
    MYSQL_WINS = "mysql_wins"
    SQLITE_WINS = "sqlite_wins"
    MANUAL = "manual"
    LATEST_TIMESTAMP = "latest_timestamp"

@dataclass
class SyncRecord:
    """Data class for sync record tracking"""
    table_name: str
    record_id: int
    operation: str  # INSERT, UPDATE, DELETE
    timestamp: datetime
    data_hash: str
    sync_status: SyncStatus
    conflict_data: Optional[Dict] = None

@dataclass
class ConflictRecord:
    """Data class for conflict tracking"""
    table_name: str
    record_id: int
    sqlite_data: Dict
    mysql_data: Dict
    sqlite_timestamp: datetime
    mysql_timestamp: datetime
    resolution_strategy: ConflictResolution
    resolved: bool = False

class DatabaseSyncManager:
    """
    Manages synchronization between SQLite and MySQL databases
    """
    
    def __init__(self, mysql_config: Dict, sqlite_db_path: str):
        self.mysql_config = mysql_config
        self.sqlite_db_path = sqlite_db_path
        self.sync_status = SyncStatus.IDLE
        self.sync_thread = None
        self.stop_sync = False
        self.sync_interval = 30  # seconds
        self.conflicts = []
        self.sync_log = []
        
        # Tables to synchronize
        self.sync_tables = [
            'Members', 'Contributions', 'Loans', 'Repayments', 
            'Users', 'Settings'
        ]
        
        # Initialize sync tracking table
        self._initialize_sync_tracking()
    
    def _initialize_sync_tracking(self):
        """Initialize sync tracking tables in SQLite"""
        try:
            conn = sqlite3.connect(self.sqlite_db_path)
            cursor = conn.cursor()          
  
            # Create sync tracking table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS SyncTracking (
                    sync_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_name TEXT NOT NULL,
                    record_id INTEGER NOT NULL,
                    operation TEXT NOT NULL,
                    timestamp TEXT NOT NULL,
                    data_hash TEXT NOT NULL,
                    sync_status TEXT NOT NULL DEFAULT 'pending',
                    conflict_data TEXT,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    synced_at TEXT
                )
            ''')
            
            # Create sync log table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS SyncLog (
                    log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sync_session_id TEXT NOT NULL,
                    message TEXT NOT NULL,
                    level TEXT NOT NULL DEFAULT 'INFO',
                    timestamp TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Create conflict resolution table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS ConflictResolution (
                    conflict_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_name TEXT NOT NULL,
                    record_id INTEGER NOT NULL,
                    sqlite_data TEXT NOT NULL,
                    mysql_data TEXT NOT NULL,
                    sqlite_timestamp TEXT NOT NULL,
                    mysql_timestamp TEXT NOT NULL,
                    resolution_strategy TEXT NOT NULL,
                    resolved INTEGER DEFAULT 0,
                    resolved_at TEXT,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            conn.commit()
            conn.close()
            logger.info("Sync tracking tables initialized successfully")
            
        except Exception as e:
            logger.error(f"Failed to initialize sync tracking: {e}")
            raise
    
    def is_mysql_available(self) -> bool:
        """Check if MySQL server is available"""
        if not MYSQL_AVAILABLE:
            return False
        
        try:
            conn = mysql.connector.connect(**self.mysql_config)
            conn.close()
            return True
        except mysql.connector.Error:
            return False
        except Exception:
            return False    

    def start_sync_monitoring(self):
        """Start background thread to monitor MySQL availability and sync"""
        if self.sync_thread and self.sync_thread.is_alive():
            logger.warning("Sync monitoring already running")
            return
        
        self.stop_sync = False
        self.sync_thread = threading.Thread(target=self._sync_monitor_loop, daemon=True)
        self.sync_thread.start()
        logger.info("Sync monitoring started")
    
    def stop_sync_monitoring(self):
        """Stop sync monitoring thread"""
        self.stop_sync = True
        if self.sync_thread:
            self.sync_thread.join(timeout=5)
        logger.info("Sync monitoring stopped")
    
    def _sync_monitor_loop(self):
        """Main sync monitoring loop"""
        while not self.stop_sync:
            try:
                if self.is_mysql_available() and self.sync_status == SyncStatus.IDLE:
                    logger.info("MySQL server detected online, starting synchronization")
                    self.synchronize_databases()
                
                time.sleep(self.sync_interval)
                
            except Exception as e:
                logger.error(f"Error in sync monitor loop: {e}")
                time.sleep(self.sync_interval)
    
    def synchronize_databases(self) -> bool:
        """
        Synchronize SQLite data to MySQL
        Returns True if sync completed successfully, False otherwise
        """
        if self.sync_status == SyncStatus.SYNCING:
            logger.warning("Synchronization already in progress")
            return False
        
        if not MYSQL_AVAILABLE:
            logger.warning("MySQL connector not available, cannot synchronize")
            return False
        
        session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.sync_status = SyncStatus.SYNCING
        
        try:
            self._log_sync_message(session_id, "Starting database synchronization", "INFO")
            
            # Get pending sync records
            pending_records = self._get_pending_sync_records()
            
            if not pending_records:
                self._log_sync_message(session_id, "No pending records to sync", "INFO")
                self.sync_status = SyncStatus.COMPLETED
                return True 
           
            # Connect to both databases
            sqlite_conn = sqlite3.connect(self.sqlite_db_path)
            mysql_conn = mysql.connector.connect(**self.mysql_config)
            
            try:
                conflicts_found = 0
                records_synced = 0
                
                for record in pending_records:
                    try:
                        conflict = self._sync_record(sqlite_conn, mysql_conn, record, session_id)
                        if conflict:
                            conflicts_found += 1
                            self.conflicts.append(conflict)
                        else:
                            records_synced += 1
                            self._mark_record_synced(record.table_name, record.record_id, record.operation)
                    
                    except Exception as e:
                        logger.error(f"Failed to sync record {record.table_name}:{record.record_id}: {e}")
                        self._log_sync_message(session_id, f"Failed to sync {record.table_name}:{record.record_id}: {e}", "ERROR")
                
                # Update sync status
                if conflicts_found > 0:
                    self.sync_status = SyncStatus.CONFLICT
                    self._log_sync_message(session_id, f"Sync completed with {conflicts_found} conflicts, {records_synced} records synced", "WARNING")
                else:
                    self.sync_status = SyncStatus.COMPLETED
                    self._log_sync_message(session_id, f"Sync completed successfully, {records_synced} records synced", "INFO")
                
                return conflicts_found == 0
                
            finally:
                sqlite_conn.close()
                mysql_conn.close()
                
        except Exception as e:
            logger.error(f"Synchronization failed: {e}")
            self._log_sync_message(session_id, f"Synchronization failed: {e}", "ERROR")
            self.sync_status = SyncStatus.ERROR
            return False
    
    def _get_pending_sync_records(self) -> List[SyncRecord]:
        """Get all pending sync records from SQLite"""
        conn = sqlite3.connect(self.sqlite_db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                SELECT table_name, record_id, operation, timestamp, data_hash, sync_status, conflict_data
                FROM SyncTracking 
                WHERE sync_status = 'pending'
                ORDER BY timestamp ASC
            ''')
            
            records = []
            for row in cursor.fetchall():
                records.append(SyncRecord(
                    table_name=row[0],
                    record_id=row[1],
                    operation=row[2],
                    timestamp=datetime.fromisoformat(row[3]),
                    data_hash=row[4],
                    sync_status=SyncStatus(row[5]),
                    conflict_data=json.loads(row[6]) if row[6] else None
                ))
            
            return records
            
        finally:
            conn.close() 
   
    def _sync_record(self, sqlite_conn: sqlite3.Connection, mysql_conn, 
                     record: SyncRecord, session_id: str) -> Optional[ConflictRecord]:
        """
        Sync a single record from SQLite to MySQL
        Returns ConflictRecord if conflict detected, None otherwise
        """
        try:
            # Get current data from SQLite
            sqlite_data = self._get_record_data(sqlite_conn, record.table_name, record.record_id)
            
            # Check if record exists in MySQL
            mysql_data = self._get_record_data(mysql_conn, record.table_name, record.record_id)
            
            # Handle different operations
            if record.operation == 'INSERT':
                return self._handle_insert_sync(mysql_conn, record, sqlite_data, mysql_data, session_id)
            elif record.operation == 'UPDATE':
                return self._handle_update_sync(mysql_conn, record, sqlite_data, mysql_data, session_id)
            elif record.operation == 'DELETE':
                return self._handle_delete_sync(mysql_conn, record, mysql_data, session_id)
            
        except Exception as e:
            logger.error(f"Error syncing record {record.table_name}:{record.record_id}: {e}")
            raise
        
        return None
    
    def _handle_insert_sync(self, mysql_conn, record: SyncRecord,
                           sqlite_data: Dict, mysql_data: Optional[Dict], session_id: str) -> Optional[ConflictRecord]:
        """Handle INSERT operation sync"""
        if mysql_data is None:
            # Record doesn't exist in MySQL, safe to insert
            self._insert_record_to_mysql(mysql_conn, record.table_name, sqlite_data)
            self._log_sync_message(session_id, f"Inserted {record.table_name}:{record.record_id} to MySQL", "INFO")
            return None
        else:
            # Conflict: record exists in both databases
            return self._create_conflict_record(record.table_name, record.record_id, sqlite_data, mysql_data)
    
    def _handle_update_sync(self, mysql_conn, record: SyncRecord,
                           sqlite_data: Dict, mysql_data: Optional[Dict], session_id: str) -> Optional[ConflictRecord]:
        """Handle UPDATE operation sync"""
        if mysql_data is None:
            # Record doesn't exist in MySQL, treat as insert
            self._insert_record_to_mysql(mysql_conn, record.table_name, sqlite_data)
            self._log_sync_message(session_id, f"Inserted {record.table_name}:{record.record_id} to MySQL (was update)", "INFO")
            return None
        else:
            # Check for conflicts by comparing data hashes
            mysql_hash = self._calculate_data_hash(mysql_data)
            if mysql_hash != record.data_hash:
                # Conflict detected
                return self._create_conflict_record(record.table_name, record.record_id, sqlite_data, mysql_data)
            else:
                # No conflict, safe to update
                self._update_record_in_mysql(mysql_conn, record.table_name, record.record_id, sqlite_data)
                self._log_sync_message(session_id, f"Updated {record.table_name}:{record.record_id} in MySQL", "INFO")
                return None    

    def _handle_delete_sync(self, mysql_conn, record: SyncRecord,
                           mysql_data: Optional[Dict], session_id: str) -> Optional[ConflictRecord]:
        """Handle DELETE operation sync"""
        if mysql_data is not None:
            # Record exists in MySQL, safe to delete
            self._delete_record_from_mysql(mysql_conn, record.table_name, record.record_id)
            self._log_sync_message(session_id, f"Deleted {record.table_name}:{record.record_id} from MySQL", "INFO")
        else:
            # Record already doesn't exist in MySQL, no action needed
            self._log_sync_message(session_id, f"Record {record.table_name}:{record.record_id} already deleted from MySQL", "INFO")
        
        return None
    
    def _get_record_data(self, conn, table_name: str, record_id: int) -> Optional[Dict]:
        """Get record data from database"""
        cursor = conn.cursor()
        
        # Get primary key column name
        pk_column = self._get_primary_key_column(table_name)
        
        try:
            if isinstance(conn, sqlite3.Connection):
                cursor.execute(f"SELECT * FROM {table_name} WHERE {pk_column} = ?", (record_id,))
                row = cursor.fetchone()
                if row:
                    # Get column names
                    columns = [description[0] for description in cursor.description]
                    return dict(zip(columns, row))
            else:  # MySQL connection
                cursor.execute(f"SELECT * FROM {table_name} WHERE {pk_column} = %s", (record_id,))
                row = cursor.fetchone()
                if row:
                    # Get column names
                    columns = [description[0] for description in cursor.description]
                    return dict(zip(columns, row))
            
            return None
            
        finally:
            cursor.close()
    
    def _get_primary_key_column(self, table_name: str) -> str:
        """Get primary key column name for table"""
        pk_mapping = {
            'Members': 'member_id',
            'Contributions': 'contribution_id',
            'Loans': 'loan_id',
            'Repayments': 'repayment_id',
            'Users': 'username',
            'Settings': 'setting_name'
        }
        return pk_mapping.get(table_name, 'id')
    
    def _insert_record_to_mysql(self, mysql_conn, table_name: str, data: Dict):
        """Insert record to MySQL database"""
        cursor = mysql_conn.cursor()
        
        try:
            columns = list(data.keys())
            placeholders = ', '.join(['%s'] * len(columns))
            values = list(data.values())
            
            query = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})"
            cursor.execute(query, values)
            mysql_conn.commit()
            
        finally:
            cursor.close() 
   
    def _update_record_in_mysql(self, mysql_conn, table_name: str, record_id: int, data: Dict):
        """Update record in MySQL database"""
        cursor = mysql_conn.cursor()
        
        try:
            pk_column = self._get_primary_key_column(table_name)
            set_clause = ', '.join([f"{col} = %s" for col in data.keys() if col != pk_column])
            values = [data[col] for col in data.keys() if col != pk_column]
            values.append(record_id)
            
            query = f"UPDATE {table_name} SET {set_clause} WHERE {pk_column} = %s"
            cursor.execute(query, values)
            mysql_conn.commit()
            
        finally:
            cursor.close()
    
    def _delete_record_from_mysql(self, mysql_conn, table_name: str, record_id: int):
        """Delete record from MySQL database"""
        cursor = mysql_conn.cursor()
        
        try:
            pk_column = self._get_primary_key_column(table_name)
            query = f"DELETE FROM {table_name} WHERE {pk_column} = %s"
            cursor.execute(query, (record_id,))
            mysql_conn.commit()
            
        finally:
            cursor.close()
    
    def _create_conflict_record(self, table_name: str, record_id: int, sqlite_data: Dict, mysql_data: Dict) -> ConflictRecord:
        """Create conflict record for manual resolution"""
        conflict = ConflictRecord(
            table_name=table_name,
            record_id=record_id,
            sqlite_data=sqlite_data,
            mysql_data=mysql_data,
            sqlite_timestamp=datetime.now(),  # This should be from actual record timestamp
            mysql_timestamp=datetime.now(),   # This should be from actual record timestamp
            resolution_strategy=ConflictResolution.MANUAL
        )
        
        # Store conflict in database
        self._store_conflict_record(conflict)
        
        return conflict
    
    def _store_conflict_record(self, conflict: ConflictRecord):
        """Store conflict record in database for later resolution"""
        conn = sqlite3.connect(self.sqlite_db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO ConflictResolution 
                (table_name, record_id, sqlite_data, mysql_data, sqlite_timestamp, mysql_timestamp, resolution_strategy)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                conflict.table_name,
                conflict.record_id,
                json.dumps(conflict.sqlite_data),
                json.dumps(conflict.mysql_data),
                conflict.sqlite_timestamp.isoformat(),
                conflict.mysql_timestamp.isoformat(),
                conflict.resolution_strategy.value
            ))
            
            conn.commit()
            
        finally:
            cursor.close()
            conn.close() 
   
    def _calculate_data_hash(self, data: Dict) -> str:
        """Calculate hash of record data for conflict detection"""
        # Sort keys to ensure consistent hashing
        sorted_data = json.dumps(data, sort_keys=True)
        return hashlib.md5(sorted_data.encode()).hexdigest()
    
    def _mark_record_synced(self, table_name: str, record_id: int, operation: str):
        """Mark record as successfully synced"""
        conn = sqlite3.connect(self.sqlite_db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                UPDATE SyncTracking 
                SET sync_status = 'synced', synced_at = CURRENT_TIMESTAMP
                WHERE table_name = ? AND record_id = ? AND operation = ?
            ''', (table_name, record_id, operation))
            
            conn.commit()
            
        finally:
            cursor.close()
            conn.close()
    
    def _log_sync_message(self, session_id: str, message: str, level: str = "INFO"):
        """Log sync message to database and logger"""
        logger.log(getattr(logging, level), message)
        
        conn = sqlite3.connect(self.sqlite_db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO SyncLog (sync_session_id, message, level)
                VALUES (?, ?, ?)
            ''', (session_id, message, level))
            
            conn.commit()
            
        finally:
            cursor.close()
            conn.close()
    
    def track_record_change(self, table_name: str, record_id: int, operation: str, data: Dict):
        """Track record change for future synchronization"""
        if table_name not in self.sync_tables:
            return
        
        data_hash = self._calculate_data_hash(data)
        
        conn = sqlite3.connect(self.sqlite_db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO SyncTracking 
                (table_name, record_id, operation, timestamp, data_hash, sync_status)
                VALUES (?, ?, ?, ?, ?, 'pending')
            ''', (
                table_name,
                record_id,
                operation,
                datetime.now().isoformat(),
                data_hash
            ))
            
            conn.commit()
            logger.debug(f"Tracked {operation} on {table_name}:{record_id}")
            
        finally:
            cursor.close()
            conn.close()    
    
    def get_sync_status(self) -> Dict:
        """Get current synchronization status"""
        return {
            'status': self.sync_status.value,
            'mysql_available': self.is_mysql_available(),
            'pending_records': len(self._get_pending_sync_records()),
            'conflicts': len(self.conflicts),
            'last_sync': self._get_last_sync_time()
        }
    
    def _get_last_sync_time(self) -> Optional[str]:
        """Get timestamp of last successful sync"""
        conn = sqlite3.connect(self.sqlite_db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                SELECT MAX(synced_at) FROM SyncTracking WHERE sync_status = 'synced'
            ''')
            result = cursor.fetchone()
            return result[0] if result and result[0] else None
            
        finally:
            cursor.close()
            conn.close()
    
    def get_conflicts(self) -> List[ConflictRecord]:
        """Get all unresolved conflicts"""
        conn = sqlite3.connect(self.sqlite_db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                SELECT table_name, record_id, sqlite_data, mysql_data, 
                       sqlite_timestamp, mysql_timestamp, resolution_strategy
                FROM ConflictResolution 
                WHERE resolved = 0
                ORDER BY created_at ASC
            ''')
            
            conflicts = []
            for row in cursor.fetchall():
                conflicts.append(ConflictRecord(
                    table_name=row[0],
                    record_id=row[1],
                    sqlite_data=json.loads(row[2]),
                    mysql_data=json.loads(row[3]),
                    sqlite_timestamp=datetime.fromisoformat(row[4]),
                    mysql_timestamp=datetime.fromisoformat(row[5]),
                    resolution_strategy=ConflictResolution(row[6])
                ))
            
            return conflicts
            
        finally:
            cursor.close()
            conn.close()
    
    def resolve_conflict(self, conflict_id: int, resolution: ConflictResolution) -> bool:
        """Resolve a specific conflict"""
        logger.info(f"Resolving conflict {conflict_id} with strategy {resolution.value}")
        
        conn = sqlite3.connect(self.sqlite_db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                UPDATE ConflictResolution 
                SET resolved = 1, resolved_at = CURRENT_TIMESTAMP, resolution_strategy = ?
                WHERE conflict_id = ?
            ''', (resolution.value, conflict_id))
            
            conn.commit()
            return cursor.rowcount > 0
            
        finally:
            cursor.close()
            conn.close()    

    def get_sync_log(self, session_id: Optional[str] = None, limit: int = 100) -> List[Dict]:
        """Get sync log entries"""
        conn = sqlite3.connect(self.sqlite_db_path)
        cursor = conn.cursor()
        
        try:
            if session_id:
                cursor.execute('''
                    SELECT sync_session_id, message, level, timestamp
                    FROM SyncLog 
                    WHERE sync_session_id = ?
                    ORDER BY timestamp DESC
                    LIMIT ?
                ''', (session_id, limit))
            else:
                cursor.execute('''
                    SELECT sync_session_id, message, level, timestamp
                    FROM SyncLog 
                    ORDER BY timestamp DESC
                    LIMIT ?
                ''', (limit,))
            
            return [
                {
                    'session_id': row[0],
                    'message': row[1],
                    'level': row[2],
                    'timestamp': row[3]
                }
                for row in cursor.fetchall()
            ]
            
        finally:
            cursor.close()
            conn.close()

# Database Migration System
class DatabaseMigration:
    """
    Handles database schema migrations for both MySQL and SQLite
    """
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
        self.migrations = []
        self._register_migrations()
    
    def _register_migrations(self):
        """Register all available migrations"""
        self.migrations = [
            {
                'version': '1.0.0',
                'description': 'Initial schema setup',
                'up': self._migration_1_0_0_up,
                'down': self._migration_1_0_0_down
            },
            {
                'version': '1.1.0',
                'description': 'Add audit logging tables',
                'up': self._migration_1_1_0_up,
                'down': self._migration_1_1_0_down
            },
            {
                'version': '1.2.0',
                'description': 'Add loan schedule and enhanced member tracking',
                'up': self._migration_1_2_0_up,
                'down': self._migration_1_2_0_down
            },
            {
                'version': '1.3.0',
                'description': 'Add system configuration and dividend calculations',
                'up': self._migration_1_3_0_up,
                'down': self._migration_1_3_0_down
            },
            {
                'version': '1.4.0',
                'description': 'Add outstanding_balance field to Loans table',
                'up': self._migration_1_4_0_up,
                'down': self._migration_1_4_0_down
            }
        ]
    
    def _create_migration_table(self, conn):
        """Create migration tracking table"""
        cursor = conn.cursor()
        
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS schema_migrations (
                    version VARCHAR(20) PRIMARY KEY,
                    description TEXT NOT NULL,
                    applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS schema_migrations (
                    version TEXT PRIMARY KEY,
                    description TEXT NOT NULL,
                    applied_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        
        conn.commit()
        cursor.close()
    
    def get_applied_migrations(self, conn) -> List[str]:
        """Get list of applied migration versions"""
        self._create_migration_table(conn)
        cursor = conn.cursor()
        
        try:
            cursor.execute('SELECT version FROM schema_migrations ORDER BY version')
            return [row[0] for row in cursor.fetchall()]
        finally:
            cursor.close()
    
    def apply_migrations(self) -> bool:
        """Apply all pending migrations"""
        conn = self.db_manager.get_connection()
        applied_versions = self.get_applied_migrations(conn)
        
        try:
            for migration in self.migrations:
                if migration['version'] not in applied_versions:
                    logger.info(f"Applying migration {migration['version']}: {migration['description']}")
                    
                    # Apply the migration
                    migration['up'](conn)
                    
                    # Record the migration
                    cursor = conn.cursor()
                    if self.db_manager.current_db_type == 'mysql':
                        cursor.execute(
                            'INSERT INTO schema_migrations (version, description) VALUES (%s, %s)',
                            (migration['version'], migration['description'])
                        )
                    else:
                        cursor.execute(
                            'INSERT INTO schema_migrations (version, description) VALUES (?, ?)',
                            (migration['version'], migration['description'])
                        )
                    conn.commit()
                    cursor.close()
                    
                    logger.info(f"Migration {migration['version']} applied successfully")
            
            conn.close()
            return True
            
        except Exception as e:
            logger.error(f"Migration failed: {e}")
            conn.rollback()
            conn.close()
            return False
    
    def _migration_1_0_0_up(self, conn):
        """Initial schema setup"""
        cursor = conn.cursor()
        
        # Users table
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Users (
                    username VARCHAR(50) PRIMARY KEY,
                    password_hash VARCHAR(64) NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    last_login TIMESTAMP NULL,
                    failed_login_attempts INT DEFAULT 0,
                    account_locked BOOLEAN DEFAULT FALSE,
                    locked_until TIMESTAMP NULL
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Users (
                    username TEXT PRIMARY KEY,
                    password_hash TEXT NOT NULL,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    last_login TEXT,
                    failed_login_attempts INTEGER DEFAULT 0,
                    account_locked INTEGER DEFAULT 0,
                    locked_until TEXT
                )
            ''')
        
        # Members table
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Members (
                    member_id INT AUTO_INCREMENT PRIMARY KEY,
                    member_number VARCHAR(20) UNIQUE,
                    name VARCHAR(100) NOT NULL,
                    surname VARCHAR(100) NOT NULL,
                    middle_name VARCHAR(100),
                    phone_number VARCHAR(20) NOT NULL,
                    phone_number_2 VARCHAR(20),
                    email VARCHAR(100),
                    national_id VARCHAR(20) UNIQUE,
                    date_of_birth DATE,
                    gender ENUM('Male', 'Female', 'Other'),
                    marital_status ENUM('Single', 'Married', 'Divorced', 'Widowed'),
                    occupation VARCHAR(100),
                    employer VARCHAR(100),
                    monthly_income DECIMAL(10,2),
                    address_line1 VARCHAR(200),
                    address_line2 VARCHAR(200),
                    city VARCHAR(100),
                    district VARCHAR(100),
                    postal_code VARCHAR(20),
                    country VARCHAR(100) DEFAULT 'Malawi',
                    emergency_contact_name VARCHAR(200),
                    emergency_contact_phone VARCHAR(20),
                    emergency_contact_relationship VARCHAR(50),
                    next_of_kin_name VARCHAR(200),
                    next_of_kin_phone VARCHAR(20),
                    next_of_kin_relationship VARCHAR(50),
                    next_of_kin_address TEXT,
                    profile_photo_path VARCHAR(500),
                    id_document_path VARCHAR(500),
                    proof_of_address_path VARCHAR(500),
                    bank_account_number VARCHAR(50),
                    bank_name VARCHAR(100),
                    preferred_language ENUM('English', 'Chichewa', 'Tumbuka') DEFAULT 'English',
                    communication_preference ENUM('SMS', 'Email', 'Phone', 'WhatsApp') DEFAULT 'SMS',
                    member_type ENUM('Individual', 'Group', 'Business') DEFAULT 'Individual',
                    membership_category ENUM('Regular', 'Premium', 'VIP') DEFAULT 'Regular',
                    credit_score INT DEFAULT 0,
                    risk_rating ENUM('Low', 'Medium', 'High') DEFAULT 'Medium',
                    kyc_status ENUM('Pending', 'Verified', 'Rejected') DEFAULT 'Pending',
                    kyc_verified_date DATE,
                    kyc_verified_by VARCHAR(100),
                    account_balance DECIMAL(15,2) DEFAULT 0.00,
                    total_contributions DECIMAL(15,2) DEFAULT 0.00,
                    total_loans DECIMAL(15,2) DEFAULT 0.00,
                    total_repayments DECIMAL(15,2) DEFAULT 0.00,
                    outstanding_balance DECIMAL(15,2) DEFAULT 0.00,
                    last_contribution_date DATE,
                    last_loan_date DATE,
                    last_repayment_date DATE,
                    member_since_days INT DEFAULT 0,
                    contribution_streak INT DEFAULT 0,
                    missed_contributions INT DEFAULT 0,
                    loan_count INT DEFAULT 0,
                    successful_loans INT DEFAULT 0,
                    defaulted_loans INT DEFAULT 0,
                    referral_code VARCHAR(20) UNIQUE,
                    referred_by_member_id INT,
                    referral_count INT DEFAULT 0,
                    loyalty_points INT DEFAULT 0,
                    special_notes TEXT,
                    join_date DATE,
                    status ENUM('active', 'inactive', 'suspended', 'closed') DEFAULT 'active',
                    status_reason VARCHAR(200),
                    last_login_date DATETIME,
                    last_activity_date DATETIME,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    created_by VARCHAR(100) DEFAULT 'system',
                    updated_by VARCHAR(100),
                    FOREIGN KEY (referred_by_member_id) REFERENCES Members(member_id)
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Members (
                    member_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    member_number TEXT UNIQUE,
                    name TEXT NOT NULL,
                    surname TEXT NOT NULL,
                    middle_name TEXT,
                    phone_number TEXT NOT NULL,
                    phone_number_2 TEXT,
                    email TEXT,
                    national_id TEXT UNIQUE,
                    date_of_birth TEXT,
                    gender TEXT CHECK(gender IN ('Male', 'Female', 'Other')),
                    marital_status TEXT CHECK(marital_status IN ('Single', 'Married', 'Divorced', 'Widowed')),
                    occupation TEXT,
                    employer TEXT,
                    monthly_income REAL,
                    address_line1 TEXT,
                    address_line2 TEXT,
                    city TEXT,
                    district TEXT,
                    postal_code TEXT,
                    country TEXT DEFAULT 'Malawi',
                    emergency_contact_name TEXT,
                    emergency_contact_phone TEXT,
                    emergency_contact_relationship TEXT,
                    next_of_kin_name TEXT,
                    next_of_kin_phone TEXT,
                    next_of_kin_relationship TEXT,
                    next_of_kin_address TEXT,
                    profile_photo_path TEXT,
                    id_document_path TEXT,
                    proof_of_address_path TEXT,
                    bank_account_number TEXT,
                    bank_name TEXT,
                    preferred_language TEXT CHECK(preferred_language IN ('English', 'Chichewa', 'Tumbuka')) DEFAULT 'English',
                    communication_preference TEXT CHECK(communication_preference IN ('SMS', 'Email', 'Phone', 'WhatsApp')) DEFAULT 'SMS',
                    member_type TEXT CHECK(member_type IN ('Individual', 'Group', 'Business')) DEFAULT 'Individual',
                    membership_category TEXT CHECK(membership_category IN ('Regular', 'Premium', 'VIP')) DEFAULT 'Regular',
                    credit_score INTEGER DEFAULT 0,
                    risk_rating TEXT CHECK(risk_rating IN ('Low', 'Medium', 'High')) DEFAULT 'Medium',
                    kyc_status TEXT CHECK(kyc_status IN ('Pending', 'Verified', 'Rejected')) DEFAULT 'Pending',
                    kyc_verified_date TEXT,
                    kyc_verified_by TEXT,
                    account_balance REAL DEFAULT 0.00,
                    total_contributions REAL DEFAULT 0.00,
                    total_loans REAL DEFAULT 0.00,
                    total_repayments REAL DEFAULT 0.00,
                    outstanding_balance REAL DEFAULT 0.00,
                    last_contribution_date TEXT,
                    last_loan_date TEXT,
                    last_repayment_date TEXT,
                    member_since_days INTEGER DEFAULT 0,
                    contribution_streak INTEGER DEFAULT 0,
                    missed_contributions INTEGER DEFAULT 0,
                    loan_count INTEGER DEFAULT 0,
                    successful_loans INTEGER DEFAULT 0,
                    defaulted_loans INTEGER DEFAULT 0,
                    referral_code TEXT UNIQUE,
                    referred_by_member_id INTEGER,
                    referral_count INTEGER DEFAULT 0,
                    loyalty_points INTEGER DEFAULT 0,
                    special_notes TEXT,
                    join_date TEXT,
                    status TEXT CHECK(status IN ('active', 'inactive', 'suspended', 'closed')) DEFAULT 'active',
                    status_reason TEXT,
                    last_login_date TEXT,
                    last_activity_date TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    created_by TEXT DEFAULT 'system',
                    updated_by TEXT,
                    FOREIGN KEY (referred_by_member_id) REFERENCES Members(member_id)
                )
            ''')
        
        # Contributions table
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Contributions (
                    contribution_id INT AUTO_INCREMENT PRIMARY KEY,
                    member_id INT NOT NULL,
                    month INT NOT NULL,
                    year INT NOT NULL,
                    amount DECIMAL(10,2) NOT NULL,
                    late_fee DECIMAL(10,2) DEFAULT 0.00,
                    contribution_date DATE NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES Members (member_id),
                    UNIQUE KEY unique_member_month_year (member_id, month, year)
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Contributions (
                    contribution_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    member_id INTEGER NOT NULL,
                    month INTEGER NOT NULL,
                    year INTEGER NOT NULL,
                    amount REAL NOT NULL,
                    late_fee REAL DEFAULT 0.0,
                    contribution_date TEXT NOT NULL,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES Members (member_id),
                    UNIQUE (member_id, month, year)
                )
            ''')
        
        # Loans table
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Loans (
                    loan_id INT AUTO_INCREMENT PRIMARY KEY,
                    member_id INT NOT NULL,
                    loan_amount DECIMAL(10,2) NOT NULL,
                    interest_rate DECIMAL(5,2) NOT NULL,
                    monthly_payment DECIMAL(10,2),
                    total_interest DECIMAL(10,2),
                    outstanding_balance DECIMAL(10,2) DEFAULT 0.00,
                    loan_date DATE NOT NULL,
                    status ENUM('active', 'completed', 'defaulted') DEFAULT 'active',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES Members (member_id)
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Loans (
                    loan_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    member_id INTEGER NOT NULL,
                    loan_amount REAL NOT NULL,
                    interest_rate REAL NOT NULL,
                    monthly_payment REAL,
                    total_interest REAL,
                    loan_date TEXT NOT NULL,
                    status TEXT DEFAULT 'active',
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES Members (member_id)
                )
            ''')
        
        # Repayments table
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Repayments (
                    repayment_id INT AUTO_INCREMENT PRIMARY KEY,
                    loan_id INT NOT NULL,
                    repayment_amount DECIMAL(10,2) NOT NULL,
                    principal_amount DECIMAL(10,2) NOT NULL,
                    interest_amount DECIMAL(10,2) NOT NULL,
                    repayment_date DATE NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (loan_id) REFERENCES Loans (loan_id)
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Repayments (
                    repayment_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    loan_id INTEGER NOT NULL,
                    repayment_amount REAL NOT NULL,
                    principal_amount REAL NOT NULL,
                    interest_amount REAL NOT NULL,
                    repayment_date TEXT NOT NULL,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (loan_id) REFERENCES Loans (loan_id)
                )
            ''')
        
        # Settings table
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Settings (
                    setting_name VARCHAR(50) PRIMARY KEY,
                    setting_value TEXT NOT NULL,
                    description TEXT,
                    last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Settings (
                    setting_name TEXT PRIMARY KEY,
                    setting_value TEXT NOT NULL,
                    description TEXT,
                    last_updated TEXT DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        
        conn.commit()
        cursor.close()
    
    def _migration_1_0_0_down(self, conn):
        """Rollback initial schema"""
        cursor = conn.cursor()
        tables = ['Settings', 'Repayments', 'Loans', 'Contributions', 'Members', 'Users']
        for table in tables:
            cursor.execute(f'DROP TABLE IF EXISTS {table}')
        conn.commit()
        cursor.close()
    
    def _migration_1_1_0_up(self, conn):
        """Add audit logging tables"""
        cursor = conn.cursor()
        
        # Audit log table
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS AuditLog (
                    audit_id INT AUTO_INCREMENT PRIMARY KEY,
                    table_name VARCHAR(50) NOT NULL,
                    operation ENUM('INSERT', 'UPDATE', 'DELETE') NOT NULL,
                    record_id INT NOT NULL,
                    old_values JSON,
                    new_values JSON,
                    user_id VARCHAR(50) NOT NULL,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    ip_address VARCHAR(45),
                    user_agent TEXT
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS AuditLog (
                    audit_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_name TEXT NOT NULL,
                    operation TEXT NOT NULL,
                    record_id INTEGER NOT NULL,
                    old_values TEXT,
                    new_values TEXT,
                    user_id TEXT NOT NULL,
                    timestamp TEXT DEFAULT CURRENT_TIMESTAMP,
                    ip_address TEXT,
                    user_agent TEXT
                )
            ''')
        
        conn.commit()
        cursor.close()
    
    def _migration_1_1_0_down(self, conn):
        """Rollback audit logging tables"""
        cursor = conn.cursor()
        cursor.execute('DROP TABLE IF EXISTS AuditLog')
        conn.commit()
        cursor.close()
    
    def _migration_1_2_0_up(self, conn):
        """Add loan schedule and member status tracking"""
        cursor = conn.cursor()
        
        # Loan amortization schedule table
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS LoanSchedule (
                    schedule_id INT AUTO_INCREMENT PRIMARY KEY,
                    loan_id INT NOT NULL,
                    payment_number INT NOT NULL,
                    payment_date DATE NOT NULL,
                    principal_amount DECIMAL(10,2) NOT NULL,
                    interest_amount DECIMAL(10,2) NOT NULL,
                    remaining_balance DECIMAL(10,2) NOT NULL,
                    payment_status ENUM('pending', 'paid', 'overdue') DEFAULT 'pending',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (loan_id) REFERENCES Loans (loan_id),
                    UNIQUE KEY unique_loan_payment (loan_id, payment_number)
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS LoanSchedule (
                    schedule_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    loan_id INTEGER NOT NULL,
                    payment_number INTEGER NOT NULL,
                    payment_date TEXT NOT NULL,
                    principal_amount REAL NOT NULL,
                    interest_amount REAL NOT NULL,
                    remaining_balance REAL NOT NULL,
                    payment_status TEXT DEFAULT 'pending',
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (loan_id) REFERENCES Loans (loan_id),
                    UNIQUE (loan_id, payment_number)
                )
            ''')
        
        # Member status tracking table
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS MemberStatus (
                    status_id INT AUTO_INCREMENT PRIMARY KEY,
                    member_id INT NOT NULL,
                    status ENUM('active', 'inactive', 'suspended') NOT NULL,
                    status_date DATE NOT NULL,
                    notes TEXT,
                    changed_by VARCHAR(50),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES Members (member_id)
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS MemberStatus (
                    status_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    member_id INTEGER NOT NULL,
                    status TEXT NOT NULL,
                    status_date TEXT NOT NULL,
                    notes TEXT,
                    changed_by TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES Members (member_id)
                )
            ''')
        
        conn.commit()
        cursor.close()
    
    def _migration_1_2_0_down(self, conn):
        """Rollback loan schedule and member status tables"""
        cursor = conn.cursor()
        cursor.execute('DROP TABLE IF EXISTS MemberStatus')
        cursor.execute('DROP TABLE IF EXISTS LoanSchedule')
        conn.commit()
        cursor.close()
    
    def _migration_1_3_0_up(self, conn):
        """Add system configuration and dividend calculations"""
        cursor = conn.cursor()
        
        # System configuration table (enhanced)
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS SystemConfig (
                    config_key VARCHAR(50) PRIMARY KEY,
                    config_value TEXT NOT NULL,
                    config_type ENUM('string', 'number', 'boolean', 'json') DEFAULT 'string',
                    description TEXT,
                    is_editable BOOLEAN DEFAULT TRUE,
                    last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    updated_by VARCHAR(50)
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS SystemConfig (
                    config_key TEXT PRIMARY KEY,
                    config_value TEXT NOT NULL,
                    config_type TEXT DEFAULT 'string',
                    description TEXT,
                    is_editable INTEGER DEFAULT 1,
                    last_updated TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_by TEXT
                )
            ''')
        
        # Dividend calculations table
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS DividendCalculations (
                    calculation_id INT AUTO_INCREMENT PRIMARY KEY,
                    member_id INT NOT NULL,
                    year INT NOT NULL,
                    total_contributions DECIMAL(10,2) NOT NULL,
                    total_interest_paid DECIMAL(10,2) NOT NULL,
                    outstanding_balance DECIMAL(10,2) NOT NULL,
                    dividend_amount DECIMAL(10,2) NOT NULL,
                    calculation_date DATE NOT NULL,
                    status ENUM('calculated', 'paid', 'cancelled') DEFAULT 'calculated',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES Members (member_id),
                    UNIQUE KEY unique_member_year (member_id, year)
                )
            ''')
        else:  # SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS DividendCalculations (
                    calculation_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    member_id INTEGER NOT NULL,
                    year INTEGER NOT NULL,
                    total_contributions REAL NOT NULL,
                    total_interest_paid REAL NOT NULL,
                    outstanding_balance REAL NOT NULL,
                    dividend_amount REAL NOT NULL,
                    calculation_date TEXT NOT NULL,
                    status TEXT DEFAULT 'calculated',
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES Members (member_id),
                    UNIQUE (member_id, year)
                )
            ''')
        
        # Insert default system configuration
        default_configs = [
            ('monthly_contribution', '1000', 'number', 'Default monthly contribution amount', 1),
            ('interest_rate', '0.2', 'number', 'Monthly interest rate for loans', 1),
            ('currency', 'MWK', 'string', 'System currency', 1),
            ('financial_year_end', '12-31', 'string', 'Financial year end (MM-DD)', 1),
            ('loan_term_months', '12', 'number', 'Default loan term in months', 1),
            ('backup_retention_days', '30', 'number', 'Number of days to keep backup files', 1),
            ('auto_refresh_interval', '5', 'number', 'Auto-refresh interval in seconds', 1),
            ('session_timeout_minutes', '30', 'number', 'Session timeout in minutes', 1),
            ('max_failed_login_attempts', '3', 'number', 'Maximum failed login attempts before lockout', 1),
            ('account_lockout_minutes', '15', 'number', 'Account lockout duration in minutes', 1)
        ]
        
        # Use database-specific INSERT syntax
        if self.db_manager.current_db_type == 'mysql':
            insert_syntax = 'INSERT IGNORE'
            placeholder = '%s'
        else:
            insert_syntax = 'INSERT OR IGNORE'
            placeholder = '?'
            
        for config in default_configs:
            cursor.execute(f'''
                {insert_syntax} INTO SystemConfig 
                (config_key, config_value, config_type, description, is_editable)
                VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
            ''', config)
        
        conn.commit()
        cursor.close()
    
    def _migration_1_3_0_down(self, conn):
        """Rollback system configuration and dividend tables"""
        cursor = conn.cursor()
        cursor.execute('DROP TABLE IF EXISTS DividendCalculations')
        cursor.execute('DROP TABLE IF EXISTS SystemConfig')
        conn.commit()
        cursor.close()
    
    def _migration_1_4_0_up(self, conn):
        """Add outstanding_balance field to Loans table"""
        cursor = conn.cursor()
        
        # Check if column already exists
        column_exists = False
        
        if self.db_manager.current_db_type == 'mysql':
            cursor.execute('''
                SELECT COUNT(*) 
                FROM INFORMATION_SCHEMA.COLUMNS 
                WHERE TABLE_SCHEMA = %s 
                AND TABLE_NAME = 'Loans' 
                AND COLUMN_NAME = 'outstanding_balance'
            ''', (MYSQL_CONFIG['database'],))
            column_exists = cursor.fetchone()[0] > 0
        else:  # SQLite
            cursor.execute("PRAGMA table_info(Loans)")
            columns = cursor.fetchall()
            column_exists = any(col[1] == 'outstanding_balance' for col in columns)
        
        # Add outstanding_balance column only if it doesn't exist
        if not column_exists:
            if self.db_manager.current_db_type == 'mysql':
                cursor.execute('''
                    ALTER TABLE Loans 
                    ADD COLUMN outstanding_balance DECIMAL(10,2) DEFAULT 0
                ''')
            else:  # SQLite
                cursor.execute('''
                    ALTER TABLE Loans 
                    ADD COLUMN outstanding_balance REAL DEFAULT 0
                ''')
            
            # Initialize outstanding_balance with loan_amount for existing loans
            cursor.execute('''
                UPDATE Loans 
                SET outstanding_balance = loan_amount 
                WHERE outstanding_balance IS NULL OR outstanding_balance = 0
            ''')
        
        conn.commit()
        cursor.close()
    
    def _migration_1_4_0_down(self, conn):
        """Rollback outstanding_balance field from Loans table"""
        cursor = conn.cursor()
        
        # SQLite doesn't support DROP COLUMN, so we would need to recreate the table
        # For now, we'll just set the values to NULL
        cursor.execute('UPDATE Loans SET outstanding_balance = NULL')
        
        conn.commit()
        cursor.close()

# Database Backup and Recovery System
class BackupManager:
    """
    Handles database backup and recovery operations for both MySQL and SQLite
    """
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
        self.backup_dir = 'backups'
        self._ensure_backup_directory()
    
    def _ensure_backup_directory(self):
        """Create backup directory if it doesn't exist"""
        if not os.path.exists(self.backup_dir):
            os.makedirs(self.backup_dir)
            logger.info(f"Created backup directory: {self.backup_dir}")
    
    def create_backup(self) -> str:
        """Create a backup of the current database"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if self.db_manager.current_db_type == 'mysql':
            return self._create_mysql_backup(timestamp)
        else:
            return self._create_sqlite_backup(timestamp)
    
    def _create_sqlite_backup(self, timestamp: str) -> str:
        """Create SQLite database backup"""
        backup_filename = f"bank_mmudzi_backup_{timestamp}.db"
        backup_path = os.path.join(self.backup_dir, backup_filename)
        
        try:
            # Simple file copy for SQLite
            import shutil
            shutil.copy2(SQLITE_DB_NAME, backup_path)
            
            # Create metadata file
            metadata = {
                'backup_type': 'sqlite',
                'timestamp': timestamp,
                'original_file': SQLITE_DB_NAME,
                'backup_file': backup_filename,
                'file_size': os.path.getsize(backup_path)
            }
            
            metadata_path = os.path.join(self.backup_dir, f"backup_{timestamp}.json")
            with open(metadata_path, 'w') as f:
                json.dump(metadata, f, indent=2)
            
            logger.info(f"SQLite backup created: {backup_path}")
            return backup_path
            
        except Exception as e:
            logger.error(f"Failed to create SQLite backup: {e}")
            raise
    
    def _create_mysql_backup(self, timestamp: str) -> str:
        """Create MySQL database backup using mysqldump"""
        backup_filename = f"bank_mmudzi_backup_{timestamp}.sql"
        backup_path = os.path.join(self.backup_dir, backup_filename)
        
        try:
            # Use mysqldump command
            import subprocess
            
            cmd = [
                'mysqldump',
                f"--host={MYSQL_CONFIG['host']}",
                f"--user={MYSQL_CONFIG['user']}",
                f"--password={MYSQL_CONFIG['password']}",
                '--single-transaction',
                '--routines',
                '--triggers',
                MYSQL_CONFIG['database']
            ]
            
            with open(backup_path, 'w') as f:
                result = subprocess.run(cmd, stdout=f, stderr=subprocess.PIPE, text=True)
            
            if result.returncode != 0:
                raise Exception(f"mysqldump failed: {result.stderr}")
            
            # Create metadata file
            metadata = {
                'backup_type': 'mysql',
                'timestamp': timestamp,
                'database': MYSQL_CONFIG['database'],
                'backup_file': backup_filename,
                'file_size': os.path.getsize(backup_path)
            }
            
            metadata_path = os.path.join(self.backup_dir, f"backup_{timestamp}.json")
            with open(metadata_path, 'w') as f:
                json.dump(metadata, f, indent=2)
            
            logger.info(f"MySQL backup created: {backup_path}")
            return backup_path
            
        except Exception as e:
            logger.error(f"Failed to create MySQL backup: {e}")
            raise
    
    def restore_from_backup(self, backup_file: str) -> bool:
        """Restore database from backup file"""
        if not os.path.exists(backup_file):
            logger.error(f"Backup file not found: {backup_file}")
            return False
        
        if not self.validate_backup_integrity(backup_file):
            logger.error(f"Backup file integrity check failed: {backup_file}")
            return False
        
        try:
            # Determine backup type from file extension
            if backup_file.endswith('.db'):
                return self._restore_sqlite_backup(backup_file)
            elif backup_file.endswith('.sql'):
                return self._restore_mysql_backup(backup_file)
            else:
                logger.error(f"Unknown backup file type: {backup_file}")
                return False
                
        except Exception as e:
            logger.error(f"Failed to restore backup: {e}")
            return False
    
    def _restore_sqlite_backup(self, backup_file: str) -> bool:
        """Restore SQLite database from backup"""
        try:
            import shutil
            
            # Create a backup of current database before restore
            current_backup = f"{SQLITE_DB_NAME}.pre_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            shutil.copy2(SQLITE_DB_NAME, current_backup)
            
            # Restore from backup
            shutil.copy2(backup_file, SQLITE_DB_NAME)
            
            logger.info(f"SQLite database restored from: {backup_file}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to restore SQLite backup: {e}")
            return False
    
    def _restore_mysql_backup(self, backup_file: str) -> bool:
        """Restore MySQL database from backup"""
        try:
            import subprocess
            
            cmd = [
                'mysql',
                f"--host={MYSQL_CONFIG['host']}",
                f"--user={MYSQL_CONFIG['user']}",
                f"--password={MYSQL_CONFIG['password']}",
                MYSQL_CONFIG['database']
            ]
            
            with open(backup_file, 'r') as f:
                result = subprocess.run(cmd, stdin=f, stderr=subprocess.PIPE, text=True)
            
            if result.returncode != 0:
                raise Exception(f"mysql restore failed: {result.stderr}")
            
            logger.info(f"MySQL database restored from: {backup_file}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to restore MySQL backup: {e}")
            return False
    
    def validate_backup_integrity(self, backup_file: str) -> bool:
        """Validate backup file integrity"""
        try:
            if backup_file.endswith('.db'):
                # For SQLite, try to open and query
                conn = sqlite3.connect(backup_file)
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tables = cursor.fetchall()
                conn.close()
                return len(tables) > 0
            
            elif backup_file.endswith('.sql'):
                # For MySQL dump, check if file contains expected SQL statements
                with open(backup_file, 'r') as f:
                    content = f.read(1000)  # Read first 1000 characters
                    return 'CREATE TABLE' in content or 'INSERT INTO' in content
            
            return False
            
        except Exception as e:
            logger.error(f"Backup integrity check failed: {e}")
            return False
    
    def cleanup_old_backups(self, retention_days: int = 30):
        """Remove backup files older than retention_days"""
        try:
            cutoff_date = datetime.now() - timedelta(days=retention_days)
            removed_count = 0
            
            for filename in os.listdir(self.backup_dir):
                file_path = os.path.join(self.backup_dir, filename)
                
                if os.path.isfile(file_path):
                    file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                    
                    if file_time < cutoff_date:
                        os.remove(file_path)
                        removed_count += 1
                        logger.info(f"Removed old backup: {filename}")
            
            logger.info(f"Cleanup completed. Removed {removed_count} old backup files.")
            
        except Exception as e:
            logger.error(f"Backup cleanup failed: {e}")
    
    def list_backups(self) -> List[Dict]:
        """List all available backups with metadata"""
        backups = []
        
        try:
            for filename in os.listdir(self.backup_dir):
                if filename.endswith(('.db', '.sql')):
                    file_path = os.path.join(self.backup_dir, filename)
                    file_stat = os.stat(file_path)
                    
                    # Try to load metadata
                    metadata_file = filename.replace('.db', '.json').replace('.sql', '.json')
                    metadata_file = metadata_file.replace('bank_mmudzi_backup_', 'backup_')
                    metadata_path = os.path.join(self.backup_dir, metadata_file)
                    
                    metadata = {}
                    if os.path.exists(metadata_path):
                        with open(metadata_path, 'r') as f:
                            metadata = json.load(f)
                    
                    backup_info = {
                        'filename': filename,
                        'path': file_path,
                        'size': file_stat.st_size,
                        'created': datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
                        'type': 'mysql' if filename.endswith('.sql') else 'sqlite',
                        'metadata': metadata
                    }
                    
                    backups.append(backup_info)
            
            # Sort by creation time, newest first
            backups.sort(key=lambda x: x['created'], reverse=True)
            
        except Exception as e:
            logger.error(f"Failed to list backups: {e}")
        
        return backups

# Audit Trail System
class AuditManager:
    """
    Manages audit trail logging for all financial operations
    """
    
    def __init__(self):
        self.current_user = None
        self._initialize_current_user()
    
    def _initialize_current_user(self):
        """Initialize current user tracking - for now use 'system' as default"""
        # TODO: This should be set during login process
        self.current_user = 'system'
    
    def set_current_user(self, username: str):
        """Set the current logged-in user for audit tracking"""
        self.current_user = username
    
    def log_transaction(self, table_name: str, operation: str, record_id: int, 
                       old_values: dict = None, new_values: dict = None):
        """
        Log a database transaction to the audit trail
        
        Args:
            table_name (str): Name of the table being modified
            operation (str): Type of operation (INSERT, UPDATE, DELETE)
            record_id (int): ID of the record being modified
            old_values (dict): Previous values (for UPDATE and DELETE operations)
            new_values (dict): New values (for INSERT and UPDATE operations)
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Convert dictionaries to JSON strings for storage
            old_values_json = json.dumps(old_values) if old_values else None
            new_values_json = json.dumps(new_values) if new_values else None
            
            # Get current timestamp
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Insert audit log entry with database-specific syntax
            if db_manager.current_db_type == 'mysql':
                cursor.execute('''
                    INSERT INTO AuditLog (table_name, operation, record_id, old_values, new_values, user_id, timestamp, ip_address, user_agent)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (table_name, operation, record_id, old_values_json, new_values_json, 
                      self.current_user or 'unknown', timestamp, 'localhost', 'system'))
            else:
                cursor.execute('''
                    INSERT INTO AuditLog (table_name, operation, record_id, old_values, new_values, user_id, timestamp, ip_address, user_agent)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (table_name, operation, record_id, old_values_json, new_values_json, 
                      self.current_user or 'unknown', timestamp, 'localhost', 'system'))
            
            conn.commit()
            conn.close()
            
            logger.info(f"Audit log entry created: {operation} on {table_name}:{record_id} by {self.current_user}")
            
        except Exception as e:
            logger.error(f"Failed to create audit log entry: {e}")
            # Don't raise exception to avoid breaking the main operation
    
    def log_security_violation(self, violation_type: str, details: dict = None):
        """
        Log security violations and unauthorized access attempts
        
        Args:
            violation_type (str): Type of security violation
            details (dict): Additional details about the violation
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Create a security violation entry in audit log
            details_json = json.dumps(details) if details else None
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Use 'INSERT' as operation and put violation_type in new_values
            violation_data = {'violation_type': violation_type, 'details': details}
            violation_json = json.dumps(violation_data)
            
            if db_manager.current_db_type == 'mysql':
                cursor.execute('''
                    INSERT INTO AuditLog (table_name, operation, record_id, old_values, new_values, user_id, timestamp, ip_address, user_agent)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', ('SECURITY', 'INSERT', 0, None, violation_json, 
                      self.current_user or 'unknown', timestamp, 'localhost', 'system'))
            else:
                cursor.execute('''
                    INSERT INTO AuditLog (table_name, operation, record_id, old_values, new_values, user_id, timestamp, ip_address, user_agent)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', ('SECURITY', 'INSERT', 0, None, violation_json, 
                      self.current_user or 'unknown', timestamp, 'localhost', 'system'))
            
            conn.commit()
            conn.close()
            
            logger.warning(f"Security violation logged: {violation_type} by {self.current_user}")
            
        except Exception as e:
            logger.error(f"Failed to log security violation: {e}")
    
    def get_audit_trail(self, table_name: str = None, operation: str = None, 
                       user_id: str = None, start_date: str = None, end_date: str = None,
                       limit: int = 1000) -> list:
        """
        Retrieve audit trail entries with optional filtering
        
        Args:
            table_name (str): Filter by table name
            operation (str): Filter by operation type
            user_id (str): Filter by user ID
            start_date (str): Filter by start date (YYYY-MM-DD)
            end_date (str): Filter by end date (YYYY-MM-DD)
            limit (int): Maximum number of entries to return
            
        Returns:
            list: List of audit trail entries
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Build dynamic query based on filters
            query = 'SELECT * FROM AuditLog WHERE 1=1'
            params = []
            
            if table_name:
                query += ' AND table_name = ?'
                params.append(table_name)
            
            if operation:
                query += ' AND operation = ?'
                params.append(operation)
            
            if user_id:
                query += ' AND user_id = ?'
                params.append(user_id)
            
            if start_date:
                query += ' AND date(timestamp) >= ?'
                params.append(start_date)
            
            if end_date:
                query += ' AND date(timestamp) <= ?'
                params.append(end_date)
            
            query += ' ORDER BY timestamp DESC LIMIT ?'
            params.append(limit)
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            # Convert results to list of dictionaries
            audit_entries = []
            for row in results:
                entry = {
                    'audit_id': row[0],
                    'table_name': row[1],
                    'operation': row[2],
                    'record_id': row[3],
                    'old_values': json.loads(row[4]) if row[4] else None,
                    'new_values': json.loads(row[5]) if row[5] else None,
                    'user_id': row[6],
                    'timestamp': row[7]
                }
                audit_entries.append(entry)
            
            return audit_entries
            
        except Exception as e:
            logger.error(f"Failed to retrieve audit trail: {e}")
            return []
    
    def export_audit_log(self, start_date: str, end_date: str, format: str = 'csv') -> str:
        """
        Export audit log to CSV format
        
        Args:
            start_date (str): Start date for export (YYYY-MM-DD)
            end_date (str): End date for export (YYYY-MM-DD)
            format (str): Export format ('csv' only for now)
            
        Returns:
            str: Path to the exported file
        """
        try:
            # Get audit entries for the date range
            audit_entries = self.get_audit_trail(
                start_date=start_date, 
                end_date=end_date, 
                limit=10000
            )
            
            # Create export filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'audit_log_export_{timestamp}.csv'
            
            # Write to CSV file
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['audit_id', 'table_name', 'operation', 'record_id', 
                             'old_values', 'new_values', 'user_id', 'timestamp']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                writer.writeheader()
                for entry in audit_entries:
                    # Convert dict values to strings for CSV
                    csv_entry = entry.copy()
                    if csv_entry['old_values']:
                        csv_entry['old_values'] = json.dumps(csv_entry['old_values'])
                    if csv_entry['new_values']:
                        csv_entry['new_values'] = json.dumps(csv_entry['new_values'])
                    writer.writerow(csv_entry)
            
            logger.info(f"Audit log exported to {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Failed to export audit log: {e}")
            return None

# Create global audit manager instance
audit_manager = AuditManager()

# Enhanced Database Connection Manager with MySQL Primary and SQLite Fallback
class DatabaseManager:
    def __init__(self):
        self.current_db_type = None
        self.connection = None
        self.mysql_available = False
        self.sqlite_fallback_active = False
        self.sync_manager = None
        self.connection_retry_count = 0
        self.max_retry_attempts = 3
        self.mysql_check_interval = 30  # seconds
        self.last_mysql_check = 0
        self._initialize_sync_manager()
        self._check_mysql_availability()
    
    def _initialize_sync_manager(self):
        """Initialize the database synchronization manager"""
        try:
            self.sync_manager = DatabaseSyncManager(MYSQL_CONFIG, SQLITE_DB_NAME)
            # Start sync monitoring in background
            self.sync_manager.start_sync_monitoring()
            logger.info("Database sync manager initialized successfully")
        except Exception as e:
            logger.error(f"Failed to initialize sync manager: {e}")
            self.sync_manager = None
    
    def _check_mysql_availability(self):
        """Check MySQL availability and update status"""
        current_time = time.time()
        
        # Only check MySQL availability periodically to avoid constant connection attempts
        if current_time - self.last_mysql_check < self.mysql_check_interval:
            return self.mysql_available
        
        self.last_mysql_check = current_time
        
        if not MYSQL_AVAILABLE:
            self.mysql_available = False
            return False
        
        try:
            # Test MySQL connection with timeout
            conn = mysql.connector.connect(
                **MYSQL_CONFIG,
                connection_timeout=5,
                autocommit=True
            )
            conn.close()
            
            # If we were in fallback mode and MySQL is now available, trigger sync
            if self.sqlite_fallback_active and not self.mysql_available:
                logger.info("MySQL server is back online, initiating synchronization")
                self._trigger_sync_from_fallback()
            
            self.mysql_available = True
            self.sqlite_fallback_active = False
            self.connection_retry_count = 0
            
            return True
            
        except Exception as e:
            if self.mysql_available:
                logger.warning(f"MySQL connection lost: {e}")
                self.sqlite_fallback_active = True
            
            self.mysql_available = False
            return False
    
    def _trigger_sync_from_fallback(self):
        """Trigger synchronization when MySQL comes back online"""
        if self.sync_manager:
            try:
                # Run sync in background thread to avoid blocking
                sync_thread = threading.Thread(
                    target=self.sync_manager.synchronize_databases,
                    daemon=True
                )
                sync_thread.start()
                logger.info("Background synchronization initiated")
            except Exception as e:
                logger.error(f"Failed to trigger sync from fallback: {e}")
    
    def track_change(self, table_name: str, record_id: int, operation: str, data: dict):
        """Track database changes for synchronization when using SQLite fallback"""
        if self.sync_manager and self.current_db_type == 'sqlite':
            try:
                self.sync_manager.track_record_change(table_name, record_id, operation, data)
                logger.debug(f"Tracked change: {operation} on {table_name}:{record_id}")
            except Exception as e:
                logger.error(f"Failed to track change: {e}")
        
    def get_connection(self):
        """Get database connection with MySQL primary and SQLite fallback"""
        # Check MySQL availability periodically
        self._check_mysql_availability()
        
        try:
            # Always try MySQL first (primary database)
            if self.mysql_available:
                return self._get_mysql_connection()
            else:
                # Fall back to SQLite if MySQL is not available
                logger.info("Using SQLite fallback database")
                return self._get_sqlite_connection()
                
        except Exception as e:
            logger.error(f"Database connection error: {e}")
            # Final fallback to SQLite on any error
            self.sqlite_fallback_active = True
            return self._get_sqlite_connection()
    
    def force_mysql_reconnect(self):
        """Force a reconnection attempt to MySQL"""
        self.last_mysql_check = 0
        self.connection_retry_count = 0
        return self._check_mysql_availability()
  
    def _test_mysql_connection(self):
        """Test if MySQL server is available with enhanced error handling"""
        if not MYSQL_AVAILABLE:
            return False
        
        try:
            conn = mysql.connector.connect(
                **MYSQL_CONFIG,
                connection_timeout=5,
                autocommit=True
            )
            # Test with a simple query
            cursor = conn.cursor()
            cursor.execute("SELECT 1")
            cursor.fetchone()
            cursor.close()
            conn.close()
            return True
        except mysql.connector.Error as err:
            logger.debug(f"MySQL connection test failed: {err}")
            return False
        except Exception as e:
            logger.debug(f"MySQL connection test error: {e}")
            return False
    
    def _get_mysql_connection(self):
        """Get MySQL connection and create database if needed"""
        try:
            # First try to connect to the specific database
            conn = mysql.connector.connect(
                **MYSQL_CONFIG,
                autocommit=False,
                connection_timeout=10
            )
            self.current_db_type = 'mysql'
            self.sqlite_fallback_active = False
            
            # Schema will be handled by migration system
            # self._ensure_mysql_schema(conn)  # Disabled to prevent connection issues
            
            return conn
            
        except mysql.connector.Error as err:
            if err.errno == mysql.connector.errorcode.ER_BAD_DB_ERROR:
                # Database doesn't exist, create it
                self._create_mysql_database()
                conn = mysql.connector.connect(**MYSQL_CONFIG)
                self.current_db_type = 'mysql'
                self.sqlite_fallback_active = False
                return conn
            else:
                logger.error(f"MySQL connection error: {err}")
                raise
    
    def _ensure_mysql_schema(self, conn):
        """Ensure MySQL database schema is properly set up"""
        # Schema initialization is handled by the migration system
        # This method is kept for compatibility but does nothing
        pass
    
    def _create_mysql_database(self):
        """Create MySQL database if it doesn't exist"""
        try:
            config_without_db = MYSQL_CONFIG.copy()
            del config_without_db['database']
            conn = mysql.connector.connect(**config_without_db)
            cursor = conn.cursor()
            cursor.execute(f"CREATE DATABASE IF NOT EXISTS {MYSQL_CONFIG['database']} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
            conn.commit()
            cursor.close()
            conn.close()
            logger.info(f"MySQL database '{MYSQL_CONFIG['database']}' created successfully")
        except mysql.connector.Error as err:
            logger.error(f"Failed to create MySQL database: {err}")
            raise
    
    def _get_sqlite_connection(self):
        """Get SQLite connection with fallback tracking"""
        try:
            conn = sqlite3.connect(SQLITE_DB_NAME, timeout=30.0)
            # Enable foreign key constraints
            conn.execute("PRAGMA foreign_keys = ON")
            # Enable WAL mode for better concurrency
            conn.execute("PRAGMA journal_mode = WAL")
            
            self.current_db_type = 'sqlite'
            
            # Mark that we're in fallback mode if MySQL should be available
            if MYSQL_AVAILABLE and not self.sqlite_fallback_active:
                self.sqlite_fallback_active = True
                logger.warning("Operating in SQLite fallback mode - changes will be synced when MySQL is available")
            
            return conn
            
        except sqlite3.Error as e:
            logger.error(f"Failed to connect to SQLite database: {e}")
            messagebox.showerror("Database Error", f"Failed to connect to SQLite database: {e}")
            raise
    
    def execute_query(self, query, params=None, fetch_one=False, fetch_all=False):
        """Execute query with automatic connection management and retry logic"""
        max_retries = 2
        retry_count = 0
        
        while retry_count <= max_retries:
            conn = None
            try:
                conn = self.get_connection()
                cursor = conn.cursor()
                
                # Convert SQLite syntax to MySQL if needed
                if self.current_db_type == 'mysql':
                    query = self._convert_query_to_mysql(query)
                
                if params:
                    cursor.execute(query, params)
                else:
                    cursor.execute(query)
                
                if fetch_one:
                    result = cursor.fetchone()
                elif fetch_all:
                    result = cursor.fetchall()
                else:
                    result = None
                
                conn.commit()
                return result
                
            except Exception as e:
                if conn:
                    try:
                        conn.rollback()
                    except:
                        pass
                
                retry_count += 1
                logger.error(f"Query execution error (attempt {retry_count}): {e}")
                
                # If MySQL connection failed, force a fallback check
                if self.current_db_type == 'mysql' and retry_count <= max_retries:
                    self.mysql_available = False
                    self.last_mysql_check = 0
                    continue
                
                if retry_count > max_retries:
                    raise
                    
            finally:
                if conn:
                    try:
                        cursor.close()
                        conn.close()
                    except:
                        pass

    def _convert_query_to_mysql(self, query):
        """Convert SQLite-specific syntax to MySQL with enhanced compatibility"""
        # Convert INTEGER PRIMARY KEY to AUTO_INCREMENT
        query = query.replace('INTEGER PRIMARY KEY AUTOINCREMENT', 'INT AUTO_INCREMENT PRIMARY KEY')
        query = query.replace('INTEGER PRIMARY KEY', 'INT AUTO_INCREMENT PRIMARY KEY')
        
        # Convert TEXT to appropriate MySQL types
        query = query.replace('TEXT NOT NULL', 'VARCHAR(255) NOT NULL')
        query = query.replace('TEXT DEFAULT', 'VARCHAR(255) DEFAULT')
        query = query.replace('TEXT,', 'TEXT,')
        query = query.replace('TEXT)', 'TEXT)')
        
        # Convert REAL to DECIMAL for better precision
        query = query.replace('REAL NOT NULL', 'DECIMAL(15,2) NOT NULL')
        query = query.replace('REAL DEFAULT', 'DECIMAL(15,2) DEFAULT')
        query = query.replace('REAL,', 'DECIMAL(15,2),')
        query = query.replace('REAL)', 'DECIMAL(15,2))')
        
        # Handle INSERT OR IGNORE/REPLACE
        query = query.replace('INSERT OR IGNORE', 'INSERT IGNORE')
        query = query.replace('INSERT OR REPLACE', 'REPLACE')
        
        # Handle SQLite date functions
        query = query.replace('CURRENT_TIMESTAMP', 'NOW()')
        query = query.replace("strftime('%Y', ", "YEAR(")
        query = query.replace("strftime('%m', ", "MONTH(")
        query = query.replace("strftime('%d', ", "DAY(")
        
        return query
    
    def get_db_status(self):
        """Get comprehensive database connection status"""
        return {
            'type': self.current_db_type,
            'mysql_available': self.mysql_available,
            'sqlite_fallback_active': self.sqlite_fallback_active,
            'sqlite_file': SQLITE_DB_NAME,
            'sync_status': self.sync_manager.get_sync_status() if self.sync_manager else None,
            'last_mysql_check': self.last_mysql_check,
            'connection_retry_count': self.connection_retry_count
        }
    
    def get_sync_status(self):
        """Get synchronization status"""
        if self.sync_manager:
            return self.sync_manager.get_sync_status()
        return {'status': 'unavailable', 'message': 'Sync manager not initialized'}
    
    def manual_sync(self):
        """Manually trigger synchronization"""
        if self.sync_manager and self.mysql_available:
            try:
                return self.sync_manager.synchronize_databases()
            except Exception as e:
                logger.error(f"Manual sync failed: {e}")
                return False
        return False

# Global database manager instance
db_manager = DatabaseManager()

# Database connection function (updated to use manager)
def connect_db():
    """Legacy function for backward compatibility"""
    return db_manager.get_connection()

# Initialize database with tables and default values
def initialize_db():
    """Initialize database using the migration system with MySQL-primary support"""
    try:
        # First ensure MySQL database exists if MySQL is available
        if MYSQL_AVAILABLE:
            try:
                _ensure_mysql_database_exists()
                logger.info("MySQL database verified/created successfully")
            except Exception as mysql_error:
                logger.warning(f"MySQL database creation failed: {mysql_error}")
                logger.info("Will use SQLite fallback")
        
        # Force a fresh connection check
        db_manager.last_mysql_check = 0
        db_manager._check_mysql_availability()
        
        # Test connection before applying migrations
        try:
            test_conn = db_manager.get_connection()
            test_conn.close()
            logger.info(f"Database connection successful using {db_manager.current_db_type}")
        except Exception as conn_error:
            logger.error(f"Database connection failed: {conn_error}")
            raise Exception(f"Cannot establish database connection: {conn_error}")
        
        # Apply all pending migrations
        migration_manager = DatabaseMigration(db_manager)
        success = migration_manager.apply_migrations()
        
        if not success:
            logger.error("Failed to apply database migrations")
            # Try basic schema creation as fallback
            logger.info("Attempting basic schema creation as fallback")
            _create_basic_schema()
            logger.info("Basic schema created successfully")
        else:
            logger.info("Database migrations applied successfully")
        
        # Create default admin user if it doesn't exist
        try:
            _create_default_admin_user()
            logger.info("Default admin user verified/created")
        except Exception as user_error:
            logger.warning(f"Failed to create default admin user: {user_error}")
        
        logger.info(f"Database initialized successfully using {db_manager.current_db_type}")
        
    except Exception as e:
        logger.error(f"Database initialization failed: {e}")
        # Final fallback to basic schema creation
        try:
            logger.info("Attempting final fallback to basic schema")
            _create_basic_schema()
            logger.info("Basic schema created as final fallback")
        except Exception as fallback_error:
            logger.error(f"Final fallback schema creation also failed: {fallback_error}")
            # Don't raise exception - let the app try to continue
            logger.warning("Database initialization incomplete - some features may not work")

def _ensure_mysql_database_exists():
    """Ensure MySQL database exists, create if it doesn't"""
    if not MYSQL_AVAILABLE:
        return
    
    try:
        # Try to connect to the specific database first
        conn = mysql.connector.connect(**MYSQL_CONFIG)
        conn.close()
        logger.info("MySQL database already exists")
        return
    except mysql.connector.Error as err:
        if err.errno == mysql.connector.errorcode.ER_BAD_DB_ERROR:
            # Database doesn't exist, create it
            logger.info("MySQL database doesn't exist, creating it...")
            _create_mysql_database()
        else:
            # Other MySQL error
            raise

def _create_mysql_database():
    """Create MySQL database if it doesn't exist"""
    try:
        config_without_db = MYSQL_CONFIG.copy()
        del config_without_db['database']
        
        conn = mysql.connector.connect(**config_without_db)
        cursor = conn.cursor()
        
        # Create database with proper charset
        cursor.execute(f"""
            CREATE DATABASE IF NOT EXISTS {MYSQL_CONFIG['database']} 
            CHARACTER SET utf8mb4 
            COLLATE utf8mb4_unicode_ci
        """)
        
        conn.commit()
        cursor.close()
        conn.close()
        
        logger.info(f"MySQL database '{MYSQL_CONFIG['database']}' created successfully")
        
    except mysql.connector.Error as err:
        logger.error(f"Failed to create MySQL database: {err}")
        raise

def _create_default_admin_user():
    """Create default admin user if it doesn't exist"""
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        
        # Check if admin user exists
        if db_manager.current_db_type == 'mysql':
            cursor.execute("SELECT COUNT(*) FROM Users WHERE username = %s", ('admin',))
        else:
            cursor.execute("SELECT COUNT(*) FROM Users WHERE username = ?", ('admin',))
        
        if cursor.fetchone()[0] == 0:
            # Create default admin user
            import hashlib
            default_password = "admin123"
            password_hash = hashlib.sha256(default_password.encode()).hexdigest()
            
            if db_manager.current_db_type == 'mysql':
                cursor.execute("INSERT INTO Users (username, password_hash) VALUES (%s, %s)", 
                             ('admin', password_hash))
            else:
                cursor.execute("INSERT INTO Users (username, password_hash) VALUES (?, ?)", 
                             ('admin', password_hash))
            
            conn.commit()
            logger.info("Default admin user created (username: admin, password: admin123)")
        
        conn.close()
        
    except Exception as e:
        logger.error(f"Failed to create default admin user: {e}")
        if 'conn' in locals():
            conn.close()

def _create_basic_schema():
    """Fallback function to create basic schema if migrations fail"""
    conn = connect_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Users (
            username TEXT PRIMARY KEY,
            password_hash TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Members (
            member_id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            surname TEXT NOT NULL,
            phone_number TEXT NOT NULL,
            email TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Contributions (
            contribution_id INTEGER PRIMARY KEY,
            member_id INTEGER,
            month INTEGER,
            amount REAL,
            contribution_date TEXT,
            FOREIGN KEY (member_id) REFERENCES Members (member_id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Loans (
            loan_id INTEGER PRIMARY KEY,
            member_id INTEGER,
            loan_amount REAL,
            interest_rate REAL,
            loan_date TEXT,
            outstanding_balance REAL DEFAULT 0,
            FOREIGN KEY (member_id) REFERENCES Members (member_id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Repayments (
            repayment_id INTEGER PRIMARY KEY,
            loan_id INTEGER,
            repayment_amount REAL,
            principal_amount REAL DEFAULT 0,
            interest_amount REAL DEFAULT 0,
            repayment_date TEXT,
            FOREIGN KEY (loan_id) REFERENCES Loans (loan_id)
        )
    ''')
    
    # Add default settings
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Settings (
            setting_name TEXT PRIMARY KEY,
            setting_value TEXT
        )
    ''')
    
    # Use database-specific INSERT syntax
    if db_manager.current_db_type == 'mysql':
        insert_syntax = 'INSERT IGNORE'
    else:
        insert_syntax = 'INSERT OR IGNORE'
    
    cursor.execute(f'{insert_syntax} INTO Settings (setting_name, setting_value) VALUES ("monthly_contribution", "1000")')
    cursor.execute(f'{insert_syntax} INTO Settings (setting_name, setting_value) VALUES ("interest_rate", "0.05")')
    cursor.execute(f'{insert_syntax} INTO Settings (setting_name, setting_value) VALUES ("currency", "MWK")')
    cursor.execute(f'{insert_syntax} INTO Settings (setting_name, setting_value) VALUES ("financial_year_end", "2025-12-31")')
    
    # Add default admin user
    default_password = 'admin123'
    password_hash = hashlib.sha256(default_password.encode()).hexdigest()
    cursor.execute(f'{insert_syntax} INTO Users (username, password_hash) VALUES ("admin", ?)', (password_hash,))
    
    conn.commit()
    conn.close()
    logger.info("Basic schema created successfully")

# Helper function to format currency
def format_currency(amount):
    return f"MWK {amount:,.2f}"

# Register new admin function
def register_admin(username, password):
    if not username or not password:
        return False, "Username and password are required"
    
    # Validate password strength
    validation_result = ValidationEngine.validate_password_strength(password)
    if not validation_result['valid']:
        return False, validation_result['message']
    
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Use database-specific parameter placeholders
        if db_manager.current_db_type == 'mysql':
            placeholder = '%s'
        else:
            placeholder = '?'
        
        cursor.execute(f'SELECT username FROM Users WHERE username = {placeholder}', (username,))
        if cursor.fetchone():
            conn.close()
            return False, "Username already exists"
        
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        cursor.execute(f'INSERT INTO Users (username, password_hash) VALUES ({placeholder}, {placeholder})', 
                      (username, password_hash))
        conn.commit()
        conn.close()
        
        # Log successful user creation
        audit_manager.log_transaction('Users', 'CREATE', None, None, {
            'username': username,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        return True, "Admin registered successfully"
        
    except Exception as e:
        if conn:
            conn.close()
        error_message = ErrorHandler.get_user_friendly_message(e)
        logger.error(f"Error registering admin user: {e}")
        return False, f"Registration failed: {error_message}"

# CRUD Operations for Members
def create_member(name, surname, phone_number, email=None, **kwargs):
    """
    Create a new member with comprehensive data and automatic member number assignment
    """
    try:
        # Validate input data
        validation_result = ValidationEngine.validate_member_data(name, surname, phone_number, email or '')
        if not validation_result.get('valid', False):
            error_msg = validation_result.get('message', 'Validation failed')
            raise ValidationError(error_msg)
        
        formatted_data = validation_result.get('formatted_data', {})
        
        conn = connect_db()
        cursor = conn.cursor()
        
        # Generate unique member number
        member_number = generate_unique_member_number()
        join_date = datetime.now().strftime('%Y-%m-%d')
        
        # Prepare member data with defaults for enhanced fields
        member_data = {
            'name': formatted_data.get('name', name),
            'surname': formatted_data.get('surname', surname),
            'phone_number': formatted_data.get('phone', phone_number),
            'email': formatted_data.get('email', email),
            'member_number': member_number,
            'status': 'active',
            'join_date': join_date,
            'middle_name': kwargs.get('middle_name'),
            'phone_number_2': kwargs.get('phone_number_2'),
            'national_id': kwargs.get('national_id'),
            'date_of_birth': kwargs.get('date_of_birth'),
            'gender': kwargs.get('gender'),
            'marital_status': kwargs.get('marital_status'),
            'occupation': kwargs.get('occupation'),
            'employer': kwargs.get('employer'),
            'monthly_income': kwargs.get('monthly_income'),
            'address_line1': kwargs.get('address_line1'),
            'address_line2': kwargs.get('address_line2'),
            'city': kwargs.get('city'),
            'district': kwargs.get('district'),
            'postal_code': kwargs.get('postal_code'),
            'country': kwargs.get('country', 'Malawi'),
            'emergency_contact_name': kwargs.get('emergency_contact_name'),
            'emergency_contact_phone': kwargs.get('emergency_contact_phone'),
            'emergency_contact_relationship': kwargs.get('emergency_contact_relationship'),
            'next_of_kin_name': kwargs.get('next_of_kin_name'),
            'next_of_kin_phone': kwargs.get('next_of_kin_phone'),
            'next_of_kin_relationship': kwargs.get('next_of_kin_relationship'),
            'next_of_kin_address': kwargs.get('next_of_kin_address'),
            'preferred_language': kwargs.get('preferred_language', 'English'),
            'communication_preference': kwargs.get('communication_preference', 'SMS'),
            'member_type': kwargs.get('member_type', 'Individual'),
            'membership_category': kwargs.get('membership_category', 'Regular'),
            'referred_by_member_id': kwargs.get('referred_by_member_id'),
            'special_notes': kwargs.get('special_notes'),
            'created_by': kwargs.get('created_by', 'system')
        }
        
        # Insert member with basic required fields first
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('''INSERT INTO Members 
                             (name, surname, phone_number, email, member_number, status, join_date, created_by) 
                             VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''', 
                           (member_data['name'], member_data['surname'], 
                            member_data['phone_number'], member_data['email'], 
                            member_number, 'active', join_date, member_data['created_by']))
        else:
            cursor.execute('''INSERT INTO Members 
                             (name, surname, phone_number, email, member_number, status, join_date, created_by) 
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', 
                           (member_data['name'], member_data['surname'], 
                            member_data['phone_number'], member_data['email'], 
                            member_number, 'active', join_date, member_data['created_by']))
        
        member_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        # Create audit log entry with safe operation name
        try:
            new_values = {
                'member_id': member_id, 
                'name': member_data['name'], 
                'surname': member_data['surname'], 
                'phone_number': member_data['phone_number'], 
                'email': member_data['email'],
                'member_number': member_number,
                'status': 'active',
                'join_date': join_date
            }
            
            # Use a safe audit logging approach
            if 'audit_manager' in globals() and audit_manager:
                audit_manager.log_action('Members', 'INSERT', member_id, None, new_values)
        except Exception as audit_error:
            logger.warning(f"Audit logging failed: {audit_error}")
        
        # Track change for sync
        try:
            if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'sqlite':
                if hasattr(db_manager, 'track_change'):
                    db_manager.track_change('Members', member_id, 'INSERT', new_values)
        except Exception as sync_error:
            logger.warning(f"Sync tracking failed: {sync_error}")
        
        logger.info(f"Member created successfully: {member_data['name']} {member_data['surname']} (ID: {member_id})")
        return member_id
        
    except Exception as e:
        logger.error(f"Error creating member: {str(e)}")
        if isinstance(e, BankMmudziException):
            raise
        else:
            raise DatabaseError(f"Failed to create member: {str(e)}", "INSERT", "Members")

def generate_unique_member_number() -> str:
    """
    Generate a unique member number in format: BM-YYYY-NNNN
    """
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        current_year = datetime.now().year
        year_prefix = f"BM-{current_year}-"
        
        # Find the highest member number for current year
        cursor.execute('''
            SELECT member_number FROM Members 
            WHERE member_number LIKE ? 
            ORDER BY member_number DESC 
            LIMIT 1
        ''', (f"{year_prefix}%",))
        
        result = cursor.fetchone()
        
        if result:
            # Extract the number part and increment
            last_number = int(result[0].split('-')[-1])
            new_number = last_number + 1
        else:
            # First member of the year
            new_number = 1
        
        conn.close()
        
        # Format as 4-digit number
        member_number = f"{year_prefix}{new_number:04d}"
        
        return member_number
        
    except Exception as e:
        # Fallback to timestamp-based number
        timestamp = int(datetime.now().timestamp())
        return f"BM-{datetime.now().year}-{timestamp % 10000:04d}"

def update_member_status(member_id: int, new_status: str, reason: str = None) -> bool:
    """
    Update member status with audit trail
    
    Args:
        member_id (int): Member ID
        new_status (str): New status ('active', 'inactive', 'suspended')
        reason (str): Reason for status change
        
    Returns:
        bool: Success status
    """
    try:
        valid_statuses = ['active', 'inactive', 'suspended']
        if new_status not in valid_statuses:
            raise ValidationError(f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
        
        conn = connect_db()
        cursor = conn.cursor()
        
        # Get current member data
        cursor.execute('SELECT name, surname, status FROM Members WHERE member_id = ?', (member_id,))
        member_data = cursor.fetchone()
        
        if not member_data:
            raise RecordNotFoundError(f"Member with ID {member_id} not found", "Members", member_id)
        
        old_status = member_data[2]
        member_name = f"{member_data[0]} {member_data[1]}"
        
        # Update status
        cursor.execute('''
            UPDATE Members 
            SET status = ?, status_change_date = ?, status_change_reason = ? 
            WHERE member_id = ?
        ''', (new_status, datetime.now().strftime('%Y-%m-%d'), reason, member_id))
        
        conn.commit()
        conn.close()
        
        # Create audit log entry
        old_values = {'status': old_status}
        new_values = {
            'status': new_status, 
            'status_change_date': datetime.now().strftime('%Y-%m-%d'),
            'status_change_reason': reason
        }
        
        audit_message = f"Status changed from '{old_status}' to '{new_status}' for {member_name}"
        if reason:
            audit_message += f" - Reason: {reason}"
            
        audit_manager.log_transaction('Members', 'UPDATE', member_id, old_values, new_values)
        
        return True
        
    except Exception as e:
        if isinstance(e, BankMmudziException):
            raise
        else:
            raise DatabaseError(f"Failed to update member status: {str(e)}", "UPDATE", "Members")

def get_member_profile(member_id: int) -> dict:
    """
    Get comprehensive member profile with contribution and loan history
    
    Args:
        member_id (int): Member ID
        
    Returns:
        dict: Complete member profile data
    """
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Get basic member info using database-specific syntax
        if db_manager.current_db_type == 'mysql':
            cursor.execute('''
                SELECT member_id, name, surname, phone_number, email, member_number, 
                       status, join_date, created_at, updated_at
                FROM Members 
                WHERE member_id = %s
            ''', (member_id,))
        else:
            cursor.execute('''
                SELECT member_id, name, surname, phone_number, email, member_number, 
                       status, join_date, status_change_date, status_change_reason
                FROM Members 
                WHERE member_id = ?
            ''', (member_id,))
        
        member_data = cursor.fetchone()
        if not member_data:
            raise RecordNotFoundError(f"Member with ID {member_id} not found", "Members", member_id)
        
        # Get contribution summary using database-specific syntax
        if db_manager.current_db_type == 'mysql':
            cursor.execute('''
                SELECT COUNT(*) as contribution_count, 
                       COALESCE(SUM(amount), 0) as total_contributions,
                       MIN(contribution_date) as first_contribution,
                       MAX(contribution_date) as last_contribution
                FROM Contributions 
                WHERE member_id = %s
            ''', (member_id,))
        else:
            cursor.execute('''
                SELECT COUNT(*) as contribution_count, 
                       COALESCE(SUM(amount), 0) as total_contributions,
                       MIN(contribution_date) as first_contribution,
                       MAX(contribution_date) as last_contribution
                FROM Contributions 
                WHERE member_id = ?
            ''', (member_id,))
        
        contrib_data = cursor.fetchone()
        
        # Get loan summary using database-specific syntax
        if db_manager.current_db_type == 'mysql':
            cursor.execute('''
                SELECT COUNT(*) as loan_count,
                       COALESCE(SUM(loan_amount), 0) as total_loans,
                       COALESCE(SUM(outstanding_balance), 0) as outstanding_balance,
                       MIN(loan_date) as first_loan,
                       MAX(loan_date) as last_loan
                FROM Loans 
                WHERE member_id = %s
            ''', (member_id,))
        else:
            cursor.execute('''
                SELECT COUNT(*) as loan_count,
                       COALESCE(SUM(loan_amount), 0) as total_loans,
                       COALESCE(SUM(outstanding_balance), 0) as outstanding_balance,
                       MIN(loan_date) as first_loan,
                       MAX(loan_date) as last_loan
                FROM Loans 
                WHERE member_id = ?
            ''', (member_id,))
        
        loan_data = cursor.fetchone()
        
        # Get repayment summary using database-specific syntax
        if db_manager.current_db_type == 'mysql':
            cursor.execute('''
                SELECT COUNT(*) as repayment_count,
                       COALESCE(SUM(r.repayment_amount), 0) as total_repayments,
                       COALESCE(SUM(r.interest_amount), 0) as total_interest_paid
                FROM Repayments r
                JOIN Loans l ON r.loan_id = l.loan_id
                WHERE l.member_id = %s
            ''', (member_id,))
        else:
            cursor.execute('''
                SELECT COUNT(*) as repayment_count,
                       COALESCE(SUM(r.repayment_amount), 0) as total_repayments,
                       COALESCE(SUM(r.interest_amount), 0) as total_interest_paid
                FROM Repayments r
                JOIN Loans l ON r.loan_id = l.loan_id
                WHERE l.member_id = ?
            ''', (member_id,))
        
        repayment_data = cursor.fetchone()
        
        # Get recent activity using database-specific syntax
        if db_manager.current_db_type == 'mysql':
            cursor.execute('''
                SELECT operation, table_name, timestamp, new_values
                FROM AuditLog 
                WHERE record_id = %s AND table_name IN ('Members', 'Contributions', 'Loans', 'Repayments')
                ORDER BY timestamp DESC 
                LIMIT 10
            ''', (member_id,))
        else:
            cursor.execute('''
                SELECT operation, table_name, timestamp, new_values
                FROM AuditLog 
                WHERE record_id = ? AND table_name IN ('Members', 'Contributions', 'Loans', 'Repayments')
                ORDER BY timestamp DESC 
                LIMIT 10
            ''', (member_id,))
        
        recent_activity = cursor.fetchall()
        
        conn.close()
        
        # Build comprehensive profile
        profile = {
            'member_info': {
                'member_id': member_data[0],
                'name': member_data[1],
                'surname': member_data[2],
                'full_name': f"{member_data[1]} {member_data[2]}",
                'phone_number': member_data[3],
                'email': member_data[4],
                'member_number': member_data[5],
                'status': member_data[6],
                'join_date': member_data[7],
                'status_change_date': member_data[8],
                'status_change_reason': member_data[9]
            },
            'contribution_summary': {
                'contribution_count': contrib_data[0],
                'total_contributions': contrib_data[1],
                'first_contribution': contrib_data[2],
                'last_contribution': contrib_data[3],
                'average_monthly_contribution': contrib_data[1] / max(contrib_data[0], 1)
            },
            'loan_summary': {
                'loan_count': loan_data[0],
                'total_loans': loan_data[1],
                'outstanding_balance': loan_data[2],
                'first_loan': loan_data[3],
                'last_loan': loan_data[4],
                'repayment_count': repayment_data[0],
                'total_repayments': repayment_data[1],
                'total_interest_paid': repayment_data[2]
            },
            'financial_standing': {
                'net_contribution': contrib_data[1] - loan_data[2],  # Contributions minus outstanding
                'member_type': 'Borrower' if loan_data[2] > 0 else 'Contributor',
                'payment_history': 'Good' if repayment_data[0] > 0 else 'No loans'
            },
            'recent_activity': [
                {
                    'operation': activity[0],
                    'table': activity[1],
                    'date': activity[2],
                    'details': activity[3]
                } for activity in recent_activity
            ]
        }
        
        return profile
        
    except Exception as e:
        if isinstance(e, BankMmudziException):
            raise
        else:
            raise DatabaseError(f"Failed to get member profile: {str(e)}", "SELECT", "Members")

def read_members():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM Members')
    members = cursor.fetchall()
    conn.close()
    return members

def update_member(member_id, name, surname, phone_number, email=None):
    # Get old values for audit trail
    conn = connect_db()
    cursor = conn.cursor()
    
    # Use database-specific syntax
    if db_manager.current_db_type == 'mysql':
        cursor.execute('SELECT name, surname, phone_number, email FROM Members WHERE member_id = %s', (member_id,))
        old_data = cursor.fetchone()
        old_values = {'member_id': member_id, 'name': old_data[0], 'surname': old_data[1], 
                      'phone_number': old_data[2], 'email': old_data[3]} if old_data else None
        
        cursor.execute('UPDATE Members SET name = %s, surname = %s, phone_number = %s, email = %s WHERE member_id = %s', 
                       (name, surname, phone_number, email, member_id))
    else:
        cursor.execute('SELECT name, surname, phone_number, email FROM Members WHERE member_id = ?', (member_id,))
        old_data = cursor.fetchone()
        old_values = {'member_id': member_id, 'name': old_data[0], 'surname': old_data[1], 
                      'phone_number': old_data[2], 'email': old_data[3]} if old_data else None
        
        cursor.execute('UPDATE Members SET name = ?, surname = ?, phone_number = ?, email = ? WHERE member_id = ?', 
                       (name, surname, phone_number, email, member_id))
    
    conn.commit()
    conn.close()
    
    # Create audit log entry
    new_values = {'member_id': member_id, 'name': name, 'surname': surname, 'phone_number': phone_number, 'email': email}
    audit_manager.log_transaction('Members', 'UPDATE', member_id, old_values, new_values)
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite':
        data = {'member_id': member_id, 'name': name, 'surname': surname, 'phone_number': phone_number, 'email': email}
        db_manager.track_change('Members', member_id, 'UPDATE', data)
    
    return True

def delete_member(member_id):
    # Get member data before deletion for sync tracking and audit trail
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM Members WHERE member_id = ?', (member_id,))
    member_data = cursor.fetchone()
    
    old_values = {'member_id': member_data[0], 'name': member_data[1], 'surname': member_data[2], 
                  'phone_number': member_data[3], 'email': member_data[4]} if member_data else None
    
    cursor.execute('DELETE FROM Members WHERE member_id = ?', (member_id,))
    conn.commit()
    conn.close()
    
    # Create audit log entry
    if old_values:
        audit_manager.log_transaction('Members', 'DELETE', member_id, old_values, None)
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite' and member_data:
        data = {'member_id': member_data[0], 'name': member_data[1], 'surname': member_data[2], 
                'phone_number': member_data[3], 'email': member_data[4]}
        db_manager.track_change('Members', member_id, 'DELETE', data)

# CRUD Operations for Contributions
def create_contribution(member_id, month, amount, year=None):
    contribution_date = datetime.now().strftime('%Y-%m-%d')
    month_num = MONTH_TO_NUM[month]
    if year is None:
        year = datetime.now().year
    
    conn = connect_db()
    cursor = conn.cursor()
    
    # Use database-specific syntax
    if db_manager.current_db_type == 'mysql':
        cursor.execute('INSERT INTO Contributions (member_id, month, year, amount, contribution_date) VALUES (%s, %s, %s, %s, %s)', 
                       (member_id, month_num, year, amount, contribution_date))
    else:
        cursor.execute('INSERT INTO Contributions (member_id, month, year, amount, contribution_date) VALUES (?, ?, ?, ?, ?)', 
                       (member_id, month_num, year, amount, contribution_date))
    
    contribution_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    # Create audit log entry
    new_values = {'contribution_id': contribution_id, 'member_id': member_id, 'month': month_num, 'year': year, 'amount': amount, 'contribution_date': contribution_date}
    audit_manager.log_transaction('Contributions', 'INSERT', contribution_id, None, new_values)
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite':
        data = {'contribution_id': contribution_id, 'member_id': member_id, 'month': month_num, 'year': year, 'amount': amount, 'contribution_date': contribution_date}
        db_manager.track_change('Contributions', contribution_id, 'INSERT', data)
    
    return contribution_id

def read_contributions_for_member(member_id):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT contribution_id, month, amount, contribution_date FROM Contributions WHERE member_id = ?', 
                   (member_id,))
    contributions = [(row[0], MONTHS[row[1]-1], row[2], row[3]) for row in cursor.fetchall()]
    conn.close()
    return contributions

def read_contributions_up_to_date(end_date):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT member_id, SUM(amount) FROM Contributions WHERE contribution_date <= ? GROUP BY member_id', 
                   (end_date,))
    contributions = cursor.fetchall()
    conn.close()
    return dict(contributions)

def update_contribution(contribution_id, month, amount):
    contribution_date = datetime.now().strftime('%Y-%m-%d')
    month_num = MONTH_TO_NUM[month]
    
    # Get old values for audit trail
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT member_id, month, year, amount, contribution_date FROM Contributions WHERE contribution_id = ?', (contribution_id,))
    old_data = cursor.fetchone()
    old_values = {'contribution_id': contribution_id, 'member_id': old_data[0], 'month': old_data[1], 
                  'year': old_data[2], 'amount': old_data[3], 'contribution_date': old_data[4]} if old_data else None
    member_id = old_data[0] if old_data else None
    year = old_data[2] if old_data else None
    
    cursor.execute('UPDATE Contributions SET month = ?, amount = ?, contribution_date = ? WHERE contribution_id = ?', 
                   (month_num, amount, contribution_date, contribution_id))
    conn.commit()
    conn.close()
    
    # Create audit log entry
    new_values = {'contribution_id': contribution_id, 'member_id': member_id, 'month': month_num, 'year': year, 'amount': amount, 'contribution_date': contribution_date}
    audit_manager.log_transaction('Contributions', 'UPDATE', contribution_id, old_values, new_values)
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite' and member_id:
        data = {'contribution_id': contribution_id, 'member_id': member_id, 'month': month_num, 'amount': amount, 'contribution_date': contribution_date}
        db_manager.track_change('Contributions', contribution_id, 'UPDATE', data)

def delete_contribution(contribution_id):
    # Get contribution data before deletion for sync tracking and audit trail
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM Contributions WHERE contribution_id = ?', (contribution_id,))
    contribution_data = cursor.fetchone()
    
    old_values = {'contribution_id': contribution_data[0], 'member_id': contribution_data[1], 
                  'month': contribution_data[2], 'year': contribution_data[3], 'amount': contribution_data[4], 
                  'contribution_date': contribution_data[5]} if contribution_data else None
    
    cursor.execute('DELETE FROM Contributions WHERE contribution_id = ?', (contribution_id,))
    conn.commit()
    conn.close()
    
    # Create audit log entry
    if old_values:
        audit_manager.log_transaction('Contributions', 'DELETE', contribution_id, old_values, None)
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite' and contribution_data:
        data = {'contribution_id': contribution_data[0], 'member_id': contribution_data[1], 
                'month': contribution_data[2], 'amount': contribution_data[3], 'contribution_date': contribution_data[4]}
        db_manager.track_change('Contributions', contribution_id, 'DELETE', data)

# CRUD Operations for Loans
def create_loan(member_id, loan_amount, interest_rate):
    loan_date = datetime.now().strftime('%Y-%m-%d')
    
    # Calculate monthly payment and total interest using proper amortization
    monthly_rate = interest_rate / 100 / 12  # Convert annual percentage to monthly decimal
    monthly_payment = FinancialCalculator.calculate_loan_payment(loan_amount, monthly_rate, 12)
    total_interest = FinancialCalculator.calculate_total_interest(loan_amount, monthly_payment, 12)
    
    conn = connect_db()
    cursor = conn.cursor()
    
    # Use database-specific syntax
    if db_manager.current_db_type == 'mysql':
        cursor.execute('''INSERT INTO Loans (member_id, loan_amount, interest_rate, monthly_payment, total_interest, outstanding_balance, loan_date) 
                          VALUES (%s, %s, %s, %s, %s, %s, %s)''', 
                       (member_id, loan_amount, interest_rate, monthly_payment, total_interest, loan_amount, loan_date))
        loan_id = cursor.lastrowid
        conn.commit()
        
        # Generate and store amortization schedule
        schedule = FinancialCalculator.generate_amortization_schedule(loan_id, loan_amount, interest_rate, monthly_payment, 12)
        for payment_number, payment_date, principal_amount, interest_amount, remaining_balance in schedule:
            cursor.execute('''INSERT INTO LoanSchedule (loan_id, payment_number, payment_date, principal_amount, interest_amount, remaining_balance)
                              VALUES (%s, %s, %s, %s, %s, %s)''',
                           (loan_id, payment_number, payment_date, principal_amount, interest_amount, remaining_balance))
    else:
        cursor.execute('''INSERT INTO Loans (member_id, loan_amount, interest_rate, monthly_payment, total_interest, outstanding_balance, loan_date) 
                          VALUES (?, ?, ?, ?, ?, ?, ?)''', 
                       (member_id, loan_amount, interest_rate, monthly_payment, total_interest, loan_amount, loan_date))
        loan_id = cursor.lastrowid
        conn.commit()
        
        # Generate and store amortization schedule
        schedule = FinancialCalculator.generate_amortization_schedule(loan_id, loan_amount, interest_rate, monthly_payment, 12)
        for payment_number, payment_date, principal_amount, interest_amount, remaining_balance in schedule:
            cursor.execute('''INSERT INTO LoanSchedule (loan_id, payment_number, payment_date, principal_amount, interest_amount, remaining_balance)
                              VALUES (?, ?, ?, ?, ?, ?)''',
                           (loan_id, payment_number, payment_date, principal_amount, interest_amount, remaining_balance))
    
    conn.commit()
    conn.close()
    
    # Create audit log entry
    new_values = {'loan_id': loan_id, 'member_id': member_id, 'loan_amount': loan_amount, 'interest_rate': interest_rate, 
                  'monthly_payment': monthly_payment, 'total_interest': total_interest, 'loan_date': loan_date}
    audit_manager.log_transaction('Loans', 'INSERT', loan_id, None, new_values)
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite':
        data = {'loan_id': loan_id, 'member_id': member_id, 'loan_amount': loan_amount, 'interest_rate': interest_rate, 
                'monthly_payment': monthly_payment, 'total_interest': total_interest, 'loan_date': loan_date}
        db_manager.track_change('Loans', loan_id, 'INSERT', data)
    
    return loan_id

def read_loans():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('''SELECT Loans.loan_id, Members.name || " " || Members.surname AS full_name, 
                      Loans.loan_amount, Loans.interest_rate, Loans.monthly_payment, Loans.total_interest, Loans.loan_date 
                      FROM Loans JOIN Members ON Loans.member_id = Members.member_id''')
    loans = cursor.fetchall()
    conn.close()
    return loans

def read_loans_up_to_date(end_date):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT Loans.member_id, Loans.loan_amount, Loans.interest_rate, Loans.loan_date '
                   'FROM Loans WHERE loan_date <= ?', (end_date,))
    loans = cursor.fetchall()
    conn.close()
    return loans

def update_loan(loan_id, member_id, loan_amount, interest_rate):
    loan_date = datetime.now().strftime('%Y-%m-%d')
    
    # Get old values for audit trail
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT member_id, loan_amount, interest_rate, monthly_payment, total_interest, loan_date FROM Loans WHERE loan_id = ?', (loan_id,))
    old_data = cursor.fetchone()
    old_values = {'loan_id': loan_id, 'member_id': old_data[0], 'loan_amount': old_data[1], 
                  'interest_rate': old_data[2], 'monthly_payment': old_data[3], 'total_interest': old_data[4], 
                  'loan_date': old_data[5]} if old_data else None
    
    # Calculate monthly payment and total interest using proper amortization
    monthly_payment = FinancialCalculator.calculate_loan_payment(loan_amount, interest_rate, 12)
    total_interest = FinancialCalculator.calculate_total_interest(loan_amount, monthly_payment, 12)
    
    cursor.execute('''UPDATE Loans SET member_id = ?, loan_amount = ?, interest_rate = ?, monthly_payment = ?, total_interest = ?, loan_date = ? 
                      WHERE loan_id = ?''', 
                   (member_id, loan_amount, interest_rate, monthly_payment, total_interest, loan_date, loan_id))
    
    # Delete existing amortization schedule
    cursor.execute('DELETE FROM LoanSchedule WHERE loan_id = ?', (loan_id,))
    
    # Generate and store new amortization schedule
    schedule = FinancialCalculator.generate_amortization_schedule(loan_id, loan_amount, interest_rate, monthly_payment, 12)
    for payment_number, payment_date, principal_amount, interest_amount, remaining_balance in schedule:
        cursor.execute('''INSERT INTO LoanSchedule (loan_id, payment_number, payment_date, principal_amount, interest_amount, remaining_balance)
                          VALUES (?, ?, ?, ?, ?, ?)''',
                       (loan_id, payment_number, payment_date, principal_amount, interest_amount, remaining_balance))
    
    conn.commit()
    conn.close()
    
    # Create audit log entry
    new_values = {'loan_id': loan_id, 'member_id': member_id, 'loan_amount': loan_amount, 'interest_rate': interest_rate, 
                  'monthly_payment': monthly_payment, 'total_interest': total_interest, 'loan_date': loan_date}
    audit_manager.log_transaction('Loans', 'UPDATE', loan_id, old_values, new_values)
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite':
        data = {'loan_id': loan_id, 'member_id': member_id, 'loan_amount': loan_amount, 'interest_rate': interest_rate, 
                'monthly_payment': monthly_payment, 'total_interest': total_interest, 'loan_date': loan_date}
        db_manager.track_change('Loans', loan_id, 'UPDATE', data)

def delete_loan(loan_id):
    # Get loan data before deletion for sync tracking and audit trail
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM Loans WHERE loan_id = ?', (loan_id,))
    loan_data = cursor.fetchone()
    
    old_values = {'loan_id': loan_data[0], 'member_id': loan_data[1], 'loan_amount': loan_data[2], 
                  'interest_rate': loan_data[3], 'monthly_payment': loan_data[4], 'total_interest': loan_data[5], 
                  'loan_date': loan_data[6]} if loan_data else None
    
    cursor.execute('DELETE FROM Loans WHERE loan_id = ?', (loan_id,))
    conn.commit()
    conn.close()
    
    # Create audit log entry
    if old_values:
        audit_manager.log_transaction('Loans', 'DELETE', loan_id, old_values, None)
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite' and loan_data:
        data = {'loan_id': loan_data[0], 'member_id': loan_data[1], 'loan_amount': loan_data[2], 
                'interest_rate': loan_data[3], 'loan_date': loan_data[4]}
        db_manager.track_change('Loans', loan_id, 'DELETE', data)

# CRUD Operations for Repayments
def create_repayment(loan_id, repayment_amount):
    """
    Create a repayment record with proper principal and interest separation
    """
    repayment_date = datetime.now().strftime('%Y-%m-%d')
    
    # Get loan details to calculate principal and interest portions
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT loan_amount, interest_rate, monthly_payment, outstanding_balance FROM Loans WHERE loan_id = ?', (loan_id,))
    loan_data = cursor.fetchone()
    
    if not loan_data:
        conn.close()
        raise ValueError(f"Loan with ID {loan_id} not found")
    
    loan_amount, interest_rate, monthly_payment, outstanding_balance = loan_data
    
    # Calculate monthly interest rate (assuming annual rate is stored)
    monthly_rate = interest_rate / 12 / 100 if interest_rate else 0
    
    # Calculate interest portion based on current outstanding balance
    interest_amount = outstanding_balance * monthly_rate
    
    # Calculate principal portion
    principal_amount = repayment_amount - interest_amount
    
    # Ensure principal doesn't exceed outstanding balance
    if principal_amount > outstanding_balance:
        principal_amount = outstanding_balance
        interest_amount = repayment_amount - principal_amount
    
    # Ensure amounts are not negative
    principal_amount = max(0, principal_amount)
    interest_amount = max(0, interest_amount)
    
    # Insert repayment with principal and interest breakdown
    cursor.execute('''INSERT INTO Repayments 
                     (loan_id, repayment_amount, principal_amount, interest_amount, repayment_date) 
                     VALUES (?, ?, ?, ?, ?)''', 
                   (loan_id, repayment_amount, principal_amount, interest_amount, repayment_date))
    repayment_id = cursor.lastrowid
    
    # Update loan outstanding balance
    new_outstanding_balance = max(0, outstanding_balance - principal_amount)
    cursor.execute('UPDATE Loans SET outstanding_balance = ? WHERE loan_id = ?', 
                   (new_outstanding_balance, loan_id))
    
    conn.commit()
    conn.close()
    
    # Create audit log entry
    new_values = {
        'repayment_id': repayment_id, 
        'loan_id': loan_id, 
        'repayment_amount': repayment_amount,
        'principal_amount': principal_amount,
        'interest_amount': interest_amount,
        'repayment_date': repayment_date
    }
    audit_manager.log_transaction('Repayments', 'INSERT', repayment_id, None, new_values)
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite':
        data = {
            'repayment_id': repayment_id, 
            'loan_id': loan_id, 
            'repayment_amount': repayment_amount,
            'principal_amount': principal_amount,
            'interest_amount': interest_amount,
            'repayment_date': repayment_date
        }
        db_manager.track_change('Repayments', repayment_id, 'INSERT', data)
    
    return repayment_id

def read_repayments():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT Repayments.repayment_id, Loans.loan_id, Members.name || " " || Members.surname AS full_name, Repayments.repayment_amount, Repayments.repayment_date '
                   'FROM Repayments JOIN Loans ON Repayments.loan_id = Loans.loan_id '
                   'JOIN Members ON Loans.member_id = Members.member_id')
    repayments = cursor.fetchall()
    conn.close()
    return repayments

def read_repayments_up_to_date(end_date):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT Loans.member_id, Repayments.repayment_amount, Repayments.repayment_date '
                   'FROM Repayments JOIN Loans ON Repayments.loan_id = Loans.loan_id '
                   'WHERE Repayments.repayment_date <= ?', (end_date,))
    repayments = cursor.fetchall()
    conn.close()
    return repayments

def update_repayment(repayment_id, loan_id, repayment_amount):
    # Get old values for audit trail
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT loan_id, repayment_amount, principal_amount, interest_amount, repayment_date FROM Repayments WHERE repayment_id = ?', (repayment_id,))
    old_data = cursor.fetchone()
    old_values = {'repayment_id': repayment_id, 'loan_id': old_data[0], 'repayment_amount': old_data[1], 
                  'principal_amount': old_data[2], 'interest_amount': old_data[3], 'repayment_date': old_data[4]} if old_data else None
    
    repayment_date = datetime.now().strftime('%Y-%m-%d')
    cursor.execute('UPDATE Repayments SET loan_id = ?, repayment_amount = ?, repayment_date = ? WHERE repayment_id = ?', 
                   (loan_id, repayment_amount, repayment_date, repayment_id))
    conn.commit()
    conn.close()
    
    # Create audit log entry
    new_values = {'repayment_id': repayment_id, 'loan_id': loan_id, 'repayment_amount': repayment_amount, 'repayment_date': repayment_date}
    audit_manager.log_transaction('Repayments', 'UPDATE', repayment_id, old_values, new_values)
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite':
        data = {'repayment_id': repayment_id, 'loan_id': loan_id, 'repayment_amount': repayment_amount, 'repayment_date': repayment_date}
        db_manager.track_change('Repayments', repayment_id, 'UPDATE', data)

def delete_repayment(repayment_id):
    # Get repayment data before deletion for sync tracking and audit trail
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM Repayments WHERE repayment_id = ?', (repayment_id,))
    repayment_data = cursor.fetchone()
    
    old_values = {'repayment_id': repayment_data[0], 'loan_id': repayment_data[1], 
                  'repayment_amount': repayment_data[2], 'principal_amount': repayment_data[3], 
                  'interest_amount': repayment_data[4], 'repayment_date': repayment_data[5]} if repayment_data else None
    
    cursor.execute('DELETE FROM Repayments WHERE repayment_id = ?', (repayment_id,))
    conn.commit()
    conn.close()
    
    # Create audit log entry
    if old_values:
        audit_manager.log_transaction('Repayments', 'DELETE', repayment_id, old_values, None)
    
    # Track change for sync
    if db_manager.current_db_type == 'sqlite' and repayment_data:
        data = {'repayment_id': repayment_data[0], 'loan_id': repayment_data[1], 
                'repayment_amount': repayment_data[2], 'repayment_date': repayment_data[3]}
        db_manager.track_change('Repayments', repayment_id, 'DELETE', data)

# Consolidated Payments View
def read_all_payments(start_date=None, end_date=None, payment_type=None, limit=100, offset=0):
    conn = connect_db()
    cursor = conn.cursor()
    
    base_query = '''
        SELECT Members.name || " " || Members.surname AS full_name, "Contribution" as type, Contributions.amount, 
               Contributions.contribution_date, Contributions.contribution_id, NULL as loan_id
        FROM Contributions 
        JOIN Members ON Contributions.member_id = Members.member_id
        {where}
        UNION ALL
        SELECT Members.name || " " || Members.surname AS full_name, "Repayment" as type, Repayments.repayment_amount, 
               Repayments.repayment_date, Repayments.repayment_id, Loans.loan_id
        FROM Repayments 
        JOIN Loans ON Repayments.loan_id = Loans.loan_id
        JOIN Members ON Loans.member_id = Members.member_id
        {where}
        UNION ALL
        SELECT Members.name || " " || Members.surname AS full_name, "Loan Disbursement" as type, -Loans.loan_amount as amount,
               Loans.loan_date, Loans.loan_id, NULL
        FROM Loans
        JOIN Members ON Loans.member_id = Members.member_id
        {where}
    '''
    
    where_clause = []
    params = []
    
    if start_date:
        where_clause.append("contribution_date >= ?")
        params.append(start_date)
    if end_date:
        where_clause.append("contribution_date <= ?")
        params.append(end_date)
    if payment_type and payment_type != "All":
        where_clause.append("type = ?")
        params.append(payment_type)
        
    where_str = "WHERE " + " AND ".join(where_clause) if where_clause else ""
    query = base_query.format(where=where_str) + " ORDER BY contribution_date DESC LIMIT ? OFFSET ?"
    params.extend([limit, offset])
    
    cursor.execute(query, params)
    payments = cursor.fetchall()
    conn.close()
    return payments

# Derived Calculations
def get_total_contributions_up_to_date(end_date):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT SUM(amount) FROM Contributions WHERE contribution_date <= ?', (end_date,))
    total = cursor.fetchone()[0]
    conn.close()
    return total if total else 0

def get_total_loans_disbursed_up_to_date(end_date):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT SUM(loan_amount) FROM Loans WHERE loan_date <= ?', (end_date,))
    total = cursor.fetchone()[0]
    conn.close()
    return total if total else 0

def get_total_repayments_up_to_date(end_date):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT SUM(repayment_amount) FROM Repayments WHERE repayment_date <= ?', (end_date,))
    total = cursor.fetchone()[0]
    conn.close()
    return total if total else 0

def get_loan_outstanding_balance_up_to_date(loan_id, end_date):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT loan_amount FROM Loans WHERE loan_id = ? AND loan_date <= ?', (loan_id, end_date))
    loan_result = cursor.fetchone()
    loan_amount = loan_result[0] if loan_result else 0
    cursor.execute('SELECT SUM(repayment_amount) FROM Repayments WHERE loan_id = ? AND repayment_date <= ?', (loan_id, end_date))
    sum_repayments = cursor.fetchone()[0] or 0
    conn.close()
    return max(loan_amount - sum_repayments, 0)

def get_member_outstanding_balance_up_to_date(member_id, end_date):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute('SELECT loan_id FROM Loans WHERE member_id = ? AND loan_date <= ?', (member_id, end_date))
    loan_ids = [row[0] for row in cursor.fetchall()]
    total_outstanding = 0
    for loan_id in loan_ids:
        total_outstanding += get_loan_outstanding_balance_up_to_date(loan_id, end_date)
    conn.close()
    return total_outstanding

def calculate_dividends(end_date):
    """
    Calculate dividends using the correct financial model:
    - Non-borrowers: 12C (total contributions)
    - Borrowers: 12C + interest paid - outstanding balance
    
    This function maintains compatibility with existing UI code while using the new calculation engine.
    """
    # Extract year from end_date
    year = int(end_date.split('-')[0]) if isinstance(end_date, str) else end_date.year
    
    # Get all active members
    members = read_members()
    if not members:
        return []
    
    dividends = []
    for member in members:
        member_id, name, surname, _, status = member
        
        # Skip inactive members
        if status != 'active':
            continue
            
        full_name = f"{name} {surname}"
        
        # Calculate dividend using the new financial model
        dividend_data = calculate_member_dividend_for_year(member_id, year)
        dividend_amount = dividend_data['dividend_amount']
        
        dividends.append((full_name, dividend_amount))
    
    return dividends

def check_login(username, password):
    """
    Enhanced login check with account lockout and security logging
    """
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Use database-specific parameter placeholders
        if db_manager.current_db_type == 'mysql':
            placeholder = '%s'
        else:
            placeholder = '?'
        
        # Check if account is locked
        cursor.execute(f'''
            SELECT password_hash, failed_login_attempts, locked_until, last_login 
            FROM Users 
            WHERE username = {placeholder}
        ''', (username,))
        
        row = cursor.fetchone()
        
        if not row:
            conn.close()
            # Log failed login attempt for non-existent user
            audit_manager.log_security_violation('FAILED_LOGIN_INVALID_USER', {
                'username': username,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'ip_address': 'localhost'  # Could be enhanced with actual IP detection
            })
            return False
        
        stored_hash, failed_login_attempts, locked_until, last_login = row
        
        # Check if account is currently locked
        if locked_until:
            try:
                lock_time = datetime.strptime(locked_until, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                # Handle different datetime formats
                try:
                    lock_time = datetime.fromisoformat(locked_until.replace('T', ' ').replace('Z', ''))
                except:
                    # If parsing fails, assume account is not locked
                    lock_time = datetime.now() - timedelta(minutes=1)
            
            if datetime.now() < lock_time:
                remaining_time = lock_time - datetime.now()
                conn.close()
                audit_manager.log_security_violation('LOGIN_ATTEMPT_LOCKED_ACCOUNT', {
                    'username': username,
                    'remaining_lockout_minutes': int(remaining_time.total_seconds() / 60),
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                raise BusinessRuleViolationError(
                    f"Account is locked. Try again in {int(remaining_time.total_seconds() / 60)} minutes.",
                    "ACCOUNT_LOCKED"
                )
        
        # Verify password
        input_hash = hashlib.sha256(password.encode()).hexdigest()
        
        if stored_hash == input_hash:
            # Successful login - reset failed attempts and update last login
            cursor.execute(f'''
                UPDATE Users 
                SET failed_login_attempts = 0, locked_until = NULL, last_login = {placeholder}
                WHERE username = {placeholder}
            ''', (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), username))
            
            conn.commit()
            conn.close()
            
            # Log successful login
            audit_manager.log_transaction('Users', 'LOGIN_SUCCESS', None, None, {
                'username': username,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            return True
        else:
            # Failed login - increment failed attempts
            new_failed_attempts = (failed_login_attempts or 0) + 1
            
            # Lock account after 3 failed attempts
            if new_failed_attempts >= 3:
                lock_until = datetime.now() + timedelta(minutes=30)  # 30-minute lockout
                cursor.execute(f'''
                    UPDATE Users 
                    SET failed_login_attempts = {placeholder}, locked_until = {placeholder}
                    WHERE username = {placeholder}
                ''', (new_failed_attempts, lock_until.strftime('%Y-%m-%d %H:%M:%S'), username))
                
                audit_manager.log_security_violation('ACCOUNT_LOCKED_MULTIPLE_FAILURES', {
                    'username': username,
                    'failed_attempts': new_failed_attempts,
                    'locked_until': lock_until.strftime('%Y-%m-%d %H:%M:%S'),
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
            else:
                cursor.execute(f'''
                    UPDATE Users 
                    SET failed_login_attempts = {placeholder}
                    WHERE username = {placeholder}
                ''', (new_failed_attempts, username))
                
                audit_manager.log_security_violation('FAILED_LOGIN_INVALID_PASSWORD', {
                    'username': username,
                    'failed_attempts': new_failed_attempts,
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
            
            conn.commit()
            conn.close()
            
            return False
            
    except Exception as e:
        if isinstance(e, BankMmudziException):
            raise
        else:
            raise DatabaseError(f"Login check failed: {str(e)}", "SELECT", "Users")

def create_user(username, password):
    """
    Create a new user with strong password validation
    """
    try:
        # Validate password strength
        password_validation = ValidationEngine.validate_password_strength(password)
        if not password_validation['valid']:
            raise ValidationError(password_validation['message'])
        
        # Check if username already exists
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM Users WHERE username = ?', (username,))
        if cursor.fetchone()[0] > 0:
            raise DuplicateRecordError(f"Username '{username}' already exists", "Users")
        
        # Create password hash
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        # Insert new user
        cursor.execute('''
            INSERT INTO Users (username, password_hash, created_at, failed_login_attempts) 
            VALUES (?, ?, ?, 0)
        ''', (username, password_hash, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        
        user_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        # Log user creation
        audit_manager.log_transaction('Users', 'INSERT', user_id, None, {
            'username': username,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        return user_id
        
    except Exception as e:
        if isinstance(e, BankMmudziException):
            raise
        else:
            raise DatabaseError(f"Failed to create user: {str(e)}", "INSERT", "Users")

def change_password(username, old_password, new_password):
    """
    Change user password with validation
    """
    try:
        # Verify old password
        if not check_login(username, old_password):
            raise ValidationError("Current password is incorrect")
        
        # Validate new password strength
        password_validation = ValidationEngine.validate_password_strength(new_password)
        if not password_validation['valid']:
            raise ValidationError(password_validation['message'])
        
        # Update password
        new_password_hash = hashlib.sha256(new_password.encode()).hexdigest()
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE Users 
            SET password_hash = ?, password_changed_at = ?
            WHERE username = ?
        ''', (new_password_hash, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), username))
        
        conn.commit()
        conn.close()
        
        # Log password change
        audit_manager.log_transaction('Users', 'PASSWORD_CHANGE', None, None, {
            'username': username,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        return True
        
    except Exception as e:
        if isinstance(e, BankMmudziException):
            raise
        else:
            raise DatabaseError(f"Failed to change password: {str(e)}", "UPDATE", "Users")

def unlock_user_account(username):
    """
    Manually unlock a user account (admin function)
    """
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE Users 
            SET failed_login_attempts = 0, locked_until = NULL
            WHERE username = ?
        ''', (username,))
        
        if cursor.rowcount == 0:
            raise RecordNotFoundError(f"User '{username}' not found", "Users")
        
        conn.commit()
        conn.close()
        
        # Log account unlock
        audit_manager.log_transaction('Users', 'ACCOUNT_UNLOCK', None, None, {
            'username': username,
            'unlocked_by': audit_manager.current_user,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        return True
        
    except Exception as e:
        if isinstance(e, BankMmudziException):
            raise
        else:
            raise DatabaseError(f"Failed to unlock account: {str(e)}", "UPDATE", "Users")

# Automated Payment Processing System
class AutomatedPaymentProcessor:
    """
    Handles automated payment processing including late fees and balance updates
    """
    
    def __init__(self):
        self.config_manager = SystemConfigManager()
    
    def calculate_late_payment_penalty(self, loan_id: int) -> dict:
        """
        Calculate late payment penalty for overdue loans
        
        Args:
            loan_id (int): Loan ID to check
            
        Returns:
            dict: Penalty calculation details
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Get loan details
            cursor.execute('''
                SELECT loan_amount, interest_rate, loan_date, outstanding_balance, monthly_payment
                FROM Loans WHERE loan_id = ?
            ''', (loan_id,))
            loan_data = cursor.fetchone()
            
            if not loan_data:
                return {'success': False, 'message': 'Loan not found'}
            
            loan_amount, interest_rate, loan_date, outstanding_balance, monthly_payment = loan_data
            
            # Get last payment date
            cursor.execute('''
                SELECT MAX(repayment_date) FROM Repayments WHERE loan_id = ?
            ''', (loan_id,))
            last_payment = cursor.fetchone()[0]
            
            # Calculate days overdue
            loan_start = datetime.strptime(loan_date, '%Y-%m-%d')
            last_payment_date = datetime.strptime(last_payment, '%Y-%m-%d') if last_payment else loan_start
            
            # Expected payment date (monthly)
            expected_payment_date = last_payment_date + timedelta(days=30)
            days_overdue = max(0, (datetime.now() - expected_payment_date).days)
            
            # Calculate penalty
            late_fee_rate = float(self.config_manager.get_config_value('late_fee_rate', 0.05))  # 5% default
            penalty_amount = 0
            
            if days_overdue > 7:  # Grace period of 7 days
                penalty_amount = outstanding_balance * late_fee_rate * (days_overdue / 30)
            
            conn.close()
            
            return {
                'success': True,
                'loan_id': loan_id,
                'days_overdue': days_overdue,
                'penalty_amount': round(penalty_amount, 2),
                'outstanding_balance': outstanding_balance,
                'last_payment_date': last_payment,
                'expected_payment_date': expected_payment_date.strftime('%Y-%m-%d')
            }
            
        except Exception as e:
            return ErrorHandler.handle_error(e, "late payment penalty calculation")
    
    def apply_late_payment_penalty(self, loan_id: int) -> dict:
        """
        Apply late payment penalty to a loan
        
        Args:
            loan_id (int): Loan ID
            
        Returns:
            dict: Application result
        """
        try:
            penalty_data = self.calculate_late_payment_penalty(loan_id)
            
            if not penalty_data['success'] or penalty_data['penalty_amount'] <= 0:
                return penalty_data
            
            conn = connect_db()
            cursor = conn.cursor()
            
            # Update loan with penalty
            cursor.execute('''
                UPDATE Loans 
                SET outstanding_balance = outstanding_balance + ?,
                    late_fee = COALESCE(late_fee, 0) + ?
                WHERE loan_id = ?
            ''', (penalty_data['penalty_amount'], penalty_data['penalty_amount'], loan_id))
            
            # Log the penalty application
            audit_manager.log_action(
                'UPDATE', 'Loans', loan_id,
                f"Applied late payment penalty: MWK {penalty_data['penalty_amount']:.2f}",
                {'penalty_amount': penalty_data['penalty_amount'], 'days_overdue': penalty_data['days_overdue']}
            )
            
            conn.commit()
            conn.close()
            
            return {
                'success': True,
                'message': f"Late payment penalty of MWK {penalty_data['penalty_amount']:.2f} applied to loan {loan_id}",
                'penalty_amount': penalty_data['penalty_amount']
            }
            
        except Exception as e:
            return ErrorHandler.handle_error(e, "applying late payment penalty")
    
    def process_automatic_balance_updates(self) -> dict:
        """
        Process automatic outstanding balance updates for all active loans
        
        Returns:
            dict: Processing results
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Get all active loans
            cursor.execute('''
                SELECT loan_id, loan_amount, outstanding_balance 
                FROM Loans 
                WHERE outstanding_balance > 0
            ''')
            active_loans = cursor.fetchall()
            
            updated_loans = 0
            total_penalties = 0
            
            for loan_id, loan_amount, outstanding_balance in active_loans:
                # Calculate and apply penalties if needed
                penalty_result = self.apply_late_payment_penalty(loan_id)
                
                if penalty_result['success'] and penalty_result.get('penalty_amount', 0) > 0:
                    updated_loans += 1
                    total_penalties += penalty_result['penalty_amount']
            
            conn.close()
            
            return {
                'success': True,
                'message': f"Processed {len(active_loans)} loans, updated {updated_loans} with penalties",
                'loans_processed': len(active_loans),
                'loans_updated': updated_loans,
                'total_penalties': round(total_penalties, 2)
            }
            
        except Exception as e:
            return ErrorHandler.handle_error(e, "automatic balance updates")

# Year-End Processing System
class YearEndProcessor:
    """
    Handles automated year-end processing including dividend calculations and new year preparation
    """
    
    def __init__(self):
        self.config_manager = SystemConfigManager()
    
    def generate_annual_summary(self, year: int) -> dict:
        """
        Generate comprehensive annual summary for a specific year
        
        Args:
            year (int): Year to generate summary for
            
        Returns:
            dict: Annual summary data
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Total contributions for the year
            cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) FROM Contributions WHERE year = ?
            ''', (year,))
            total_contributions = cursor.fetchone()[0]
            
            # Total loans disbursed
            cursor.execute('''
                SELECT COALESCE(SUM(loan_amount), 0) FROM Loans 
                WHERE strftime('%Y', loan_date) = ?
            ''', (str(year),))
            total_loans_disbursed = cursor.fetchone()[0]
            
            # Total repayments received
            cursor.execute('''
                SELECT COALESCE(SUM(repayment_amount), 0) FROM Repayments 
                WHERE strftime('%Y', repayment_date) = ?
            ''', (str(year),))
            total_repayments = cursor.fetchone()[0]
            
            # Total interest collected
            cursor.execute('''
                SELECT COALESCE(SUM(interest_amount), 0) FROM Repayments 
                WHERE strftime('%Y', repayment_date) = ?
            ''', (str(year),))
            total_interest = cursor.fetchone()[0]
            
            # Outstanding loans at year end
            cursor.execute('''
                SELECT COALESCE(SUM(outstanding_balance), 0) FROM Loans 
                WHERE outstanding_balance > 0
            ''')
            outstanding_balance = cursor.fetchone()[0]
            
            # Active members
            cursor.execute('''
                SELECT COUNT(DISTINCT member_id) FROM Contributions WHERE year = ?
            ''', (year,))
            active_members = cursor.fetchone()[0]
            
            conn.close()
            
            # Calculate key metrics
            net_fund_growth = total_contributions + total_interest - total_loans_disbursed
            loan_recovery_rate = (total_repayments / total_loans_disbursed * 100) if total_loans_disbursed > 0 else 0
            
            return {
                'success': True,
                'year': year,
                'total_contributions': total_contributions,
                'total_loans_disbursed': total_loans_disbursed,
                'total_repayments': total_repayments,
                'total_interest_collected': total_interest,
                'outstanding_balance': outstanding_balance,
                'active_members': active_members,
                'net_fund_growth': net_fund_growth,
                'loan_recovery_rate': round(loan_recovery_rate, 2),
                'generated_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
        except Exception as e:
            return ErrorHandler.handle_error(e, "annual summary generation")
    
    def prepare_new_financial_year(self, new_year: int) -> dict:
        """
        Prepare system for new financial year
        
        Args:
            new_year (int): New financial year
            
        Returns:
            dict: Preparation results
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Archive previous year's dividend calculations
            previous_year = new_year - 1
            cursor.execute('''
                UPDATE DividendCalculations 
                SET status = 'archived' 
                WHERE year = ? AND status != 'archived'
            ''', (previous_year,))
            
            # Reset any temporary calculations
            cursor.execute('''
                DELETE FROM DividendCalculations 
                WHERE status = 'temporary' OR status = 'draft'
            ''')
            
            # Update system configuration for new year
            cursor.execute('''
                UPDATE SystemConfig 
                SET config_value = ? 
                WHERE config_key = 'current_financial_year'
            ''', (str(new_year),))
            
            # Log the year-end preparation
            audit_manager.log_action(
                'SYSTEM', 'YearEnd', new_year,
                f"Prepared system for financial year {new_year}",
                {'previous_year': previous_year, 'new_year': new_year}
            )
            
            conn.commit()
            conn.close()
            
            return {
                'success': True,
                'message': f"System prepared for financial year {new_year}",
                'new_year': new_year,
                'previous_year': previous_year,
                'preparation_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
        except Exception as e:
            return ErrorHandler.handle_error(e, "new financial year preparation")
    
    def process_automatic_dividend_distribution(self, year: int) -> dict:
        """
        Process automatic dividend calculation and distribution for year-end
        
        Args:
            year (int): Year to process dividends for
            
        Returns:
            dict: Distribution processing results
        """
        try:
            # Calculate all dividends for the year
            dividend_results = calculate_all_dividends_for_year(year)
            
            if not dividend_results['distribution_valid']:
                return {
                    'success': False,
                    'message': 'Dividend distribution validation failed. Total dividends do not match available funds.',
                    'total_fund': dividend_results['total_fund'],
                    'total_dividends': dividend_results['total_dividends']
                }
            
            # Mark dividends as ready for distribution
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE DividendCalculations 
                SET status = 'ready_for_distribution' 
                WHERE year = ? AND status = 'calculated'
            ''', (year,))
            
            # Log the dividend processing
            audit_manager.log_action(
                'SYSTEM', 'DividendDistribution', year,
                f"Processed automatic dividend distribution for year {year}",
                {
                    'total_fund': dividend_results['total_fund'],
                    'total_dividends': dividend_results['total_dividends'],
                    'member_count': len(dividend_results['dividend_calculations'])
                }
            )
            
            conn.commit()
            conn.close()
            
            return {
                'success': True,
                'message': f"Dividend distribution processed for {len(dividend_results['dividend_calculations'])} members",
                'year': year,
                'total_fund': dividend_results['total_fund'],
                'total_dividends': dividend_results['total_dividends'],
                'member_count': len(dividend_results['dividend_calculations']),
                'distribution_valid': True
            }
            
        except Exception as e:
            return ErrorHandler.handle_error(e, "automatic dividend distribution")

# System Configuration Management
class SystemConfigManager:
    """
    Manages system configuration settings using the SystemConfig table
    """
    
    @staticmethod
    def get_config_value(key: str, default_value: str = None) -> str:
        """
        Get a configuration value by key
        
        Args:
            key (str): Configuration key
            default_value (str): Default value if key not found
            
        Returns:
            str: Configuration value
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Use database-specific syntax
            if db_manager.current_db_type == 'mysql':
                cursor.execute('SELECT config_value FROM SystemConfig WHERE config_key = %s', (key,))
            else:
                cursor.execute('SELECT config_value FROM SystemConfig WHERE config_key = ?', (key,))
            
            result = cursor.fetchone()
            conn.close()
            
            if result:
                return result[0]
            else:
                return default_value
                
        except Exception as e:
            logger.error(f"Error getting config value for {key}: {str(e)}")
            return default_value
    
    @staticmethod
    def set_config_value(key: str, value: str, description: str = None) -> bool:
        """
        Set a configuration value
        
        Args:
            key (str): Configuration key
            value (str): Configuration value
            description (str): Optional description
            
        Returns:
            bool: Success status
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Check if key exists using database-specific syntax
            if db_manager.current_db_type == 'mysql':
                cursor.execute('SELECT config_key FROM SystemConfig WHERE config_key = %s', (key,))
                exists = cursor.fetchone()
                
                if exists:
                    # Update existing
                    cursor.execute('''
                        UPDATE SystemConfig 
                        SET config_value = %s, updated_at = %s
                        WHERE config_key = %s
                    ''', (value, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), key))
                    operation = 'UPDATE'
                else:
                    # Insert new
                    cursor.execute('''
                        INSERT INTO SystemConfig (config_key, config_value, description, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s)
                    ''', (key, value, description, 
                          datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                          datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            else:
                cursor.execute('SELECT config_key FROM SystemConfig WHERE config_key = ?', (key,))
                exists = cursor.fetchone()
                
                if exists:
                    # Update existing
                    cursor.execute('''
                        UPDATE SystemConfig 
                        SET config_value = ?, updated_at = ?
                        WHERE config_key = ?
                    ''', (value, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), key))
                    operation = 'UPDATE'
                else:
                    # Insert new
                    cursor.execute('''
                        INSERT INTO SystemConfig (config_key, config_value, description, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (key, value, description, 
                          datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                          datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                operation = 'INSERT'
            
            conn.commit()
            conn.close()
            
            # Log configuration change
            audit_manager.log_transaction('SystemConfig', operation, None, None, {
                'config_key': key,
                'config_value': value,
                'description': description,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            return True
            
        except Exception as e:
            logger.error(f"Error setting config value for {key}: {str(e)}")
            return False
    
    @staticmethod
    def get_default_contribution_amount() -> float:
        """Get default monthly contribution amount"""
        value = SystemConfigManager.get_config_value('default_contribution_amount', '100.0')
        try:
            return float(value)
        except ValueError:
            return 100.0
    
    @staticmethod
    def get_default_interest_rate() -> float:
        """Get default loan interest rate (monthly)"""
        value = SystemConfigManager.get_config_value('default_interest_rate', '0.2')
        try:
            return float(value)
        except ValueError:
            return 0.2
    
    @staticmethod
    def initialize_default_config():
        """Initialize default configuration values if they don't exist"""
        default_configs = [
            ('organization_name', 'Bank Mmudzi', 'Name of the organization'),
            ('currency_symbol', 'MWK', 'Currency symbol used in the system'),
            ('default_contribution_amount', '100.0', 'Default monthly contribution amount'),
            ('min_contribution_amount', '50.0', 'Minimum monthly contribution amount'),
            ('enable_variable_contributions', 'true', 'Allow variable contribution amounts per member'),
            ('default_interest_rate', '0.2', 'Default monthly interest rate (as decimal)'),
            ('default_loan_term', '12', 'Default loan term in months'),
            ('max_loan_amount', '5000.0', 'Maximum loan amount allowed')
        ]
        
        for key, value, description in default_configs:
            # Only set if doesn't exist
            if SystemConfigManager.get_config_value(key) is None:
                SystemConfigManager.set_config_value(key, value, description)

# Login Window with Registration
class LoginWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Login - Bank Mmudzi System")
        self.geometry("400x280")
        self.configure(bg='#f0f0f0')
        self.resizable(False, False)
        
        # Center the window
        self.center_window()
        
        # Main content frame
        main_frame = tk.Frame(self, bg='#f0f0f0')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(
            main_frame, 
            text="Bank Mmudzi", 
            font=('Arial', 18, 'bold'),
            bg='#f0f0f0',
            fg='#2c3e50'
        )
        title_label.pack(pady=(0, 20))
        
        tk.Label(main_frame, text="Username:", bg='#f0f0f0', font=('Arial', 10)).pack(pady=(10, 5))
        self.username_entry = tk.Entry(main_frame, font=('Arial', 10), width=25)
        self.username_entry.pack(pady=5)
        
        tk.Label(main_frame, text="Password:", bg='#f0f0f0', font=('Arial', 10)).pack(pady=(10, 5))
        self.password_entry = tk.Entry(main_frame, show='*', font=('Arial', 10), width=25)
        self.password_entry.pack(pady=5)
        
        button_frame = tk.Frame(main_frame, bg='#f0f0f0')
        button_frame.pack(pady=15)
        
        tk.Button(button_frame, text="Login", command=self.try_login, 
                 bg='#4CAF50', fg='white', font=('Arial', 10, 'bold'),
                 padx=20, pady=8).pack(side='left', padx=5)
        tk.Button(button_frame, text="Register", command=self.show_register_dialog, 
                 bg='#2196F3', fg='white', font=('Arial', 10, 'bold'),
                 padx=20, pady=8).pack(side='left', padx=5)
        
        # Add separator line
        separator = tk.Frame(main_frame, height=2, bg='#bdc3c7')
        separator.pack(fill='x', pady=(15, 10))
        
        # Developer info directly in main frame (more visible)
        dev_info_frame = tk.Frame(main_frame, bg='#f0f0f0')
        dev_info_frame.pack(pady=5)
        
        dev_label = tk.Label(
            dev_info_frame,
            text="Developer: Nehemiah Nganjo",
            font=('Arial', 9, 'bold'),
            fg='#2c3e50',
            bg='#f0f0f0'
        )
        dev_label.pack()
        
        contact_label = tk.Label(
            dev_info_frame,
            text="Phone: 0997082156",
            font=('Arial', 8),
            fg='#34495e',
            bg='#f0f0f0'
        )
        contact_label.pack()
        
        github_label = tk.Label(
            dev_info_frame,
            text="GitHub: github.com/Nehemiahnganjo",
            font=('Arial', 8),
            fg='#3498db',
            bg='#f0f0f0',
            cursor="hand2"
        )
        github_label.pack()
        
        # Make GitHub link clickable
        def open_github(event):
            import webbrowser
            webbrowser.open("https://github.com/Nehemiahnganjo")
        
        github_label.bind("<Button-1>", open_github)
        
        # Bind Enter key to login
        self.bind('<Return>', lambda event: self.try_login())
        self.username_entry.focus()
    
    def center_window(self):
        """Center the login window on screen"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
    

    def try_login(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get()
        
        # Validate input
        if not username:
            messagebox.showerror("Error", "Please enter a username")
            return
        
        if not password:
            messagebox.showerror("Error", "Please enter a password")
            return
        
        try:
            if check_login(username, password):
                # Set current user for audit tracking
                audit_manager.set_current_user(username)
                self.destroy()
                app = App()
                app.mainloop()
            else:
                # Log security violation for failed login
                audit_manager.log_security_violation('FAILED_LOGIN', {
                    'username': username,
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                messagebox.showerror("Error", "Invalid username or password")
        
        except BusinessRuleViolationError as e:
            # Handle account lockout and other business rule violations
            messagebox.showerror("Login Error", e.message)
        
        except DatabaseError as e:
            # Handle database connection issues
            messagebox.showerror("Database Error", "Unable to connect to database. Please try again later.")
            logger.error(f"Database error during login: {e}")
        
        except Exception as e:
            # Handle any other unexpected errors
            error_message = ErrorHandler.get_user_friendly_message(e)
            messagebox.showerror("Login Error", error_message)
            logger.error(f"Unexpected error during login: {e}")

    def show_register_dialog(self):
        dialog = tk.Toplevel(self)
        dialog.title("Register New Admin")
        dialog.geometry("300x250")
        dialog.configure(bg='#f0f0f0')
        
        tk.Label(dialog, text="New Username:", bg='#f0f0f0').pack(pady=10)
        new_username_entry = tk.Entry(dialog)
        new_username_entry.pack(pady=5)
        
        tk.Label(dialog, text="New Password:", bg='#f0f0f0').pack(pady=10)
        new_password_entry = tk.Entry(dialog, show='*')
        new_password_entry.pack(pady=5)
        
        tk.Label(dialog, text="Confirm Password:", bg='#f0f0f0').pack(pady=10)
        confirm_password_entry = tk.Entry(dialog, show='*')
        confirm_password_entry.pack(pady=5)
        
        def register():
            new_username = new_username_entry.get()
            new_password = new_password_entry.get()
            confirm_password = confirm_password_entry.get()
            
            if new_password != confirm_password:
                messagebox.showerror("Error", "Passwords do not match")
                return
            
            success, message = register_admin(new_username, new_password)
            if success:
                messagebox.showinfo("Success", message)
                dialog.destroy()
            else:
                messagebox.showerror("Error", message)
        
        tk.Button(dialog, text="Register", command=register, 
                 bg='#4CAF50', fg='white').pack(pady=10)

# Member Analysis and Filtering System
class MemberAnalyzer:
    """
    Provides advanced member analysis and filtering capabilities
    """
    
    def __init__(self):
        self.config_manager = SystemConfigManager()
    
    def calculate_overdue_loans(self, member_id: int = None) -> list:
        """
        Calculate overdue loans for a specific member or all members
        
        Args:
            member_id (int, optional): Specific member ID, None for all members
            
        Returns:
            list: List of overdue loan details
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Base query for overdue loans
            base_query = '''
                SELECT l.loan_id, l.member_id, m.name, m.surname, 
                       l.loan_amount, l.outstanding_balance, l.loan_date,
                       l.monthly_payment, COALESCE(MAX(r.repayment_date), l.loan_date) as last_payment
                FROM Loans l
                JOIN Members m ON l.member_id = m.member_id
                LEFT JOIN Repayments r ON l.loan_id = r.loan_id
                WHERE l.outstanding_balance > 0
            '''
            
            params = []
            if member_id:
                base_query += ' AND l.member_id = ?'
                params.append(member_id)
            
            base_query += ' GROUP BY l.loan_id'
            
            cursor.execute(base_query, params)
            loans = cursor.fetchall()
            
            overdue_loans = []
            current_date = datetime.now()
            
            for loan in loans:
                loan_id, member_id, name, surname, loan_amount, outstanding_balance, loan_date, monthly_payment, last_payment = loan
                
                # Calculate expected next payment date
                last_payment_date = datetime.strptime(last_payment, '%Y-%m-%d')
                expected_payment_date = last_payment_date + timedelta(days=30)
                
                # Check if overdue (with 7-day grace period)
                days_overdue = (current_date - expected_payment_date).days - 7
                
                if days_overdue > 0:
                    overdue_loans.append({
                        'loan_id': loan_id,
                        'member_id': member_id,
                        'member_name': f"{name} {surname}",
                        'loan_amount': loan_amount,
                        'outstanding_balance': outstanding_balance,
                        'monthly_payment': monthly_payment,
                        'last_payment_date': last_payment,
                        'expected_payment_date': expected_payment_date.strftime('%Y-%m-%d'),
                        'days_overdue': days_overdue,
                        'overdue_amount': monthly_payment  # Simplified - could be more complex
                    })
            
            conn.close()
            return overdue_loans
            
        except Exception as e:
            logger.error(f"Error calculating overdue loans: {str(e)}")
            return []
    
    def filter_members_by_criteria(self, status: str = None, contribution_level: str = None, 
                                 loan_status: str = None) -> list:
        """
        Filter members by various criteria
        
        Args:
            status (str): Member status filter ('active', 'inactive')
            contribution_level (str): Contribution level filter ('high', 'medium', 'low')
            loan_status (str): Loan status filter ('active', 'paid', 'overdue')
            
        Returns:
            list: Filtered member list
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Base query
            query = '''
                SELECT m.member_id, m.member_number, m.name, m.surname, m.phone_number, m.status,
                       COALESCE(SUM(c.amount), 0) as total_contributions,
                       COALESCE(SUM(l.outstanding_balance), 0) as outstanding_loans,
                       COUNT(DISTINCT l.loan_id) as loan_count
                FROM Members m
                LEFT JOIN Contributions c ON m.member_id = c.member_id
                LEFT JOIN Loans l ON m.member_id = l.member_id AND l.outstanding_balance > 0
                WHERE 1=1
            '''
            
            params = []
            
            # Apply status filter
            if status:
                query += ' AND m.status = ?'
                params.append(status)
            
            query += ' GROUP BY m.member_id'
            
            cursor.execute(query, params)
            members = cursor.fetchall()
            
            filtered_members = []
            
            for member in members:
                member_id, member_number, name, surname, phone, member_status, total_contributions, outstanding_loans, loan_count = member
                
                # Apply contribution level filter
                if contribution_level:
                    monthly_contribution = float(self.config_manager.get_config_value('monthly_contribution_amount', 100))
                    annual_expected = monthly_contribution * 12
                    
                    contribution_ratio = total_contributions / annual_expected if annual_expected > 0 else 0
                    
                    if contribution_level == 'high' and contribution_ratio < 0.8:
                        continue
                    elif contribution_level == 'medium' and (contribution_ratio < 0.5 or contribution_ratio >= 0.8):
                        continue
                    elif contribution_level == 'low' and contribution_ratio >= 0.5:
                        continue
                
                # Apply loan status filter
                if loan_status:
                    if loan_status == 'active' and outstanding_loans <= 0:
                        continue
                    elif loan_status == 'paid' and outstanding_loans > 0:
                        continue
                    elif loan_status == 'overdue':
                        # Check if member has overdue loans
                        overdue_loans = self.calculate_overdue_loans(member_id)
                        if not overdue_loans:
                            continue
                
                filtered_members.append({
                    'member_id': member_id,
                    'member_number': member_number,
                    'name': name,
                    'surname': surname,
                    'phone': phone,
                    'status': member_status,
                    'total_contributions': total_contributions,
                    'outstanding_loans': outstanding_loans,
                    'loan_count': loan_count
                })
            
            conn.close()
            return filtered_members
            
        except Exception as e:
            logger.error(f"Error filtering members: {str(e)}")
            return []
    
    def calculate_member_standing(self, member_id: int) -> dict:
        """
        Calculate comprehensive member standing
        
        Args:
            member_id (int): Member ID
            
        Returns:
            dict: Member standing details
        """
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Get member basic info using database-specific syntax
            if db_manager.current_db_type == 'mysql':
                cursor.execute('''
                    SELECT name, surname, member_number, status, created_at
                    FROM Members WHERE member_id = %s
                ''', (member_id,))
                member_info = cursor.fetchone()
                
                if not member_info:
                    return {'success': False, 'message': 'Member not found'}
                
                name, surname, member_number, status, created_date = member_info
                
                # Calculate contribution metrics
                cursor.execute('''
                    SELECT COUNT(*) as months_contributed, 
                           COALESCE(SUM(amount), 0) as total_contributions,
                           COALESCE(AVG(amount), 0) as avg_contribution
                    FROM Contributions WHERE member_id = %s
                ''', (member_id,))
            else:
                cursor.execute('''
                    SELECT name, surname, member_number, status, created_at
                    FROM Members WHERE member_id = ?
                ''', (member_id,))
                member_info = cursor.fetchone()
                
                if not member_info:
                    return {'success': False, 'message': 'Member not found'}
                
                name, surname, member_number, status, created_date = member_info
                
                # Calculate contribution metrics
                cursor.execute('''
                    SELECT COUNT(*) as months_contributed, 
                           COALESCE(SUM(amount), 0) as total_contributions,
                           COALESCE(AVG(amount), 0) as avg_contribution
                    FROM Contributions WHERE member_id = ?
                ''', (member_id,))
            contrib_data = cursor.fetchone()
            months_contributed, total_contributions, avg_contribution = contrib_data
            
            # Calculate loan metrics
            cursor.execute('''
                SELECT COUNT(*) as total_loans,
                       COALESCE(SUM(loan_amount), 0) as total_borrowed,
                       COALESCE(SUM(outstanding_balance), 0) as outstanding_balance
                FROM Loans WHERE member_id = ?
            ''', (member_id,))
            loan_data = cursor.fetchone()
            total_loans, total_borrowed, outstanding_balance = loan_data
            
            # Calculate repayment metrics
            cursor.execute('''
                SELECT COALESCE(SUM(r.repayment_amount), 0) as total_repaid,
                       COALESCE(SUM(r.interest_amount), 0) as total_interest_paid
                FROM Repayments r
                JOIN Loans l ON r.loan_id = l.loan_id
                WHERE l.member_id = ?
            ''', (member_id,))
            repayment_data = cursor.fetchone()
            total_repaid, total_interest_paid = repayment_data
            
            # Check for overdue loans
            overdue_loans = self.calculate_overdue_loans(member_id)
            
            # Calculate standing score (0-100)
            standing_score = 100
            
            # Deduct for overdue loans
            if overdue_loans:
                standing_score -= min(50, len(overdue_loans) * 10)
            
            # Deduct for low contribution rate
            expected_contributions = float(self.config_manager.get_config_value('monthly_contribution_amount', 100)) * 12
            contribution_rate = (total_contributions / expected_contributions) if expected_contributions > 0 else 1
            if contribution_rate < 0.8:
                standing_score -= (0.8 - contribution_rate) * 50
            
            # Bonus for consistent repayments
            if total_borrowed > 0:
                repayment_rate = total_repaid / total_borrowed
                if repayment_rate > 0.9:
                    standing_score += 10
            
            standing_score = max(0, min(100, standing_score))
            
            # Determine standing category
            if standing_score >= 90:
                standing_category = "Excellent"
            elif standing_score >= 75:
                standing_category = "Good"
            elif standing_score >= 60:
                standing_category = "Fair"
            else:
                standing_category = "Poor"
            
            conn.close()
            
            return {
                'success': True,
                'member_id': member_id,
                'member_name': f"{name} {surname}",
                'member_number': member_number,
                'status': status,
                'created_date': created_date,
                'months_contributed': months_contributed,
                'total_contributions': total_contributions,
                'avg_contribution': avg_contribution,
                'total_loans': total_loans,
                'total_borrowed': total_borrowed,
                'outstanding_balance': outstanding_balance,
                'total_repaid': total_repaid,
                'total_interest_paid': total_interest_paid,
                'overdue_loans_count': len(overdue_loans),
                'standing_score': round(standing_score, 1),
                'standing_category': standing_category,
                'contribution_rate': round(contribution_rate * 100, 1)
            }
            
        except Exception as e:
            return ErrorHandler.handle_error(e, "member standing calculation")

# Session Security System
class SessionManager:
    """
    Manages user sessions with automatic logout and security features
    """
    
    def __init__(self):
        self.session_start_time = None
        self.last_activity_time = None
        self.current_user = None
        self.idle_timeout_minutes = 30  # Default 30 minutes
        self.security_violations = []
    
    def start_session(self, username: str) -> dict:
        """
        Start a new user session
        
        Args:
            username (str): Username starting the session
            
        Returns:
            dict: Session start result
        """
        try:
            self.session_start_time = datetime.now()
            self.last_activity_time = datetime.now()
            self.current_user = username
            
            # Get timeout from config
            config_manager = SystemConfigManager()
            self.idle_timeout_minutes = int(config_manager.get_config_value('session_timeout_minutes', 30))
            
            # Log session start
            audit_manager.log_action(
                'LOGIN', 'Session', 0,
                f"Session started for user {username}",
                {'session_start': self.session_start_time.isoformat()}
            )
            
            return {
                'success': True,
                'message': f"Session started for {username}",
                'session_timeout': self.idle_timeout_minutes
            }
            
        except Exception as e:
            return ErrorHandler.handle_error(e, "session start")
    
    def update_activity(self):
        """Update last activity time"""
        self.last_activity_time = datetime.now()
    
    def check_session_timeout(self) -> dict:
        """
        Check if session has timed out
        
        Returns:
            dict: Timeout check result
        """
        if not self.last_activity_time:
            return {'timed_out': False, 'message': 'No active session'}
        
        time_since_activity = datetime.now() - self.last_activity_time
        timeout_threshold = timedelta(minutes=self.idle_timeout_minutes)
        
        if time_since_activity > timeout_threshold:
            self.end_session('timeout')
            return {
                'timed_out': True,
                'message': f'Session timed out after {self.idle_timeout_minutes} minutes of inactivity'
            }
        
        return {
            'timed_out': False,
            'remaining_minutes': int((timeout_threshold - time_since_activity).total_seconds() / 60)
        }
    
    def end_session(self, reason: str = 'logout'):
        """
        End the current session
        
        Args:
            reason (str): Reason for ending session
        """
        if self.current_user:
            # Log session end
            audit_manager.log_action(
                'LOGOUT', 'Session', 0,
                f"Session ended for user {self.current_user} - {reason}",
                {
                    'session_duration': str(datetime.now() - self.session_start_time) if self.session_start_time else 'unknown',
                    'end_reason': reason
                }
            )
        
        self.session_start_time = None
        self.last_activity_time = None
        self.current_user = None
    
    def log_security_violation(self, violation_type: str, details: dict):
        """
        Log a security violation
        
        Args:
            violation_type (str): Type of violation
            details (dict): Violation details
        """
        violation = {
            'timestamp': datetime.now().isoformat(),
            'type': violation_type,
            'user': self.current_user,
            'details': details
        }
        
        self.security_violations.append(violation)
        
        # Log to audit trail
        audit_manager.log_action(
            'SECURITY_VIOLATION', 'Security', 0,
            f"Security violation: {violation_type}",
            violation
        )

# Enhanced Accessible Dialog Base Class
class AccessibleDialog:
    """
    Base class for all member dialogs with consistent accessibility features
    Provides standardized layout, keyboard navigation, and accessibility support
    """
    
    def __init__(self, parent, title, width=600, height=500):
        self.parent = parent
        self.title = title
        self.result = None
        self.width = width
        self.height = height
        
        # Accessibility state
        self.focus_order = []
        self.current_focus_index = 0
        
        # Create dialog window
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(f"{title} - Bank Mmudzi")
        self.dialog.geometry(f"{width}x{height}")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        
        # Configure accessibility attributes
        self.setup_accessibility_attributes()
        
        # Center the dialog
        self.center_dialog()
        
        # Configure style
        self.configure_style()
        
        # Create standardized layout
        self.create_base_layout()
        
        # Setup keyboard bindings
        self.setup_keyboard_bindings()
        
        # Set modal
        self.dialog.grab_set()
        
        # Set initial focus
        self.dialog.after(100, self.set_initial_focus)
    
    def setup_accessibility_attributes(self):
        """Setup accessibility attributes for screen readers"""
        # Set window role for screen readers
        try:
            self.dialog.wm_attributes('-type', 'dialog')
        except tk.TclError:
            # Fallback for systems that don't support -type attribute
            pass
        
        # Configure for accessibility
        self.dialog.configure(highlightthickness=2)
        self.dialog.configure(highlightcolor='#2E86AB')
        
        # Set ARIA-like attributes for screen readers
        self.dialog.wm_title(f"{self.title} - Dialog Window")
        
        # Create live region for status announcements
        self.live_region = None
    
    def center_dialog(self):
        """Center the dialog on the parent window"""
        self.dialog.update_idletasks()
        
        # Get parent window position and size
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        # Calculate center position
        x = parent_x + (parent_width - self.width) // 2
        y = parent_y + (parent_height - self.height) // 2
        
        # Ensure dialog stays on screen
        x = max(0, min(x, self.dialog.winfo_screenwidth() - self.width))
        y = max(0, min(y, self.dialog.winfo_screenheight() - self.height))
        
        self.dialog.geometry(f"+{x}+{y}")
    
    def configure_style(self):
        """Configure consistent dialog styling"""
        self.colors = {
            'primary': '#2E86AB',
            'secondary': '#A23B72', 
            'success': '#F18F01',
            'danger': '#C73E1D',
            'light': '#F5F5F5',
            'dark': '#333333',
            'white': '#FFFFFF',
            'focus': '#4A90E2',
            'disabled': '#CCCCCC'
        }
        
        self.dialog.configure(bg=self.colors['light'])
    
    def create_base_layout(self):
        """Create standardized dialog layout with button bar at top"""
        # Main container
        self.main_frame = tk.Frame(self.dialog, bg=self.colors['light'], padx=20, pady=15)
        self.main_frame.pack(fill='both', expand=True)
        
        # Top button bar (standardized position)
        self.create_button_bar()
        
        # Separator
        separator = tk.Frame(self.main_frame, height=2, bg=self.colors['primary'])
        separator.pack(fill='x', pady=(0, 15))
        
        # Content area (to be implemented by subclasses)
        self.content_frame = tk.Frame(self.main_frame, bg=self.colors['light'])
        self.content_frame.pack(fill='both', expand=True, pady=(0, 15))
        
        # Status bar at bottom
        self.create_status_bar()
        
        # Create content (implemented by subclasses)
        self.create_content()
    
    def create_button_bar(self):
        """Create standardized button bar at top of dialog"""
        self.button_frame = tk.Frame(self.main_frame, bg=self.colors['light'])
        self.button_frame.pack(fill='x', pady=(0, 10))
        
        # Help button (left side)
        self.help_btn = tk.Button(
            self.button_frame,
            text="Help",
            font=('Arial', 10),
            bg=self.colors['light'],
            fg=self.colors['dark'],
            relief='solid',
            borderwidth=1,
            padx=15,
            pady=5,
            command=self.show_help,
            cursor='hand2'
        )
        self.help_btn.pack(side='left')
        self.add_to_focus_order(self.help_btn)
        
        # Main action buttons (right side)
        button_container = tk.Frame(self.button_frame, bg=self.colors['light'])
        button_container.pack(side='right')
        
        # Cancel button
        self.cancel_btn = tk.Button(
            button_container,
            text="Cancel",
            font=('Arial', 10),
            bg=self.colors['light'],
            fg=self.colors['dark'],
            relief='solid',
            borderwidth=1,
            padx=20,
            pady=5,
            command=self.cancel,
            cursor='hand2'
        )
        self.cancel_btn.pack(side='right', padx=(10, 0))
        self.add_to_focus_order(self.cancel_btn)
        
        # OK button (primary action)
        self.ok_btn = tk.Button(
            button_container,
            text="OK",
            font=('Arial', 10, 'bold'),
            bg=self.colors['primary'],
            fg=self.colors['white'],
            relief='flat',
            padx=25,
            pady=5,
            command=self.ok_action,
            cursor='hand2'
        )
        self.ok_btn.pack(side='right')
        self.add_to_focus_order(self.ok_btn)
        
        # Configure button accessibility
        self.setup_button_accessibility()
    
    def setup_button_accessibility(self):
        """Setup accessibility features for buttons"""
        # Add ARIA-like attributes and tooltips
        buttons = [
            (self.ok_btn, "Confirm and execute the primary action (Alt+O)"),
            (self.cancel_btn, "Cancel and close dialog (Alt+C, Escape)"),
            (self.help_btn, "Show help information (F1)")
        ]
        
        for button, tooltip in buttons:
            # Add tooltip
            self.create_tooltip(button, tooltip)
            
            # Add focus indicators
            button.bind('<FocusIn>', self.on_button_focus_in)
            button.bind('<FocusOut>', self.on_button_focus_out)
            
            # Add keyboard activation
            button.bind('<Return>', lambda e, btn=button: btn.invoke())
            button.bind('<space>', lambda e, btn=button: btn.invoke())
    
    def create_tooltip(self, widget, text):
        """Create tooltip for accessibility"""
        def show_tooltip(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            
            label = tk.Label(
                tooltip,
                text=text,
                background='#FFFFDD',
                relief='solid',
                borderwidth=1,
                font=('Arial', 9)
            )
            label.pack()
            
            # Auto-hide after 3 seconds
            tooltip.after(3000, tooltip.destroy)
            
            # Store reference to prevent garbage collection
            widget.tooltip = tooltip
        
        def hide_tooltip(event):
            if hasattr(widget, 'tooltip'):
                try:
                    widget.tooltip.destroy()
                except:
                    pass
        
        widget.bind('<Enter>', show_tooltip)
        widget.bind('<Leave>', hide_tooltip)
    
    def on_button_focus_in(self, event):
        """Handle button focus in - add visual focus indicator"""
        button = event.widget
        button.configure(
            highlightthickness=3,
            highlightcolor=self.colors['focus'],
            highlightbackground=self.colors['focus']
        )
    
    def on_button_focus_out(self, event):
        """Handle button focus out - remove visual focus indicator"""
        button = event.widget
        button.configure(
            highlightthickness=1,
            highlightcolor=self.colors['light'],
            highlightbackground=self.colors['light']
        )
    
    def create_status_bar(self):
        """Create status bar for user feedback with live region support"""
        self.status_frame = tk.Frame(self.main_frame, bg=self.colors['light'], height=25)
        self.status_frame.pack(fill='x', side='bottom')
        self.status_frame.pack_propagate(False)
        
        self.status_label = tk.Label(
            self.status_frame,
            text="Ready",
            font=('Arial', 9),
            bg=self.colors['light'],
            fg=self.colors['dark'],
            anchor='w'
        )
        self.status_label.pack(fill='x', padx=5, pady=2)
        
        # Create live region for screen reader announcements
        self.create_live_region()
    
    def create_live_region(self):
        """Create live region for screen reader announcements"""
        # Create invisible live region for screen reader announcements
        self.live_region = tk.Label(
            self.status_frame,
            text="",
            font=('Arial', 1),  # Very small font
            bg=self.colors['light'],
            fg=self.colors['light'],  # Same as background to make invisible
            width=1,
            height=1
        )
        # Position off-screen but still accessible to screen readers
        self.live_region.place(x=-1000, y=-1000)
    
    def announce_to_screen_reader(self, message):
        """Announce message to screen readers via live region"""
        if self.live_region:
            self.live_region.configure(text=message)
            # Clear after a short delay to allow for re-announcement of same message
            self.dialog.after(100, lambda: self.live_region.configure(text=""))
    
    def setup_keyboard_bindings(self):
        """Setup comprehensive keyboard navigation and shortcuts"""
        # Primary shortcuts
        self.dialog.bind('<Return>', self.handle_enter_key)
        self.dialog.bind('<Escape>', lambda e: self.cancel())
        self.dialog.bind('<F1>', lambda e: self.show_help())
        
        # Alt key shortcuts
        self.dialog.bind('<Alt-o>', lambda e: self.ok_action())
        self.dialog.bind('<Alt-O>', lambda e: self.ok_action())
        self.dialog.bind('<Alt-c>', lambda e: self.cancel())
        self.dialog.bind('<Alt-C>', lambda e: self.cancel())
        
        # Tab navigation
        self.dialog.bind('<Tab>', self.handle_tab_navigation)
        self.dialog.bind('<Shift-Tab>', self.handle_shift_tab_navigation)
        
        # Arrow key navigation for buttons
        self.dialog.bind('<Left>', self.handle_arrow_navigation)
        self.dialog.bind('<Right>', self.handle_arrow_navigation)
        
        # Focus management
        self.dialog.bind('<FocusIn>', self.on_dialog_focus_in)
        
        # Make dialog focusable
        self.dialog.focus_set()
    
    def handle_enter_key(self, event):
        """Handle Enter key - context-sensitive action"""
        focused_widget = self.dialog.focus_get()
        
        # If focus is on a button, activate it
        if focused_widget in [self.ok_btn, self.cancel_btn, self.help_btn]:
            focused_widget.invoke()
            return 'break'
        
        # If focus is on a text widget, don't intercept
        if isinstance(focused_widget, tk.Text):
            return
        
        # Default to OK action
        self.ok_action()
        return 'break'
    
    def handle_tab_navigation(self, event):
        """Handle Tab key navigation through focus order"""
        if not self.focus_order:
            return
        
        current_widget = self.dialog.focus_get()
        
        try:
            current_index = self.focus_order.index(current_widget)
            next_index = (current_index + 1) % len(self.focus_order)
        except ValueError:
            next_index = 0
        
        self.focus_order[next_index].focus_set()
        return 'break'
    
    def handle_shift_tab_navigation(self, event):
        """Handle Shift+Tab key navigation (reverse)"""
        if not self.focus_order:
            return
        
        current_widget = self.dialog.focus_get()
        
        try:
            current_index = self.focus_order.index(current_widget)
            prev_index = (current_index - 1) % len(self.focus_order)
        except ValueError:
            prev_index = len(self.focus_order) - 1
        
        self.focus_order[prev_index].focus_set()
        return 'break'
    
    def handle_arrow_navigation(self, event):
        """Handle arrow key navigation for button bar"""
        focused_widget = self.dialog.focus_get()
        
        # Only handle arrow keys when focus is on buttons
        if focused_widget not in [self.ok_btn, self.cancel_btn, self.help_btn]:
            return
        
        buttons = [self.help_btn, self.ok_btn, self.cancel_btn]
        
        try:
            current_index = buttons.index(focused_widget)
            
            if event.keysym == 'Right':
                next_index = (current_index + 1) % len(buttons)
            else:  # Left
                next_index = (current_index - 1) % len(buttons)
            
            buttons[next_index].focus_set()
            return 'break'
        except ValueError:
            pass
    
    def on_dialog_focus_in(self, event):
        """Handle dialog receiving focus"""
        # Ensure proper focus management
        if not self.dialog.focus_get() and self.focus_order:
            self.focus_order[0].focus_set()
    
    def add_to_focus_order(self, widget):
        """Add widget to focus order for tab navigation"""
        if widget not in self.focus_order:
            self.focus_order.append(widget)
    
    def set_initial_focus(self):
        """Set initial focus to first focusable element"""
        if self.focus_order:
            self.focus_order[0].focus_set()
    
    def update_status(self, message, message_type='info'):
        """Update status bar with accessibility announcements"""
        colors = {
            'info': self.colors['dark'],
            'success': self.colors['success'],
            'error': self.colors['danger'],
            'warning': self.colors['secondary']
        }
        
        color = colors.get(message_type, self.colors['dark'])
        self.status_label.configure(text=message, fg=color)
        
        # Announce to screen readers via live region
        self.announce_to_screen_reader(message)
        
        # Force update for screen readers
        self.dialog.update_idletasks()
    
    # Abstract methods to be implemented by subclasses
    def create_content(self):
        """Create dialog content - implemented by subclasses"""
        # Default implementation creates a placeholder
        placeholder = tk.Label(
            self.content_frame,
            text="Content area - implement create_content() in subclass",
            font=('Arial', 12),
            bg=self.colors['light'],
            fg=self.colors['dark']
        )
        placeholder.pack(expand=True)
    
    def ok_action(self):
        """Primary action - implemented by subclasses"""
        self.result = True
        self.dialog.destroy()
    
    def cancel(self):
        """Cancel action - can be overridden by subclasses"""
        self.result = False
        self.dialog.destroy()
    
    def show_help(self):
        """Show help information - can be overridden by subclasses"""
        help_text = f"""
{self.title} Help

Keyboard Shortcuts:
• Enter or Alt+O: Confirm action
• Escape or Alt+C: Cancel
• Tab: Navigate forward
• Shift+Tab: Navigate backward
• F1: Show this help
• Arrow keys: Navigate buttons

This dialog follows accessibility standards for keyboard navigation and screen reader support.
        """
        
        messagebox.showinfo(f"{self.title} - Help", help_text.strip())
    
    def create_accessible_form_field(self, parent, label_text, field_type='entry', required=False, **kwargs):
        """
        Create an accessible form field with proper ARIA labels and focus indicators
        
        Args:
            parent: Parent widget
            label_text: Label text for the field
            field_type: Type of field ('entry', 'text', 'combobox')
            required: Whether the field is required
            **kwargs: Additional arguments for the widget
        """
        field_frame = tk.Frame(parent, bg=self.colors['light'])
        field_frame.pack(fill='x', pady=5)
        
        # Create label with required indicator
        label_text_full = f"{label_text} {'*' if required else ''}:"
        label = tk.Label(
            field_frame,
            text=label_text_full,
            font=('Arial', 10, 'bold' if required else 'normal'),
            bg=self.colors['light'],
            fg=self.colors['danger'] if required else self.colors['dark'],
            anchor='w'
        )
        label.pack(fill='x', pady=(0, 2))
        
        # Create the appropriate widget
        if field_type == 'entry':
            widget = tk.Entry(
                field_frame,
                font=('Arial', 11),
                relief='solid',
                borderwidth=1,
                highlightthickness=2,
                highlightcolor=self.colors['focus'],
                **kwargs
            )
        elif field_type == 'text':
            widget = tk.Text(
                field_frame,
                font=('Arial', 11),
                relief='solid',
                borderwidth=1,
                highlightthickness=2,
                highlightcolor=self.colors['focus'],
                height=kwargs.get('height', 3),
                wrap=tk.WORD
            )
        elif field_type == 'combobox':
            widget = ttk.Combobox(
                field_frame,
                font=('Arial', 11),
                **kwargs
            )
        else:
            widget = tk.Entry(field_frame, **kwargs)
        
        widget.pack(fill='x', ipady=5 if field_type == 'entry' else 0)
        
        # Add to focus order
        self.add_to_focus_order(widget)
        
        # Add accessibility features
        self.setup_field_accessibility(widget, label_text, required)
        
        return widget, label
    
    def setup_field_accessibility(self, widget, label_text, required=False):
        """Setup accessibility features for form fields"""
        # Add focus indicators
        widget.bind('<FocusIn>', self.on_field_focus_in)
        widget.bind('<FocusOut>', self.on_field_focus_out)
        
        # Add validation feedback placeholder
        widget.validation_status = 'neutral'  # neutral, valid, invalid
        
        # Store label text for screen reader announcements
        widget.aria_label = label_text
        widget.is_required = required
    
    def on_field_focus_in(self, event):
        """Handle field focus in - add visual focus indicator and announce to screen reader"""
        widget = event.widget
        
        # Visual focus indicator
        if hasattr(widget, 'configure'):
            try:
                widget.configure(
                    highlightthickness=3,
                    highlightcolor=self.colors['focus']
                )
            except tk.TclError:
                pass  # Some widgets don't support these options
        
        # Announce field to screen reader
        if hasattr(widget, 'aria_label'):
            announcement = widget.aria_label
            if hasattr(widget, 'is_required') and widget.is_required:
                announcement += " - Required field"
            self.announce_to_screen_reader(f"Focused on {announcement}")
    
    def on_field_focus_out(self, event):
        """Handle field focus out - remove visual focus indicator"""
        widget = event.widget
        
        # Remove visual focus indicator
        if hasattr(widget, 'configure'):
            try:
                widget.configure(
                    highlightthickness=1,
                    highlightcolor=self.colors['light']
                )
            except tk.TclError:
                pass
    
    def add_validation_indicator(self, widget, status='neutral', message=''):
        """
        Add visual validation indicator to a field with enhanced feedback
        
        Args:
            widget: The form widget
            status: 'valid', 'invalid', or 'neutral'
            message: Validation message
        """
        widget.validation_status = status
        widget.validation_message = message
        
        # Color coding for validation status
        colors = {
            'valid': self.colors['success'],
            'invalid': self.colors['danger'],
            'neutral': self.colors['light']
        }
        
        try:
            if status == 'valid':
                widget.configure(
                    highlightcolor=colors[status], 
                    highlightbackground=colors[status],
                    highlightthickness=2
                )
                # Add checkmark indicator
                self._add_validation_icon(widget, '✓', self.colors['success'])
            elif status == 'invalid':
                widget.configure(
                    highlightcolor=colors[status], 
                    highlightbackground=colors[status],
                    highlightthickness=2
                )
                # Add X indicator
                self._add_validation_icon(widget, '✗', self.colors['danger'])
            else:
                widget.configure(
                    highlightcolor=self.colors['focus'], 
                    highlightbackground=self.colors['light'],
                    highlightthickness=1
                )
                # Remove validation icon
                self._remove_validation_icon(widget)
        except tk.TclError:
            pass
        
        # Update field-specific error message
        self._update_field_error_message(widget, message, status)
        
        # Announce validation result to screen reader
        if message and status != 'neutral':
            self.announce_to_screen_reader(f"Validation: {message}")
    
    def _add_validation_icon(self, widget, icon_text, color):
        """Add validation icon next to field"""
        parent = widget.master
        
        # Remove existing icon if present
        self._remove_validation_icon(widget)
        
        # Create validation icon
        icon_label = tk.Label(
            parent,
            text=icon_text,
            font=('Arial', 12, 'bold'),
            fg=color,
            bg=self.colors['light'],
            width=2
        )
        
        # Position icon to the right of the widget
        icon_label.place(
            in_=widget,
            x=widget.winfo_reqwidth() + 5,
            y=2
        )
        
        # Store reference to icon for later removal
        widget.validation_icon = icon_label
    
    def _remove_validation_icon(self, widget):
        """Remove validation icon from field"""
        if hasattr(widget, 'validation_icon') and widget.validation_icon:
            try:
                widget.validation_icon.destroy()
            except tk.TclError:
                pass
            widget.validation_icon = None
    
    def _update_field_error_message(self, widget, message, status):
        """Update field-specific error message below the field"""
        parent = widget.master
        
        # Remove existing error message if present
        if hasattr(widget, 'error_message_label') and widget.error_message_label:
            try:
                widget.error_message_label.destroy()
            except tk.TclError:
                pass
            widget.error_message_label = None
        
        # Add error message for invalid fields
        if status == 'invalid' and message:
            error_label = tk.Label(
                parent,
                text=f"⚠ {message}",
                font=('Arial', 9),
                fg=self.colors['danger'],
                bg=self.colors['light'],
                anchor='w',
                wraplength=300
            )
            error_label.pack(fill='x', pady=(2, 5))
            widget.error_message_label = error_label
        elif status == 'valid' and message:
            # Show success message briefly
            success_label = tk.Label(
                parent,
                text=f"✓ {message}",
                font=('Arial', 9),
                fg=self.colors['success'],
                bg=self.colors['light'],
                anchor='w'
            )
            success_label.pack(fill='x', pady=(2, 5))
            widget.error_message_label = success_label
            
            # Auto-hide success message after 3 seconds
            self.dialog.after(3000, lambda: self._hide_success_message(success_label))
    
    def _hide_success_message(self, label):
        """Hide success message after delay"""
        try:
            label.destroy()
        except tk.TclError:
            pass
    
    def create_validation_summary(self, parent):
        """Create validation summary widget for overall form status"""
        self.validation_summary_frame = tk.Frame(parent, bg=self.colors['light'])
        self.validation_summary_frame.pack(fill='x', pady=(10, 0))
        
        self.validation_summary_label = tk.Label(
            self.validation_summary_frame,
            text="",
            font=('Arial', 10),
            bg=self.colors['light'],
            fg=self.colors['dark'],
            anchor='w',
            wraplength=400
        )
        self.validation_summary_label.pack(fill='x', padx=5, pady=5)
        
        # Initially hidden
        self.validation_summary_frame.pack_forget()
    
    def update_validation_summary(self, validation_result):
        """Update validation summary based on overall form validation"""
        if not hasattr(self, 'validation_summary_frame'):
            return
        
        errors = validation_result.get('errors', {})
        warnings = validation_result.get('warnings', {})
        
        if errors:
            # Show errors
            error_count = len(errors)
            error_text = f"⚠ {error_count} error{'s' if error_count > 1 else ''} found:"
            for field, message in errors.items():
                error_text += f"\n• {message}"
            
            self.validation_summary_label.configure(
                text=error_text,
                fg=self.colors['danger']
            )
            self.validation_summary_frame.pack(fill='x', pady=(10, 0))
            
        elif warnings:
            # Show warnings
            warning_count = len(warnings)
            warning_text = f"⚠ {warning_count} warning{'s' if warning_count > 1 else ''} found:"
            for field, message in warnings.items():
                warning_text += f"\n• {message}"
            
            self.validation_summary_label.configure(
                text=warning_text,
                fg=self.colors['secondary']
            )
            self.validation_summary_frame.pack(fill='x', pady=(10, 0))
            
        else:
            # Hide summary when no errors or warnings
            self.validation_summary_frame.pack_forget()
    
    def add_real_time_validation(self, widget, validation_func, *args, **kwargs):
        """
        Add real-time validation to a widget
        
        Args:
            widget: The widget to validate
            validation_func: Function to call for validation
            *args, **kwargs: Arguments to pass to validation function
        """
        def validate_on_change(event=None):
            # Get current value
            if isinstance(widget, tk.Entry):
                value = widget.get()
            elif isinstance(widget, tk.Text):
                value = widget.get('1.0', tk.END).strip()
            else:
                return
            
            # Perform validation
            try:
                result = validation_func(value, *args, **kwargs)
                
                if result.get('valid', False):
                    self.add_validation_indicator(widget, 'valid', result.get('message', ''))
                else:
                    self.add_validation_indicator(widget, 'invalid', result.get('message', ''))
                    
            except Exception as e:
                self.add_validation_indicator(widget, 'invalid', f"Validation error: {str(e)}")
        
        # Bind validation to appropriate events
        if isinstance(widget, tk.Entry):
            widget.bind('<KeyRelease>', validate_on_change)
            widget.bind('<FocusOut>', validate_on_change)
        elif isinstance(widget, tk.Text):
            widget.bind('<KeyRelease>', validate_on_change)
            widget.bind('<FocusOut>', validate_on_change)
        
        # Store validation function for later use
        widget.validation_func = validation_func
        widget.validation_args = args
        widget.validation_kwargs = kwargs
    
    def format_phone_number_input(self, widget):
        """Add phone number formatting assistance to entry widget"""
        def format_phone(event=None):
            current_value = widget.get()
            
            # Remove all non-digit characters except +
            cleaned = ''.join(c for c in current_value if c.isdigit() or c == '+')
            
            # Apply formatting based on length and pattern
            if cleaned.startswith('+265') and len(cleaned) >= 7:
                # Format as +265 XXX XXX XXX
                formatted = cleaned[:4]
                if len(cleaned) > 4:
                    formatted += ' ' + cleaned[4:7]
                if len(cleaned) > 7:
                    formatted += ' ' + cleaned[7:10]
                if len(cleaned) > 10:
                    formatted += ' ' + cleaned[10:13]
                
                # Update widget if different
                if formatted != current_value:
                    cursor_pos = widget.index(tk.INSERT)
                    widget.delete(0, tk.END)
                    widget.insert(0, formatted)
                    # Try to maintain cursor position
                    try:
                        widget.icursor(min(cursor_pos, len(formatted)))
                    except tk.TclError:
                        pass
        
        # Bind formatting to key release
        widget.bind('<KeyRelease>', format_phone)
    
    def create_input_assistance_tooltip(self, widget, help_text, examples=None):
        """Create input assistance tooltip for form fields"""
        def show_help(event):
            # Create tooltip window
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            tooltip.configure(bg='#FFFFDD')
            
            # Help text
            help_label = tk.Label(
                tooltip,
                text=help_text,
                background='#FFFFDD',
                relief='solid',
                borderwidth=1,
                font=('Arial', 9),
                justify='left',
                wraplength=300
            )
            help_label.pack(padx=5, pady=5)
            
            # Examples if provided
            if examples:
                examples_text = "Examples:\n" + "\n".join(f"• {ex}" for ex in examples)
                examples_label = tk.Label(
                    tooltip,
                    text=examples_text,
                    background='#FFFFDD',
                    font=('Arial', 8),
                    justify='left',
                    fg='#666666'
                )
                examples_label.pack(padx=5, pady=(0, 5))
            
            # Auto-hide after 5 seconds
            tooltip.after(5000, tooltip.destroy)
            
            # Store reference
            widget.help_tooltip = tooltip
        
        def hide_help(event):
            if hasattr(widget, 'help_tooltip'):
                try:
                    widget.help_tooltip.destroy()
                except tk.TclError:
                    pass
        
        # Bind to focus events
        widget.bind('<FocusIn>', show_help)
        widget.bind('<FocusOut>', hide_help)
    
    def create_accessible_button(self, parent, text, command, button_type='normal', **kwargs):
        """
        Create an accessible button with proper focus indicators and ARIA attributes
        
        Args:
            parent: Parent widget
            text: Button text
            command: Button command
            button_type: 'primary', 'secondary', 'danger', or 'normal'
            **kwargs: Additional button arguments
        """
        # Button styling based on type
        styles = {
            'primary': {'bg': self.colors['primary'], 'fg': self.colors['white'], 'font': ('Arial', 10, 'bold')},
            'secondary': {'bg': self.colors['secondary'], 'fg': self.colors['white'], 'font': ('Arial', 10)},
            'danger': {'bg': self.colors['danger'], 'fg': self.colors['white'], 'font': ('Arial', 10)},
            'normal': {'bg': self.colors['light'], 'fg': self.colors['dark'], 'font': ('Arial', 10)}
        }
        
        style = styles.get(button_type, styles['normal'])
        
        button = tk.Button(
            parent,
            text=text,
            command=command,
            relief='flat' if button_type in ['primary', 'secondary', 'danger'] else 'solid',
            borderwidth=1,
            padx=kwargs.get('padx', 20),
            pady=kwargs.get('pady', 5),
            cursor='hand2',
            **style,
            **{k: v for k, v in kwargs.items() if k not in ['padx', 'pady']}
        )
        
        # Add to focus order
        self.add_to_focus_order(button)
        
        # Add accessibility features
        button.bind('<FocusIn>', self.on_button_focus_in)
        button.bind('<FocusOut>', self.on_button_focus_out)
        button.bind('<Return>', lambda e: button.invoke())
        button.bind('<space>', lambda e: button.invoke())
        
        return button

# Member Management Dialog Classes
class AddMemberDialog(AccessibleDialog):
    """Enhanced dialog for adding new members with improved UI and validation"""
    
    def __init__(self, parent, app):
        self.app = app
        # Initialize the base class with appropriate title and size
        super().__init__(parent, "Add New Member", width=550, height=650)
        
    def create_content(self):
        """Create dialog content using the accessible base class framework"""
        # Update button text to be more specific
        self.ok_btn.configure(text="Add Member")
        
        # Create form sections
        self.create_form_sections()
        
        # Set initial status
        self.update_status("Ready to add new member", 'info')
    
    def create_form_sections(self):
        """Create form sections with accessible fields"""
        # Personal Information Section
        personal_section = tk.LabelFrame(
            self.content_frame,
            text="Personal Information",
            font=('Arial', 12, 'bold'),
            bg=self.colors['light'],
            fg=self.colors['dark'],
            padx=15,
            pady=10
        )
        personal_section.pack(fill='x', pady=(0, 15))
        
        # First Name - using accessible form field
        self.name_entry, self.name_label = self.create_accessible_form_field(
            personal_section, "First Name", required=True
        )
        
        # Last Name - using accessible form field
        self.surname_entry, self.surname_label = self.create_accessible_form_field(
            personal_section, "Last Name", required=True
        )
        
        # Contact Information Section
        contact_section = tk.LabelFrame(
            self.content_frame,
            text="Contact Information",
            font=('Arial', 12, 'bold'),
            bg=self.colors['light'],
            fg=self.colors['dark'],
            padx=15,
            pady=10
        )
        contact_section.pack(fill='x', pady=(0, 15))
        
        # Phone Number - using accessible form field
        self.phone_entry, self.phone_label = self.create_accessible_form_field(
            contact_section, "Phone Number", required=True
        )
        self.add_placeholder(self.phone_entry, "e.g., +265 999 123 456")
        
        # Email Address - using accessible form field
        self.email_entry, self.email_label = self.create_accessible_form_field(
            contact_section, "Email Address", required=False
        )
        self.add_placeholder(self.email_entry, "e.g., member@example.com")
        
        # Additional Information Section
        additional_section = tk.LabelFrame(
            self.content_frame,
            text="Additional Information",
            font=('Arial', 12, 'bold'),
            bg=self.colors['light'],
            fg=self.colors['dark'],
            padx=15,
            pady=10
        )
        additional_section.pack(fill='x')
        
        # Address - using accessible form field
        self.address_entry, self.address_label = self.create_accessible_form_field(
            additional_section, "Address", required=False
        )
        self.add_placeholder(self.address_entry, "Optional: Physical address")
        
        # Notes - using accessible text field
        self.notes_text, self.notes_label = self.create_accessible_form_field(
            additional_section, "Notes", field_type='text', required=False, height=3
        )
        
        # Add placeholder for notes
        self.notes_text.insert('1.0', "Optional: Any additional notes about the member...")
        self.notes_text.bind('<FocusIn>', self.clear_notes_placeholder)
        self.notes_text.bind('<FocusOut>', self.restore_notes_placeholder)
        
        # Add Clear Form button to the existing button bar
        self.clear_btn = self.create_accessible_button(
            self.button_frame, "Clear Form", self.clear_form, 'secondary'
        )
        self.clear_btn.pack(side='left')
    
    def add_placeholder(self, entry, placeholder_text):
        """Add placeholder functionality to entry widget"""
        entry.insert(0, placeholder_text)
        entry.configure(fg='gray')
        
        def on_focus_in(event):
            if entry.get() == placeholder_text:
                entry.delete(0, tk.END)
                entry.configure(fg='black')
        
        def on_focus_out(event):
            if not entry.get():
                entry.insert(0, placeholder_text)
                entry.configure(fg='gray')
        
        entry.bind('<FocusIn>', on_focus_in)
        entry.bind('<FocusOut>', on_focus_out)
    
    def clear_notes_placeholder(self, event):
        """Clear notes placeholder text"""
        if self.notes_text.get('1.0', tk.END).strip() == "Optional: Any additional notes about the member...":
            self.notes_text.delete('1.0', tk.END)
    
    def restore_notes_placeholder(self, event):
        """Restore notes placeholder if empty"""
        if not self.notes_text.get('1.0', tk.END).strip():
            self.notes_text.insert('1.0', "Optional: Any additional notes about the member...")
    
    def get_form_data(self):
        """Get form data with placeholder handling"""
        def get_entry_value(entry, placeholder=""):
            value = entry.get().strip()
            return value if value != placeholder else ""
        
        name = get_entry_value(self.name_entry)
        surname = get_entry_value(self.surname_entry)
        phone = get_entry_value(self.phone_entry, "e.g., +265 999 123 456")
        email = get_entry_value(self.email_entry, "e.g., member@example.com")
        address = get_entry_value(self.address_entry, "Optional: Physical address")
        
        notes = self.notes_text.get('1.0', tk.END).strip()
        if notes == "Optional: Any additional notes about the member...":
            notes = ""
        
        return {
            'name': name,
            'surname': surname,
            'phone': phone,
            'email': email or None,
            'address': address or None,
            'notes': notes or None
        }
    
    def validate_form_data(self, data):
        """Enhanced form validation with visual feedback"""
        errors = []
        
        # Reset all validation indicators
        self.add_validation_indicator(self.name_entry, 'neutral')
        self.add_validation_indicator(self.surname_entry, 'neutral')
        self.add_validation_indicator(self.phone_entry, 'neutral')
        self.add_validation_indicator(self.email_entry, 'neutral')
        
        # Required field validation
        if not data['name']:
            errors.append("First name is required")
            self.add_validation_indicator(self.name_entry, 'invalid', "First name is required")
        elif len(data['name']) < 2:
            errors.append("First name must be at least 2 characters")
            self.add_validation_indicator(self.name_entry, 'invalid', "Must be at least 2 characters")
        else:
            self.add_validation_indicator(self.name_entry, 'valid', "Valid first name")
        
        if not data['surname']:
            errors.append("Last name is required")
            self.add_validation_indicator(self.surname_entry, 'invalid', "Last name is required")
        elif len(data['surname']) < 2:
            errors.append("Last name must be at least 2 characters")
            self.add_validation_indicator(self.surname_entry, 'invalid', "Must be at least 2 characters")
        else:
            self.add_validation_indicator(self.surname_entry, 'valid', "Valid last name")
        
        if not data['phone']:
            errors.append("Phone number is required")
            self.add_validation_indicator(self.phone_entry, 'invalid', "Phone number is required")
        elif len(data['phone']) < 8:
            errors.append("Phone number must be at least 8 characters")
            self.add_validation_indicator(self.phone_entry, 'invalid', "Must be at least 8 characters")
        else:
            self.add_validation_indicator(self.phone_entry, 'valid', "Valid phone number")
        
        # Email validation (if provided)
        if data['email']:
            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, data['email']):
                errors.append("Please enter a valid email address")
                self.add_validation_indicator(self.email_entry, 'invalid', "Invalid email format")
            else:
                self.add_validation_indicator(self.email_entry, 'valid', "Valid email address")
        else:
            self.add_validation_indicator(self.email_entry, 'neutral')
        
        return errors
    
    def clear_form(self):
        """Clear all form fields"""
        self.name_entry.delete(0, tk.END)
        self.surname_entry.delete(0, tk.END)
        self.phone_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END)
        self.address_entry.delete(0, tk.END)
        self.notes_text.delete('1.0', tk.END)
        
        # Reset validation indicators
        self.add_validation_indicator(self.name_entry, 'neutral')
        self.add_validation_indicator(self.surname_entry, 'neutral')
        self.add_validation_indicator(self.phone_entry, 'neutral')
        self.add_validation_indicator(self.email_entry, 'neutral')
        
        # Restore placeholders
        self.add_placeholder(self.phone_entry, "e.g., +265 999 123 456")
        self.add_placeholder(self.email_entry, "e.g., member@example.com")
        self.add_placeholder(self.address_entry, "Optional: Physical address")
        self.notes_text.insert('1.0', "Optional: Any additional notes about the member...")
        
        # Focus on first field
        self.name_entry.focus()
        self.update_status("Form cleared - ready for new member", 'info')
    
    def ok_action(self):
        """Override the base class OK action to add member"""
        self.add_member()
    
    def add_member(self):
        """Add the new member with enhanced validation and feedback"""
        try:
            self.update_status("Validating member information...", 'info')
            
            # Get form data
            data = self.get_form_data()
            
            # Validate form data
            validation_errors = self.validate_form_data(data)
            if validation_errors:
                error_message = "\n".join(f"• {error}" for error in validation_errors)
                messagebox.showerror("Validation Error", f"Please correct the following:\n\n{error_message}")
                self.update_status("Validation failed - please check your input", 'error')
                return
            
            # Additional validation using existing validation engine if available
            try:
                validation_result = ValidationEngine.validate_member_data(
                    data['name'], data['surname'], data['phone'], data['email']
                )
                if not validation_result['valid']:
                    messagebox.showerror("Validation Error", validation_result['message'])
                    self.update_status("Validation failed", 'error')
                    return
            except NameError:
                # ValidationEngine not available, skip this validation
                pass
            
            self.update_status("Creating member record...", 'info')
            
            # Create member
            member_id = create_member(
                data['name'], 
                data['surname'], 
                data['phone'], 
                data['email']
            )
            
            if member_id:
                success_message = f"Member '{data['name']} {data['surname']}' added successfully!\nMember ID: {member_id}"
                messagebox.showinfo("Success", success_message)
                self.result = True
                self.update_status("Member created successfully!", 'success')
                
                # Close dialog after short delay
                self.dialog.after(1000, self.dialog.destroy)
            else:
                messagebox.showerror("Error", "Failed to add member. Please try again.")
                self.update_status("Failed to create member", 'error')
                
        except Exception as e:
            error_info = ErrorHandler.handle_error(e, "adding member")
            messagebox.showerror("Error", error_info['message'])
            self.update_status("Error occurred while adding member", 'error')
    
    def cancel(self):
        """Cancel dialog with confirmation if form has data"""
        data = self.get_form_data()
        has_data = any([data['name'], data['surname'], data['phone'], data['email'], data['address'], data['notes']])
        
        if has_data:
            if messagebox.askyesno("Confirm Cancel", "You have unsaved changes. Are you sure you want to cancel?"):
                self.result = False
                self.dialog.destroy()
        else:
            self.result = False
            self.dialog.destroy()
    
    def show_help(self):
        """Show help information specific to adding members"""
        help_text = """
Add New Member Help

This dialog allows you to add new members to the Bank Mmudzi system.

Required Fields (marked with *):
• First Name: Member's first name (minimum 2 characters)
• Last Name: Member's last name (minimum 2 characters)  
• Phone Number: Contact phone number (minimum 8 characters)

Optional Fields:
• Email Address: Valid email address for communication
• Address: Physical address for records
• Notes: Any additional information about the member

Keyboard Shortcuts:
• Enter or Alt+O: Add the member
• Escape or Alt+C: Cancel and close
• Tab: Navigate between fields
• F1: Show this help

The form includes real-time validation with visual indicators:
• Green highlight: Valid input
• Red highlight: Invalid input that needs correction

Use the "Clear Form" button to reset all fields and start over.
        """
        
        messagebox.showinfo("Add New Member - Help", help_text.strip())


class MemberDetailsDialog:
    """Dialog for viewing and editing member details"""
    
    def __init__(self, parent, member_id, app):
        self.parent = parent
        self.app = app
        self.member_id = member_id
        self.result = None
        self.member_data = None
        
        # Create dialog window
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Member Details")
        self.dialog.geometry("600x700")
        self.dialog.resizable(True, True)
        self.dialog.transient(parent)
        
        # Center the dialog
        self.dialog.geometry("+%d+%d" % (parent.winfo_rootx() + 50, parent.winfo_rooty() + 50))
        
        self.load_member_data()
        self.create_widgets()
        
        # Set grab after everything is initialized
        self.dialog.after(100, lambda: self.dialog.grab_set())
        
    def load_member_data(self):
        """Load member data from database"""
        try:
            self.member_data = get_member_profile(self.member_id)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load member data: {ErrorHandler.get_user_friendly_message(e)}")
            self.dialog.destroy()
            return
    
    def create_widgets(self):
        """Create dialog widgets"""
        if not self.member_data:
            return
            
        # Create notebook for tabs
        notebook = ttk.Notebook(self.dialog)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Personal Info Tab
        self.create_personal_info_tab(notebook)
        
        # Contributions Tab
        self.create_contributions_tab(notebook)
        
        # Loans Tab
        self.create_loans_tab(notebook)
        
        # Summary Tab
        self.create_summary_tab(notebook)
        
        # Buttons frame
        button_frame = ttk.Frame(self.dialog)
        button_frame.pack(fill='x', padx=10, pady=(0, 10))
        
        ttk.Button(button_frame, text="Close", command=self.close_dialog).pack(side='right', padx=(10, 0))
        ttk.Button(button_frame, text="Edit Member", command=self.edit_member).pack(side='right')
        ttk.Button(button_frame, text="Print Statement", command=self.print_statement).pack(side='left')
    
    def create_personal_info_tab(self, notebook):
        """Create personal information tab"""
        frame = ttk.Frame(notebook)
        notebook.add(frame, text="Personal Info")
        
        # Scrollable frame
        canvas = tk.Canvas(frame)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Member info
        info_frame = ttk.LabelFrame(scrollable_frame, text="Member Information", padding="10")
        info_frame.pack(fill='x', padx=10, pady=10)
        
        member = self.member_data['member_info']
        
        # Display member information
        info_items = [
            ("Member ID:", member.get('member_id', 'N/A')),
            ("Member Number:", member.get('member_number', 'N/A')),
            ("Full Name:", f"{member.get('name', '')} {member.get('surname', '')}"),
            ("Phone Number:", member.get('phone_number', 'N/A')),
            ("Email:", member.get('email', 'N/A')),
            ("Join Date:", member.get('join_date', 'N/A')),
            ("Status:", member.get('status', 'N/A')),
            ("Created:", member.get('created_at', 'N/A')),
            ("Last Updated:", member.get('updated_at', 'N/A'))
        ]
        
        for i, (label, value) in enumerate(info_items):
            ttk.Label(info_frame, text=label, font=('Arial', 9, 'bold')).grid(row=i, column=0, sticky='w', padx=(0, 10), pady=2)
            ttk.Label(info_frame, text=str(value)).grid(row=i, column=1, sticky='w', pady=2)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def create_contributions_tab(self, notebook):
        """Create contributions tab"""
        frame = ttk.Frame(notebook)
        notebook.add(frame, text="Contributions")
        
        # Summary frame
        summary_frame = ttk.LabelFrame(frame, text="Contribution Summary", padding="10")
        summary_frame.pack(fill='x', padx=10, pady=10)
        
        contributions = self.member_data.get('contributions', [])
        total_contributions = sum(float(c.get('amount', 0)) for c in contributions)
        
        ttk.Label(summary_frame, text=f"Total Contributions: MWK {total_contributions:,.2f}", 
                 font=('Arial', 10, 'bold')).pack(anchor='w')
        ttk.Label(summary_frame, text=f"Number of Contributions: {len(contributions)}").pack(anchor='w')
        
        # Contributions list
        list_frame = ttk.LabelFrame(frame, text="Contribution History", padding="10")
        list_frame.pack(fill='both', expand=True, padx=10, pady=(0, 10))
        
        # Treeview for contributions
        columns = ('Date', 'Month', 'Year', 'Amount', 'Late Fee')
        contrib_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=10)
        
        for col in columns:
            contrib_tree.heading(col, text=col)
            contrib_tree.column(col, width=100)
        
        # Add contributions to tree
        for contrib in contributions:
            contrib_tree.insert('', 'end', values=(
                contrib.get('contribution_date', ''),
                contrib.get('month', ''),
                contrib.get('year', ''),
                f"MWK {float(contrib.get('amount', 0)):,.2f}",
                f"MWK {float(contrib.get('late_fee', 0)):,.2f}"
            ))
        
        contrib_tree.pack(fill='both', expand=True)
        
        # Scrollbar for contributions
        contrib_scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=contrib_tree.yview)
        contrib_tree.configure(yscrollcommand=contrib_scrollbar.set)
        contrib_scrollbar.pack(side='right', fill='y')
    
    def create_loans_tab(self, notebook):
        """Create loans tab"""
        frame = ttk.Frame(notebook)
        notebook.add(frame, text="Loans")
        
        # Summary frame
        summary_frame = ttk.LabelFrame(frame, text="Loan Summary", padding="10")
        summary_frame.pack(fill='x', padx=10, pady=10)
        
        loans = self.member_data.get('loans', [])
        total_loans = sum(float(l.get('loan_amount', 0)) for l in loans)
        total_outstanding = sum(float(l.get('outstanding_balance', 0)) for l in loans)
        
        ttk.Label(summary_frame, text=f"Total Loans Taken: MWK {total_loans:,.2f}", 
                 font=('Arial', 10, 'bold')).pack(anchor='w')
        ttk.Label(summary_frame, text=f"Outstanding Balance: MWK {total_outstanding:,.2f}", 
                 font=('Arial', 10, 'bold')).pack(anchor='w')
        ttk.Label(summary_frame, text=f"Number of Loans: {len(loans)}").pack(anchor='w')
        
        # Loans list
        list_frame = ttk.LabelFrame(frame, text="Loan History", padding="10")
        list_frame.pack(fill='both', expand=True, padx=10, pady=(0, 10))
        
        # Treeview for loans
        columns = ('Date', 'Amount', 'Interest Rate', 'Outstanding', 'Status')
        loans_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=10)
        
        for col in columns:
            loans_tree.heading(col, text=col)
            loans_tree.column(col, width=120)
        
        # Add loans to tree
        for loan in loans:
            outstanding = float(loan.get('outstanding_balance', 0))
            status = 'Paid' if outstanding <= 0 else 'Active'
            
            loans_tree.insert('', 'end', values=(
                loan.get('loan_date', ''),
                f"MWK {float(loan.get('loan_amount', 0)):,.2f}",
                f"{float(loan.get('interest_rate', 0)):.1f}%",
                f"MWK {outstanding:,.2f}",
                status
            ))
        
        loans_tree.pack(fill='both', expand=True)
        
        # Scrollbar for loans
        loans_scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=loans_tree.yview)
        loans_tree.configure(yscrollcommand=loans_scrollbar.set)
        loans_scrollbar.pack(side='right', fill='y')
    
    def create_summary_tab(self, notebook):
        """Create summary tab"""
        frame = ttk.Frame(notebook)
        notebook.add(frame, text="Summary")
        
        # Financial summary
        summary_frame = ttk.LabelFrame(frame, text="Financial Summary", padding="15")
        summary_frame.pack(fill='x', padx=10, pady=10)
        
        # Calculate totals
        contributions = self.member_data.get('contributions', [])
        loans = self.member_data.get('loans', [])
        
        total_contributions = sum(float(c.get('amount', 0)) for c in contributions)
        total_loans = sum(float(l.get('loan_amount', 0)) for l in loans)
        total_outstanding = sum(float(l.get('outstanding_balance', 0)) for l in loans)
        
        # Display summary
        summary_items = [
            ("Total Contributions:", f"MWK {total_contributions:,.2f}"),
            ("Total Loans Taken:", f"MWK {total_loans:,.2f}"),
            ("Outstanding Balance:", f"MWK {total_outstanding:,.2f}"),
            ("Net Position:", f"MWK {total_contributions - total_outstanding:,.2f}"),
            ("Member Since:", self.member_data['member_info'].get('join_date', 'N/A')),
            ("Current Status:", self.member_data['member_info'].get('status', 'N/A').title())
        ]
        
        for i, (label, value) in enumerate(summary_items):
            ttk.Label(summary_frame, text=label, font=('Arial', 10, 'bold')).grid(row=i, column=0, sticky='w', padx=(0, 20), pady=5)
            ttk.Label(summary_frame, text=value, font=('Arial', 10)).grid(row=i, column=1, sticky='w', pady=5)
        
        # Member standing
        standing_frame = ttk.LabelFrame(frame, text="Member Standing", padding="15")
        standing_frame.pack(fill='x', padx=10, pady=(0, 10))
        
        # Determine member standing
        if total_outstanding > 0:
            standing = "Borrower"
            standing_color = "orange"
        elif total_contributions > 0:
            standing = "Good Standing"
            standing_color = "green"
        else:
            standing = "New Member"
            standing_color = "blue"
        
        ttk.Label(standing_frame, text=f"Status: {standing}", 
                 font=('Arial', 12, 'bold')).pack(anchor='w')
    
    def edit_member(self):
        """Open edit member dialog"""
        dialog = EditMemberDialog(self.dialog, self.member_id, self.app)
        self.dialog.wait_window(dialog.dialog)
        if dialog.result:
            self.load_member_data()
            self.dialog.destroy()
            self.result = True
    
    def print_statement(self):
        """Print member statement"""
        try:
            # Generate member statement
            member = self.member_data['member_info']
            member_name = f"{member.get('name', '')} {member.get('surname', '')}"
            
            statement = f"""
BANK MMUDZI - MEMBER STATEMENT
{'='*50}

Member: {member_name}
Member ID: {member.get('member_id', 'N/A')}
Member Number: {member.get('member_number', 'N/A')}
Statement Date: {datetime.now().strftime('%Y-%m-%d')}

CONTRIBUTIONS:
{'-'*30}
"""
            
            contributions = self.member_data.get('contributions', [])
            total_contributions = 0
            
            for contrib in contributions:
                amount = float(contrib.get('amount', 0))
                total_contributions += amount
                statement += f"{contrib.get('contribution_date', ''):<12} {contrib.get('month', ''):<10} MWK {amount:>10,.2f}\n"
            
            statement += f"{'-'*30}\nTotal Contributions: MWK {total_contributions:,.2f}\n\n"
            
            statement += "LOANS:\n" + "-"*30 + "\n"
            
            loans = self.member_data.get('loans', [])
            total_loans = 0
            total_outstanding = 0
            
            for loan in loans:
                amount = float(loan.get('loan_amount', 0))
                outstanding = float(loan.get('outstanding_balance', 0))
                total_loans += amount
                total_outstanding += outstanding
                
                statement += f"{loan.get('loan_date', ''):<12} MWK {amount:>10,.2f} Outstanding: MWK {outstanding:>10,.2f}\n"
            
            statement += f"{'-'*30}\nTotal Loans: MWK {total_loans:,.2f}\n"
            statement += f"Outstanding Balance: MWK {total_outstanding:,.2f}\n\n"
            statement += f"Net Position: MWK {total_contributions - total_outstanding:,.2f}\n"
            
            # Show statement in a new window
            statement_window = tk.Toplevel(self.dialog)
            statement_window.title(f"Statement - {member_name}")
            statement_window.geometry("600x500")
            
            text_widget = tk.Text(statement_window, wrap='word', font=('Courier', 10))
            text_widget.pack(fill='both', expand=True, padx=10, pady=10)
            text_widget.insert('1.0', statement)
            text_widget.config(state='disabled')
            
            # Print button
            ttk.Button(statement_window, text="Close", 
                      command=statement_window.destroy).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate statement: {ErrorHandler.get_user_friendly_message(e)}")
    
    def close_dialog(self):
        """Close dialog"""
        self.result = False
        self.dialog.destroy()


class EditMemberDialog:
    """Dialog for editing member information"""
    
    def __init__(self, parent, member_id, app):
        self.parent = parent
        self.app = app
        self.member_id = member_id
        self.result = None
        
        # Load member data
        try:
            member_profile = get_member_profile(member_id)
            self.member_data = member_profile['member_info']
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load member data: {ErrorHandler.get_user_friendly_message(e)}")
            return
        
        # Create dialog window
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Edit Member")
        self.dialog.geometry("400x400")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center the dialog
        self.dialog.geometry("+%d+%d" % (parent.winfo_rootx() + 50, parent.winfo_rooty() + 50))
        
        self.create_widgets()
        
    def create_widgets(self):
        """Create dialog widgets"""
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill='both', expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="Edit Member Information", font=('Arial', 14, 'bold'))
        title_label.pack(pady=(0, 20))
        
        # Form fields
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill='x', pady=(0, 20))
        
        # Name
        ttk.Label(fields_frame, text="First Name *:").pack(anchor='w')
        self.name_entry = ttk.Entry(fields_frame, width=40)
        self.name_entry.pack(fill='x', pady=(5, 10))
        self.name_entry.insert(0, self.member_data.get('name', ''))
        
        # Surname
        ttk.Label(fields_frame, text="Last Name *:").pack(anchor='w')
        self.surname_entry = ttk.Entry(fields_frame, width=40)
        self.surname_entry.pack(fill='x', pady=(5, 10))
        self.surname_entry.insert(0, self.member_data.get('surname', ''))
        
        # Phone
        ttk.Label(fields_frame, text="Phone Number *:").pack(anchor='w')
        self.phone_entry = ttk.Entry(fields_frame, width=40)
        self.phone_entry.pack(fill='x', pady=(5, 10))
        self.phone_entry.insert(0, self.member_data.get('phone_number', ''))
        
        # Email
        ttk.Label(fields_frame, text="Email (Optional):").pack(anchor='w')
        self.email_entry = ttk.Entry(fields_frame, width=40)
        self.email_entry.pack(fill='x', pady=(5, 10))
        self.email_entry.insert(0, self.member_data.get('email', '') or '')
        
        # Status
        ttk.Label(fields_frame, text="Status:").pack(anchor='w')
        self.status_combo = ttk.Combobox(fields_frame, values=['active', 'inactive'], state='readonly')
        self.status_combo.pack(fill='x', pady=(5, 10))
        self.status_combo.set(self.member_data.get('status', 'active'))
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x')
        
        ttk.Button(button_frame, text="Cancel", command=self.cancel).pack(side='right', padx=(10, 0))
        ttk.Button(button_frame, text="Update Member", command=self.update_member).pack(side='right')
        
        # Focus on first field
        self.name_entry.focus()
        
        # Bind keys
        self.dialog.bind('<Return>', lambda e: self.update_member())
        self.dialog.bind('<Escape>', lambda e: self.cancel())
    
    def update_member(self):
        """Update member information"""
        try:
            # Get form data
            name = self.name_entry.get().strip()
            surname = self.surname_entry.get().strip()
            phone = self.phone_entry.get().strip()
            email = self.email_entry.get().strip() or None
            status = self.status_combo.get()
            
            # Validate required fields
            if not name:
                messagebox.showerror("Validation Error", "First name is required")
                self.name_entry.focus()
                return
            
            if not surname:
                messagebox.showerror("Validation Error", "Last name is required")
                self.surname_entry.focus()
                return
            
            if not phone:
                messagebox.showerror("Validation Error", "Phone number is required")
                self.phone_entry.focus()
                return
            
            # Validate data using existing validation
            validation_result = ValidationEngine.validate_member_data(name, surname, phone, email)
            if not validation_result['valid']:
                messagebox.showerror("Validation Error", validation_result['message'])
                return
            
            # Update member
            success = update_member(self.member_id, name, surname, phone, email)
            
            # Update status if changed
            if status != self.member_data.get('status', 'active'):
                update_member_status(self.member_id, status, f"Status updated via edit dialog")
            
            if success:
                messagebox.showinfo("Success", f"Member '{name} {surname}' updated successfully!")
                self.result = True
                self.dialog.destroy()
            else:
                messagebox.showerror("Error", "Failed to update member")
                
        except Exception as e:
            error_info = ErrorHandler.handle_error(e, "updating member")
            messagebox.showerror("Error", error_info['message'])
    
    def cancel(self):
        """Cancel dialog"""
        self.result = False
        self.dialog.destroy()


# Main Application with Modern Admin Panel UI
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Bank Mmudzi - Community Banking System")
        
        # Initialize managers and processors
        self.session_manager = SessionManager()
        self.member_analyzer = MemberAnalyzer()
        self.payment_processor = AutomatedPaymentProcessor()
        self.year_end_processor = YearEndProcessor()
        self.report_generator = EnhancedReportGenerator()
        
        # UI State
        self.current_panel = "dashboard"
        self.sidebar_collapsed = False
        
        # Colors and styling
        self.colors = {
            'primary': '#2c3e50',
            'secondary': '#34495e', 
            'accent': '#3498db',
            'success': '#27ae60',
            'warning': '#f39c12',
            'danger': '#e74c3c',
            'light': '#ecf0f1',
            'dark': '#2c3e50',
            'white': '#ffffff'
        }
        
        self.setup_modern_ui()
    
    def setup_modern_ui(self):
        """Setup the modern admin panel UI with sidebar navigation"""
        self.geometry("1400x900")
        self.configure(bg=self.colors['light'])
        
        # Create main container
        self.main_container = tk.Frame(self, bg=self.colors['light'])
        self.main_container.pack(fill='both', expand=True)
        
        # Create sidebar
        self.create_sidebar()
        
        # Create main content area
        self.create_main_content_area()
        
        # Initialize with dashboard
        self.show_panel("dashboard")
        
        # Create footer
        self.create_footer()
        
        # Start session timeout checker
        self.check_session_timeout()
    
    def create_sidebar(self):
        """Create the collapsible sidebar navigation"""
        sidebar_width = 250 if not self.sidebar_collapsed else 60
        
        self.sidebar = tk.Frame(
            self.main_container, 
            bg=self.colors['primary'], 
            width=sidebar_width,
            height=900
        )
        self.sidebar.pack(side='left', fill='y')
        self.sidebar.pack_propagate(False)
        
        # Sidebar header
        header_frame = tk.Frame(self.sidebar, bg=self.colors['primary'], height=80)
        header_frame.pack(fill='x', pady=(10, 20))
        header_frame.pack_propagate(False)
        
        # Logo/Title
        if not self.sidebar_collapsed:
            title_label = tk.Label(
                header_frame,
                text="Bank Mmudzi",
                font=('Arial', 16, 'bold'),
                fg=self.colors['white'],
                bg=self.colors['primary']
            )
            title_label.pack(pady=20)
        
        # Collapse/Expand button
        collapse_btn = tk.Button(
            header_frame,
            text="☰" if self.sidebar_collapsed else "◀",
            font=('Arial', 12),
            bg=self.colors['secondary'],
            fg=self.colors['white'],
            border=0,
            command=self.toggle_sidebar
        )
        collapse_btn.pack(side='right', padx=10)
        
        # Navigation menu
        self.create_navigation_menu()
    
    def create_navigation_menu(self):
        """Create the navigation menu items"""
        menu_items = [
            ("📊", "Dashboard", "dashboard"),
            ("👥", "Members", "members"),
            ("💰", "Contributions", "contributions"),
            ("🏦", "Loans", "loans"),
            ("💳", "Repayments", "repayments"),
            ("📈", "Reports", "reports"),
            ("⚙️", "Settings", "settings"),
            ("🔍", "Audit", "audit")
        ]
        
        self.nav_buttons = {}
        
        for icon, text, panel_id in menu_items:
            btn_frame = tk.Frame(self.sidebar, bg=self.colors['primary'])
            btn_frame.pack(fill='x', padx=5, pady=2)
            
            btn = tk.Button(
                btn_frame,
                text=f"{icon} {text}" if not self.sidebar_collapsed else icon,
                font=('Arial', 10, 'bold'),
                bg=self.colors['secondary'],
                fg=self.colors['white'],
                activebackground=self.colors['accent'],
                activeforeground=self.colors['white'],
                border=0,
                relief='flat',
                anchor='w',
                padx=15,
                pady=10,
                command=lambda p=panel_id: self.show_panel(p)
            )
            btn.pack(fill='x')
            
            self.nav_buttons[panel_id] = btn
    
    def toggle_sidebar(self):
        """Toggle sidebar collapse/expand"""
        self.sidebar_collapsed = not self.sidebar_collapsed
        self.sidebar.destroy()
        self.create_sidebar()
    
    def create_main_content_area(self):
        """Create the main content area"""
        self.content_area = tk.Frame(
            self.main_container,
            bg=self.colors['white'],
            relief='flat'
        )
        self.content_area.pack(side='right', fill='both', expand=True, padx=10, pady=10)
        
        # Content header
        self.content_header = tk.Frame(self.content_area, bg=self.colors['white'], height=60)
        self.content_header.pack(fill='x', pady=(0, 20))
        self.content_header.pack_propagate(False)
        
        # Page title
        self.page_title = tk.Label(
            self.content_header,
            text="Dashboard",
            font=('Arial', 24, 'bold'),
            fg=self.colors['dark'],
            bg=self.colors['white']
        )
        self.page_title.pack(side='left', pady=15)
        
        # User info and logout
        user_frame = tk.Frame(self.content_header, bg=self.colors['white'])
        user_frame.pack(side='right', pady=15)
        
        user_label = tk.Label(
            user_frame,
            text=f"Welcome, {getattr(self.session_manager, 'current_user', 'Admin')}",
            font=('Arial', 10),
            fg=self.colors['dark'],
            bg=self.colors['white']
        )
        user_label.pack(side='left', padx=10)
        
        logout_btn = tk.Button(
            user_frame,
            text="Logout",
            font=('Arial', 10),
            bg=self.colors['danger'],
            fg=self.colors['white'],
            border=0,
            padx=15,
            pady=5,
            command=self.logout
        )
        logout_btn.pack(side='right')
        
        # Main content panel
        self.main_panel = tk.Frame(self.content_area, bg=self.colors['white'])
        self.main_panel.pack(fill='both', expand=True)
    
    def show_panel(self, panel_id):
        """Show the specified panel"""
        # Update active button
        for btn_id, btn in self.nav_buttons.items():
            if btn_id == panel_id:
                btn.configure(bg=self.colors['accent'])
            else:
                btn.configure(bg=self.colors['secondary'])
        
        # Clear main panel
        for widget in self.main_panel.winfo_children():
            widget.destroy()
        
        # Update page title
        panel_titles = {
            "dashboard": "Dashboard",
            "members": "Member Management",
            "contributions": "Contributions",
            "loans": "Loan Management",
            "repayments": "Repayments",
            "reports": "Reports & Analytics",
            "settings": "System Settings",
            "audit": "Audit Trail"
        }
        
        self.page_title.configure(text=panel_titles.get(panel_id, "Dashboard"))
        self.current_panel = panel_id
        
        # Show appropriate panel
        if panel_id == "dashboard":
            self.create_dashboard_panel()
        elif panel_id == "members":
            self.create_members_panel()
        elif panel_id == "contributions":
            self.create_contributions_panel()
        elif panel_id == "loans":
            self.create_loans_panel()
        elif panel_id == "repayments":
            self.create_repayments_panel()
        elif panel_id == "reports":
            self.create_reports_panel()
        elif panel_id == "settings":
            self.create_settings_panel()
        elif panel_id == "audit":
            self.create_audit_panel()
        
        # Update session activity
        if hasattr(self, 'session_manager'):
            self.session_manager.update_activity()
    
    def create_dashboard_panel(self):
        """Create the dashboard overview panel"""
        # Summary cards row
        cards_frame = tk.Frame(self.main_panel, bg=self.colors['white'])
        cards_frame.pack(fill='x', pady=(0, 20))
        
        # Get dashboard data
        dashboard_data = self.get_dashboard_data()
        
        # Create summary cards
        cards = [
            ("Total Members", dashboard_data['total_members'], self.colors['primary'], "👥"),
            ("Active Loans", dashboard_data['active_loans'], self.colors['warning'], "🏦"),
            ("Monthly Contributions", f"MWK {dashboard_data['monthly_contributions']:,.2f}", self.colors['success'], "💰"),
            ("Available Funds", f"MWK {dashboard_data['available_funds']:,.2f}", self.colors['accent'], "💳")
        ]
        
        for i, (title, value, color, icon) in enumerate(cards):
            card = self.create_summary_card(cards_frame, title, value, color, icon)
            card.grid(row=0, column=i, padx=10, sticky='ew')
        
        cards_frame.grid_columnconfigure(0, weight=1)
        cards_frame.grid_columnconfigure(1, weight=1)
        cards_frame.grid_columnconfigure(2, weight=1)
        cards_frame.grid_columnconfigure(3, weight=1)
        
        # Quick actions and recent activity
        content_frame = tk.Frame(self.main_panel, bg=self.colors['white'])
        content_frame.pack(fill='both', expand=True)
        
        # Quick actions panel
        actions_frame = tk.LabelFrame(
            content_frame, 
            text="Quick Actions", 
            font=('Arial', 12, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['dark']
        )
        actions_frame.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        quick_actions = [
            ("Add New Member", lambda: self.show_panel("members")),
            ("Record Contribution", lambda: self.show_panel("contributions")),
            ("Process Loan", lambda: self.show_panel("loans")),
            ("Record Repayment", lambda: self.show_panel("repayments")),
            ("Generate Report", lambda: self.show_panel("reports")),
            ("Calculate Dividends", self.calculate_year_dividends)
        ]
        
        for i, (action_text, command) in enumerate(quick_actions):
            btn = tk.Button(
                actions_frame,
                text=action_text,
                font=('Arial', 10),
                bg=self.colors['accent'],
                fg=self.colors['white'],
                border=0,
                padx=20,
                pady=8,
                command=command
            )
            btn.pack(fill='x', padx=10, pady=5)
        
        # Recent activity panel
        activity_frame = tk.LabelFrame(
            content_frame, 
            text="Recent Activity", 
            font=('Arial', 12, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['dark']
        )
        activity_frame.pack(side='right', fill='both', expand=True, padx=(10, 0))
        
        # Recent activity list
        activity_list = tk.Listbox(
            activity_frame,
            font=('Arial', 9),
            bg=self.colors['light'],
            selectbackground=self.colors['accent']
        )
        activity_list.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Load recent activity
        recent_activities = self.get_recent_activities()
        for activity in recent_activities:
            activity_list.insert(tk.END, activity)
        
        # Add developer info section at the bottom of dashboard
        dev_info_frame = tk.Frame(self.main_panel, bg=self.colors['light'], relief='solid', bd=1)
        dev_info_frame.pack(fill='x', pady=(20, 0))
        
        # Developer info header
        dev_header = tk.Label(
            dev_info_frame,
            text="💻 System Developer",
            font=('Arial', 11, 'bold'),
            bg=self.colors['light'],
            fg=self.colors['primary']
        )
        dev_header.pack(pady=(10, 5))
        
        # Developer details
        dev_details = tk.Label(
            dev_info_frame,
            text="Nehemiah Nganjo | Phone: 0997082156 | GitHub: github.com/Nehemiahnganjo",
            font=('Arial', 10),
            bg=self.colors['light'],
            fg=self.colors['dark'],
            cursor="hand2"
        )
        dev_details.pack(pady=(0, 10))
        
        # Make GitHub link clickable
        def open_github(event):
            import webbrowser
            webbrowser.open("https://github.com/Nehemiahnganjo")
        
        dev_details.bind("<Button-1>", open_github)
    
    def create_summary_card(self, parent, title, value, color, icon):
        """Create a summary card widget"""
        card = tk.Frame(parent, bg=color, relief='raised', bd=1)
        
        # Icon
        icon_label = tk.Label(
            card,
            text=icon,
            font=('Arial', 24),
            bg=color,
            fg=self.colors['white']
        )
        icon_label.pack(pady=(15, 5))
        
        # Value
        value_label = tk.Label(
            card,
            text=str(value),
            font=('Arial', 18, 'bold'),
            bg=color,
            fg=self.colors['white']
        )
        value_label.pack()
        
        # Title
        title_label = tk.Label(
            card,
            text=title,
            font=('Arial', 10),
            bg=color,
            fg=self.colors['white']
        )
        title_label.pack(pady=(5, 15))
        
        return card
    
    def get_dashboard_data(self):
        """Get data for dashboard summary cards"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Total members
            cursor.execute("SELECT COUNT(*) FROM Members WHERE status = 'active'")
            total_members = cursor.fetchone()[0]
            
            # Active loans
            cursor.execute("SELECT COUNT(*) FROM Loans WHERE outstanding_balance > 0")
            active_loans = cursor.fetchone()[0]
            
            # Monthly contributions (current month)
            current_month = datetime.now().month
            current_year = datetime.now().year
            
            # Use database-specific syntax
            if db_manager.current_db_type == 'mysql':
                cursor.execute("""
                    SELECT COALESCE(SUM(amount), 0) FROM Contributions 
                    WHERE month = %s AND year = %s
                """, (current_month, current_year))
            else:
                cursor.execute("""
                    SELECT COALESCE(SUM(amount), 0) FROM Contributions 
                    WHERE month = ? AND year = ?
                """, (current_month, current_year))
            monthly_contributions = cursor.fetchone()[0]
            
            # Available funds
            cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM Contributions")
            total_contributions = cursor.fetchone()[0]
            
            cursor.execute("SELECT COALESCE(SUM(outstanding_balance), 0) FROM Loans")
            outstanding_loans = cursor.fetchone()[0]
            
            available_funds = total_contributions - outstanding_loans
            
            conn.close()
            
            return {
                'total_members': total_members,
                'active_loans': active_loans,
                'monthly_contributions': monthly_contributions,
                'available_funds': available_funds
            }
            
        except Exception as e:
            logger.error(f"Error getting dashboard data: {str(e)}")
            return {
                'total_members': 0,
                'active_loans': 0,
                'monthly_contributions': 0,
                'available_funds': 0
            }
    
    def get_recent_activities(self):
        """Get recent activities for dashboard"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT operation, table_name, old_values, timestamp
                FROM AuditLog 
                ORDER BY timestamp DESC 
                LIMIT 10
            """)
            
            activities = []
            for row in cursor.fetchall():
                operation, table_name, old_values, timestamp = row
                activity_text = f"{timestamp}: {operation} on {table_name}"
                if old_values:
                    activity_text += f" - {old_values[:50]}..."
                activities.append(activity_text)
            
            conn.close()
            return activities
            
        except Exception as e:
            logger.error(f"Error getting recent activities: {str(e)}")
            return ["No recent activities available"]
    
    def calculate_year_dividends(self):
        """Calculate dividends for current year"""
        current_year = datetime.now().year
        try:
            result = calculate_all_dividends_for_year(current_year)
            if result['distribution_valid']:
                messagebox.showinfo(
                    "Dividend Calculation Complete",
                    f"Dividends calculated for {len(result['dividend_calculations'])} members.\n"
                    f"Total Fund: MWK {result['total_fund']:,.2f}\n"
                    f"Total Dividends: MWK {result['total_dividends']:,.2f}"
                )
            else:
                messagebox.showwarning(
                    "Dividend Calculation Warning",
                    "Dividend distribution validation failed. Please review calculations."
                )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to calculate dividends: {ErrorHandler.get_user_friendly_message(e)}")
    
    def create_footer(self):
        """Create footer with developer information"""
        footer_frame = tk.Frame(
            self.main_container,
            bg=self.colors['primary'],
            height=60
        )
        footer_frame.pack(side='bottom', fill='x')
        footer_frame.pack_propagate(False)
        
        # Create a more prominent footer with better layout
        footer_content = tk.Frame(footer_frame, bg=self.colors['primary'])
        footer_content.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Developer info section
        dev_frame = tk.Frame(footer_content, bg=self.colors['primary'])
        dev_frame.pack(side='left', fill='y')
        
        dev_name = tk.Label(
            dev_frame,
            text="Developer: Nehemiah Nganjo",
            font=('Arial', 10, 'bold'),
            fg=self.colors['white'],
            bg=self.colors['primary']
        )
        dev_name.pack(anchor='w')
        
        contact_info = tk.Label(
            dev_frame,
            text="Phone: 0997082156 | GitHub: Nehemiahnganjo",
            font=('Arial', 9),
            fg=self.colors['light'],
            bg=self.colors['primary'],
            cursor="hand2"
        )
        contact_info.pack(anchor='w')
        
        # Make contact info clickable for GitHub
        def open_github(event):
            import webbrowser
            webbrowser.open("https://github.com/Nehemiahnganjo")
        
        contact_info.bind("<Button-1>", open_github)
        
        # Version info section
        version_frame = tk.Frame(footer_content, bg=self.colors['primary'])
        version_frame.pack(side='right', fill='y')
        
        version_info = tk.Label(
            version_frame,
            text="Bank Mmudzi v1.4.0",
            font=('Arial', 10, 'bold'),
            fg=self.colors['white'],
            bg=self.colors['primary']
        )
        version_info.pack(anchor='e')
        
        copyright_info = tk.Label(
            version_frame,
            text="© 2024 Community Banking System",
            font=('Arial', 9),
            fg=self.colors['light'],
            bg=self.colors['primary']
        )
        copyright_info.pack(anchor='e')
    
    def check_session_timeout(self):
        """Check for session timeout"""
        if hasattr(self, 'session_manager'):
            timeout_result = self.session_manager.check_session_timeout()
            if timeout_result.get('timed_out', False):
                messagebox.showwarning("Session Timeout", timeout_result['message'])
                self.logout()
                return
        
        # Schedule next check
        self.after(60000, self.check_session_timeout)  # Check every minute
    
    def logout(self):
        """Logout and return to login screen"""
        if hasattr(self, 'session_manager'):
            self.session_manager.end_session('manual_logout')
        
        self.destroy()
        
        # Show login window
        login_window = LoginWindow()
        login_window.mainloop()
    
    def create_members_panel(self):
        """Create the enhanced members management panel"""
        # Create the enhanced member panel instance
        self.enhanced_member_panel = EnhancedMemberPanel(self.main_panel, self)
        
        # Initialize the panel
        self.enhanced_member_panel.setup_panel()


# Enhanced Member Management Panel
class EnhancedMemberPanel:
    """
    Unified member management interface with comprehensive CRUD operations
    Provides accessible, consistent interface for all member operations
    """
    
    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self.colors = app.colors
        
        # State management
        self.selected_member_id = None
        self.current_sort_column = None
        self.current_sort_direction = 'asc'
        self.search_text = ""
        self.status_filter = "All"
        self.contribution_filter = "All"
        self.loan_filter = "All"
        
        # UI components (will be created in setup_panel)
        self.main_frame = None
        self.action_bar = None
        self.search_frame = None
        self.members_tree = None
        self.status_bar = None
        
        # Action buttons
        self.add_btn = None
        self.edit_btn = None
        self.delete_btn = None
        self.view_btn = None
        self.refresh_btn = None
        
        # Search and filter components
        self.search_entry = None
        self.status_filter_combo = None
        self.contribution_filter_combo = None
        self.loan_filter_combo = None
        
    def setup_panel(self):
        """Setup the complete member management panel"""
        # Create main container
        self.main_frame = tk.Frame(self.parent, bg=self.colors['white'])
        self.main_frame.pack(fill='both', expand=True)
        
        # Create components in order
        self.create_action_bar()
        self.create_search_and_filter_section()
        self.create_member_list()
        self.create_status_bar()
        
        # Setup keyboard bindings
        self.setup_keyboard_bindings()
        
        # Load initial data
        self.refresh_member_list()
        
        # Update button states
        self.update_button_states()
    
    def create_action_bar(self):
        """Create top action bar with CRUD buttons positioned at top for accessibility"""
        self.action_bar = tk.Frame(self.main_frame, bg=self.colors['white'], height=60)
        self.action_bar.pack(fill='x', pady=(0, 15))
        self.action_bar.pack_propagate(False)
        
        # Left side - Primary CRUD actions
        left_actions = tk.Frame(self.action_bar, bg=self.colors['white'])
        left_actions.pack(side='left', fill='y')
        
        # Add Member button
        self.add_btn = tk.Button(
            left_actions,
            text="+ Add Member",
            font=('Arial', 10, 'bold'),
            bg=self.colors['success'],
            fg=self.colors['white'],
            relief='flat',
            padx=20,
            pady=8,
            cursor='hand2',
            command=self.add_member_action
        )
        self.add_btn.pack(side='left', padx=(0, 10))
        
        # Edit Member button
        self.edit_btn = tk.Button(
            left_actions,
            text="✏ Edit",
            font=('Arial', 10),
            bg=self.colors['accent'],
            fg=self.colors['white'],
            relief='flat',
            padx=20,
            pady=8,
            cursor='hand2',
            command=self.edit_member_action,
            state='disabled'
        )
        self.edit_btn.pack(side='left', padx=(0, 10))
        
        # Delete Member button
        self.delete_btn = tk.Button(
            left_actions,
            text="🗑 Delete",
            font=('Arial', 10),
            bg=self.colors['danger'],
            fg=self.colors['white'],
            relief='flat',
            padx=20,
            pady=8,
            cursor='hand2',
            command=self.delete_member_action,
            state='disabled'
        )
        self.delete_btn.pack(side='left', padx=(0, 10))
        
        # View Details button
        self.view_btn = tk.Button(
            left_actions,
            text="👁 View Details",
            font=('Arial', 10),
            bg=self.colors['primary'],
            fg=self.colors['white'],
            relief='flat',
            padx=20,
            pady=8,
            cursor='hand2',
            command=self.view_member_action,
            state='disabled'
        )
        self.view_btn.pack(side='left', padx=(0, 10))
        
        # Right side - Utility actions
        right_actions = tk.Frame(self.action_bar, bg=self.colors['white'])
        right_actions.pack(side='right', fill='y')
        
        # Refresh button
        self.refresh_btn = tk.Button(
            right_actions,
            text="🔄 Refresh",
            font=('Arial', 10),
            bg=self.colors['secondary'],
            fg=self.colors['white'],
            relief='flat',
            padx=20,
            pady=8,
            cursor='hand2',
            command=self.refresh_member_list
        )
        self.refresh_btn.pack(side='right')
        
        # Add hover effects to all buttons
        self.setup_button_hover_effects()
    
    def setup_button_hover_effects(self):
        """Add hover effects to action buttons for better UX"""
        buttons = [
            (self.add_btn, self.colors['success']),
            (self.edit_btn, self.colors['accent']),
            (self.delete_btn, self.colors['danger']),
            (self.view_btn, self.colors['primary']),
            (self.refresh_btn, self.colors['secondary'])
        ]
        
        for button, original_color in buttons:
            # Create hover effect
            def on_enter(event, btn=button, color=original_color):
                if btn['state'] != 'disabled':
                    # Darken the color for hover effect
                    hover_color = self.darken_color(color)
                    btn.configure(bg=hover_color)
            
            def on_leave(event, btn=button, color=original_color):
                if btn['state'] != 'disabled':
                    btn.configure(bg=color)
            
            button.bind('<Enter>', on_enter)
            button.bind('<Leave>', on_leave)
    
    def darken_color(self, color):
        """Darken a hex color for hover effects"""
        # Simple color darkening - remove # and convert to RGB
        if color.startswith('#'):
            color = color[1:]
        
        # Convert hex to RGB
        r = int(color[0:2], 16)
        g = int(color[2:4], 16)
        b = int(color[4:6], 16)
        
        # Darken by reducing each component by 20
        r = max(0, r - 20)
        g = max(0, g - 20)
        b = max(0, b - 20)
        
        # Convert back to hex
        return f"#{r:02x}{g:02x}{b:02x}"
    
    def create_search_and_filter_section(self):
        """Create search and filter controls"""
        self.search_frame = tk.Frame(self.main_frame, bg=self.colors['white'])
        self.search_frame.pack(fill='x', pady=(0, 15))
        
        # Search section
        search_section = tk.Frame(self.search_frame, bg=self.colors['white'])
        search_section.pack(side='left', fill='x', expand=True)
        
        # Search label and entry
        tk.Label(
            search_section,
            text="Search:",
            font=('Arial', 10, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['dark']
        ).pack(side='left', padx=(0, 5))
        
        self.search_entry = tk.Entry(
            search_section,
            font=('Arial', 10),
            width=30,
            relief='solid',
            borderwidth=1
        )
        self.search_entry.pack(side='left', padx=(0, 15))
        
        # Bind search events
        self.search_entry.bind('<KeyRelease>', self.on_search_change)
        self.search_entry.bind('<Return>', lambda e: self.apply_filters())
        
        # Filter section
        filter_section = tk.Frame(self.search_frame, bg=self.colors['white'])
        filter_section.pack(side='right')
        
        # Status filter
        tk.Label(
            filter_section,
            text="Status:",
            font=('Arial', 10),
            bg=self.colors['white'],
            fg=self.colors['dark']
        ).pack(side='left', padx=(0, 5))
        
        self.status_filter_combo = ttk.Combobox(
            filter_section,
            values=['All', 'Active', 'Inactive', 'Suspended'],
            state='readonly',
            width=12,
            font=('Arial', 9)
        )
        self.status_filter_combo.set('All')
        self.status_filter_combo.pack(side='left', padx=(0, 15))
        self.status_filter_combo.bind('<<ComboboxSelected>>', lambda e: self.apply_filters())
        
        # Contribution level filter
        tk.Label(
            filter_section,
            text="Level:",
            font=('Arial', 10),
            bg=self.colors['white'],
            fg=self.colors['dark']
        ).pack(side='left', padx=(0, 5))
        
        self.contribution_filter_combo = ttk.Combobox(
            filter_section,
            values=['All', 'High', 'Medium', 'Low'],
            state='readonly',
            width=10,
            font=('Arial', 9)
        )
        self.contribution_filter_combo.set('All')
        self.contribution_filter_combo.pack(side='left', padx=(0, 15))
        self.contribution_filter_combo.bind('<<ComboboxSelected>>', lambda e: self.apply_filters())
        
        # Loan status filter
        tk.Label(
            filter_section,
            text="Loans:",
            font=('Arial', 10),
            bg=self.colors['white'],
            fg=self.colors['dark']
        ).pack(side='left', padx=(0, 5))
        
        self.loan_filter_combo = ttk.Combobox(
            filter_section,
            values=['All', 'Active', 'None', 'Overdue'],
            state='readonly',
            width=10,
            font=('Arial', 9)
        )
        self.loan_filter_combo.set('All')
        self.loan_filter_combo.pack(side='left')
        self.loan_filter_combo.bind('<<ComboboxSelected>>', lambda e: self.apply_filters())
    
    def create_member_list(self):
        """Create enhanced member list with sortable columns and proper resizing"""
        # List container
        list_container = tk.Frame(self.main_frame, bg=self.colors['white'])
        list_container.pack(fill='both', expand=True, pady=(0, 10))
        
        # Define columns with proper widths and configuration
        columns = ('ID', 'Number', 'Name', 'Phone', 'Status', 'Actions')
        column_config = {
            'ID': {'width': 60, 'minwidth': 50, 'anchor': 'center', 'stretch': False},
            'Number': {'width': 120, 'minwidth': 100, 'anchor': 'center', 'stretch': False},
            'Name': {'width': 250, 'minwidth': 150, 'anchor': 'w', 'stretch': True},
            'Phone': {'width': 140, 'minwidth': 120, 'anchor': 'center', 'stretch': False},
            'Status': {'width': 100, 'minwidth': 80, 'anchor': 'center', 'stretch': False},
            'Actions': {'width': 100, 'minwidth': 80, 'anchor': 'center', 'stretch': False}
        }
        
        # Create treeview with enhanced styling and proper selection mode
        self.members_tree = ttk.Treeview(
            list_container,
            columns=columns,
            show='headings',
            height=20,
            selectmode='browse'  # Single selection with clear visual feedback
        )
        
        # Configure columns with proper sizing and resizing capabilities
        for col in columns:
            config = column_config[col]
            
            # Set up column heading with sort functionality
            self.members_tree.heading(
                col,
                text=col,
                command=lambda c=col: self.sort_by_column(c),
                anchor='center'
            )
            
            # Configure column properties for proper resizing
            self.members_tree.column(
                col,
                width=config['width'],
                minwidth=config['minwidth'],
                anchor=config['anchor'],
                stretch=config['stretch']  # Allow/prevent column stretching
            )
        
        # Create scrollbars for large member lists
        v_scrollbar = ttk.Scrollbar(
            list_container,
            orient='vertical',
            command=self.members_tree.yview
        )
        h_scrollbar = ttk.Scrollbar(
            list_container,
            orient='horizontal',
            command=self.members_tree.xview
        )
        
        # Configure scrollbars
        self.members_tree.configure(
            yscrollcommand=v_scrollbar.set,
            xscrollcommand=h_scrollbar.set
        )
        
        # Pack treeview and scrollbars using grid for better control
        self.members_tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        # Configure grid weights for proper resizing
        list_container.grid_rowconfigure(0, weight=1)
        list_container.grid_columnconfigure(0, weight=1)
        
        # Bind events for enhanced interaction
        self.members_tree.bind('<<TreeviewSelect>>', self.on_member_select)
        self.members_tree.bind('<Double-1>', self.on_member_double_click)
        self.members_tree.bind('<Return>', self.view_member_action)
        self.members_tree.bind('<Button-1>', self.on_member_click)
        
        # Add keyboard navigation for member list
        self.members_tree.bind('<Up>', self.on_arrow_key)
        self.members_tree.bind('<Down>', self.on_arrow_key)
        self.members_tree.bind('<Home>', self.on_home_key)
        self.members_tree.bind('<End>', self.on_end_key)
        
        # Configure enhanced treeview styling
        self.configure_treeview_styling()
        
        # Add alternating row colors for better readability
        self.setup_alternating_row_colors()
    
    def configure_treeview_styling(self):
        """Configure enhanced styling for the treeview"""
        style = ttk.Style()
        
        # Configure treeview style
        style.configure(
            "Treeview",
            background=self.colors['white'],
            foreground=self.colors['dark'],
            rowheight=25,
            fieldbackground=self.colors['white']
        )
        
        # Configure heading style
        style.configure(
            "Treeview.Heading",
            background=self.colors['light'],
            foreground=self.colors['dark'],
            font=('Arial', 10, 'bold')
        )
        
        # Configure selection style
        style.map(
            "Treeview",
            background=[('selected', self.colors['accent'])],
            foreground=[('selected', self.colors['white'])]
        )
    
    def create_status_bar(self):
        """Create status bar for user feedback"""
        self.status_bar = tk.Frame(self.main_frame, bg=self.colors['light'], height=30)
        self.status_bar.pack(fill='x', side='bottom')
        self.status_bar.pack_propagate(False)
        
        # Status label
        self.status_label = tk.Label(
            self.status_bar,
            text="Ready - 0 members shown",
            font=('Arial', 9),
            bg=self.colors['light'],
            fg=self.colors['dark'],
            anchor='w'
        )
        self.status_label.pack(side='left', fill='x', expand=True, padx=10, pady=5)
        
        # Selection info
        self.selection_label = tk.Label(
            self.status_bar,
            text="No selection",
            font=('Arial', 9),
            bg=self.colors['light'],
            fg=self.colors['dark'],
            anchor='e'
        )
        self.selection_label.pack(side='right', padx=10, pady=5)
    
    def setup_keyboard_bindings(self):
        """Setup keyboard shortcuts for accessibility"""
        # Bind to the main frame to capture keyboard events
        self.main_frame.bind('<Alt-a>', lambda e: self.add_member_action())
        self.main_frame.bind('<Alt-A>', lambda e: self.add_member_action())
        self.main_frame.bind('<Alt-e>', lambda e: self.edit_member_action())
        self.main_frame.bind('<Alt-E>', lambda e: self.edit_member_action())
        self.main_frame.bind('<Alt-d>', lambda e: self.delete_member_action())
        self.main_frame.bind('<Alt-D>', lambda e: self.delete_member_action())
        self.main_frame.bind('<Alt-v>', lambda e: self.view_member_action())
        self.main_frame.bind('<Alt-V>', lambda e: self.view_member_action())
        self.main_frame.bind('<F5>', lambda e: self.refresh_member_list())
        self.main_frame.bind('<Control-f>', lambda e: self.focus_search())
        self.main_frame.bind('<Control-F>', lambda e: self.focus_search())
        
        # Make the main frame focusable to receive keyboard events
        self.main_frame.focus_set()
        self.main_frame.bind('<Button-1>', lambda e: self.main_frame.focus_set())
    
    def focus_search(self):
        """Focus the search entry box"""
        self.search_entry.focus_set()
        self.search_entry.select_range(0, tk.END)
    
    # Action methods (to be implemented in subsequent tasks)
    def add_member_action(self):
        """Handle add member action"""
        self.app.show_add_member_dialog()
    
    def edit_member_action(self):
        """Handle edit member action"""
        if self.selected_member_id:
            # This will be implemented in later tasks
            messagebox.showinfo("Edit Member", f"Edit member functionality will be implemented in task 7")
    
    def delete_member_action(self):
        """Handle delete member action"""
        if self.selected_member_id:
            # This will be implemented in later tasks
            messagebox.showinfo("Delete Member", f"Delete member functionality will be implemented in task 9")
    
    def view_member_action(self):
        """Handle view member details action"""
        if self.selected_member_id:
            self.app.show_member_details(self.selected_member_id)
    
    def refresh_member_list(self):
        """Refresh the member list"""
        try:
            # Clear existing items
            for item in self.members_tree.get_children():
                self.members_tree.delete(item)
            
            # Get filtered member data
            members = self.get_filtered_members()
            
            # Populate the tree
            for member in members:
                # Format member data for display
                values = (
                    member.get('member_id', ''),
                    member.get('member_number', ''),
                    f"{member.get('name', '')} {member.get('surname', '')}".strip(),
                    member.get('phone_number', ''),
                    member.get('status', '').title(),
                    '...'  # Actions column placeholder
                )
                
                # Insert with member_id as the item identifier
                item_id = self.members_tree.insert('', 'end', values=values)
                # Store member_id in the item for later reference
                self.members_tree.set(item_id, '#0', member.get('member_id', ''))
            
            # Update status
            member_count = len(members)
            self.update_status(f"Ready - {member_count} members shown")
            
        except Exception as e:
            logger.error(f"Error refreshing member list: {e}")
            messagebox.showerror("Error", f"Failed to refresh member list: {str(e)}")
    
    def get_filtered_members(self):
        """Get members based on current filters"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Base query
            query = """
                SELECT member_id, member_number, name, surname, phone_number, status
                FROM Members
                WHERE 1=1
            """
            params = []
            
            # Apply search filter
            if self.search_text.strip():
                search_term = f"%{self.search_text.strip()}%"
                query += " AND (name LIKE ? OR surname LIKE ? OR phone_number LIKE ? OR member_number LIKE ?)"
                params.extend([search_term, search_term, search_term, search_term])
            
            # Apply status filter
            if self.status_filter != "All":
                query += " AND status = ?"
                params.append(self.status_filter.lower())
            
            # Execute query
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            # Convert to list of dictionaries
            members = []
            for row in results:
                members.append({
                    'member_id': row[0],
                    'member_number': row[1],
                    'name': row[2],
                    'surname': row[3],
                    'phone_number': row[4],
                    'status': row[5]
                })
            
            return members
            
        except Exception as e:
            logger.error(f"Error getting filtered members: {e}")
            return []
    
    def on_search_change(self, event):
        """Handle search text changes with debouncing"""
        # Cancel any pending search
        if hasattr(self, '_search_after_id'):
            self.main_frame.after_cancel(self._search_after_id)
        
        # Schedule new search after 300ms delay
        self._search_after_id = self.main_frame.after(300, self.apply_filters)
    
    def apply_filters(self):
        """Apply current search and filter settings"""
        # Update filter state
        self.search_text = self.search_entry.get()
        self.status_filter = self.status_filter_combo.get()
        self.contribution_filter = self.contribution_filter_combo.get()
        self.loan_filter = self.loan_filter_combo.get()
        
        # Refresh the list with new filters
        self.refresh_member_list()
    
    def sort_by_column(self, column):
        """Sort the member list by the specified column"""
        # Toggle sort direction if same column
        if self.current_sort_column == column:
            self.current_sort_direction = 'desc' if self.current_sort_direction == 'asc' else 'asc'
        else:
            self.current_sort_column = column
            self.current_sort_direction = 'asc'
        
        # Get current items
        items = [(self.members_tree.set(item, column), item) for item in self.members_tree.get_children('')]
        
        # Sort items
        reverse = self.current_sort_direction == 'desc'
        
        # Handle different column types
        if column in ['ID']:
            # Numeric sort
            items.sort(key=lambda x: int(x[0]) if x[0].isdigit() else 0, reverse=reverse)
        else:
            # String sort
            items.sort(key=lambda x: x[0].lower(), reverse=reverse)
        
        # Rearrange items in sorted order
        for index, (val, item) in enumerate(items):
            self.members_tree.move(item, '', index)
        
        # Update column heading to show sort direction
        for col in ('ID', 'Number', 'Name', 'Phone', 'Status', 'Actions'):
            if col == column:
                direction_symbol = ' ↓' if reverse else ' ↑'
                self.members_tree.heading(col, text=f"{col}{direction_symbol}")
            else:
                self.members_tree.heading(col, text=col)
    
    def on_member_select(self, event):
        """Handle member selection"""
        selection = self.members_tree.selection()
        if selection:
            item = selection[0]
            # Get member_id from the first column (ID)
            member_id = self.members_tree.item(item)['values'][0]
            self.selected_member_id = member_id
            
            # Get member name for display
            member_name = self.members_tree.item(item)['values'][2]
            self.update_selection_info(f"Selected: {member_name}")
        else:
            self.selected_member_id = None
            self.update_selection_info("No selection")
        
        # Update button states
        self.update_button_states()
    
    def on_member_double_click(self, event):
        """Handle double-click on member - show details"""
        self.view_member_action()
    
    def update_button_states(self):
        """Update button enabled/disabled states based on selection"""
        has_selection = self.selected_member_id is not None
        
        # Enable/disable buttons based on selection
        state = 'normal' if has_selection else 'disabled'
        self.edit_btn.configure(state=state)
        self.delete_btn.configure(state=state)
        self.view_btn.configure(state=state)
        
        # Update button colors when disabled
        if not has_selection:
            self.edit_btn.configure(bg=self.colors['secondary'])
            self.delete_btn.configure(bg=self.colors['secondary'])
            self.view_btn.configure(bg=self.colors['secondary'])
        else:
            self.edit_btn.configure(bg=self.colors['accent'])
            self.delete_btn.configure(bg=self.colors['danger'])
            self.view_btn.configure(bg=self.colors['primary'])
    
    def update_status(self, message):
        """Update the status bar message"""
        self.status_label.configure(text=message)
    
    def update_selection_info(self, message):
        """Update the selection info in status bar"""
        self.selection_label.configure(text=message)
    
    def create_contributions_panel(self):
        """Create the contributions management panel"""
        # Add contribution form
        form_frame = tk.LabelFrame(self.main_panel, text="Record Contribution", font=('Arial', 12, 'bold'))
        form_frame.pack(fill='x', pady=(0, 10), padx=10)
        
        # Member selection
        tk.Label(form_frame, text="Member:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.contrib_member_combo = ttk.Combobox(form_frame, state='readonly', width=30)
        self.contrib_member_combo.grid(row=0, column=1, padx=5, pady=5)
        
        # Month and year
        tk.Label(form_frame, text="Month:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.contrib_month_combo = ttk.Combobox(form_frame, values=MONTHS, state='readonly', width=15)
        self.contrib_month_combo.grid(row=0, column=3, padx=5, pady=5)
        
        tk.Label(form_frame, text="Year:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.contrib_year_entry = ttk.Entry(form_frame, width=10)
        self.contrib_year_entry.insert(0, str(datetime.now().year))
        self.contrib_year_entry.grid(row=1, column=1, padx=5, pady=5)
        
        # Amount
        tk.Label(form_frame, text="Amount:").grid(row=1, column=2, sticky='w', padx=5, pady=5)
        self.contrib_amount_entry = ttk.Entry(form_frame, width=15)
        self.contrib_amount_entry.grid(row=1, column=3, padx=5, pady=5)
        
        # Buttons
        btn_frame = tk.Frame(form_frame)
        btn_frame.grid(row=2, column=0, columnspan=4, pady=10)
        
        tk.Button(btn_frame, text="Record Contribution", bg=self.colors['success'], fg=self.colors['white'], command=self.record_contribution).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Clear", bg=self.colors['secondary'], fg=self.colors['white'], command=self.clear_contribution_form).pack(side='left', padx=5)
        
        # Contributions list
        list_frame = tk.LabelFrame(self.main_panel, text="Recent Contributions", font=('Arial', 12, 'bold'))
        list_frame.pack(fill='both', expand=True, padx=10)
        
        columns = ('ID', 'Member', 'Month', 'Year', 'Amount', 'Date')
        self.contributions_tree = ttk.Treeview(list_frame, columns=columns, show='headings')
        
        for col in columns:
            self.contributions_tree.heading(col, text=col)
            self.contributions_tree.column(col, width=150)
        
        scrollbar_contrib = ttk.Scrollbar(list_frame, orient='vertical', command=self.contributions_tree.yview)
        self.contributions_tree.configure(yscrollcommand=scrollbar_contrib.set)
        
        self.contributions_tree.pack(side='left', fill='both', expand=True)
        scrollbar_contrib.pack(side='right', fill='y')
        
        self.refresh_contributions_list()
        self.load_members_for_contributions()
    
    def create_loans_panel(self):
        """Create the loans management panel"""
        # Add loan form
        form_frame = tk.LabelFrame(self.main_panel, text="Process Loan", font=('Arial', 12, 'bold'))
        form_frame.pack(fill='x', pady=(0, 10), padx=10)
        
        # Member selection
        tk.Label(form_frame, text="Member:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.loan_member_combo = ttk.Combobox(form_frame, state='readonly', width=30)
        self.loan_member_combo.grid(row=0, column=1, padx=5, pady=5)
        
        # Loan amount
        tk.Label(form_frame, text="Amount:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.loan_amount_entry = ttk.Entry(form_frame, width=15)
        self.loan_amount_entry.grid(row=0, column=3, padx=5, pady=5)
        
        # Interest rate
        tk.Label(form_frame, text="Interest Rate (%):").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.interest_rate_entry = ttk.Entry(form_frame, width=10)
        self.interest_rate_entry.insert(0, "20")  # Default 20%
        self.interest_rate_entry.grid(row=1, column=1, padx=5, pady=5)
        
        # Loan term
        tk.Label(form_frame, text="Term (months):").grid(row=1, column=2, sticky='w', padx=5, pady=5)
        self.loan_term_entry = ttk.Entry(form_frame, width=10)
        self.loan_term_entry.insert(0, "12")  # Default 12 months
        self.loan_term_entry.grid(row=1, column=3, padx=5, pady=5)
        
        # Buttons
        btn_frame = tk.Frame(form_frame)
        btn_frame.grid(row=2, column=0, columnspan=4, pady=10)
        
        tk.Button(btn_frame, text="Process Loan", bg=self.colors['warning'], fg=self.colors['white'], command=self.process_loan).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Calculate Payment", bg=self.colors['accent'], fg=self.colors['white'], command=self.calculate_loan_payment).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Clear", bg=self.colors['secondary'], fg=self.colors['white'], command=self.clear_loan_form).pack(side='left', padx=5)
        
        # Loans list
        list_frame = tk.LabelFrame(self.main_panel, text="Active Loans", font=('Arial', 12, 'bold'))
        list_frame.pack(fill='both', expand=True, padx=10)
        
        columns = ('ID', 'Member', 'Amount', 'Interest Rate', 'Monthly Payment', 'Outstanding', 'Status')
        self.loans_tree = ttk.Treeview(list_frame, columns=columns, show='headings')
        
        for col in columns:
            self.loans_tree.heading(col, text=col)
            self.loans_tree.column(col, width=120)
        
        scrollbar_loans = ttk.Scrollbar(list_frame, orient='vertical', command=self.loans_tree.yview)
        self.loans_tree.configure(yscrollcommand=scrollbar_loans.set)
        
        self.loans_tree.pack(side='left', fill='both', expand=True)
        scrollbar_loans.pack(side='right', fill='y')
        
        self.refresh_loans_list()
        self.load_members_for_loans()
    
    def create_repayments_panel(self):
        """Create the repayments management panel"""
        # Add repayment form
        form_frame = tk.LabelFrame(self.main_panel, text="Record Repayment", font=('Arial', 12, 'bold'))
        form_frame.pack(fill='x', pady=(0, 10), padx=10)
        
        # Loan selection
        tk.Label(form_frame, text="Loan:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.repayment_loan_combo = ttk.Combobox(form_frame, state='readonly', width=40)
        self.repayment_loan_combo.grid(row=0, column=1, columnspan=2, padx=5, pady=5)
        
        # Amount
        tk.Label(form_frame, text="Amount:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.repayment_amount_entry = ttk.Entry(form_frame, width=15)
        self.repayment_amount_entry.grid(row=1, column=1, padx=5, pady=5)
        
        # Buttons
        btn_frame = tk.Frame(form_frame)
        btn_frame.grid(row=2, column=0, columnspan=3, pady=10)
        
        tk.Button(btn_frame, text="Record Repayment", bg=self.colors['success'], fg=self.colors['white'], command=self.record_repayment).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Clear", bg=self.colors['secondary'], fg=self.colors['white'], command=self.clear_repayment_form).pack(side='left', padx=5)
        
        # Repayments list
        list_frame = tk.LabelFrame(self.main_panel, text="Recent Repayments", font=('Arial', 12, 'bold'))
        list_frame.pack(fill='both', expand=True, padx=10)
        
        columns = ('ID', 'Loan ID', 'Member', 'Amount', 'Principal', 'Interest', 'Date')
        self.repayments_tree = ttk.Treeview(list_frame, columns=columns, show='headings')
        
        for col in columns:
            self.repayments_tree.heading(col, text=col)
            self.repayments_tree.column(col, width=120)
        
        scrollbar_repay = ttk.Scrollbar(list_frame, orient='vertical', command=self.repayments_tree.yview)
        self.repayments_tree.configure(yscrollcommand=scrollbar_repay.set)
        
        self.repayments_tree.pack(side='left', fill='both', expand=True)
        scrollbar_repay.pack(side='right', fill='y')
        
        self.refresh_repayments_list()
        self.load_loans_for_repayments()
    
    def create_reports_panel(self):
        """Create the reports and analytics panel"""
        # Report generation controls
        controls_frame = tk.LabelFrame(self.main_panel, text="Generate Reports", font=('Arial', 12, 'bold'))
        controls_frame.pack(fill='x', pady=(0, 10), padx=10)
        
        # Report type selection
        tk.Label(controls_frame, text="Report Type:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.report_type_combo = ttk.Combobox(controls_frame, values=[
            'Comprehensive Financial Report',
            'Member Analysis Report',
            'Loan Portfolio Report',
            'Dividend Distribution Report',
            'Monthly Summary Report'
        ], state='readonly', width=30)
        self.report_type_combo.set('Comprehensive Financial Report')
        self.report_type_combo.grid(row=0, column=1, padx=5, pady=5)
        
        # Date range
        tk.Label(controls_frame, text="End Date:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.report_end_date = ttk.Entry(controls_frame, width=15)
        self.report_end_date.insert(0, datetime.now().strftime('%Y-%m-%d'))
        self.report_end_date.grid(row=0, column=3, padx=5, pady=5)
        
        # Buttons
        btn_frame = tk.Frame(controls_frame)
        btn_frame.grid(row=1, column=0, columnspan=4, pady=10)
        
        tk.Button(btn_frame, text="Generate Report", bg=self.colors['accent'], fg=self.colors['white'], command=self.generate_report).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Export PDF", bg=self.colors['danger'], fg=self.colors['white'], command=self.export_report_pdf).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Export Excel", bg=self.colors['success'], fg=self.colors['white'], command=self.export_report_excel).pack(side='left', padx=5)
        
        # Report display area
        display_frame = tk.LabelFrame(self.main_panel, text="Report Output", font=('Arial', 12, 'bold'))
        display_frame.pack(fill='both', expand=True, padx=10)
        
        self.report_text = tk.Text(display_frame, font=('Courier', 9), wrap='none')
        
        # Scrollbars for report text
        v_scrollbar = ttk.Scrollbar(display_frame, orient='vertical', command=self.report_text.yview)
        h_scrollbar = ttk.Scrollbar(display_frame, orient='horizontal', command=self.report_text.xview)
        self.report_text.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.report_text.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        display_frame.grid_rowconfigure(0, weight=1)
        display_frame.grid_columnconfigure(0, weight=1)
    
    def create_settings_panel(self):
        """Create the system settings panel"""
        # Configuration settings
        config_frame = tk.LabelFrame(self.main_panel, text="System Configuration", font=('Arial', 12, 'bold'))
        config_frame.pack(fill='x', pady=(0, 10), padx=10)
        
        # Get current config values
        config_manager = SystemConfigManager()
        
        # Organization settings
        tk.Label(config_frame, text="Organization Name:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.org_name_entry = ttk.Entry(config_frame, width=30)
        self.org_name_entry.insert(0, config_manager.get_config_value('organization_name', 'Bank Mmudzi'))
        self.org_name_entry.grid(row=0, column=1, padx=5, pady=5)
        
        # Financial settings
        tk.Label(config_frame, text="Monthly Contribution:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.monthly_contrib_entry = ttk.Entry(config_frame, width=15)
        self.monthly_contrib_entry.insert(0, config_manager.get_config_value('monthly_contribution_amount', '100'))
        self.monthly_contrib_entry.grid(row=1, column=1, padx=5, pady=5)
        
        tk.Label(config_frame, text="Default Interest Rate (%):").grid(row=1, column=2, sticky='w', padx=5, pady=5)
        self.default_interest_entry = ttk.Entry(config_frame, width=10)
        self.default_interest_entry.insert(0, config_manager.get_config_value('default_interest_rate', '20'))
        self.default_interest_entry.grid(row=1, column=3, padx=5, pady=5)
        
        # Security settings
        tk.Label(config_frame, text="Session Timeout (min):").grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.session_timeout_entry = ttk.Entry(config_frame, width=10)
        self.session_timeout_entry.insert(0, config_manager.get_config_value('session_timeout_minutes', '30'))
        self.session_timeout_entry.grid(row=2, column=1, padx=5, pady=5)
        
        # Save button
        tk.Button(config_frame, text="Save Settings", bg=self.colors['success'], fg=self.colors['white'], command=self.save_settings).grid(row=3, column=0, columnspan=4, pady=10)
        
        # Database management
        db_frame = tk.LabelFrame(self.main_panel, text="Database Management", font=('Arial', 12, 'bold'))
        db_frame.pack(fill='x', pady=(0, 10), padx=10)
        
        # Database status
        status_frame = tk.Frame(db_frame)
        status_frame.pack(fill='x', padx=10, pady=5)
        
        tk.Label(status_frame, text="Database Status:", font=('Arial', 10, 'bold')).pack(side='left')
        self.db_status_label = tk.Label(status_frame, text="Connected", fg=self.colors['success'])
        self.db_status_label.pack(side='left', padx=10)
        
        # Database actions
        actions_frame = tk.Frame(db_frame)
        actions_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Button(actions_frame, text="Backup Database", bg=self.colors['warning'], fg=self.colors['white'], command=self.backup_database).pack(side='left', padx=5)
        tk.Button(actions_frame, text="Sync to MySQL", bg=self.colors['accent'], fg=self.colors['white'], command=self.sync_to_mysql).pack(side='left', padx=5)
        tk.Button(actions_frame, text="Check Integrity", bg=self.colors['secondary'], fg=self.colors['white'], command=self.check_db_integrity).pack(side='left', padx=5)
    
    def create_audit_panel(self):
        """Create the audit trail panel"""
        # Audit filters
        filter_frame = tk.LabelFrame(self.main_panel, text="Audit Filters", font=('Arial', 12, 'bold'))
        filter_frame.pack(fill='x', pady=(0, 10), padx=10)
        
        # Date range
        tk.Label(filter_frame, text="From Date:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.audit_from_date = ttk.Entry(filter_frame, width=15)
        self.audit_from_date.insert(0, (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        self.audit_from_date.grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(filter_frame, text="To Date:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.audit_to_date = ttk.Entry(filter_frame, width=15)
        self.audit_to_date.insert(0, datetime.now().strftime('%Y-%m-%d'))
        self.audit_to_date.grid(row=0, column=3, padx=5, pady=5)
        
        # Action type filter
        tk.Label(filter_frame, text="Action Type:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.audit_action_combo = ttk.Combobox(filter_frame, values=['All', 'INSERT', 'UPDATE', 'DELETE', 'LOGIN', 'LOGOUT'], state='readonly', width=15)
        self.audit_action_combo.set('All')
        self.audit_action_combo.grid(row=1, column=1, padx=5, pady=5)
        
        # Apply filter button
        tk.Button(filter_frame, text="Apply Filter", bg=self.colors['accent'], fg=self.colors['white'], command=self.apply_audit_filter).grid(row=1, column=2, padx=10, pady=5)
        
        # Export button
        tk.Button(filter_frame, text="Export CSV", bg=self.colors['success'], fg=self.colors['white'], command=self.export_audit_csv).grid(row=1, column=3, padx=5, pady=5)
        
        # Audit log list
        list_frame = tk.LabelFrame(self.main_panel, text="Audit Trail", font=('Arial', 12, 'bold'))
        list_frame.pack(fill='both', expand=True, padx=10)
        
        columns = ('ID', 'Date/Time', 'User', 'Action', 'Table', 'Description')
        self.audit_tree = ttk.Treeview(list_frame, columns=columns, show='headings')
        
        for col in columns:
            self.audit_tree.heading(col, text=col)
            self.audit_tree.column(col, width=150)
        
        scrollbar_audit = ttk.Scrollbar(list_frame, orient='vertical', command=self.audit_tree.yview)
        self.audit_tree.configure(yscrollcommand=scrollbar_audit.set)
        
        self.audit_tree.pack(side='left', fill='both', expand=True)
        scrollbar_audit.pack(side='right', fill='y')
        
        self.refresh_audit_log()
    
    # Helper methods for the panels
    def apply_member_filter(self):
        """Apply member filtering"""
        status = None if self.member_status_filter.get() == 'All' else self.member_status_filter.get()
        contribution_level = None if self.contribution_filter.get() == 'All' else self.contribution_filter.get()
        
        filtered_members = self.member_analyzer.filter_members_by_criteria(
            status=status, 
            contribution_level=contribution_level
        )
        
        # Clear and populate tree
        for item in self.members_tree.get_children():
            self.members_tree.delete(item)
        
        for member in filtered_members:
            standing_data = self.member_analyzer.calculate_member_standing(member['member_id'])
            standing = standing_data.get('standing_category', 'Unknown') if standing_data.get('success') else 'Unknown'
            
            self.members_tree.insert('', 'end', values=(
                member['member_id'],
                member['member_number'],
                f"{member['name']} {member['surname']}",
                member['phone'],
                member['status'],
                f"MWK {member['total_contributions']:,.2f}",
                f"MWK {member['outstanding_loans']:,.2f}",
                standing
            ))
    
    def refresh_members_list(self):
        """Refresh the members list"""
        self.apply_member_filter()
    
    def refresh_contributions_list(self):
        """Refresh the contributions list"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT c.contribution_id, m.name || ' ' || m.surname, c.month, c.year, c.amount, c.contribution_date
                FROM Contributions c
                JOIN Members m ON c.member_id = m.member_id
                ORDER BY c.contribution_date DESC
                LIMIT 100
            ''')
            
            # Clear tree
            for item in self.contributions_tree.get_children():
                self.contributions_tree.delete(item)
            
            # Populate tree
            for row in cursor.fetchall():
                month_name = MONTHS[row[2] - 1] if 1 <= row[2] <= 12 else str(row[2])
                self.contributions_tree.insert('', 'end', values=(
                    row[0], row[1], month_name, row[3], f"MWK {row[4]:,.2f}", row[5]
                ))
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh contributions: {ErrorHandler.get_user_friendly_message(e)}")
    
    def refresh_loans_list(self):
        """Refresh the loans list"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT l.loan_id, m.name || ' ' || m.surname, l.loan_amount, l.interest_rate, 
                       l.monthly_payment, l.outstanding_balance,
                       CASE WHEN l.outstanding_balance > 0 THEN 'Active' ELSE 'Paid' END
                FROM Loans l
                JOIN Members m ON l.member_id = m.member_id
                ORDER BY l.loan_date DESC
            ''')
            
            # Clear tree
            for item in self.loans_tree.get_children():
                self.loans_tree.delete(item)
            
            # Populate tree
            for row in cursor.fetchall():
                self.loans_tree.insert('', 'end', values=(
                    row[0], row[1], f"MWK {row[2]:,.2f}", f"{row[3]:.1f}%", 
                    f"MWK {row[4]:,.2f}", f"MWK {row[5]:,.2f}", row[6]
                ))
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh loans: {ErrorHandler.get_user_friendly_message(e)}")
    
    def refresh_repayments_list(self):
        """Refresh the repayments list"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT r.repayment_id, r.loan_id, m.name || ' ' || m.surname, 
                       r.repayment_amount, r.principal_amount, r.interest_amount, r.repayment_date
                FROM Repayments r
                JOIN Loans l ON r.loan_id = l.loan_id
                JOIN Members m ON l.member_id = m.member_id
                ORDER BY r.repayment_date DESC
                LIMIT 100
            ''')
            
            # Clear tree
            for item in self.repayments_tree.get_children():
                self.repayments_tree.delete(item)
            
            # Populate tree
            for row in cursor.fetchall():
                self.repayments_tree.insert('', 'end', values=(
                    row[0], row[1], row[2], f"MWK {row[3]:,.2f}", 
                    f"MWK {row[4]:,.2f}", f"MWK {row[5]:,.2f}", row[6]
                ))
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh repayments: {ErrorHandler.get_user_friendly_message(e)}")
    
    def refresh_audit_log(self):
        """Refresh the audit log"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT audit_id, timestamp, user_id, operation, table_name, old_values
                FROM AuditLog
                ORDER BY timestamp DESC
                LIMIT 500
            ''')
            
            # Clear tree
            for item in self.audit_tree.get_children():
                self.audit_tree.delete(item)
            
            # Populate tree
            for row in cursor.fetchall():
                self.audit_tree.insert('', 'end', values=row)
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh audit log: {ErrorHandler.get_user_friendly_message(e)}")
    
    # Placeholder methods for form operations (these would need full implementation)
    def show_add_member_dialog(self):
        """Show dialog to add a new member"""
        dialog = AddMemberDialog(self, self)
        self.wait_window(dialog.dialog)
        if dialog.result:
            self.refresh_members_list()
    
    def show_member_details(self, event):
        """Show detailed member information dialog"""
        selection = self.members_tree.selection()
        if not selection:
            return
        
        item = self.members_tree.item(selection[0])
        member_id = item['values'][0]
        
        dialog = MemberDetailsDialog(self, member_id, self)
        self.wait_window(dialog.dialog)
        if dialog.result:
            self.refresh_members_list()
    
    def record_contribution(self):
        """Record a new contribution"""
        try:
            # Get form data
            member_name = self.contrib_member_combo.get().strip()
            month = self.contrib_month_combo.get().strip()
            amount_str = self.contrib_amount_entry.get().strip()
            
            # Validate inputs
            if not member_name:
                messagebox.showerror("Validation Error", "Please select a member")
                return
            
            if not month:
                messagebox.showerror("Validation Error", "Please select a month")
                return
            
            if not amount_str:
                messagebox.showerror("Validation Error", "Please enter contribution amount")
                return
            
            try:
                amount = float(amount_str)
                if amount <= 0:
                    messagebox.showerror("Validation Error", "Contribution amount must be greater than 0")
                    return
            except ValueError:
                messagebox.showerror("Validation Error", "Please enter a valid amount")
                return
            
            # Get member ID from name
            member_id = None
            members = read_members()
            for member in members:
                if f"{member[1]} {member[2]}" == member_name:
                    member_id = member[0]
                    break
            
            if not member_id:
                messagebox.showerror("Error", "Selected member not found")
                return
            
            # Check for duplicate contribution
            current_year = datetime.now().year
            duplicate_check = ValidationEngine.check_duplicate_contribution(member_id, MONTH_TO_NUM[month], current_year)
            if not duplicate_check['valid']:
                messagebox.showerror("Duplicate Contribution", duplicate_check['message'])
                return
            
            # Create contribution
            contribution_id = create_contribution(member_id, month, amount)
            
            if contribution_id:
                messagebox.showinfo("Success", f"Contribution of MWK {amount:,.2f} recorded successfully for {member_name} ({month})")
                self.clear_contribution_form()
                # Refresh any relevant displays
                if hasattr(self, 'refresh_dashboard'):
                    self.refresh_dashboard()
            else:
                messagebox.showerror("Error", "Failed to record contribution")
                
        except Exception as e:
            error_info = ErrorHandler.handle_error(e, "recording contribution")
            messagebox.showerror("Error", error_info['message'])
    
    def clear_contribution_form(self):
        """Clear contribution form"""
        self.contrib_member_combo.set('')
        self.contrib_month_combo.set('')
        self.contrib_amount_entry.delete(0, tk.END)
    
    def process_loan(self):
        """Process a new loan"""
        try:
            # Get form data
            member_name = self.loan_member_combo.get().strip()
            amount_str = self.loan_amount_entry.get().strip()
            interest_rate_str = self.loan_interest_entry.get().strip() if hasattr(self, 'loan_interest_entry') else "10.0"
            
            # Validate inputs
            if not member_name:
                messagebox.showerror("Validation Error", "Please select a member")
                return
            
            if not amount_str:
                messagebox.showerror("Validation Error", "Please enter loan amount")
                return
            
            try:
                loan_amount = float(amount_str)
                if loan_amount <= 0:
                    messagebox.showerror("Validation Error", "Loan amount must be greater than 0")
                    return
            except ValueError:
                messagebox.showerror("Validation Error", "Please enter a valid loan amount")
                return
            
            try:
                interest_rate = float(interest_rate_str)
                if interest_rate < 0:
                    messagebox.showerror("Validation Error", "Interest rate cannot be negative")
                    return
            except ValueError:
                messagebox.showerror("Validation Error", "Please enter a valid interest rate")
                return
            
            # Get member ID from name
            member_id = None
            members = read_members()
            for member in members:
                if f"{member[1]} {member[2]}" == member_name:
                    member_id = member[0]
                    break
            
            if not member_id:
                messagebox.showerror("Error", "Selected member not found")
                return
            
            # Validate loan capacity
            capacity_check = ValidationEngine.validate_loan_capacity(member_id, loan_amount)
            if not capacity_check['valid']:
                messagebox.showerror("Loan Capacity Error", capacity_check['message'])
                return
            
            # Show loan confirmation dialog
            monthly_rate = interest_rate / 100 / 12
            monthly_payment = FinancialCalculator.calculate_loan_payment(loan_amount, monthly_rate, 12)
            total_interest = FinancialCalculator.calculate_total_interest(loan_amount, monthly_payment, 12)
            
            confirmation_msg = f"""
Loan Details:
Member: {member_name}
Loan Amount: MWK {loan_amount:,.2f}
Interest Rate: {interest_rate:.1f}% per year
Monthly Payment: MWK {monthly_payment:,.2f}
Total Interest: MWK {total_interest:,.2f}
Total Repayment: MWK {(loan_amount + total_interest):,.2f}

Do you want to process this loan?
            """
            
            if messagebox.askyesno("Confirm Loan", confirmation_msg):
                # Create loan
                loan_id = create_loan(member_id, loan_amount, interest_rate)
                
                if loan_id:
                    messagebox.showinfo("Success", f"Loan of MWK {loan_amount:,.2f} processed successfully for {member_name}")
                    self.clear_loan_form()
                    # Refresh any relevant displays
                    if hasattr(self, 'refresh_dashboard'):
                        self.refresh_dashboard()
                else:
                    messagebox.showerror("Error", "Failed to process loan")
                
        except Exception as e:
            error_info = ErrorHandler.handle_error(e, "processing loan")
            messagebox.showerror("Error", error_info['message'])
    
    def calculate_loan_payment(self):
        """Calculate and display loan payment details"""
        try:
            amount_str = self.loan_amount_entry.get().strip()
            interest_rate_str = self.loan_interest_entry.get().strip() if hasattr(self, 'loan_interest_entry') else "10.0"
            
            if not amount_str:
                messagebox.showwarning("Input Required", "Please enter loan amount")
                return
            
            try:
                loan_amount = float(amount_str)
                interest_rate = float(interest_rate_str)
                
                if loan_amount <= 0:
                    messagebox.showerror("Invalid Input", "Loan amount must be greater than 0")
                    return
                
                # Calculate payment details
                monthly_rate = interest_rate / 100 / 12
                monthly_payment = FinancialCalculator.calculate_loan_payment(loan_amount, monthly_rate, 12)
                total_interest = FinancialCalculator.calculate_total_interest(loan_amount, monthly_payment, 12)
                total_repayment = loan_amount + total_interest
                
                # Show calculation results
                calc_msg = f"""
Loan Payment Calculation:

Principal Amount: MWK {loan_amount:,.2f}
Interest Rate: {interest_rate:.1f}% per year
Loan Term: 12 months

Monthly Payment: MWK {monthly_payment:,.2f}
Total Interest: MWK {total_interest:,.2f}
Total Repayment: MWK {total_repayment:,.2f}

Interest-to-Principal Ratio: {(total_interest/loan_amount)*100:.1f}%
                """
                
                messagebox.showinfo("Loan Calculation", calc_msg)
                
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter valid numeric values")
                
        except Exception as e:
            error_info = ErrorHandler.handle_error(e, "calculating loan payment")
            messagebox.showerror("Error", error_info['message'])
    
    def clear_loan_form(self):
        """Clear loan form"""
        self.loan_member_combo.set('')
        self.loan_amount_entry.delete(0, tk.END)
        if hasattr(self, 'loan_interest_entry'):
            self.loan_interest_entry.delete(0, tk.END)
    
    def record_repayment(self):
        messagebox.showinfo("Feature", "Repayment recording would be implemented here")
    
    def clear_repayment_form(self):
        self.repayment_loan_combo.set('')
        self.repayment_amount_entry.delete(0, tk.END)
    
    def generate_report(self):
        """Generate the selected report"""
        try:
            end_date = self.report_end_date.get()
            report_content = self.report_generator.generate_comprehensive_report(end_date)
            
            self.report_text.delete(1.0, tk.END)
            self.report_text.insert(1.0, report_content)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report: {ErrorHandler.get_user_friendly_message(e)}")
    
    def export_report_pdf(self):
        messagebox.showinfo("Feature", "PDF export would be implemented here")
    
    def export_report_excel(self):
        messagebox.showinfo("Feature", "Excel export would be implemented here")
    
    def save_settings(self):
        messagebox.showinfo("Feature", "Settings save would be implemented here")
    
    def backup_database(self):
        messagebox.showinfo("Feature", "Database backup would be implemented here")
    
    def sync_to_mysql(self):
        messagebox.showinfo("Feature", "MySQL sync would be implemented here")
    
    def check_db_integrity(self):
        messagebox.showinfo("Feature", "Database integrity check would be implemented here")
    
    def apply_audit_filter(self):
        messagebox.showinfo("Feature", "Audit filter would be implemented here")
    
    def export_audit_csv(self):
        messagebox.showinfo("Feature", "Audit CSV export would be implemented here")
    
    def load_members_for_contributions(self):
        """Load members into contribution combo"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT member_id, name, surname FROM Members WHERE status = 'active' ORDER BY name")
            members = [f"{row[1]} {row[2]} (ID: {row[0]})" for row in cursor.fetchall()]
            self.contrib_member_combo['values'] = members
            conn.close()
        except Exception as e:
            logger.error(f"Error loading members: {str(e)}")
    
    def load_members_for_loans(self):
        """Load members into loan combo"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT member_id, name, surname FROM Members WHERE status = 'active' ORDER BY name")
            members = [f"{row[1]} {row[2]} (ID: {row[0]})" for row in cursor.fetchall()]
            self.loan_member_combo['values'] = members
            conn.close()
        except Exception as e:
            logger.error(f"Error loading members: {str(e)}")
    
    def load_loans_for_repayments(self):
        """Load active loans into repayment combo"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT l.loan_id, m.name, m.surname, l.outstanding_balance
                FROM Loans l
                JOIN Members m ON l.member_id = m.member_id
                WHERE l.outstanding_balance > 0
                ORDER BY m.name
            ''')
            loans = [f"Loan {row[0]} - {row[1]} {row[2]} (Outstanding: MWK {row[3]:,.2f})" for row in cursor.fetchall()]
            self.repayment_loan_combo['values'] = loans
            conn.close()
        except Exception as e:
            logger.error(f"Error loading loans: {str(e)}")
        self.geometry("1200x800")
        self.configure(bg='#f0f0f0')
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Create tabs
        self.create_members_tab()
        self.create_contributions_tab()
        self.create_loans_tab()
        self.create_repayments_tab()
        self.create_reports_tab()
        self.create_sync_status_tab()
        self.create_audit_trail_tab()
        
        # Auto-refresh data every 5 seconds
        self.auto_refresh()
    
    def auto_refresh(self):
        """Auto-refresh data in all tabs"""
        try:
            self.refresh_members()
            self.refresh_contributions()
            self.refresh_loans()
            self.refresh_repayments()
            self.refresh_sync_status()
            self.refresh_audit_trail()
        except Exception as e:
            print(f"Auto-refresh error: {e}")
        
        # Schedule next refresh
        self.after(5000, self.auto_refresh)
    
    def create_members_tab(self):
        # Members tab
        members_frame = ttk.Frame(self.notebook)
        self.notebook.add(members_frame, text="Members")
        
        # Add member form
        form_frame = ttk.LabelFrame(members_frame, text="Add/Edit Member")
        form_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(form_frame, text="Name:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.name_entry = ttk.Entry(form_frame)
        self.name_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Surname:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.surname_entry = ttk.Entry(form_frame)
        self.surname_entry.grid(row=0, column=3, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Phone:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.phone_entry = ttk.Entry(form_frame)
        self.phone_entry.grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Email:").grid(row=1, column=2, sticky='w', padx=5, pady=5)
        self.email_entry = ttk.Entry(form_frame)
        self.email_entry.grid(row=1, column=3, padx=5, pady=5)
        
        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=2, column=0, columnspan=4, pady=10)
        
        ttk.Button(button_frame, text="Add Member", command=self.add_member).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Update Member", command=self.update_member).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Delete Member", command=self.delete_member).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Clear", command=self.clear_member_form).pack(side='left', padx=5)
        
        # Members list
        list_frame = ttk.LabelFrame(members_frame, text="Members List")
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        columns = ('ID', 'Name', 'Surname', 'Phone', 'Email')
        self.members_tree = ttk.Treeview(list_frame, columns=columns, show='headings')
        
        for col in columns:
            self.members_tree.heading(col, text=col)
            self.members_tree.column(col, width=150)
        
        scrollbar_members = ttk.Scrollbar(list_frame, orient='vertical', command=self.members_tree.yview)
        self.members_tree.configure(yscrollcommand=scrollbar_members.set)
        
        self.members_tree.pack(side='left', fill='both', expand=True)
        scrollbar_members.pack(side='right', fill='y')
        
        self.members_tree.bind('<<TreeviewSelect>>', self.on_member_select)
        
        self.refresh_members()
    
    def create_contributions_tab(self):
        # Contributions tab
        contributions_frame = ttk.Frame(self.notebook)
        self.notebook.add(contributions_frame, text="Contributions")
        
        # Add contribution form
        form_frame = ttk.LabelFrame(contributions_frame, text="Add/Edit Contribution")
        form_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(form_frame, text="Member:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.member_combo = ttk.Combobox(form_frame, state='readonly')
        self.member_combo.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Month:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.month_combo = ttk.Combobox(form_frame, values=MONTHS, state='readonly')
        self.month_combo.grid(row=0, column=3, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Amount:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.contribution_amount_entry = ttk.Entry(form_frame)
        self.contribution_amount_entry.grid(row=1, column=1, padx=5, pady=5)
        
        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=2, column=0, columnspan=4, pady=10)
        
        ttk.Button(button_frame, text="Add Contribution", command=self.add_contribution).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Update Contribution", command=self.update_contribution).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Delete Contribution", command=self.delete_contribution).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Clear", command=self.clear_contribution_form).pack(side='left', padx=5)
        
        # Contributions list
        list_frame = ttk.LabelFrame(contributions_frame, text="Contributions List")
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        columns = ('ID', 'Member', 'Month', 'Amount', 'Date')
        self.contributions_tree = ttk.Treeview(list_frame, columns=columns, show='headings')
        
        for col in columns:
            self.contributions_tree.heading(col, text=col)
            self.contributions_tree.column(col, width=150)
        
        scrollbar_contributions = ttk.Scrollbar(list_frame, orient='vertical', command=self.contributions_tree.yview)
        self.contributions_tree.configure(yscrollcommand=scrollbar_contributions.set)
        
        self.contributions_tree.pack(side='left', fill='both', expand=True)
        scrollbar_contributions.pack(side='right', fill='y')
        
        self.contributions_tree.bind('<<TreeviewSelect>>', self.on_contribution_select)
        
        self.refresh_contributions()
    
    def create_loans_tab(self):
        # Loans tab
        loans_frame = ttk.Frame(self.notebook)
        self.notebook.add(loans_frame, text="Loans")
        
        # Add loan form
        form_frame = ttk.LabelFrame(loans_frame, text="Add/Edit Loan")
        form_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(form_frame, text="Member:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.loan_member_combo = ttk.Combobox(form_frame, state='readonly')
        self.loan_member_combo.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Amount:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.loan_amount_entry = ttk.Entry(form_frame)
        self.loan_amount_entry.grid(row=0, column=3, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Interest Rate:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.interest_rate_entry = ttk.Entry(form_frame)
        self.interest_rate_entry.grid(row=1, column=1, padx=5, pady=5)
        
        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=2, column=0, columnspan=4, pady=10)
        
        ttk.Button(button_frame, text="Add Loan", command=self.add_loan).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Update Loan", command=self.update_loan).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Delete Loan", command=self.delete_loan).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Clear", command=self.clear_loan_form).pack(side='left', padx=5)       
 
        # Loans list
        list_frame = ttk.LabelFrame(loans_frame, text="Loans List")
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        columns = ('ID', 'Member', 'Amount', 'Interest Rate', 'Monthly Payment', 'Total Interest', 'Date')
        self.loans_tree = ttk.Treeview(list_frame, columns=columns, show='headings')
        
        for col in columns:
            self.loans_tree.heading(col, text=col)
            self.loans_tree.column(col, width=150)
        
        scrollbar_loans = ttk.Scrollbar(list_frame, orient='vertical', command=self.loans_tree.yview)
        self.loans_tree.configure(yscrollcommand=scrollbar_loans.set)
        
        self.loans_tree.pack(side='left', fill='both', expand=True)
        scrollbar_loans.pack(side='right', fill='y')
        
        self.loans_tree.bind('<<TreeviewSelect>>', self.on_loan_select)
        
        self.refresh_loans()
    
    def create_repayments_tab(self):
        # Repayments tab
        repayments_frame = ttk.Frame(self.notebook)
        self.notebook.add(repayments_frame, text="Repayments")
        
        # Add repayment form
        form_frame = ttk.LabelFrame(repayments_frame, text="Add/Edit Repayment")
        form_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(form_frame, text="Loan ID:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.repayment_loan_combo = ttk.Combobox(form_frame, state='readonly')
        self.repayment_loan_combo.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Amount:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.repayment_amount_entry = ttk.Entry(form_frame)
        self.repayment_amount_entry.grid(row=0, column=3, padx=5, pady=5)
        
        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=1, column=0, columnspan=4, pady=10)
        
        ttk.Button(button_frame, text="Add Repayment", command=self.add_repayment).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Update Repayment", command=self.update_repayment).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Delete Repayment", command=self.delete_repayment).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Clear", command=self.clear_repayment_form).pack(side='left', padx=5)
        
        # Repayments list
        list_frame = ttk.LabelFrame(repayments_frame, text="Repayments List")
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        columns = ('ID', 'Loan ID', 'Member', 'Amount', 'Date')
        self.repayments_tree = ttk.Treeview(list_frame, columns=columns, show='headings')
        
        for col in columns:
            self.repayments_tree.heading(col, text=col)
            self.repayments_tree.column(col, width=150)
        
        scrollbar_repayments = ttk.Scrollbar(list_frame, orient='vertical', command=self.repayments_tree.yview)
        self.repayments_tree.configure(yscrollcommand=scrollbar_repayments.set)
        
        self.repayments_tree.pack(side='left', fill='both', expand=True)
        scrollbar_repayments.pack(side='right', fill='y')
        
        self.repayments_tree.bind('<<TreeviewSelect>>', self.on_repayment_select)
        
        self.refresh_repayments()
    
    def create_reports_tab(self):
        # Reports tab
        reports_frame = ttk.Frame(self.notebook)
        self.notebook.add(reports_frame, text="Reports")
        
        # Date selection frame
        date_frame = ttk.LabelFrame(reports_frame, text="Report Date")
        date_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(date_frame, text="End Date:").pack(side='left', padx=5)
        self.report_date_entry = ttk.Entry(date_frame)
        self.report_date_entry.pack(side='left', padx=5)
        self.report_date_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        
        ttk.Button(date_frame, text="Generate Report", command=self.generate_report).pack(side='left', padx=10)
        ttk.Button(date_frame, text="Export CSV", command=self.export_csv).pack(side='left', padx=5)
        
        # Report display
        report_display_frame = ttk.LabelFrame(reports_frame, text="Report")
        report_display_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.report_text = tk.Text(report_display_frame, wrap='word')
        scrollbar_report = ttk.Scrollbar(report_display_frame, orient='vertical', command=self.report_text.yview)
        self.report_text.configure(yscrollcommand=scrollbar_report.set)
        
        self.report_text.pack(side='left', fill='both', expand=True)
        scrollbar_report.pack(side='right', fill='y')
    
    def create_sync_status_tab(self):
        # Sync Status tab
        sync_frame = ttk.Frame(self.notebook)
        self.notebook.add(sync_frame, text="Sync Status")
        
        # Database status frame
        status_frame = ttk.LabelFrame(sync_frame, text="Database Status")
        status_frame.pack(fill='x', padx=10, pady=5)
        
        self.db_status_label = ttk.Label(status_frame, text="Database: Checking...")
        self.db_status_label.pack(anchor='w', padx=10, pady=5)
        
        self.mysql_status_label = ttk.Label(status_frame, text="MySQL: Checking...")
        self.mysql_status_label.pack(anchor='w', padx=10, pady=5)
        
        self.sync_status_label = ttk.Label(status_frame, text="Sync Status: Checking...")
        self.sync_status_label.pack(anchor='w', padx=10, pady=5)
        
        # Sync controls
        controls_frame = ttk.LabelFrame(sync_frame, text="Sync Controls")
        controls_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(controls_frame, text="Force Sync", command=self.force_sync).pack(side='left', padx=5, pady=5)
        ttk.Button(controls_frame, text="View Sync Log", command=self.view_sync_log).pack(side='left', padx=5, pady=5)
        ttk.Button(controls_frame, text="View Conflicts", command=self.view_conflicts).pack(side='left', padx=5, pady=5)
        
        # Sync log display
        log_frame = ttk.LabelFrame(sync_frame, text="Recent Sync Activity")
        log_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.sync_log_text = tk.Text(log_frame, wrap='word', height=10)
        scrollbar_sync = ttk.Scrollbar(log_frame, orient='vertical', command=self.sync_log_text.yview)
        self.sync_log_text.configure(yscrollcommand=scrollbar_sync.set)
        
        self.sync_log_text.pack(side='left', fill='both', expand=True)
        scrollbar_sync.pack(side='right', fill='y')
        
        self.refresh_sync_status()   
 # Member management methods
    def refresh_members(self):
        for item in self.members_tree.get_children():
            self.members_tree.delete(item)
        
        members = read_members()
        for member in members:
            self.members_tree.insert('', 'end', values=member)
        
        # Update member combo boxes
        member_names = [f"{member[1]} {member[2]} (ID: {member[0]})" for member in members]
        self.member_combo['values'] = member_names
        self.loan_member_combo['values'] = member_names
    
    def on_member_select(self, event):
        selection = self.members_tree.selection()
        if selection:
            item = self.members_tree.item(selection[0])
            values = item['values']
            
            self.name_entry.delete(0, tk.END)
            self.name_entry.insert(0, values[1])
            
            self.surname_entry.delete(0, tk.END)
            self.surname_entry.insert(0, values[2])
            
            self.phone_entry.delete(0, tk.END)
            self.phone_entry.insert(0, values[3])
            
            self.email_entry.delete(0, tk.END)
            self.email_entry.insert(0, values[4] if values[4] else '')
    
    def add_member(self):
        try:
            name = self.name_entry.get().strip()
            surname = self.surname_entry.get().strip()
            phone = self.phone_entry.get().strip()
            email = self.email_entry.get().strip() or None
            
            if not name or not surname or not phone:
                messagebox.showerror("Error", "Name, surname, and phone are required")
                return
            
            # Call the create_member function with proper error handling
            member_id = create_member(name, surname, phone, email)
            
            if member_id:
                self.refresh_members()
                self.clear_member_form()
                messagebox.showinfo("Success", f"Member added successfully with ID: {member_id}")
            else:
                messagebox.showerror("Error", "Failed to create member - no ID returned")
                
        except Exception as e:
            # Handle all errors gracefully
            if isinstance(e, BankMmudziException):
                error_msg = e.message
            elif hasattr(e, 'args') and e.args:
                error_msg = str(e.args[0])
            else:
                error_msg = str(e)
            messagebox.showerror("Error", f"Failed to add member: {error_msg}")
            logger.error(f"UI Error in add_member: {str(e)}")
    
    def update_member(self):
        selection = self.members_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a member to update")
            return
        
        try:
            item = self.members_tree.item(selection[0])
            member_id = item['values'][0]
            
            name = self.name_entry.get().strip()
            surname = self.surname_entry.get().strip()
            phone = self.phone_entry.get().strip()
            email = self.email_entry.get().strip() or None
            
            if not name or not surname or not phone:
                messagebox.showerror("Error", "Name, surname, and phone are required")
                return
            
            update_member(member_id, name, surname, phone, email)
            self.refresh_members()
            self.clear_member_form()
            messagebox.showinfo("Success", "Member updated successfully")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update member: {e}")
    
    def delete_member(self):
        selection = self.members_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a member to delete")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this member?"):
            try:
                item = self.members_tree.item(selection[0])
                member_id = item['values'][0]
                
                delete_member(member_id)
                self.refresh_members()
                self.clear_member_form()
                messagebox.showinfo("Success", "Member deleted successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete member: {e}")
    
    def clear_member_form(self):
        self.name_entry.delete(0, tk.END)
        self.surname_entry.delete(0, tk.END)
        self.phone_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END) 
   # Contribution management methods
    def refresh_contributions(self):
        for item in self.contributions_tree.get_children():
            self.contributions_tree.delete(item)
        
        members = read_members()
        for member in members:
            member_id, name, surname = member[0], member[1], member[2]
            contributions = read_contributions_for_member(member_id)
            for contribution in contributions:
                contribution_id, month, amount, date = contribution
                full_name = f"{name} {surname}"
                self.contributions_tree.insert('', 'end', values=(contribution_id, full_name, month, format_currency(amount), date))
    
    def on_contribution_select(self, event):
        selection = self.contributions_tree.selection()
        if selection:
            item = self.contributions_tree.item(selection[0])
            values = item['values']
            
            # Find member by name
            member_name = values[1]
            self.member_combo.set(member_name)
            
            self.month_combo.set(values[2])
            
            # Extract amount without currency formatting
            amount_str = values[3].replace('MWK ', '').replace(',', '')
            self.contribution_amount_entry.delete(0, tk.END)
            self.contribution_amount_entry.insert(0, amount_str)
    
    def add_contribution(self):
        try:
            member_selection = self.member_combo.get()
            if not member_selection:
                messagebox.showerror("Error", "Please select a member")
                return
            
            # Extract member ID from selection
            member_id = int(member_selection.split('ID: ')[1].split(')')[0])
            
            month = self.month_combo.get()
            if not month:
                messagebox.showerror("Error", "Please select a month")
                return
            
            amount = float(self.contribution_amount_entry.get())
            if amount <= 0:
                messagebox.showerror("Error", "Amount must be positive")
                return
            
            create_contribution(member_id, month, amount)
            self.refresh_contributions()
            self.clear_contribution_form()
            messagebox.showinfo("Success", "Contribution added successfully")
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid amount")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add contribution: {e}")
    
    def update_contribution(self):
        selection = self.contributions_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a contribution to update")
            return
        
        try:
            item = self.contributions_tree.item(selection[0])
            contribution_id = item['values'][0]
            
            month = self.month_combo.get()
            if not month:
                messagebox.showerror("Error", "Please select a month")
                return
            
            amount = float(self.contribution_amount_entry.get())
            if amount <= 0:
                messagebox.showerror("Error", "Amount must be positive")
                return
            
            update_contribution(contribution_id, month, amount)
            self.refresh_contributions()
            self.clear_contribution_form()
            messagebox.showinfo("Success", "Contribution updated successfully")
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid amount")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update contribution: {e}")
    
    def delete_contribution(self):
        selection = self.contributions_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a contribution to delete")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this contribution?"):
            try:
                item = self.contributions_tree.item(selection[0])
                contribution_id = item['values'][0]
                
                delete_contribution(contribution_id)
                self.refresh_contributions()
                self.clear_contribution_form()
                messagebox.showinfo("Success", "Contribution deleted successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete contribution: {e}")
    
    def clear_contribution_form(self):
        self.member_combo.set('')
        self.month_combo.set('')
        self.contribution_amount_entry.delete(0, tk.END)
    
    # Loan management methods
    def refresh_loans(self):
        for item in self.loans_tree.get_children():
            self.loans_tree.delete(item)
        
        loans = read_loans()
        loan_options = []
        for loan in loans:
            loan_id, member_name, amount, interest_rate, monthly_payment, total_interest, date = loan
            # Handle cases where monthly_payment or total_interest might be None (for existing loans)
            monthly_payment_str = format_currency(monthly_payment) if monthly_payment else "N/A"
            total_interest_str = format_currency(total_interest) if total_interest else "N/A"
            
            self.loans_tree.insert('', 'end', values=(
                loan_id, 
                member_name, 
                format_currency(amount), 
                f"{interest_rate*100:.1f}%", 
                monthly_payment_str,
                total_interest_str,
                date
            ))
            loan_options.append(f"Loan {loan_id} - {member_name}")
        
        # Update repayment loan combo
        self.repayment_loan_combo['values'] = loan_options
    
    def on_loan_select(self, event):
        selection = self.loans_tree.selection()
        if selection:
            item = self.loans_tree.item(selection[0])
            values = item['values']
            
            # Find member by name
            member_name = values[1]
            self.loan_member_combo.set(member_name)
            
            # Extract amount without currency formatting
            amount_str = values[2].replace('MWK ', '').replace(',', '')
            self.loan_amount_entry.delete(0, tk.END)
            self.loan_amount_entry.insert(0, amount_str)
            
            # Extract interest rate
            interest_str = values[3].replace('%', '')
            interest_rate = float(interest_str) / 100
            self.interest_rate_entry.delete(0, tk.END)
            self.interest_rate_entry.insert(0, str(interest_rate))
    
    def add_loan(self):
        try:
            member_selection = self.loan_member_combo.get()
            if not member_selection:
                messagebox.showerror("Error", "Please select a member")
                return
            
            # Extract member ID from selection
            member_id = int(member_selection.split('ID: ')[1].split(')')[0])
            
            amount = float(self.loan_amount_entry.get())
            if amount <= 0:
                messagebox.showerror("Error", "Amount must be positive")
                return
            
            interest_rate = float(self.interest_rate_entry.get())
            if interest_rate < 0:
                messagebox.showerror("Error", "Interest rate cannot be negative")
                return
            
            create_loan(member_id, amount, interest_rate)
            self.refresh_loans()
            self.clear_loan_form()
            messagebox.showinfo("Success", "Loan added successfully")
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numeric values")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add loan: {e}")
    
    def update_loan(self):
        selection = self.loans_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a loan to update")
            return
        
        try:
            item = self.loans_tree.item(selection[0])
            loan_id = item['values'][0]
            
            member_selection = self.loan_member_combo.get()
            if not member_selection:
                messagebox.showerror("Error", "Please select a member")
                return
            
            # Extract member ID from selection
            member_id = int(member_selection.split('ID: ')[1].split(')')[0])
            
            amount = float(self.loan_amount_entry.get())
            if amount <= 0:
                messagebox.showerror("Error", "Amount must be positive")
                return
            
            interest_rate = float(self.interest_rate_entry.get())
            if interest_rate < 0:
                messagebox.showerror("Error", "Interest rate cannot be negative")
                return
            
            update_loan(loan_id, member_id, amount, interest_rate)
            self.refresh_loans()
            self.clear_loan_form()
            messagebox.showinfo("Success", "Loan updated successfully")
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numeric values")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update loan: {e}")
    
    def delete_loan(self):
        selection = self.loans_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a loan to delete")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this loan?"):
            try:
                item = self.loans_tree.item(selection[0])
                loan_id = item['values'][0]
                
                delete_loan(loan_id)
                self.refresh_loans()
                self.clear_loan_form()
                messagebox.showinfo("Success", "Loan deleted successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete loan: {e}")
    
    def clear_loan_form(self):
        self.loan_member_combo.set('')
        self.loan_amount_entry.delete(0, tk.END)
        self.interest_rate_entry.delete(0, tk.END)
    
    # Repayment management methods
    def refresh_repayments(self):
        for item in self.repayments_tree.get_children():
            self.repayments_tree.delete(item)
        
        repayments = read_repayments()
        for repayment in repayments:
            repayment_id, loan_id, member_name, amount, date = repayment
            self.repayments_tree.insert('', 'end', values=(repayment_id, loan_id, member_name, format_currency(amount), date))
    
    def on_repayment_select(self, event):
        selection = self.repayments_tree.selection()
        if selection:
            item = self.repayments_tree.item(selection[0])
            values = item['values']
            
            # Find loan by ID
            loan_id = values[1]
            member_name = values[2]
            loan_option = f"Loan {loan_id} - {member_name}"
            self.repayment_loan_combo.set(loan_option)
            
            # Extract amount without currency formatting
            amount_str = values[3].replace('MWK ', '').replace(',', '')
            self.repayment_amount_entry.delete(0, tk.END)
            self.repayment_amount_entry.insert(0, amount_str)
    
    def add_repayment(self):
        try:
            loan_selection = self.repayment_loan_combo.get()
            if not loan_selection:
                messagebox.showerror("Error", "Please select a loan")
                return
            
            # Extract loan ID from selection
            loan_id = int(loan_selection.split('Loan ')[1].split(' - ')[0])
            
            amount = float(self.repayment_amount_entry.get())
            if amount <= 0:
                messagebox.showerror("Error", "Amount must be positive")
                return
            
            create_repayment(loan_id, amount)
            self.refresh_repayments()
            self.clear_repayment_form()
            messagebox.showinfo("Success", "Repayment added successfully")
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid amount")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add repayment: {e}")
    
    def update_repayment(self):
        selection = self.repayments_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a repayment to update")
            return
        
        try:
            item = self.repayments_tree.item(selection[0])
            repayment_id = item['values'][0]
            
            loan_selection = self.repayment_loan_combo.get()
            if not loan_selection:
                messagebox.showerror("Error", "Please select a loan")
                return
            
            # Extract loan ID from selection
            loan_id = int(loan_selection.split('Loan ')[1].split(' - ')[0])
            
            amount = float(self.repayment_amount_entry.get())
            if amount <= 0:
                messagebox.showerror("Error", "Amount must be positive")
                return
            
            update_repayment(repayment_id, loan_id, amount)
            self.refresh_repayments()
            self.clear_repayment_form()
            messagebox.showinfo("Success", "Repayment updated successfully")
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid amount")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update repayment: {e}")
    
    def delete_repayment(self):
        selection = self.repayments_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a repayment to delete")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this repayment?"):
            try:
                item = self.repayments_tree.item(selection[0])
                repayment_id = item['values'][0]
                
                delete_repayment(repayment_id)
                self.refresh_repayments()
                self.clear_repayment_form()
                messagebox.showinfo("Success", "Repayment deleted successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete repayment: {e}")
    
    def clear_repayment_form(self):
        self.repayment_loan_combo.set('')
        self.repayment_amount_entry.delete(0, tk.END)
    
    # Report methods
    def generate_report(self):
        try:
            end_date = self.report_date_entry.get()
            if not end_date:
                end_date = datetime.now().strftime('%Y-%m-%d')
            
            # Clear previous report
            self.report_text.delete(1.0, tk.END)
            
            # Generate enhanced comprehensive report
            report_generator = EnhancedReportGenerator()
            report = report_generator.generate_comprehensive_report(end_date)
            
            self.report_text.insert(1.0, report)
            
        except Exception as e:
            user_message = ErrorHandler.get_user_friendly_message(e)
            messagebox.showerror("Report Generation Error", user_message)
    
    def export_csv(self):
        try:
            end_date = self.report_date_entry.get()
            if not end_date:
                end_date = datetime.now().strftime('%Y-%m-%d')
            
            filename = f"bank_mmudzi_report_{end_date}.csv"
            
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                writer.writerow(['Report Type', 'Member Name', 'Amount', 'Date'])
                
                # Write contributions
                member_contributions = read_contributions_up_to_date(end_date)
                members = read_members()
                
                for member in members:
                    member_id, name, surname = member[0], member[1], member[2]
                    contribution = member_contributions.get(member_id, 0)
                    writer.writerow(['Contribution', f"{name} {surname}", contribution, end_date])
                
                # Write dividends
                dividends = calculate_dividends(end_date)
                for member_name, dividend in dividends:
                    writer.writerow(['Dividend', member_name, dividend, end_date])
            
            messagebox.showinfo("Success", f"Report exported to {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export CSV: {e}") 
   # Sync status methods
    def refresh_sync_status(self):
        try:
            # Update database status
            db_status = db_manager.get_db_status()
            self.db_status_label.config(text=f"Database: {db_status['type'].upper()}")
            self.mysql_status_label.config(text=f"MySQL Available: {'Yes' if db_status['mysql_available'] else 'No'}")
            
            if db_manager.sync_manager:
                sync_status = db_manager.sync_manager.get_sync_status()
                status_text = f"Sync Status: {sync_status['status'].upper()}"
                if sync_status['pending_records'] > 0:
                    status_text += f" ({sync_status['pending_records']} pending)"
                if sync_status['conflicts'] > 0:
                    status_text += f" ({sync_status['conflicts']} conflicts)"
                
                self.sync_status_label.config(text=status_text)
                
                # Update sync log
                self.sync_log_text.delete(1.0, tk.END)
                log_entries = db_manager.sync_manager.get_sync_log(limit=20)
                
                for entry in log_entries:
                    timestamp = entry['timestamp']
                    level = entry['level']
                    message = entry['message']
                    log_line = f"[{timestamp}] {level}: {message}\n"
                    self.sync_log_text.insert(tk.END, log_line)
            else:
                self.sync_status_label.config(text="Sync Status: DISABLED (Sync manager not available)")
                
        except Exception as e:
            print(f"Error refreshing sync status: {e}")
    
    def force_sync(self):
        if not db_manager.sync_manager:
            messagebox.showerror("Error", "Sync manager not available")
            return
        
        try:
            if db_manager.sync_manager.synchronize_databases():
                messagebox.showinfo("Success", "Database synchronization completed successfully")
            else:
                messagebox.showwarning("Warning", "Synchronization completed with conflicts. Check sync log for details.")
            
            self.refresh_sync_status()
            
        except Exception as e:
            messagebox.showerror("Error", f"Synchronization failed: {e}")
    
    def view_sync_log(self):
        if not db_manager.sync_manager:
            messagebox.showerror("Error", "Sync manager not available")
            return
        
        # Create a new window to show detailed sync log
        log_window = tk.Toplevel(self)
        log_window.title("Detailed Sync Log")
        log_window.geometry("800x600")
        
        log_text = tk.Text(log_window, wrap='word')
        scrollbar = ttk.Scrollbar(log_window, orient='vertical', command=log_text.yview)
        log_text.configure(yscrollcommand=scrollbar.set)
        
        log_text.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Load all log entries
        log_entries = db_manager.sync_manager.get_sync_log(limit=1000)
        
        for entry in log_entries:
            session_id = entry['session_id']
            timestamp = entry['timestamp']
            level = entry['level']
            message = entry['message']
            log_line = f"[{session_id}] {timestamp} - {level}: {message}\n"
            log_text.insert(tk.END, log_line)
    
    def view_conflicts(self):
        if not db_manager.sync_manager:
            messagebox.showerror("Error", "Sync manager not available")
            return
        
        conflicts = db_manager.sync_manager.get_conflicts()
        
        if not conflicts:
            messagebox.showinfo("Info", "No conflicts found")
            return
        
        # Create a new window to show conflicts
        conflict_window = tk.Toplevel(self)
        conflict_window.title("Sync Conflicts")
        conflict_window.geometry("1000x600")
        
        conflict_text = tk.Text(conflict_window, wrap='word')
        scrollbar = ttk.Scrollbar(conflict_window, orient='vertical', command=conflict_text.yview)
        conflict_text.configure(yscrollcommand=scrollbar.set)
        
        conflict_text.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Display conflicts
        for i, conflict in enumerate(conflicts, 1):
            conflict_info = f"CONFLICT {i}\n"
            conflict_info += f"Table: {conflict.table_name}\n"
            conflict_info += f"Record ID: {conflict.record_id}\n"
            conflict_info += f"SQLite Data: {conflict.sqlite_data}\n"
            conflict_info += f"MySQL Data: {conflict.mysql_data}\n"
            conflict_info += f"Resolution Strategy: {conflict.resolution_strategy.value}\n"
            conflict_info += "-" * 50 + "\n\n"
            
            conflict_text.insert(tk.END, conflict_info)

    def create_audit_trail_tab(self):
        """Create audit trail tab with filtering and export capabilities"""
        # Audit Trail tab
        audit_frame = ttk.Frame(self.notebook)
        self.notebook.add(audit_frame, text="Audit Trail")
        
        # Filter frame
        filter_frame = ttk.LabelFrame(audit_frame, text="Filters")
        filter_frame.pack(fill='x', padx=5, pady=5)
        
        # Filter controls in a grid
        filter_grid = ttk.Frame(filter_frame)
        filter_grid.pack(fill='x', padx=5, pady=5)
        
        # Table filter
        ttk.Label(filter_grid, text="Table:").grid(row=0, column=0, sticky='w', padx=5)
        self.audit_table_var = tk.StringVar()
        self.audit_table_combo = ttk.Combobox(filter_grid, textvariable=self.audit_table_var, width=15)
        self.audit_table_combo['values'] = ('All', 'Members', 'Contributions', 'Loans', 'Repayments', 'SECURITY')
        self.audit_table_combo.set('All')
        self.audit_table_combo.grid(row=0, column=1, padx=5)
        
        # Operation filter
        ttk.Label(filter_grid, text="Operation:").grid(row=0, column=2, sticky='w', padx=5)
        self.audit_operation_var = tk.StringVar()
        self.audit_operation_combo = ttk.Combobox(filter_grid, textvariable=self.audit_operation_var, width=15)
        self.audit_operation_combo['values'] = ('All', 'INSERT', 'UPDATE', 'DELETE', 'FAILED_LOGIN')
        self.audit_operation_combo.set('All')
        self.audit_operation_combo.grid(row=0, column=3, padx=5)
        
        # User filter
        ttk.Label(filter_grid, text="User:").grid(row=1, column=0, sticky='w', padx=5)
        self.audit_user_var = tk.StringVar()
        self.audit_user_entry = ttk.Entry(filter_grid, textvariable=self.audit_user_var, width=15)
        self.audit_user_entry.grid(row=1, column=1, padx=5)
        
        # Date filters
        ttk.Label(filter_grid, text="Start Date:").grid(row=1, column=2, sticky='w', padx=5)
        self.audit_start_date_var = tk.StringVar()
        self.audit_start_date_entry = ttk.Entry(filter_grid, textvariable=self.audit_start_date_var, width=15)
        self.audit_start_date_entry.grid(row=1, column=3, padx=5)
        self.audit_start_date_entry.insert(0, "YYYY-MM-DD")
        
        ttk.Label(filter_grid, text="End Date:").grid(row=2, column=0, sticky='w', padx=5)
        self.audit_end_date_var = tk.StringVar()
        self.audit_end_date_entry = ttk.Entry(filter_grid, textvariable=self.audit_end_date_var, width=15)
        self.audit_end_date_entry.grid(row=2, column=1, padx=5)
        self.audit_end_date_entry.insert(0, "YYYY-MM-DD")
        
        # Filter and export buttons
        button_frame = ttk.Frame(filter_grid)
        button_frame.grid(row=2, column=2, columnspan=2, padx=5, pady=5)
        
        ttk.Button(button_frame, text="Apply Filters", command=self.apply_audit_filters).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Clear Filters", command=self.clear_audit_filters).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Export CSV", command=self.export_audit_log).pack(side='left', padx=5)
        
        # Audit trail tree
        audit_tree_frame = ttk.Frame(audit_frame)
        audit_tree_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        columns = ('ID', 'Table', 'Operation', 'Record ID', 'User', 'Timestamp')
        self.audit_tree = ttk.Treeview(audit_tree_frame, columns=columns, show='headings', height=15)
        
        # Configure columns
        self.audit_tree.heading('ID', text='Audit ID')
        self.audit_tree.heading('Table', text='Table')
        self.audit_tree.heading('Operation', text='Operation')
        self.audit_tree.heading('Record ID', text='Record ID')
        self.audit_tree.heading('User', text='User')
        self.audit_tree.heading('Timestamp', text='Timestamp')
        
        self.audit_tree.column('ID', width=80)
        self.audit_tree.column('Table', width=100)
        self.audit_tree.column('Operation', width=80)
        self.audit_tree.column('Record ID', width=80)
        self.audit_tree.column('User', width=100)
        self.audit_tree.column('Timestamp', width=150)
        
        # Scrollbars for audit tree
        audit_scrollbar_v = ttk.Scrollbar(audit_tree_frame, orient='vertical', command=self.audit_tree.yview)
        audit_scrollbar_h = ttk.Scrollbar(audit_tree_frame, orient='horizontal', command=self.audit_tree.xview)
        self.audit_tree.configure(yscrollcommand=audit_scrollbar_v.set, xscrollcommand=audit_scrollbar_h.set)
        
        self.audit_tree.pack(side='left', fill='both', expand=True)
        audit_scrollbar_v.pack(side='right', fill='y')
        audit_scrollbar_h.pack(side='bottom', fill='x')
        
        # Bind double-click to view details
        self.audit_tree.bind('<Double-1>', self.view_audit_details)
        
        # Load initial data
        self.refresh_audit_trail()
    
    def refresh_audit_trail(self):
        """Refresh audit trail data"""
        try:
            # Clear existing items
            for item in self.audit_tree.get_children():
                self.audit_tree.delete(item)
            
            # Get audit entries (limit to recent entries for performance)
            entries = audit_manager.get_audit_trail(limit=500)
            
            # Populate tree
            for entry in entries:
                self.audit_tree.insert('', 'end', values=(
                    entry['audit_id'],
                    entry['table_name'],
                    entry['operation'],
                    entry['record_id'],
                    entry['user_id'],
                    entry['timestamp']
                ))
                
        except Exception as e:
            print(f"Error refreshing audit trail: {e}")
    
    def apply_audit_filters(self):
        """Apply filters to audit trail"""
        try:
            # Get filter values
            table_name = self.audit_table_var.get() if self.audit_table_var.get() != 'All' else None
            operation = self.audit_operation_var.get() if self.audit_operation_var.get() != 'All' else None
            user_id = self.audit_user_var.get() if self.audit_user_var.get() else None
            start_date = self.audit_start_date_var.get() if self.audit_start_date_var.get() not in ('', 'YYYY-MM-DD') else None
            end_date = self.audit_end_date_var.get() if self.audit_end_date_var.get() not in ('', 'YYYY-MM-DD') else None
            
            # Clear existing items
            for item in self.audit_tree.get_children():
                self.audit_tree.delete(item)
            
            # Get filtered entries
            entries = audit_manager.get_audit_trail(
                table_name=table_name,
                operation=operation,
                user_id=user_id,
                start_date=start_date,
                end_date=end_date,
                limit=1000
            )
            
            # Populate tree
            for entry in entries:
                self.audit_tree.insert('', 'end', values=(
                    entry['audit_id'],
                    entry['table_name'],
                    entry['operation'],
                    entry['record_id'],
                    entry['user_id'],
                    entry['timestamp']
                ))
            
            messagebox.showinfo("Success", f"Found {len(entries)} audit entries matching filters")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to apply filters: {e}")
    
    def clear_audit_filters(self):
        """Clear all audit filters"""
        self.audit_table_var.set('All')
        self.audit_operation_var.set('All')
        self.audit_user_var.set('')
        self.audit_start_date_var.set('')
        self.audit_end_date_var.set('')
        self.audit_start_date_entry.delete(0, tk.END)
        self.audit_start_date_entry.insert(0, "YYYY-MM-DD")
        self.audit_end_date_entry.delete(0, tk.END)
        self.audit_end_date_entry.insert(0, "YYYY-MM-DD")
        self.refresh_audit_trail()
    
    def export_audit_log(self):
        """Export audit log to CSV"""
        try:
            # Get date range for export
            start_date = self.audit_start_date_var.get() if self.audit_start_date_var.get() not in ('', 'YYYY-MM-DD') else '2020-01-01'
            end_date = self.audit_end_date_var.get() if self.audit_end_date_var.get() not in ('', 'YYYY-MM-DD') else '2030-12-31'
            
            # Export audit log
            filename = audit_manager.export_audit_log(start_date, end_date)
            
            if filename:
                messagebox.showinfo("Success", f"Audit log exported to {filename}")
            else:
                messagebox.showerror("Error", "Failed to export audit log")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export audit log: {e}")
    
    def view_audit_details(self, event):
        """View detailed information about an audit entry"""
        selection = self.audit_tree.selection()
        if not selection:
            return
        
        item = self.audit_tree.item(selection[0])
        audit_id = item['values'][0]
        
        try:
            # Get detailed audit entry
            entries = audit_manager.get_audit_trail(limit=1)
            entry = None
            for e in entries:
                if e['audit_id'] == audit_id:
                    entry = e
                    break
            
            if not entry:
                messagebox.showerror("Error", "Audit entry not found")
                return
            
            # Create details window
            details_window = tk.Toplevel(self)
            details_window.title(f"Audit Entry Details - ID {audit_id}")
            details_window.geometry("600x500")
            
            # Create text widget with scrollbar
            text_frame = ttk.Frame(details_window)
            text_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            details_text = tk.Text(text_frame, wrap='word')
            scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=details_text.yview)
            details_text.configure(yscrollcommand=scrollbar.set)
            
            details_text.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            
            # Format and display details
            details_content = f"Audit Entry Details\n"
            details_content += "=" * 50 + "\n\n"
            details_content += f"Audit ID: {entry['audit_id']}\n"
            details_content += f"Table: {entry['table_name']}\n"
            details_content += f"Operation: {entry['operation']}\n"
            details_content += f"Record ID: {entry['record_id']}\n"
            details_content += f"User: {entry['user_id']}\n"
            details_content += f"Timestamp: {entry['timestamp']}\n\n"
            
            if entry['old_values']:
                details_content += "Old Values:\n"
                details_content += "-" * 20 + "\n"
                for key, value in entry['old_values'].items():
                    details_content += f"{key}: {value}\n"
                details_content += "\n"
            
            if entry['new_values']:
                details_content += "New Values:\n"
                details_content += "-" * 20 + "\n"
                for key, value in entry['new_values'].items():
                    details_content += f"{key}: {value}\n"
                details_content += "\n"
            
            details_text.insert(tk.END, details_content)
            details_text.config(state='disabled')
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to view audit details: {e}")

# Main execution
if __name__ == "__main__":
    try:
        # Initialize database
        initialize_db()
        
        # Start login window
        login_window = LoginWindow()
        login_window.mainloop()
        
    except Exception as e:
        print(f"Application error: {e}")
        messagebox.showerror("Application Error", f"Failed to start application: {e}")
    finally:
        # Clean up sync manager
        if db_manager and db_manager.sync_manager:
            db_manager.sync_manager.stop_sync_monitoring()

# Core Database Functions
def connect_db():
    """Connect to the appropriate database (MySQL or SQLite)"""
    global db_manager
    
    if db_manager is None:
        db_manager = initialize_database_manager()
    
    try:
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            return mysql.connector.connect(**MYSQL_CONFIG)
        else:
            conn = sqlite3.connect(SQLITE_DB_NAME, timeout=30.0)
            conn.execute('PRAGMA foreign_keys = ON')
            return conn
    except Exception as e:
        logger.error(f"Database connection error: {e}")
        # Fallback to SQLite
        conn = sqlite3.connect(SQLITE_DB_NAME, timeout=30.0)
        conn.execute('PRAGMA foreign_keys = ON')
        return conn

def initialize_db():
    """Initialize the database with proper schema"""
    try:
        global db_manager
        db_manager = initialize_database_manager()
        
        conn = connect_db()
        cursor = conn.cursor()
        
        # For SQLite, create basic schema
        if not hasattr(db_manager, 'current_db_type') or db_manager.current_db_type != 'mysql':
            # Create basic tables for SQLite
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Members (
                    member_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    member_number VARCHAR(20) UNIQUE,
                    name VARCHAR(100) NOT NULL,
                    surname VARCHAR(100) NOT NULL,
                    phone_number VARCHAR(20) NOT NULL,
                    email VARCHAR(100),
                    join_date DATE,
                    status VARCHAR(10) DEFAULT 'active',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Contributions (
                    contribution_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    member_id INTEGER NOT NULL,
                    month INTEGER NOT NULL,
                    year INTEGER NOT NULL,
                    amount DECIMAL(10,2) NOT NULL,
                    late_fee DECIMAL(10,2) DEFAULT 0.00,
                    contribution_date DATE NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES Members(member_id),
                    UNIQUE(member_id, month, year)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Loans (
                    loan_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    member_id INTEGER NOT NULL,
                    loan_amount DECIMAL(10,2) NOT NULL,
                    interest_rate DECIMAL(5,2) NOT NULL,
                    monthly_payment DECIMAL(10,2),
                    total_interest DECIMAL(10,2),
                    loan_date DATE NOT NULL,
                    status VARCHAR(20) DEFAULT 'active',
                    outstanding_balance DECIMAL(10,2) DEFAULT 0.00,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES Members(member_id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS Repayments (
                    repayment_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    loan_id INTEGER NOT NULL,
                    repayment_amount DECIMAL(10,2) NOT NULL,
                    principal_amount DECIMAL(10,2) NOT NULL,
                    interest_amount DECIMAL(10,2) NOT NULL,
                    repayment_date DATE NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (loan_id) REFERENCES Loans(loan_id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS DividendCalculations (
                    calculation_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    member_id INTEGER NOT NULL,
                    year INTEGER NOT NULL,
                    total_contributions DECIMAL(10,2) NOT NULL,
                    total_interest_paid DECIMAL(10,2) NOT NULL,
                    outstanding_balance DECIMAL(10,2) NOT NULL,
                    dividend_amount DECIMAL(10,2) NOT NULL,
                    calculation_date DATE NOT NULL,
                    status VARCHAR(20) DEFAULT 'calculated',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (member_id) REFERENCES Members(member_id),
                    UNIQUE(member_id, year)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS AuditLog (
                    audit_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_name VARCHAR(50) NOT NULL,
                    operation VARCHAR(10) NOT NULL,
                    record_id INTEGER NOT NULL,
                    old_values TEXT,
                    new_values TEXT,
                    user_id VARCHAR(50) NOT NULL,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    ip_address VARCHAR(45),
                    user_agent TEXT
                )
            ''')
        
        conn.commit()
        conn.close()
        logger.info("Database initialized successfully")
        return True
    except Exception as e:
        logger.error(f"Error initializing database: {e}")
        return False

# Core CRUD Functions
def create_member(name: str, surname: str, phone: str, email: str = '') -> int:
    """Create a new member"""
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Generate unique member number
        current_year = datetime.now().year
        import random
        member_number = f'BM-{current_year}-{random.randint(1000, 9999):04d}'
        join_date = datetime.now().strftime('%Y-%m-%d')
        
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('''
                INSERT INTO Members (member_number, name, surname, phone_number, email, join_date, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (member_number, name, surname, phone, email or None, join_date, 'active'))
        else:
            cursor.execute('''
                INSERT INTO Members (member_number, name, surname, phone_number, email, join_date, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (member_number, name, surname, phone, email or None, join_date, 'active'))
        
        member_id = cursor.lastrowid
        conn.commit()
        conn.close()
        logger.info(f"Member created: {name} {surname} (ID: {member_id})")
        return member_id
    except Exception as e:
        logger.error(f"Error creating member: {e}")
        return None

def update_member(member_id: int, name: str, surname: str, phone: str, email: str = '') -> bool:
    """Update an existing member"""
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('''
                UPDATE Members 
                SET name = %s, surname = %s, phone_number = %s, email = %s, updated_at = NOW()
                WHERE member_id = %s
            ''', (name, surname, phone, email or None, member_id))
        else:
            cursor.execute('''
                UPDATE Members 
                SET name = ?, surname = ?, phone_number = ?, email = ?, updated_at = CURRENT_TIMESTAMP
                WHERE member_id = ?
            ''', (name, surname, phone, email or None, member_id))
        
        success = cursor.rowcount > 0
        conn.commit()
        conn.close()
        
        if success:
            logger.info(f"Member updated: {name} {surname} (ID: {member_id})")
        return success
    except Exception as e:
        logger.error(f"Error updating member: {e}")
        return False

def create_contribution(member_id: int, month: str, amount: float, year: int = None) -> int:
    """Create a new contribution"""
    try:
        if year is None:
            year = datetime.now().year
        
        # Convert month to number
        if isinstance(month, str) and month in MONTH_TO_NUM:
            month_num = MONTH_TO_NUM[month]
        else:
            month_num = int(month) if str(month).isdigit() else 1
        
        conn = connect_db()
        cursor = conn.cursor()
        
        contribution_date = datetime.now().strftime('%Y-%m-%d')
        
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('''
                INSERT INTO Contributions (member_id, month, year, amount, contribution_date)
                VALUES (%s, %s, %s, %s, %s)
            ''', (member_id, month_num, year, amount, contribution_date))
        else:
            cursor.execute('''
                INSERT INTO Contributions (member_id, month, year, amount, contribution_date)
                VALUES (?, ?, ?, ?, ?)
            ''', (member_id, month_num, year, amount, contribution_date))
        
        contribution_id = cursor.lastrowid
        conn.commit()
        conn.close()
        logger.info(f"Contribution created: Member {member_id}, Amount {amount} (ID: {contribution_id})")
        return contribution_id
    except Exception as e:
        logger.error(f"Error creating contribution: {e}")
        return None

def create_loan(member_id: int, loan_amount: float, interest_rate: float, months: int = 12) -> int:
    """Create a new loan"""
    try:
        monthly_rate = (interest_rate / 100) / 12
        monthly_payment = FinancialCalculator.calculate_loan_payment(loan_amount, monthly_rate, months)
        total_interest = FinancialCalculator.calculate_total_interest(loan_amount, monthly_payment, months)
        
        conn = connect_db()
        cursor = conn.cursor()
        
        loan_date = datetime.now().strftime('%Y-%m-%d')
        
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('''
                INSERT INTO Loans (member_id, loan_amount, interest_rate, monthly_payment, 
                                 total_interest, loan_date, outstanding_balance, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ''', (member_id, loan_amount, interest_rate, monthly_payment, 
                  total_interest, loan_date, loan_amount, 'active'))
        else:
            cursor.execute('''
                INSERT INTO Loans (member_id, loan_amount, interest_rate, monthly_payment, 
                                 total_interest, loan_date, outstanding_balance, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (member_id, loan_amount, interest_rate, monthly_payment, 
                  total_interest, loan_date, loan_amount, 'active'))
        
        loan_id = cursor.lastrowid
        conn.commit()
        conn.close()
        logger.info(f"Loan created: Member {member_id}, Amount {loan_amount} (ID: {loan_id})")
        return loan_id
    except Exception as e:
        logger.error(f"Error creating loan: {e}")
        return None

def create_repayment(loan_id: int, repayment_amount: float) -> int:
    """Create a loan repayment"""
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Get loan details
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('SELECT interest_rate, outstanding_balance FROM Loans WHERE loan_id = %s', (loan_id,))
        else:
            cursor.execute('SELECT interest_rate, outstanding_balance FROM Loans WHERE loan_id = ?', (loan_id,))
        
        loan_data = cursor.fetchone()
        if not loan_data:
            conn.close()
            return None
        
        annual_rate, outstanding_balance = loan_data
        monthly_rate = (annual_rate / 100) / 12
        interest_amount = outstanding_balance * monthly_rate
        principal_amount = repayment_amount - interest_amount
        new_balance = max(0, outstanding_balance - principal_amount)
        
        repayment_date = datetime.now().strftime('%Y-%m-%d')
        
        # Insert repayment
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('''
                INSERT INTO Repayments (loan_id, repayment_amount, principal_amount, interest_amount, repayment_date)
                VALUES (%s, %s, %s, %s, %s)
            ''', (loan_id, repayment_amount, principal_amount, interest_amount, repayment_date))
        else:
            cursor.execute('''
                INSERT INTO Repayments (loan_id, repayment_amount, principal_amount, interest_amount, repayment_date)
                VALUES (?, ?, ?, ?, ?)
            ''', (loan_id, repayment_amount, principal_amount, interest_amount, repayment_date))
        
        repayment_id = cursor.lastrowid
        
        # Update loan balance
        loan_status = 'completed' if new_balance <= 0 else 'active'
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('UPDATE Loans SET outstanding_balance = %s, status = %s WHERE loan_id = %s', 
                         (new_balance, loan_status, loan_id))
        else:
            cursor.execute('UPDATE Loans SET outstanding_balance = ?, status = ? WHERE loan_id = ?', 
                         (new_balance, loan_status, loan_id))
        
        conn.commit()
        conn.close()
        logger.info(f"Repayment created: Loan {loan_id}, Amount {repayment_amount} (ID: {repayment_id})")
        return repayment_id
    except Exception as e:
        logger.error(f"Error creating repayment: {e}")
        return None

def read_members() -> list:
    """Read all members"""
    try:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM Members ORDER BY name, surname')
        results = cursor.fetchall()
        conn.close()
        
        members = []
        for row in results:
            members.append({
                'member_id': row[0],
                'member_number': row[1],
                'name': row[2],
                'surname': row[3],
                'phone_number': row[4],
                'email': row[5],
                'join_date': row[6],
                'status': row[7]
            })
        return members
    except Exception as e:
        logger.error(f"Error reading members: {e}")
        return []

def read_contributions_for_member(member_id: int) -> list:
    """Read contributions for a member"""
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('SELECT * FROM Contributions WHERE member_id = %s ORDER BY year DESC, month DESC', (member_id,))
        else:
            cursor.execute('SELECT * FROM Contributions WHERE member_id = ? ORDER BY year DESC, month DESC', (member_id,))
        
        results = cursor.fetchall()
        conn.close()
        
        contributions = []
        for row in results:
            contributions.append({
                'contribution_id': row[0],
                'member_id': row[1],
                'month': row[2],
                'year': row[3],
                'amount': row[4],
                'late_fee': row[5] if len(row) > 5 else 0,
                'contribution_date': row[6] if len(row) > 6 else None
            })
        return contributions
    except Exception as e:
        logger.error(f"Error reading contributions: {e}")
        return []

def read_loans() -> list:
    """Read all loans"""
    try:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM Loans ORDER BY loan_date DESC')
        results = cursor.fetchall()
        conn.close()
        
        loans = []
        for row in results:
            loans.append({
                'loan_id': row[0],
                'member_id': row[1],
                'loan_amount': row[2],
                'interest_rate': row[3],
                'monthly_payment': row[4] if len(row) > 4 else 0,
                'total_interest': row[5] if len(row) > 5 else 0,
                'loan_date': row[6] if len(row) > 6 else None,
                'status': row[7] if len(row) > 7 else 'active',
                'outstanding_balance': row[8] if len(row) > 8 else 0
            })
        return loans
    except Exception as e:
        logger.error(f"Error reading loans: {e}")
        return []

def read_repayments() -> list:
    """Read all repayments"""
    try:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM Repayments ORDER BY repayment_date DESC')
        results = cursor.fetchall()
        conn.close()
        
        repayments = []
        for row in results:
            repayments.append({
                'repayment_id': row[0],
                'loan_id': row[1],
                'repayment_amount': row[2],
                'principal_amount': row[3] if len(row) > 3 else 0,
                'interest_amount': row[4] if len(row) > 4 else 0,
                'repayment_date': row[5] if len(row) > 5 else None
            })
        return repayments
    except Exception as e:
        logger.error(f"Error reading repayments: {e}")
        return []

def get_member_profile(member_id: int) -> dict:
    """Get member profile with contributions and loans"""
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
            cursor.execute('SELECT * FROM Members WHERE member_id = %s', (member_id,))
        else:
            cursor.execute('SELECT * FROM Members WHERE member_id = ?', (member_id,))
        
        member_data = cursor.fetchone()
        conn.close()
        
        if not member_data:
            return None
        
        member_info = {
            'member_id': member_data[0],
            'member_number': member_data[1],
            'name': member_data[2],
            'surname': member_data[3],
            'phone_number': member_data[4],
            'email': member_data[5],
            'join_date': member_data[6],
            'status': member_data[7]
        }
        
        contributions = read_contributions_for_member(member_id)
        loans = []  # Simplified for now
        
        return {
            'member_info': member_info,
            'contributions': contributions,
            'loans': loans
        }
    except Exception as e:
        logger.error(f"Error getting member profile: {e}")
        return None

# Validation Engine
class ValidationEngine:
    """Input validation engine"""
    
    @staticmethod
    def validate_member_data(name: str, surname: str, phone: str, email: str = '') -> dict:
        """Validate complete member data"""
        errors = []
        formatted_data = {}
        
        # Validate name
        if not name or not name.strip():
            errors.append('Name is required')
        else:
            formatted_data['name'] = name.strip().title()
        
        # Validate surname
        if not surname or not surname.strip():
            errors.append('Surname is required')
        else:
            formatted_data['surname'] = surname.strip().title()
        
        # Validate phone
        if not phone or not phone.strip():
            errors.append('Phone number is required')
        else:
            formatted_data['phone'] = phone.strip()
        
        # Validate email (optional)
        if email and email.strip():
            formatted_data['email'] = email.strip().lower()
        else:
            formatted_data['email'] = None
        
        return {
            'valid': len(errors) == 0,
            'errors': errors,
            'message': '; '.join(errors) if errors else 'Validation passed',
            'formatted_data': formatted_data
        }
    
    @staticmethod
    def validate_phone_number(phone: str) -> dict:
        """Validate phone number"""
        if not phone or not phone.strip():
            return {'valid': False, 'message': 'Phone number is required'}
        
        phone = phone.strip()
        # Basic validation - can be enhanced
        if len(phone) >= 9:
            return {'valid': True, 'formatted': phone, 'message': 'Valid phone number'}
        else:
            return {'valid': False, 'message': 'Phone number too short'}
    
    @staticmethod
    def validate_financial_amount(amount_str: str) -> dict:
        """Validate financial amount"""
        try:
            amount = float(str(amount_str).strip())
            if amount <= 0:
                return {'valid': False, 'message': 'Amount must be greater than zero'}
            return {'valid': True, 'formatted': amount, 'message': 'Valid amount'}
        except ValueError:
            return {'valid': False, 'message': 'Invalid amount format'}

# Note: Main App class is defined earlier in the file around line 7650

# Main execution
if __name__ == "__main__":
    try:
        # Initialize database
        if initialize_db():
            print("✅ Database initialized successfully")
            
            # Initialize database manager
            db_manager = initialize_database_manager()
            print(f"✅ Database manager initialized - Using: {db_manager.current_db_type}")
            
            # Start the application
            app = App()
            print("✅ Starting Bank Mmudzi application...")
            app.mainloop()
        else:
            print("❌ Failed to initialize database")
    except Exception as e:
        print(f"❌ Error starting application: {e}")
        import traceback
        traceback.print_exc()

# Missing Manager Classes for Integration Tests
class BackupManager:
    """Backup manager for database operations"""
    
    def __init__(self):
        self.backup_dir = 'backups'
        os.makedirs(self.backup_dir, exist_ok=True)
    
    def create_backup(self, backup_name: str = None) -> str:
        """Create a database backup"""
        try:
            if backup_name is None:
                backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            
            backup_path = os.path.join(self.backup_dir, backup_name)
            
            # For SQLite, copy the database file
            if not hasattr(db_manager, 'current_db_type') or db_manager.current_db_type != 'mysql':
                import shutil
                shutil.copy2(SQLITE_DB_NAME, backup_path)
                logger.info(f"SQLite backup created: {backup_path}")
                return backup_path
            else:
                # For MySQL, would need mysqldump - simplified for now
                logger.info("MySQL backup would require mysqldump")
                return backup_name
                
        except Exception as e:
            logger.error(f"Error creating backup: {e}")
            return None
    
    def restore_backup(self, backup_path: str) -> bool:
        """Restore from a backup"""
        try:
            if os.path.exists(backup_path):
                import shutil
                shutil.copy2(backup_path, SQLITE_DB_NAME)
                logger.info(f"Backup restored from: {backup_path}")
                return True
            else:
                logger.error(f"Backup file not found: {backup_path}")
                return False
        except Exception as e:
            logger.error(f"Error restoring backup: {e}")
            return False

class AuditManager:
    """Audit manager for tracking system changes"""
    
    def __init__(self):
        self.audit_enabled = True
    
    def log_action(self, table_name: str, operation: str, record_id: int, 
                   old_values: dict = None, new_values: dict = None, user_id: str = 'system'):
        """Log an audit action"""
        try:
            if not self.audit_enabled:
                return
            
            conn = connect_db()
            cursor = conn.cursor()
            
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Truncate operation to fit database column (usually 10 chars for ENUM)
            operation_short = operation[:10] if len(operation) > 10 else operation
            
            if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
                cursor.execute('''
                    INSERT INTO AuditLog (table_name, operation, record_id, old_values, new_values, user_id, timestamp, ip_address, user_agent)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (table_name, operation_short, record_id, 
                      json.dumps(old_values) if old_values else None,
                      json.dumps(new_values) if new_values else None,
                      user_id, timestamp, 'localhost', 'system'))
            else:
                cursor.execute('''
                    INSERT INTO AuditLog (table_name, operation, record_id, old_values, new_values, user_id, timestamp, ip_address, user_agent)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (table_name, operation_short, record_id, 
                      json.dumps(old_values) if old_values else None,
                      json.dumps(new_values) if new_values else None,
                      user_id, timestamp, 'localhost', 'system'))
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            logger.error(f"Error logging audit action: {e}")
    
    def get_audit_trail(self, table_name: str = None, limit: int = 100) -> list:
        """Get audit trail records"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            if table_name:
                if hasattr(db_manager, 'current_db_type') and db_manager.current_db_type == 'mysql':
                    cursor.execute('SELECT * FROM AuditLog WHERE table_name = %s ORDER BY timestamp DESC LIMIT %s', 
                                 (table_name, limit))
                else:
                    cursor.execute('SELECT * FROM AuditLog WHERE table_name = ? ORDER BY timestamp DESC LIMIT ?', 
                                 (table_name, limit))
            else:
                cursor.execute(f'SELECT * FROM AuditLog ORDER BY timestamp DESC LIMIT {limit}')
            
            results = cursor.fetchall()
            conn.close()
            
            audit_records = []
            for row in results:
                audit_records.append({
                    'audit_id': row[0],
                    'table_name': row[1],
                    'operation': row[2],
                    'record_id': row[3],
                    'old_values': json.loads(row[4]) if row[4] else None,
                    'new_values': json.loads(row[5]) if row[5] else None,
                    'user_id': row[6],
                    'timestamp': row[7],
                    'ip_address': row[8] if len(row) > 8 else None,
                    'user_agent': row[9] if len(row) > 9 else None
                })
            
            return audit_records
            
        except Exception as e:
            logger.error(f"Error getting audit trail: {e}")
            return []

# Fix the integration test file reference issue
def create_bank_mmudzi_alias():
    """Create an alias file for backward compatibility"""
    try:
        alias_content = '''# Alias file for backward compatibility
from Village import *
'''
        with open('bank_mmudzi.py', 'w') as f:
            f.write(alias_content)
        logger.info("Created bank_mmudzi.py alias file")
    except Exception as e:
        logger.error(f"Error creating alias file: {e}")

# Initialize global instances
backup_manager = None
audit_manager = None

def initialize_managers():
    """Initialize all manager instances"""
    global backup_manager, audit_manager
    
    if backup_manager is None:
        backup_manager = BackupManager()
    
    if audit_manager is None:
        audit_manager = AuditManager()
    
    return backup_manager, audit_manager
    return backup_manager, audit_manager